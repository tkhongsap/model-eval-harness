# Process into User requirement file after append_csv
import io
import os
import re
import pandas as pd
import pandasql as ps
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
import requests
from src.log import Logger
from src.sharepoint import (
    list_folders_in_folder, 
    list_files_in_folder, 
    get_item_download_url_by_path,
    upload_file_to_sharepoint
)
from openpyxl import Workbook
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
from google import genai
from google.genai import types

logger = Logger(__name__, env=os.environ.get("LOG_ENVIRONMENT", "prod"))

def upload_master_files(data_list: list[dict], access_token, **args):
    """Appends a list of data dicts to XLSX master files based on input folder structure."""
    logger.info("Update Master file")
    PROCESSING_PREFIX = args.get("PROCESSING_PREFIX")
    OUTPUT_EXCEL_PREFIX = args.get("OUTPUT_EXCEL_PREFIX")

    VERINT_SITE_BASE_ROOT = args.get("VERINT_SITE_BASE_ROOT")
    PRODUCT = args.get("OUTPUT_PRODUCT")

    if not data_list:
        logger.info("No data to append to Master Excel.")
        return
    
    # Group by output folder path
    grouped_data = {}
    for item in data_list:
        file_uri = item.get('input_folder') or item.get('file_uri') or ''

        # Extract the date folder from URI (e.g., processing/202511)
        match = re.search(fr'{re.escape(PROCESSING_PREFIX)}(\d{{6}})', file_uri)

        if match:
            date_folder = match.group(1)  # e.g. 202511
            folder_path = f"{date_folder}"
        else:
            folder_path = f"unknown_date"

        grouped_data.setdefault(folder_path, []).append(item)

    # Header grouping definitions
    group_headers = [
        ("Customer_Information", ["call_id", "phone_number", "call_month", "call_date", "call_time", "cost(thb)", "call_duration(min)", "file_name", "user"]),
        ("Call_Summary", ["call_result"]),
        ("Reason", ["main", "secondary", "third"]),
        ("", ["keyword(main)", "keyword(secondary)", "keyword(third)", "call_event_detection", "AI_recommendation"]),
        ("", ["input_folder", "output_folder", "run_date", "status", "message"])
    ]

    # Flatten all expected column names
    all_columns = [col for _, cols in group_headers for col in cols]

    # print(f"grouped_data: {grouped_data}")
    master_df = pd.DataFrame()
    for folder_path, records in grouped_data.items(): # yyyymm, [records]
        df = pd.DataFrame(records)

        # ✅ Ensure consistent column order
        for col in all_columns:
            if col not in df.columns:
                df[col] = ""  # fill missing columns

        df = df[all_columns]  # reorder columns

        master_path = f"{VERINT_SITE_BASE_ROOT}/{OUTPUT_EXCEL_PREFIX}{PRODUCT}/{folder_path}/master.xlsx"
        try:
            # Try downloading existing XLSX
            master_download_url = get_item_download_url_by_path("verint", master_path)
            master_file = requests.get(master_download_url, headers={"Authorization": f"Bearer {access_token}"})
            master_file.raise_for_status()
            master_file_bytes = master_file.content

            with io.BytesIO(master_file_bytes) as existing_file:
                existing_wb = load_workbook(existing_file)
                ws = existing_wb.active
                start_row = ws.max_row + 1
                for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=start_row):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
        except Exception:
            # Create a new workbook
            existing_wb = Workbook()
            ws = existing_wb.active

            # First row (group headers)
            col_idx = 1
            for group_name, cols in group_headers:
                start_col = col_idx
                col_idx += len(cols)
                end_col = col_idx - 1
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws.cell(row=1, column=start_col, value=group_name)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = Font(bold=True)

            # Second row (actual headers)
            col_idx = 1
            for _, cols in group_headers:
                for col_name in cols:
                    cell = ws.cell(row=2, column=col_idx, value=col_name)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.font = Font(bold=True)
                    col_idx += 1

            # Data rows start from row 3
            for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=3):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        # Auto-fit column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)

        # ✅ MOVED INSIDE LOOP - Save and upload for EACH folder
        output_stream = io.BytesIO()
        existing_wb.save(output_stream)
        output_stream.seek(0)

        upload_file_to_sharepoint("verint", output_stream, master_path)

        output_ws = existing_wb.active

        # Convert existing sheet into DataFrame (skip header rows)
        output_data = output_ws.values
        next(output_data)  # skip group headers row 1
        cols = next(output_data)  # get column headers row 2
        output_df = pd.DataFrame(output_data, columns=cols).dropna(how="all")
        master_df = pd.concat([master_df, output_df], ignore_index=True)
    return master_df

def upload_daily_files(data_list: list[dict], access_token, **args):
    """Appends a list of data dicts to XLSX daily files based on input folder structure."""
    logger.info("Update Daily file")
    PROCESSING_PREFIX = args.get("PROCESSING_PREFIX")
    OUTPUT_EXCEL_PREFIX = args.get("OUTPUT_EXCEL_PREFIX")
    OUTPUT_EXCEL_DAILY_PREFIX_FILENAME = args.get("OUTPUT_EXCEL_DAILY_PREFIX_FILENAME")

    VERINT_SITE_BASE_ROOT = args.get("VERINT_SITE_BASE_ROOT")
    PRODUCT = args.get("OUTPUT_PRODUCT")

    if not data_list:
        logger.info("No data to append to Daily Excel.")
        return

    # Group by output folder path
    grouped_data = {}
    for item in data_list:
        file_uri = item.get('input_folder') or item.get('file_uri') or ''

        # Extract the date folder from URI (e.g., processing/202511/20251107)
        match = re.search(fr'{re.escape(PROCESSING_PREFIX)}(\d{{6}}/\d{{8}})', file_uri)

        if match:
            date_folder = match.group(1)  # e.g. 202511/20251107
            folder_path = f"{date_folder}"
        else:
            folder_path = f"unknown_date"

        grouped_data.setdefault(folder_path, []).append(item)

    # Header grouping definitions
    group_headers = [
        ("Customer_Information", ["call_id", "phone_number", "call_month", "call_date", "call_time", "cost(thb)", "call_duration(min)", "file_name", "user"]),
        ("Call_Summary", ["call_result"]),
        ("Reason", ["main", "secondary", "third"]),
        ("", ["keyword(main)", "keyword(secondary)", "keyword(third)", "call_event_detection", "AI_recommendation"]),
        ("", ["input_folder", "output_folder", "run_date", "status", "message"])
    ]

    # Flatten all expected column names
    all_columns = [col for _, cols in group_headers for col in cols]

    for folder_path, records in grouped_data.items():
        df_new = pd.DataFrame(records)

        # ✅ Ensure consistent column order
        for col in all_columns:
            if col not in df_new.columns:
                df_new[col] = ""  # fill missing columns

        df_new = df_new[all_columns]  # reorder columns

        daily_path = f"{VERINT_SITE_BASE_ROOT}/{OUTPUT_EXCEL_PREFIX}{PRODUCT}/{folder_path}/{OUTPUT_EXCEL_DAILY_PREFIX_FILENAME}_{folder_path.split('/')[-1]}.xlsx"
        
        try:
            # Load existing Excel file
            download_url = get_item_download_url_by_path("verint", daily_path)
            resp = requests.get(download_url, headers={"Authorization": f"Bearer {access_token}"})
            resp.raise_for_status()
            existing_wb = load_workbook(io.BytesIO(resp.content))
            ws = existing_wb.active
            # print(1)
            # Convert existing sheet into DataFrame (skip header rows)
            data = ws.values
            next(data)  # skip group headers row 1
            cols = next(data)  # get column headers row 2
            df_existing = pd.DataFrame(data, columns=cols).dropna(how="all")
            # print(2)
            # Combine existing and new data
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)

            # Normalize run_date
            df_combined["run_date"] = pd.to_datetime(df_combined["run_date"], errors="coerce", utc=True).dt.tz_localize(None)

            # Deduplication: Keep latest record per call_id, product
            df_combined = (
                df_combined.drop_duplicates(subset=["call_id","phone_number","call_date"], keep="last")
            )
            # print(3)
            df_final = df_combined[all_columns]

            # Clear existing data rows (keep headers)
            ws.delete_rows(3, ws.max_row - 2)
            # print(4)
            # Append deduplicated data starting from row 3
            for r_idx, row in enumerate(df_final.itertuples(index=False, name=None), start=3):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            # print(5)
            logger.info(f"✅ Appended {len(df_new)} new records, {len(df_final)} total after deduplication, {daily_path}")
            
        except Exception as e:
            logger.warning(f"File not found or error reading: {e}. Creating new file.")
            
            # Create a new workbook
            existing_wb = Workbook()
            ws = existing_wb.active

            # First row (group headers)
            col_idx = 1
            for group_name, cols in group_headers:
                start_col = col_idx
                col_idx += len(cols)
                end_col = col_idx - 1
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws.cell(row=1, column=start_col, value=group_name)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = Font(bold=True)

            # Second row (actual headers)
            col_idx = 1
            for _, cols in group_headers:
                for col_name in cols:
                    cell = ws.cell(row=2, column=col_idx, value=col_name)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.font = Font(bold=True)
                    col_idx += 1

            # Data rows start from row 3
            for r_idx, row in enumerate(df_new.itertuples(index=False, name=None), start=3):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        # Auto-fit column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)

        # Save and upload
        output_stream = io.BytesIO()
        existing_wb.save(output_stream)
        output_stream.seek(0)

        upload_file_to_sharepoint("verint", output_stream, daily_path)

def upload_daily_files_with_additional(data_list: list[dict], access_token, **args):
    """Appends a list of data dicts to XLSX daily files based on input folder structure."""
    logger.info("Update Daily file")
    PROCESSING_PREFIX = args.get("PROCESSING_PREFIX")
    OUTPUT_EXCEL_PREFIX = args.get("OUTPUT_EXCEL_PREFIX")
    OUTPUT_EXCEL_DAILY_PREFIX_FILENAME = args.get("OUTPUT_EXCEL_DAILY_PREFIX_FILENAME")

    PLAYGROUND_SITE_BASE_ROOT = args.get("PLAYGROUND_SITE_BASE_ROOT")
    PLAYGROUND_NETWORK_PATH = args.get("PLAYGROUND_NETWORK_PATH")
    PRODUCT = args.get("OUTPUT_PRODUCT")

    if not data_list:
        logger.info("No data to append to Daily Excel.")
        return

    # Group by output folder path
    grouped_data = {}

    for item in data_list:
        # Flatten additional_fields into main dict
        additional_fields = item.pop('additional_fields', {
            'issue_type': None,
            'sub_reason': None,
            'problem_statement': None,
            'churn_probability': None,
            'area_tag_province': None,
            'area_tag_district': None,
            'area_tag_sub_district': None,
            'area_tag_landmark': None,
            'product_type': None,
        })
        if additional_fields and isinstance(additional_fields, dict):
            item.update(additional_fields)

        file_uri = item.get('input_folder') or item.get('file_uri') or ''

        # Extract the date folder from URI (e.g., processing/202511/20251107)
        match = re.search(fr'{re.escape(PROCESSING_PREFIX)}(\d{{6}}/\d{{8}})', file_uri)

        if match:
            date_folder = match.group(1)  # e.g. 202511/20251107
            folder_path = f"{date_folder}"
        else:
            folder_path = f"unknown_date"

        grouped_data.setdefault(folder_path, []).append(item)

    # Header grouping definitions
    group_headers = [
        ("Customer_Information", ["call_id", "phone_number", "call_month", "call_date", "call_time", "cost(thb)", "call_duration(min)", "file_name", "user"]),
        ("Call_Summary", ["call_result"]),
        ("Reason", ["main", "secondary", "third"]),
        ("", ["keyword(main)", "keyword(secondary)", "keyword(third)", "call_event_detection", "AI_recommendation"]),
        ("", ["input_folder", "output_folder", "run_date", "status", "message"]),
        ("Additional", ['issue_type', 'sub_reason', 'problem_statement', 'churn_probability', 'area_tag_province', 'area_tag_district', 'area_tag_sub_district', 'area_tag_landmark', 'product_type'])
    ]

    # Flatten all expected column names
    all_columns = [col for _, cols in group_headers for col in cols]

    for folder_path, records in grouped_data.items():
        df_new = pd.DataFrame(records)

        # ✅ Ensure consistent column order
        for col in all_columns:
            if col not in df_new.columns:
                df_new[col] = ""  # fill missing columns

        df_new = df_new[all_columns]  # reorder columns

        daily_path = f"{PLAYGROUND_SITE_BASE_ROOT}/{PLAYGROUND_NETWORK_PATH}/{folder_path.split('/')[0]}/{OUTPUT_EXCEL_DAILY_PREFIX_FILENAME}_{folder_path.split('/')[-1]}.xlsx"
        # daily_path = f"test_sentiment_batch_mnp/output/MNP Retention/{folder_path}/Prediction_Result_with_additional_{folder_path.split('/')[-1]}.xlsx"

        try:
            # Load existing Excel file
            download_url = get_item_download_url_by_path("playground", daily_path)
            resp = requests.get(download_url, headers={"Authorization": f"Bearer {access_token}"})
            resp.raise_for_status()
            existing_wb = load_workbook(io.BytesIO(resp.content))
            ws = existing_wb.active
            # print(1)
            # Convert existing sheet into DataFrame (skip header rows)
            data = ws.values
            next(data)  # skip group headers row 1
            cols = next(data)  # get column headers row 2
            df_existing = pd.DataFrame(data, columns=cols).dropna(how="all")
            # print(2)
            # Combine existing and new data
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)

            # Normalize run_date
            df_combined["run_date"] = pd.to_datetime(df_combined["run_date"], errors="coerce", utc=True).dt.tz_localize(None)

            # Deduplication: Keep latest record per call_id, product
            df_combined = (
                df_combined.drop_duplicates(subset=["call_id","phone_number","call_date"], keep="last")
            )
            # print(3)
            df_final = df_combined[all_columns]

            # Clear existing data rows (keep headers)
            ws.delete_rows(3, ws.max_row - 2)
            # print(4)
            # Append deduplicated data starting from row 3
            for r_idx, row in enumerate(df_final.itertuples(index=False, name=None), start=3):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            # print(5)
            logger.info(f"✅ Appended {len(df_new)} new records, {len(df_final)} total after deduplication, {daily_path}")
            
        except Exception as e:
            logger.warning(f"File not found or error reading: {e}. Creating new file.")
            
            # Create a new workbook
            existing_wb = Workbook()
            ws = existing_wb.active

            # First row (group headers)
            col_idx = 1
            for group_name, cols in group_headers:
                start_col = col_idx
                col_idx += len(cols)
                end_col = col_idx - 1
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws.cell(row=1, column=start_col, value=group_name)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = Font(bold=True)

            # Second row (actual headers)
            col_idx = 1
            for _, cols in group_headers:
                for col_name in cols:
                    cell = ws.cell(row=2, column=col_idx, value=col_name)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.font = Font(bold=True)
                    col_idx += 1

            # Data rows start from row 3
            for r_idx, row in enumerate(df_new.itertuples(index=False, name=None), start=3):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        # Auto-fit column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)

        # Save and upload
        output_stream = io.BytesIO()
        existing_wb.save(output_stream)
        output_stream.seek(0)

        upload_file_to_sharepoint("playground", output_stream, daily_path)

def create_daily_bar_chart_daily(df, grouped_series, **args):
    logger.info("Update Daily bar chart")
    PROCESSING_PREFIX = args.get("PROCESSING_PREFIX")
    OUTPUT_EXCEL_PREFIX = args.get("OUTPUT_EXCEL_PREFIX")

    VERINT_SITE_BASE_ROOT = args.get("VERINT_SITE_BASE_ROOT")
    PRODUCT = args.get("OUTPUT_PRODUCT")

    reason_colors = {
        'network': '#A6CEE3',
        'promotion related': "#1F78B4",
        'device promotion related': '#B2DF8A',
        'save cost': '#33A02C',
        'contract end': '#FB9A99',
        'sale upsell problem': '#E31A1C',
        'dissatisfied service': '#FDBF6F',
        'post to pre': "#CAB2D6",
        'customer reason': '#6A3D9A',
        'true point, dtac reward': '#FFFF99',
        'down sell not success': '#B15928',
        'other': "#FF7F00",
    }

    try:
        # ========= Setup Thai Font Support (Set size once globally) =========
        font_path = 'fonts/Sarabun-Regular.ttf'
        matplotlib.font_manager.fontManager.addfont(font_path)
        matplotlib.rc('font', family='Sarabun', size=16) 
        
        # ========= Rest of your code (Data Prep) =========
        # print(f"len df : {len(df)}")
        # print(f"columns df : {df.columns}")

        # Normalize date
        df["run_date"] = pd.to_datetime(df["run_date"], errors="coerce").dt.tz_localize(None)

        # Deduplication: Keep latest record per call_id and filter successful status
        df = (
            df.sort_values(by=["run_date"], ascending=False)
                .drop_duplicates(subset=["call_id","phone_number","call_date"], keep="first")
                .query("status == 'success'")
        ).copy()
        
        df['call_month'] = df['call_month'].astype(str)
        df['call_date'] = df['call_date'].astype(str)
        
        for yyyymm in sorted(grouped_series.keys()):
            for yyyymmdd in sorted(grouped_series[yyyymm]):
                # Filter cumulative data up to current date
                accum_graph_df = df[(df['call_month'] == yyyymm) & (df['call_date'] <= yyyymmdd)].copy()
                if len(accum_graph_df) == 0:
                    logger.warning(f"No data for {yyyymm}/{yyyymmdd}, skipping...")
                    continue

                # Group by the original date and reason, then sum the count, sort
                accum_graph_melted_df = accum_graph_df.melt(id_vars=['call_date', 'call_result'], value_vars=['main', 'secondary', 'third'], value_name='all_reason')
                accum_graph_melted_df = accum_graph_melted_df[accum_graph_melted_df['all_reason'].astype(str).str.strip() != '']
                accum_graph_group_df = accum_graph_melted_df.groupby(['call_date', 'all_reason']).size().unstack(fill_value=0)
                accum_graph_group_df = accum_graph_group_df.sort_index(axis=1)
                existing_reasons = accum_graph_group_df.columns
                plot_colors = [reason_colors.get(reason, 'gray') for reason in existing_reasons] 
                # ======== gen accum stacked bar chart image save rate by date, reason ===============================
                # Melt the 'save' data to count 'save' reasons per date
                call_result_melted_df = accum_graph_df.melt(id_vars=['call_date', 'call_result', 'call_id'], 
                                                value_vars=['main', 'secondary', 'third'], 
                                                value_name='reason')
                # Filter out rows where the 'reason' is empty
                call_result_melted_df = call_result_melted_df[call_result_melted_df['reason'].astype(str).str.strip() != '']
                # Calculate total count for each output_date and reason
                total_counts = call_result_melted_df.groupby(['call_date', 'reason']).size().reset_index(name='total_count')              
                # Calculate 'save' count for each output_date and reason
                save_counts = call_result_melted_df[call_result_melted_df['call_result'] == 'save'].groupby(['call_date', 'reason']).size().reset_index(name='save_count')
                # Merge the two dataframes
                result_df = pd.merge(total_counts, save_counts, on=['call_date', 'reason'], how='left').fillna(0)
                # Calculate the save percentage
                result_df['save_percentage'] = np.where(result_df['total_count'] > 0, (result_df['save_count'] / result_df['total_count']) * 100, 0)
                # Prepare data for stacked bar chart
                plot_df = result_df.pivot_table(index='call_date', columns='reason', values='save_percentage').fillna(0)
                # Create a list of colors for the existing reasons in the plot_df
                existing_reasons = plot_df.columns
                plot_colors = [reason_colors.get(reason, 'gray') for reason in existing_reasons]             
                # ============ plot 2 chart =========================================================
                num_days = len(accum_graph_group_df.index)
                fig_width = max(num_days * 2, 12)
                
                fig1 = plt.figure(figsize=(fig_width, 17)) 
                ax1 = plt.gca() 
                # Plot 1: Counts
                accum_graph_group_df.plot(kind='bar', color=plot_colors, stacked=True, ax=ax1, width=0.5)
                containers_list = [c for c in ax1.containers if len(c) > 0]
                if containers_list:
                    n_bars = len(containers_list[0])
                    totals = []
                    for i in range(n_bars):
                        s = 0.0
                        for c in containers_list:
                            try:
                                s += float(c[i].get_height())
                            except Exception:
                                pass
                        totals.append(s)

                    # annotate each segment with "count (pct%)"
                    for c in containers_list:
                        for i, bar in enumerate(c):
                            h = bar.get_height()
                            if h > 0:
                                total = totals[i] if i < len(totals) else 0
                                pct = (h / total * 100) if total else 0
                                ax1.text(
                                    bar.get_x() + bar.get_width() / 2,
                                    bar.get_y() + h / 2,
                                    f'{int(h)} ({pct:.1f}%)',
                                    ha='center', va='center', color='black', fontsize=18
                                )
                ax1.set_title('Count of Reasons by Date', fontsize=20)
                ax1.set_xlabel('Date')
                ax1.set_ylabel('Count of Reasons')
                ax1.legend(title='Reason', bbox_to_anchor=(1.01, 1), loc='upper left')
                
                original_dates = accum_graph_group_df.index.to_series()
                cleaned_dates = original_dates.str.split('/').str[-1]
                ax1.set_xticklabels(cleaned_dates, rotation=45, ha='right')
                plt.tight_layout()
                output_buffer_1 = io.BytesIO()
                plt.savefig(output_buffer_1, format='png', bbox_inches='tight')
                plt.close(fig1) 
                output_buffer_1.seek(0)

                # Plot 2: Percentages
                fig2 = plt.figure(figsize=(fig_width, 17)) 
                ax2 = plt.gca()
                plot_df.plot(kind='bar', stacked=True, color=plot_colors, ax=ax2, width=0.5)
                for bar_container in ax2.containers:
                    for bar in bar_container:
                        height = bar.get_height()
                        if height > 0:
                            ax2.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + height / 2, f'{int(height)}%', ha='center', va='center', color='black', fontsize=18)
                ax2.set_title('Save Rate Percentage by Reason and Date', fontsize=20)
                ax2.set_xlabel('Date')
                ax2.set_ylabel('Save Rate Percentage')
                ax2.legend(title='Reason', bbox_to_anchor=(1.01, 1), loc='upper left')
                ax2.set_xticklabels(cleaned_dates, rotation=45, ha='right')
                plt.tight_layout()
                output_buffer_2 = io.BytesIO()
                plt.savefig(output_buffer_2, format='png', bbox_inches='tight')
                plt.close(fig2) 
                output_buffer_2.seek(0)


                # --- Upload logic starts here ---
                daily_path = f"{VERINT_SITE_BASE_ROOT}/{OUTPUT_EXCEL_PREFIX}{PRODUCT}/{yyyymm}/{yyyymmdd}"
                
                # upload requests_by_reason
                requests_by_reason_path = f"{daily_path}/requests_by_reason.png"
                upload_file_to_sharepoint("verint", output_buffer_1, requests_by_reason_path)
                # logger.info(f"Uploading chart image 'requests_by_reason' to sharepoint folder '{requests_by_reason_path}'")

                # upload downsell_by_reason
                save_rate_by_reason_path = f"{daily_path}/save_rate_by_reason.png"
                upload_file_to_sharepoint("verint", output_buffer_2, save_rate_by_reason_path)
                # logger.info(f"Uploading chart image 'Save Rate Percentage by Reason and Date' to sharepoint folder '{save_rate_by_reason_path}'")

    except Exception as e:
        logger.error(f"Error in create_daily_bar_chart_daily: {str(e)}")
        import traceback
        traceback.print_exc()

def summarize(user_prompt: str, system_prompt: str, args: dict):
    logger.info(f"Gemini start processing summary")
    try:
        PROJECT_ID = args.get("PROJECT_ID")
        LOCATION = args.get("LOCATION")
        MODEL_NAME = args.get("MODEL_NAME")

        client = genai.Client(
            vertexai=True,
            project=PROJECT_ID,
            location=LOCATION,
        )

        generate_content_config = types.GenerateContentConfig(
            system_instruction=system_prompt,
            temperature = 0,
            top_p = 1,
            seed = 0,
            max_output_tokens = 25000,
            safety_settings = [types.SafetySetting(
                category="HARM_CATEGORY_HATE_SPEECH",
                threshold="OFF"
            ),types.SafetySetting(
                category="HARM_CATEGORY_DANGEROUS_CONTENT",
                threshold="OFF"
            ),types.SafetySetting(
                category="HARM_CATEGORY_SEXUALLY_EXPLICIT",
                threshold="OFF"
            ),types.SafetySetting(
                category="HARM_CATEGORY_HARASSMENT",
                threshold="OFF"
            )],
            thinking_config=types.ThinkingConfig(
                include_thoughts=True,
                thinking_budget=512,
            ),
        )
        # Construct the content for the API request
        contents = [
            types.Content(
                role="user",
                parts=[
                    types.Part.from_text(text=user_prompt),
                ]
            )
        ]
        response = client.models.generate_content(
            model=MODEL_NAME, contents=contents, config = generate_content_config
        )
        logger.info(f"Gemini finished processing summary")
        return response
    except Exception as e:
        logger.error(f"Gemini error processing summary", extra={'json_payload': str(e)}) 
        raise Exception(f"Gemini error processing summary: {str(e)}")

def df_to_text(df):
    if df is None or df.shape[0] == 0:
        return "No data available."

    # Convert df to readable aligned text table
    return df.to_string(index=False)
        
def create_daily_summarize_files(master_df: pd.DataFrame, data_list: list[dict], control_dict, verint_dict, **args):
    def _calculate_cost(llm_response: types.GenerateContentResponse) -> float:
        # gemini cost calculation
        # input = prompt token
        # output = candidate token + thought token
        usage_res = llm_response.model_dump(mode="python").get('usage_metadata', {})
        input_audio_token=0
        input_text_token=0
        thought_token=0
        ouput_token=0
        for detail in usage_res.get('prompt_tokens_details', []):
            if detail['modality'].value == 'AUDIO':
                input_audio_token += detail.get('token_count', 0)
            else:
                input_text_token += detail.get('token_count', 0)
        for detail in usage_res.get('candidates_tokens_details', []):
            if detail['modality'].value == 'TEXT':
                ouput_token += detail.get('token_count', 0)
        thought_token = usage_res.get('thoughts_token_count', 0)
        report_cost_cal = ((input_audio_token)*(1/1000000))+ ((input_text_token)*(0.3/1000000)) + ((ouput_token + thought_token)*(0.25/1000000))
        cost_thb = report_cost_cal * 33 # 1 USD = 33 THB
        return round(cost_thb, 4)

    CONTROL_SITE_BASE_ROOT = args.get("CONTROL_SITE_BASE_ROOT")
    CONTROL_SITE_PROMPTS_ROOT = args.get("CONTROL_SITE_PROMPTS_ROOT")
    
    VERINT_SITE_BASE_ROOT = args.get("VERINT_SITE_BASE_ROOT")
    OUTPUT_EXCEL_PREFIX = args.get("OUTPUT_EXCEL_PREFIX")
    PRODUCT = args.get("OUTPUT_PRODUCT")

    control_site_id = control_dict.get("control_site_id")
    control_access_token = control_dict.get("control_access_token")
    # Download summary prompt
    summary_prompt_url = f"{CONTROL_SITE_BASE_ROOT}/{CONTROL_SITE_PROMPTS_ROOT}/summary.txt"
    summary_prompt_url = get_item_download_url_by_path("control", summary_prompt_url)
    summary_prompt_file = requests.get(summary_prompt_url, headers={"Authorization": f"Bearer {control_access_token}"})
    summary_prompt_file.raise_for_status()
    # decode byte to string 
    summary_prompt = summary_prompt_file.text
    
    # Pre-Processing
    df = pd.DataFrame(data_list)
    # Normalize date
    df["run_date"] = pd.to_datetime(df["run_date"], errors="coerce").dt.tz_localize(None)

    # Deduplication: Keep latest record per call_id (prioritize non-fail status)
    df = (
        df.sort_values(["run_date"], ascending=[False])
        .drop_duplicates(subset=["call_id","phone_number", "call_date"], keep="first")
    )

    # List processing dates
    proc_date_list = df["call_date"].drop_duplicates().values.tolist()
    for proc_date in sorted(proc_date_list):
        mount = proc_date[:6]
        date = proc_date
        yyyymmdd_df = df[(df['call_month'] == mount) & (df['call_date'] == date)].copy()
        if len(yyyymmdd_df) == 0:
            logger.info(f"No data for {mount}/{date}, skipping...")
            continue

        logger.info(f"process date : {date} with {len(yyyymmdd_df)} records")   

        new_daily_melt_df = yyyymmdd_df.melt(
            id_vars=['call_date', 'call_id', 'call_result'],
            value_vars=['main', 'secondary', 'third'],
            value_name='all_reason'
        )

        reason_df = (
            new_daily_melt_df.query("all_reason.notnull() and all_reason != ''", engine='python')
            .groupby('all_reason')
            .size()
            .reset_index(name='count')
            .sort_values(by='count', ascending=False)
        )
        reason_df['percentage'] = (
            reason_df['count'] / reason_df['count'].sum() * 100
        ).round(2)

        call_result_df = (
                yyyymmdd_df.query("call_result.notnull() and call_result != ''", engine='python')
                .groupby('call_result')
                .size()
                .reset_index(name='count')
                .sort_values(by='count', ascending=False)
        )
        call_result_df['percentage'] = (
            call_result_df['count'] / call_result_df['count'].sum() * 100
        ).round(2)

        event_detection_df = (
                yyyymmdd_df.query("call_event_detection.notnull() and call_event_detection != ''", engine='python')
                .groupby('call_event_detection')
                .size()
                .reset_index(name='count')
                .sort_values(by='count', ascending=False)
        )
        event_detection_df['percentage'] = (
            event_detection_df['count'] / event_detection_df['count'].sum() * 100
        ).round(2)

        keyword_main_df = yyyymmdd_df[['main', 'call_result', 'keyword(main)', 'keyword(secondary)', 'keyword(third)']].query("call_result == 'churn'")

        # Convert DF → text
        reason_text = reason_df.to_markdown(index=False)
        call_result_text = call_result_df.to_markdown(index=False)
        event_detection_text = event_detection_df.to_markdown(index=False)
        keyword_main_text = keyword_main_df.to_markdown(index=False)

        # 🔥 very important → reset prompt
        from src.input_template import input_template
        template = input_template

        # Replace summary body
        template = (
            template.replace("${date}", date)
                .replace("${reason}", reason_text)
                .replace("${call_result}", call_result_text)
                .replace("${call_event_detection}", event_detection_text)
                .replace("${keyword_main}", keyword_main_text)
        )
        res_summary = summarize(
            user_prompt=template, system_prompt=summary_prompt, args=args
        )
        pre_summary = res_summary.text.replace("*","")

        # Calculate cost per round
        res_cost = pd.to_numeric(yyyymmdd_df['cost(thb)'], errors='coerce').fillna(0).sum()
        summary_cost =  _calculate_cost(res_summary)
        cost_per_day = round(res_cost + summary_cost, 4)
        accum_cost_mtd = round(pd.to_numeric(master_df['cost(thb)'], errors='coerce').fillna(0).sum(), 4)

        # Build final summary text
        summary = (
            f"{pre_summary}\n\n"
            f"Date: {date}\n"
            f"Number of call transcripts processed: {str(yyyymmdd_df.shape[0])}\n"
            f"Number of successful call transcripts: {str(yyyymmdd_df[(yyyymmdd_df['status'] == 'success')].shape[0])}\n"
            f"Number of failed call transcripts: {str(yyyymmdd_df[(yyyymmdd_df['status'] != 'success')].shape[0])}\n"
            f"Summary cost (THB): {summary_cost}\n"
            f"Total cost for {date} (result cost + summary cost) (THB): {cost_per_day}\n"
            f"Accumulated result cost MTD (THB): {accum_cost_mtd}\n"
        )

        # Upload to sharepoint
        summary_buffer = io.BytesIO()
        summary_buffer.write(summary.encode("utf-8"))
        summary_buffer.seek(0)
        summary_path = f"{VERINT_SITE_BASE_ROOT}/{OUTPUT_EXCEL_PREFIX}{PRODUCT}/{mount}/{date}/summary_{date}.txt"
        upload_file_to_sharepoint("verint", summary_buffer, summary_path)
        logger.info(f"✅ Uploaded summary for {date} to {summary_path}")