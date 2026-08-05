# Library imports
import asyncio
import contextlib
import copy
import io
import json
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from functools import partial
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.core.task_interface import TaskInterface

# Source code imports
from src.core.task_registry import task_registry
from src.modules.audit_log.batch_processing_log import BatchProcessingLogSchema, BatchProcessingPayload
from src.modules.audit_log.log_time_stamper import LogTimeStamper
from src.modules.audit_log.transaction_log import (
    TransactionLogSchema,
    TransactionPayload,
)
from src.modules.google.gcs import GCSModule
from src.modules.google.gemini_batch import GeminiBatchModule
from src.modules.microsoft.msgraph import MSGraphModule
from src.modules.microsoft.sharepoint import SharePointModule
from src.utils.common import (
    get_value_by_path,
    logging_ai_operation,
    recursive_dict_value_by_key,
    resolve_date,
    resolve_env,
    safe_cast_value,
    safe_list_get,
)
from src.utils.date_utils import (
    get_current_datetime,
)
from src.utils.file_utils import (
    load_yaml,
    read_xlsx,
)
from src.utils.logger import Logger
from src.utils.pandas_utils import (
    clean_invalid_xml_chars,
    ensure_df_schema,
    replace_nan_with_default,
)
from src.utils.token_utils import gemini_cost
from tasks.sentiment_qa.export_output_result_task import ExportOutputResultTask
from tasks.sentiment_qa.prep_payload_task import PrepPayloadTask

logger = Logger(__name__)


@task_registry.register("QAUserPlaygroundTask")
class UserPlaygroundTask(TaskInterface):
    """
    Combined fact-check batch task: upload voice files, prepare Gemini batch payloads,
    and submit the batch job in a single execution.

    Unlike the normal qa pipeline, commission_skill_code is read directly from the
    filename list Excel file — no agent master (emp_id) lookup is required.

    Each run produces output under an execution-datetime-stamped GCS prefix so that
    multiple executions never overwrite each other.
    """

    COMMON_CONFIG_PATH = "config/common.yml"
    DEFAULT_FILENAME_LIST_SHEET_NAME = "FilenameList"
    DEFAULT_PAYLOAD_FILE = "payloads.jsonl"
    DEFAULT_JSONL_PREDICTION_FILE = "predictions.jsonl"
    DEFAULT_MODEL_VERSION = "gemini-2.5-flash"
    DEFAULT_RECORD_DATE = "99991231"
    DEFAULT_COST_TYPE = "batch"

    TRANSACTION_LOG_SCHEMA = [
        "data_date",
        "start_time",
        "end_time",
        "total_time_mins",
        "type",
        "gcp_project_id",
        "gcp_project_name",
        "user_id",
        "source",
        "storage_path",
        "folder",
        "filename",
        "file_metadata_min",
        "status_pass_failed_retry",
        "error_log_if",
        "latency_ms",
        "token_usage_input",
        "token_usage_output",
        "total_cost_usd",
        "load_dt",
        "log_id",
        "log_type",
        "action",
        "status",
        "error_message",
        "created_dt",
        "updated_dt",
        "duration_seconds",
    ]

    GROUP_HEADER_DAILY_SCHEMA = [
        (
            "General",
            [
                "call_date",
                "agent_id",
                "call_id",
                "phone_number",
                "service_number",
                "call_duration_sec",
                "full_path",
                "department",
                "filename",
                "call_direction",
                "call_type",
                "call_type_confident",
            ],
        ),
        (
            "Customer Insight",
            [
                "summary_story",
                "product_category",
                "product_map",
                "repeat_call",
                "fcr",
                "churn_probability",
                "churn_reason",
                "customer_insight_summary",
                "standard_gsd_name",
            ],
        ),
        ("Service Quality", []),
        (
            "Sales Opportunities",
            [
                "opportunity_recognition_in_conversation",
                "product_suggested_by_ai",
                "agent_offer_product_presentation_&_explanation",
                "product_offer_by_agent",
                "sales_outcome_&_customer_decision",
                "sales_opportunities_performance_insight",
            ],
        ),
        (
            "Customer Sentiment",
            [
                "overall_sentiment",
                "initial_sentiment",
                "final_sentiment",
                "primary_sentiment_driver",
                "csat",
                "cs_performance_insight",
            ],
        ),
        (
            "Customer Experience",
            [
                "agent_communication_&_attitude",
                "agent_communication_&_attitude_reason",
                "agent_understanding_&_resolution",
                "agent_understanding_&_resolution_reason",
                "agent_responsiveness",
                "agent_responsiveness_reason",
                "system_accessibility",
                "system_accessibility_reason",
                "ivr_usability_&_design",
                "ivr_usability_&_design_reason",
                "ces",
                "self_service_readiness",
                "cx_performance_insight",
            ],
        ),
        (
            "Network",
            [
                "issue_type",
                "problem_statement",
                "area_tag_province",
                "area_tag_district",
                "area_tag_sub_district",
                "area_tag_landmark",
            ],
        ),
    ]

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.gcs = self.get_config("gcs", {})
        self.gcp = self.get_config("gcp", {})
        self.vertexai = self.get_config("vertexai", {})
        self.sharepoint = self.get_config("sharepoint", {})
        self.framework = self.get_config("framework", {})

        common_config = load_yaml(self.COMMON_CONFIG_PATH)
        self.control_access = common_config.get("control", {})
        self.msgraph_access = common_config.get("msgraph", {})

        self.project_id = resolve_env(self.gcs.get("project_id"))

        self.msgraph_sender_email = resolve_env(self.msgraph_access.get("sender_email"))
        self.msgraph_receiver_email = resolve_env(self.msgraph_access.get("receiver_email"))
        self.msgraph_cc_email = resolve_env(self.msgraph_access.get("cc_email"))

        self.task_sender_email = resolve_env(self.framework.get("sender_email"))
        self.task_receiver_email = resolve_env(self.framework.get("receiver_email"))
        self.task_cc_email = resolve_env(self.framework.get("cc_email"))

        self._cache_oper_log = {"transaction_df": pd.DataFrame(columns=self.TRANSACTION_LOG_SCHEMA), "process_date": []}

        # Caches for dynamic validation model construction
        self._cache_literal_checklist = {}
        self._cache_mapping_dict = {}

        # Per-metric thresholds from config (fallback to class-level STATUS_THRESHOLDS)
        self.metric_thresholds = self.get_config("metric_thresholds", {})

        self.batch_run_datetime = None  # YYYYMMDDHHMMSS

    def pre_execute(self):
        """Initialize SharePoint Control, GCS, and Gemini Batch modules."""
        logger.info("Initializing modules")

        # SharePoint Control
        try:
            self.control_site = resolve_env(self.control_access.get("site_domain"))
            self.sharepoint_control = SharePointModule(
                client_id=resolve_env(self.control_access.get("client_id")),
                client_secret=resolve_env(self.control_access.get("client_secret")),
                tenant_id=resolve_env(self.control_access.get("tenant_id")),
                site_domain=self.control_site,
                site_path=resolve_env(self.control_access.get("site_path")),
            )
            logger.debug(f"SharePoint Control: {self.control_site}")
        except Exception as e:
            logger.error(f"Failed to initialize SharePoint Control: {e}", exc_info=True)
            raise

        # GCS
        try:
            project_id = resolve_env(self.gcs.get("project_id"))
            bucket_name = resolve_env(self.gcs.get("bucket_name"))
            self.gcs_module = GCSModule(project_id=project_id, bucket_name=bucket_name)
            logger.debug(f"GCS: {project_id}/{bucket_name}")
        except Exception as e:
            logger.error(f"Failed to initialize GCS: {e}", exc_info=True)
            raise

        # Gemini Batch
        try:
            project_id = resolve_env(self.vertexai.get("project_id"))
            location = resolve_env(self.vertexai.get("location"))
            self.gemini_batch_module = GeminiBatchModule(
                genai_project_id=project_id,
                genai_location=location,
            )
            logger.debug(f"Gemini Batch: {project_id}/{location}")
        except Exception as e:
            logger.error(f"Failed to initialize Gemini Batch: {e}", exc_info=True)
            raise

        # Initialize Microsoft Graph client for email notifications
        try:
            self.msgraph_module = MSGraphModule(
                tenant_id=resolve_env(self.msgraph_access.get("tenant_id")),
                client_id=resolve_env(self.msgraph_access.get("client_id")),
                client_secret=resolve_env(self.msgraph_access.get("client_secret")),
            )
            logger.debug("Microsoft Graph client initialized for email notifications")
        except Exception as e:
            logger.error(f"Failed to initialize Microsoft Graph client: {e}", exc_info=True)
            raise

    def execute_task(self) -> Any:
        """
        Main execution: check for existing predictions and retrieve, or submit a new batch job.
        Returns:
            list: Processed batch results, or None if a new batch job was submitted.
        """
        execution_dt = self.get_package("execution_dt", None)
        rerun_date = self.get_package("rerun_data_dt", None)
        datadate = rerun_date if rerun_date else execution_dt.strftime("%Y-%m-%d")

        output_path = "/".join(resolve_env(self.gcs.get("output_folder")).split("/")[:-1])
        existing_files = self.gcs_module.list_files(prefix=output_path)
        prediction_output = [f for f in existing_files if self.DEFAULT_JSONL_PREDICTION_FILE in f]

        if prediction_output:
            logger.info(f"Found {len(prediction_output)} existing prediction file(s) — retrieving results.")
            return self._retrieve_prediction_step(prediction_output, execution_dt)
        logger.info("No prediction output found. Submitting new batch job.")
        self._submit_job_step(execution_dt, datadate)
        return None

    def _upload_daily_files(self, daily_df: pd.DataFrame, daily_output_path: str, sharepoint_site) -> None:
        """Appends a list of data dicts to XLSX daily files based on input folder structure."""

        group_header_daily_schema = copy.deepcopy(self.GROUP_HEADER_DAILY_SCHEMA)
        logger.info("Update Daily file")

        # ====== Grouping ======
        try:
            weight_score_path = resolve_env(get_value_by_path(self.sharepoint, "control.user_config_path"))
            weight_score_content = self.sharepoint_control.get_item_by_path(weight_score_path)
            with io.BytesIO(weight_score_content.content) as weight_score_file_buffer:
                service_quality_group_df = pd.read_excel(weight_score_file_buffer, sheet_name="service_quality_group")
        except Exception as e:
            logger.error(f"Failed to fetch weight score file from SharePoint: {e}", exc_info=True)
            weight_score_path = resolve_env(self.framework.get("user_config_path"))
            logger.info(f"Read file from config: {weight_score_path}")
            service_quality_group_df = read_xlsx(weight_score_path, sheet_name="service_quality_group")

        service_quality_group_df["sub_category"] = service_quality_group_df["sub_category"].str.strip()

        group_names = service_quality_group_df["sub_category"].dropna().loc[lambda s: s != ""].unique()

        service_quality_schema = []

        # Group
        for group_name in group_names:
            matching_items = (
                service_quality_group_df[service_quality_group_df["sub_category"] == group_name]["item"]
                .str.strip()
                .tolist()
            )
            existing_columns = [col for col in matching_items if col in daily_df.columns]

            service_quality_schema += matching_items

            if existing_columns:
                daily_df[group_name] = daily_df.apply(
                    lambda row, _cols=existing_columns: self.export_output_task_instance._calculate_category(
                        [str(row[col]) for col in _cols]
                    ),
                    axis=1,
                )

                group_reason = f"{group_name}_reason"
                daily_df[group_reason] = daily_df.apply(
                    lambda row, _cols=existing_columns: ", \n".join(
                        [f"{col}: {row[f'{col}_reason']}" for col in _cols if pd.notna(row[f"{col}_reason"])]
                    ),
                    axis=1,
                )
            else:
                daily_df[group_name] = ""
                daily_df[group_reason] = ""

            service_quality_schema.append(group_name)
            service_quality_schema.append(group_reason)
        # Non Group
        na_items = (
            service_quality_group_df[service_quality_group_df["sub_category"].isna()]["item"].str.strip().tolist()
        )
        service_quality_schema += na_items

        service_quality_schema.append("service_quality_score")
        service_quality_schema.append("service_quality_performance_insight")
        next(cols for name, cols in group_header_daily_schema if name == "Service Quality").extend(
            service_quality_schema
        )
        # ========================

        # Flatten all expected column names
        all_columns = [col for _, cols in group_header_daily_schema for col in cols]

        daily_df = daily_df.reindex(columns=all_columns)

        existing_wb = Workbook()
        ws = existing_wb.active

        # First row (group headers)
        col_idx = 1
        for group_name, cols in group_header_daily_schema:
            start_col = col_idx
            col_idx += len(cols)
            end_col = col_idx - 1
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=group_name)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(bold=True)

        # Second row (actual headers)
        col_idx = 1
        for _, cols in group_header_daily_schema:
            for col_name in cols:
                cell = ws.cell(row=2, column=col_idx, value=col_name)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = Font(bold=True)
                col_idx += 1

        # Data rows start from row 3
        for r_idx, row in enumerate(daily_df.itertuples(index=False, name=None), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        logger.info(f"Write file with {len(daily_df)} records: {daily_output_path}")

        # Auto-fit column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)

        # Save and upload
        output_stream = io.BytesIO()
        existing_wb.save(output_stream)
        output_stream.seek(0)

        sharepoint_site.upload_file(
            upload_path=daily_output_path,
            content=output_stream.read(),
        )

    def _retrieve_prediction_step(self, prediction_output: list, execution_dt: datetime) -> dict:
        """
        Retrieve and process batch prediction files from GCS.
        Parameters:
            prediction_output (list): List of GCS paths to predictions.jsonl files.
            execution_dt (datetime): Pipeline execution datetime for consistent load_dt stamping.
        Returns:
            list: Processed batch results.
        """
        all_results: list[dict] = []
        batch_run_datetime = ""
        for idx, path in enumerate(prediction_output, 1):
            logger.info(f"Processing batch [{idx}/{len(prediction_output)}]: {path}")
            try:
                raw_jsonl = GeminiBatchModule.retrieve_batch_results(gcs_module=self.gcs_module, batch_output_path=path)
                logger.info(f"Retrieved {len(raw_jsonl)} prediction records from batch")
                batch_results = self._proc_raw_prediction(batch=path, raw_jsonl=raw_jsonl, execution_dt=execution_dt)
                batch_run_datetime = path.split(
                    "/"
                )[
                    -3
                ]  #'sentiment_qa/user_playground/output/20260511090050/prediction-model-2026-05-11T02:01:19.257300Z/predictions.jsonl'  # noqa: E501
                all_results.extend(batch_results)
            except Exception as retrieve_err:
                logger.error(f"Failed to retrieve batch results from {path}: {retrieve_err}", exc_info=True)
                continue

        if not all_results:
            raise Exception("All batch files failed to retrieve — no prediction results available.")

        all_failed = all(r.get("prediction", {}).get("status") == "FAILED" for r in all_results)
        if all_failed:
            logger.error(f"All {len(all_results)} prediction(s) have FAILED status.")
            raise Exception(f"All {len(all_results)} prediction(s) failed — check batch output for errors.")

        self.batch_run_datetime = batch_run_datetime
        failed_count = sum(1 for r in all_results if r.get("prediction", {}).get("status") == "FAILED")
        logger.info(
            f"Total predictions processed: {len(all_results)} "
            f"({failed_count} failed, {len(all_results) - failed_count} succeeded)"
        )

        # Export Excel result
        try:
            export_output_task_instance = ExportOutputResultTask()
            export_output_task_instance.framework = self.framework
            export_output_task_instance.sharepoint = self.sharepoint
            export_output_task_instance.gcs = self.gcs
            export_output_task_instance.pre_execute()

            result = export_output_task_instance._format_output(all_results)
            result_df = pd.DataFrame(result)
            result_df = result_df.map(clean_invalid_xml_chars)

            self.export_output_task_instance = export_output_task_instance

            try:
                daily_output_path = resolve_date(
                    resolve_env(self.sharepoint.get("control").get("daily_output_file")), self.batch_run_datetime
                )
                logger.debug(f"Daily output path: {daily_output_path}")
            except Exception as e:
                logger.error(f"Failed to resolve output file path: {e}", exc_info=True)
                raise Exception(f"Cannot determine output file path: {e}") from e

            self._upload_daily_files(result_df, daily_output_path, self.sharepoint_control)
        except Exception as eval_err:
            logger.error(f"Export result step failed (non-fatal): {eval_err}", exc_info=True)
            raise Exception(f"Export result step failed: {eval_err}") from eval_err

        return all_results

    def _submit_job_step(self, execution_dt: datetime, datadate: str) -> tuple:
        """
        Orchestrate upload → payload preparation → batch job submission.
        Parameters:
            execution_dt (datetime): The execution datetime for the current run.
            datadate (str): The date string for the current run.
        Returns:
            tuple: (payload_path, output_folder, processing_log_list)
        """
        # Resolve per-run paths using execution datetime for isolation
        input_folder = resolve_env(self.gcs.get("input_folder"))
        processing_voice_folder = resolve_date(
            text=resolve_env(self.gcs.get("processing_voice_folder")), replace_date=execution_dt
        )
        processing_batch_folder = resolve_date(
            text=resolve_env(self.gcs.get("processing_batch_folder")), replace_date=execution_dt
        )
        output_folder = resolve_date(text=resolve_env(self.gcs.get("output_folder")), replace_date=execution_dt)
        payload_path = processing_batch_folder + "/" + self.DEFAULT_PAYLOAD_FILE

        logger.info(
            f"Run paths — input: {input_folder}, "
            f"processing_voice: {processing_voice_folder}, "
            f"batch: {processing_batch_folder}, "
            f"output: {output_folder}"
        )

        # --- Phase 1: Upload voice files from SharePoint → GCS input folder ---
        self._upload_voice_files()

        # --- Phase 2: List files from GCS input folder, filter by filename list ---
        try:
            input_files = self.gcs_module.list_files(prefix=input_folder)
            logger.info(f"Found {len(input_files)} file(s) in GCS {input_folder} folder")
        except Exception as e:
            logger.error(f"Failed to list GCS input files: {e}", exc_info=True)
            raise

        # --- Phase 3: Copy to per-run processing folder ---
        try:
            copy_result = asyncio.run(
                self.gcs_module.copy_files_batch(
                    source_files=input_files,
                    destination_prefix=processing_voice_folder,
                    max_concurrent_copies=int(resolve_env(self.framework["concurrency_upload"])),
                )
            )
            if copy_result["failed"] > 0:
                logger.warning(
                    f"Copy: {copy_result['success']} succeeded, "
                    f"{copy_result['failed']} failed — {copy_result.get('errors', [])}"
                )
            logger.info(f"Copied {copy_result['success']} file(s) to {processing_voice_folder}")
        except Exception as e:
            logger.error(f"Failed to copy files to processing folder: {e}", exc_info=True)
            raise

        # --- Phase 4: List files in processing folder ---
        try:
            processing_files = self.gcs_module.list_files(prefix=processing_voice_folder)
            if not processing_files:
                raise Exception(f"No files in processing folder: {processing_voice_folder}")
            logger.info(f"Processing {len(processing_files)} file(s)")
        except Exception as e:
            logger.error(f"Failed to list processing files: {e}", exc_info=True)
            raise

        # --- Phase 5: Build JSONL payloads ---
        gemini_generation_config = get_value_by_path(self.vertexai, "generation_config", {})
        mapping_file_type = {
            ".txt": "text/plain",
            ".wav": "audio/wav",
            ".jsonl": "application/jsonl",
        }

        payloads = []
        processing_log_list = []
        files_added = 0
        files_skipped = 0

        # Load prompt
        try:
            prep_task_instance = PrepPayloadTask()
            prep_task_instance.framework = self.framework
            prep_task_instance.sharepoint = self.sharepoint
            prep_task_instance.gcs = self.gcs

            prep_task_instance.pre_execute()

            prompt_inbound = prep_task_instance._prepare_prompt(execution_dt, "user_prompt_inbound")
            prompt_outbound = prep_task_instance._prepare_prompt(execution_dt, "user_prompt_outbound")

            prompt_inbound = prompt_inbound.replace("{date}", datadate.replace("-", ""))
            prompt_outbound = prompt_outbound.replace("{date}", datadate.replace("-", ""))

            output_schema = prep_task_instance._get_analysis_schema()
        except Exception as e:
            logger.error(f"Failed to prepare prompt: {e}", exc_info=True)
            raise Exception(f"Cannot prepare prompt: {e}") from e

        for processing_file in processing_files:
            try:
                file_uri = f"gs://{self.gcs_module.bucket_name}/{processing_file}"
                action = "PREP_PAYLOAD_BATCH_PROCESSING"
                file_name = os.path.basename(processing_file)
                file_name_without_ext = file_name.split(".")[0]
                processing_log = BatchProcessingLogSchema.from_dict(
                    BatchProcessingPayload(
                        data_date=datadate,
                        gcp_project_id=self.gcs_module.project_id,
                        gcp_project_name=self.gcs_module.project_id,
                        gcs_bucket_name=self.gcs_module.bucket_name,
                        source_path=file_uri,
                        filename=file_name,
                        prediction_payload_path=None,
                    )
                )
                call_direction = safe_list_get(file_name_without_ext.split("_"), 9, None)

                if call_direction == "IN":
                    prompt = prompt_inbound
                elif call_direction == "OUT":
                    prompt = prompt_outbound
                else:
                    raise ValueError(
                        f"Invalid call direction '{call_direction}' extracted from filename '{file_name}'. "
                        f"Expected 'IN' or 'OUT'."
                    )

                file_extension = Path(processing_file).suffix.lower()
                mime_type = mapping_file_type.get(file_extension)

                temp_generation_config = gemini_generation_config.copy()

                if not mime_type:
                    logger.debug(f"Skipping unsupported file type {file_extension}: {processing_file}")
                    files_skipped += 1
                    continue

                payload = {
                    "request": {
                        "contents": [
                            {
                                "role": "user",
                                "parts": [{"text": prompt}, {"fileData": {"fileUri": file_uri, "mimeType": mime_type}}],
                            }
                        ],
                        "generationConfig": output_schema,
                    }
                }
                payload["request"]["generationConfig"].update(temp_generation_config)
                payloads.append(payload)
                processing_log.stamp_payload_path(prediction_payload_path=payload_path)
                processing_log.stamp_completion(action=action)
                processing_log_list.append(processing_log)
                files_added += 1

            except Exception as e:
                logger.error(f"Failed to create payload for {os.path.basename(processing_file)}: {e}")
                processing_log.stamp_error(action=action, error_message=f"Failed to create payload: {e}")
                processing_log_list.append(processing_log)
                files_skipped += 1

        logger.info(f"Payloads: {files_added} created, {files_skipped} skipped")
        self.processing_log_list = processing_log_list

        if not payloads:
            logger.warning("No payloads created — skipping batch job submission")
            return None, None, processing_log_list

        # --- Phase 6: Upload merged JSONL to GCS ---
        try:
            merged_payload = ("\n".join(json.dumps(line, ensure_ascii=False) for line in payloads)).encode("utf-8")
            self.gcs_module.update_content_to_gcs(
                content=merged_payload, mime_type=mapping_file_type[".jsonl"], destination_path=payload_path
            )
            logger.info(f"Uploaded {len(payloads)} payloads ({len(merged_payload) / 1024:.2f} KB) to {payload_path}")
        except Exception as e:
            logger.error(f"Failed to upload payload: {e}", exc_info=True)
            raise Exception(f"Cannot upload payload to GCS: {e}") from e

        # --- Phase 7: Submit batch job ---
        try:
            if not self.gcs_module.is_file_exists(payload_path):
                raise FileNotFoundError(f"Payload file not found: {payload_path}")

            full_source_path = f"gs://{self.gcs_module.bucket_name}/{payload_path}"
            full_output_path = f"gs://{self.gcs_module.bucket_name}/{output_folder}"
            model_name = resolve_env(self.vertexai.get("model"))
            display_name = resolve_date(
                text=self.vertexai.get("batch_job_name", "sentiment-fc-qa-batch-job-%{DATA_DATE_YYYYMMDDHHMMSS}"),
                replace_date=execution_dt,
            )

            logger.info(f"Creating batch job: {display_name}")
            logger.debug(f"Model: {model_name}, Source: {full_source_path}, Output: {full_output_path}")

            batch_job = self.gemini_batch_module.create_batch_job(
                model_nm=model_name,
                src_uri=full_source_path,
                config={"dest": full_output_path, "display_name": display_name},
            )
            logger.info(f"Batch job created: {batch_job.name}")
        except Exception as e:
            logger.error(f"Failed to create batch job: {e}", exc_info=True)
            raise Exception(f"Batch job creation failed: {e}") from e

        # Initial status check
        try:
            time.sleep(5)
            job_status = self.gemini_batch_module.status_check_batch_job(job_name=batch_job.name)
            if job_status in ["JOB_STATE_FAILED", "JOB_STATE_CANCELLED", "JOB_STATE_EXPIRED"]:
                raise Exception(f"Batch job failed immediately with status: {job_status}")
            logger.info(f"Batch job submitted, current status: {job_status}")
        except Exception as e:
            logger.error(f"Batch job status check failed: {e}", exc_info=True)
            raise

        # Stamp batch job info to processing logs
        try:
            for log in processing_log_list:
                log.batch_job_id = str(batch_job.name.split("/")[-1])
                log.batch_job_display_name = batch_job.display_name
                log.model_name = model_name
        except Exception as e:
            logger.error(f"Failed to stamp batch job info to logs: {e}", exc_info=True)
            for log in processing_log_list:
                log.batch_job_id = None
                log.batch_job_display_name = display_name
                log.model_name = model_name

        self.processing_log_list = processing_log_list
        return payload_path, output_folder, processing_log_list

    def _upload_voice_files(self) -> None:
        """
        Upload voice files from SharePoint to GCS for the specified date range.
        Parameters:
            existing_df (pd.DataFrame): Existing DataFrame containing control log records
            start_date (str): Start date in 'YYYYMMDD'
            end_date (str): End date in 'YYYYMMDD'
        Returns:
            str: Uploaded payload path in GCS
        """

        filtered_voice_files = []

        source_path = resolve_env(get_value_by_path(self.sharepoint, "control.source_folder", ""))

        if not self.sharepoint_control.is_item_exists(source_path):
            logger.debug(f"Source folder not found for {source_path}")  # debug
            return

        voice_files = self.sharepoint_control.list_files(folder_path=source_path)
        if len(voice_files) == 0:
            logger.debug(f"No voice files for {source_path}")
            return
        logger.info(f"Processing {source_path}: {len(voice_files)} voice files")

        for file in voice_files:
            file_name = file.get("name", None)
            file_id = file.get("id", None)
            file_path = file.get("parentReference", {}).get("path", None)
            file_created_datetime = file.get("createdDateTime", None)
            file_extension = Path(file_name).suffix.lower() if file_name else None

            # Filter by file extension
            if file_extension and file_extension not in [".wav"]:
                continue

            filtered_voice_files.append(
                {
                    "file_name": file_name,
                    "file_extension": file_extension,
                    "file_id": file_id,
                    "file_path": file_path,
                    "file_created_datetime": file_created_datetime,
                }
            )

        stream_list = []
        for item in filtered_voice_files:
            input_voice_path = self.gcs.get("input_folder")

            stream_list.append(
                {
                    "download": item["file_path"].replace("/drive/root:", "") + "/" + item["file_name"],
                    "upload": input_voice_path + "/" + item["file_name"],
                    "mime_type": "audio/wav",
                }
            )

        asyncio.run(
            self.gcs_module.upload_sharepoint_to_gcs(
                sharepoint_object=self.sharepoint_control,
                stream_list=stream_list,
                max_concurrent_uploads=int(resolve_env(self.framework["concurrency_upload"])),
            )
        )

    def _proc_raw_prediction(self, batch: str, raw_jsonl: list[dict], execution_dt: datetime) -> list[dict]:
        """
        Process raw prediction records into structured format and append to batch_results.
        Parameters:
            batch (str): GCS path of the batch file (for logging).
            raw_jsonl (list[dict]): List of raw prediction records from Gemini batch output.
            execution_dt (datetime): Pipeline execution datetime for consistent load_dt stamping.
        Returns:
            list[dict]: List of processed prediction records.
        """

        def add_additional_info(line: dict) -> dict:
            model_version = get_value_by_path(line, "response.modelVersion", None)
            if model_version is None:
                logger.warning("Model version missing, using default")
                model_version = self.DEFAULT_MODEL_VERSION
            return {
                "create_time": get_value_by_path(line, "response.createTime", None),
                "processed_time": get_value_by_path(line, "processed_time", None),
                "model_version": model_version,
            }

        load_dt_str = execution_dt.strftime("%Y-%m-%d %H:%M:%S")
        processed_count = 0
        skipped_count = 0
        batch_results = []

        for line_idx, line in enumerate(raw_jsonl):
            try:
                voice_processed_path = safe_list_get(
                    recursive_dict_value_by_key(data=line, target_key="fileUri"), 0, None
                )
                if voice_processed_path is None:
                    logger.warning(f"No 'fileUri' in line {line_idx + 1}, skipping")
                    skipped_count += 1
                    continue

                file_name = os.path.splitext(os.path.basename(voice_processed_path))[0]
                logger.debug(f"Processing prediction for: {file_name}")

                try:
                    file_name_components = file_name.split("_")
                    try:
                        record_date = safe_list_get(file_name_components, 7, None)
                        if not record_date or not re.match(r"^\d{8}$", record_date or ""):
                            regex_match = re.findall(r"(?<=\/)\d{8}(?=\/)", voice_processed_path)
                            if regex_match:
                                record_date = regex_match[-1]
                                logger.warning(f"Record date extracted from path: {record_date}")
                            else:
                                record_date = self.DEFAULT_RECORD_DATE
                                logger.warning(f"Record date not found, using default: {record_date}")
                    except Exception as rec_err:
                        logger.error(f"Error extracting record date for '{file_name}': {rec_err}")
                        record_date = self.DEFAULT_RECORD_DATE

                    voice_info = {
                        "file_uri": voice_processed_path,
                        "file_name": file_name,
                        "file_ext": os.path.splitext(os.path.basename(voice_processed_path))[1],
                        "call_id": safe_list_get(file_name_components, 0, None),
                        "phone_number": safe_list_get(file_name_components, 1, None),
                        "call_time": safe_list_get(file_name_components, 2, None),
                        "agent_id": safe_list_get(file_name_components, 3, None),
                        "first_name": safe_list_get(file_name_components, 4, "").capitalize() or None,
                        "last_name": safe_list_get(file_name_components, 5, "").capitalize() or None,
                        "provider": safe_list_get(file_name_components, 6, None),
                        "record_date": record_date,
                        "duration": safe_list_get(file_name_components, 8, None),
                        "call_direction": safe_list_get(file_name_components, 9, None),
                    }
                except Exception as parse_err:
                    logger.error(f"Error parsing file name '{file_name}': {parse_err}")
                    voice_info = {
                        "file_uri": voice_processed_path,
                        "file_name": file_name,
                        "file_ext": None,
                        "call_id": None,
                        "phone_number": None,
                        "call_time": None,
                        "agent_id": None,
                        "first_name": None,
                        "last_name": None,
                        "provider": None,
                        "record_date": self.DEFAULT_RECORD_DATE,
                        "duration": None,
                    }

                # Error status in batch result
                if line.get("status", "") != "":
                    logger.warning(f"Batch result error status: {line['status']}")
                    payload = self._prepare_prediction_payload(voice_info, prediction=line["status"], err_flag=True)
                    payload["prediction"].update(add_additional_info(line))
                    payload["load_dt"] = load_dt_str
                    try:
                        usage_summary = GeminiBatchModule.sum_tokens_usage_for_billing(
                            get_value_by_path(line, "response.usageMetadata", {})
                        )
                        payload["prediction"].update(usage_summary)
                    except Exception:
                        pass
                    batch_results.append(payload)
                    processed_count += 1
                    continue

                # Missing prediction text
                prediction_str = get_value_by_path(line, "response.candidates.0.content.parts.0.text", None)
                if prediction_str is None:
                    warning_msg = f"No prediction found for file {file_name}"
                    logger.warning(warning_msg)
                    payload = self._prepare_prediction_payload(voice_info, prediction=warning_msg, err_flag=True)
                    payload["prediction"].update(add_additional_info(line))
                    payload["load_dt"] = load_dt_str
                    with contextlib.suppress(Exception):
                        payload["prediction"].update(
                            GeminiBatchModule.sum_tokens_usage_for_billing(
                                get_value_by_path(line, "response.usageMetadata", {})
                            )
                        )
                    batch_results.append(payload)
                    processed_count += 1
                    continue

                # Successful prediction
                payload = self._prepare_prediction_payload(voice_info, prediction=prediction_str)
                payload["prediction"].update(add_additional_info(line))
                payload["load_dt"] = load_dt_str
                try:
                    usage_summary = GeminiBatchModule.sum_tokens_usage_for_billing(
                        get_value_by_path(line, "response.usageMetadata", {})
                    )
                    payload["prediction"].update(usage_summary)
                    logger.debug(f"Usage: {usage_summary}")
                except Exception as usage_err:
                    logger.warning(f"Could not extract usage metadata for {file_name}: {usage_err}")
                    payload["prediction"].update(
                        {"token_input": {"text": 0, "audio": 0}, "token_output": {"text": 0}, "token_cached": 0}
                    )

                batch_results.append(payload)
                processed_count += 1

            except Exception as line_err:
                logger.error(f"Error processing line {line_idx + 1} in batch {batch}: {line_err}", exc_info=True)
                skipped_count += 1

        logger.info(f"Batch processing complete: {processed_count} processed, {skipped_count} skipped")
        return batch_results

    def _prepare_prediction_payload(self, voice_info: dict, prediction: str, err_flag: bool = False) -> dict:
        """
        Build a structured prediction payload.
        Parameters:
            voice_info (dict): File metadata extracted from the voice file path.
            prediction (str): Raw prediction JSON string or error message.
            err_flag (bool): True if this is an error record.
        Returns:
            dict: Structured payload with file_metadata and prediction fields.
        """
        logger.debug(f"Preparing payload for: {voice_info.get('file_name')}")
        fw_schema = {
            "file_metadata": voice_info,
            "prediction": {
                "raw_prediction": None,
                "status": None,
                "message": None,
            },
        }

        if err_flag:
            fw_schema["prediction"]["status"] = "FAILED"
            fw_schema["prediction"]["message"] = prediction
            logger.warning(f"Marking record as FAILED: {voice_info.get('file_name')}")
            return fw_schema

        try:
            logger.debug(f"Parsing prediction JSON for: {voice_info.get('file_name')}")
            parsed_data = json.loads(prediction)
            fw_schema["prediction"]["raw_prediction"] = parsed_data
            fw_schema["prediction"]["status"] = "SUCCESS"
            fw_schema["prediction"]["message"] = None
            logger.info(f"Successfully parsed prediction for: {voice_info.get('file_name')}")
        except Exception as e:
            logger.error(f"Failed to parse prediction for '{voice_info.get('file_name')}': {e}", exc_info=True)
            fw_schema["prediction"]["status"] = "FAILED"
            fw_schema["prediction"]["message"] = f"Failed to parse prediction JSON: {str(e)}"

        return fw_schema

    def post_execute(self, result: Any) -> Any:
        """
        Archive batch output files, clean up GCS working directories, and insert
        transaction log. Only runs when predictions have been retrieved (result is not None).
        """
        if not result:
            logger.info("No batch results (batch job submitted, not retrieved) — skipping archive and transaction log")
            return result

        execution_dt = self.get_package("execution_dt", None)

        # Build DataFrame from batch results
        after_upload_list = []
        for record in result:
            try:
                file_name = get_value_by_path(record, "file_metadata.file_name", None)
                file_ext = get_value_by_path(record, "file_metadata.file_ext", "") or ""
                full_path = (
                    resolve_env(get_value_by_path(self.sharepoint, "control.source_folder"))
                    + "/"
                    + str(file_name or "")
                    + file_ext
                )

                after_upload_list.append(
                    {
                        "file_name": file_name,
                        "full_path": full_path,
                        "folder": None,
                        "record_date": get_value_by_path(record, "file_metadata.record_date", self.DEFAULT_RECORD_DATE),
                        "duration": get_value_by_path(record, "file_metadata.duration", None),
                        "token_input": get_value_by_path(record, "prediction.token_input", None),
                        "token_cached": get_value_by_path(record, "prediction.token_cached", None),
                        "token_output": get_value_by_path(record, "prediction.token_output", None),
                        "status": get_value_by_path(record, "prediction.status", None),
                        "message": get_value_by_path(record, "prediction.message", None),
                        "processed_time": get_value_by_path(record, "prediction.processed_time", None),
                        "create_time": get_value_by_path(record, "prediction.create_time", None),
                        "model_version": get_value_by_path(record, "prediction.model_version", None),
                        "load_dt": record.get("load_dt", None),
                    }
                )
            except Exception as e:
                logger.error(f"Error building upload dict for record: {e}", exc_info=True)
                continue

        df = pd.DataFrame(after_upload_list)

        logger.info(f"Built DataFrame with {len(df)} records for archive and logging")

        # --- Phase 1: Archive file in GCS ---
        try:
            self._archive_and_cleanup(execution_dt)
        except Exception as archive_err:
            logger.error(f"Archive and cleanup failed: {archive_err}", exc_info=True)
            raise Exception(f"Archive and cleanup process failed: {archive_err}") from archive_err

        # --- Phase 2: Insert transaction log records ---
        if df.empty:
            logger.warning("DataFrame is empty — all records failed to build. Skipping transaction log.")
        else:
            try:
                self._insert_log_record(df)
            except Exception as log_err:
                logger.error(f"Failed to insert transaction log records: {log_err}", exc_info=True)
                raise Exception(f"Critical error: Transaction log insertion failed: {log_err}") from log_err

        # --- Phase 3: Archive file in SharePoint ---
        try:
            source_path = resolve_env(get_value_by_path(self.sharepoint, "control.source_folder", ""))
            destination_folder = resolve_date(
                resolve_env(self.sharepoint.get("control").get("archive_folder")), self.batch_run_datetime
            )
            daily_output_file_path = resolve_date(
                resolve_env(self.sharepoint.get("control").get("daily_output_file")), self.batch_run_datetime
            )
            user_config_file_path = resolve_env(get_value_by_path(self.sharepoint, "control.user_config_path", ""))
        except Exception as e:
            logger.error(f"Failed to resolve output file path: {e}", exc_info=True)
            raise Exception(f"Cannot determine output file path: {e}") from e

        # archive output file
        try:
            daily_output_filename = daily_output_file_path.split("/")[-1]
            self.sharepoint_control.copy_file(
                source_path=daily_output_file_path, destination_path=f"{destination_folder}/{daily_output_filename}"
            )
        except Exception as e:
            logger.error(f"Failed to archive excel output file to SharePoint: {e}", exc_info=True)

        # archive user_config.xlsx
        try:
            self.sharepoint_control.copy_file(
                source_path=user_config_file_path, destination_path=f"{destination_folder}/user_config.xlsx"
            )
        except Exception as e:
            logger.error(f"Failed to archive excel output file to SharePoint: {e}", exc_info=True)

        # archive file in input folder
        if not self.sharepoint_control.is_item_exists(source_path):
            logger.warning(f"Source path does not exist: {source_path}")
            return result

        voice_files = self.sharepoint_control.list_files(folder_path=source_path)
        if not voice_files:
            logger.warning(f"No voice files found in source path: {source_path}")
            return result

        tasks_paths = []
        for file in voice_files:
            file_name = file.get("name", "")
            tasks_paths.append({"src": f"{source_path}/{file_name}", "dest": f"{destination_folder}/input/{file_name}"})
        max_workers = int(resolve_env(self.framework.get("concurrency_upload", 5)))

        logger.info(f"Starting parallel copy for {len(tasks_paths)} files with {max_workers} workers.")

        # Run the synchronous copy_file calls in a thread pool to achieve parallelism
        async def run_parallel_copies():
            loop = asyncio.get_event_loop()
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                tasks = [
                    loop.run_in_executor(
                        executor, partial(self.sharepoint_control.copy_file, path_pair["src"], path_pair["dest"])
                    )
                    for path_pair in tasks_paths
                ]
                return await asyncio.gather(*tasks)

        results = asyncio.run(run_parallel_copies())

        success_count = sum(1 for r in results if r is True)
        logger.info(f"Parallel copy finished. Success: {success_count}/{len(results)}")

        # --- Phase 4: Print log to Splunk ---
        logger.info("Stamping AI-Operation logs")
        if not self._cache_oper_log.get("process_date"):
            logger.warning("No process_date found for stamping logs - skipping logging")
            return result

        transaction_df = self._cache_oper_log.get("transaction_df")
        if transaction_df is None or transaction_df.empty:
            logger.warning("No transaction log DataFrame found for stamping logs - skipping logging")
            return result
        try:
            required_columns = ["start_time", "end_time", "gcp_project_id", "status_pass_failed_retry", "latency_ms"]
            transaction_df = ensure_df_schema(transaction_df, required_columns).copy()
            transaction_df["start_time"] = pd.to_datetime(
                transaction_df["start_time"], errors="coerce", utc=True
            ).dt.tz_convert(LogTimeStamper.CONFIG_TIMEZONE)
            transaction_df["end_time"] = pd.to_datetime(
                transaction_df["end_time"], errors="coerce", utc=True
            ).dt.tz_convert(LogTimeStamper.CONFIG_TIMEZONE)
            transaction_df["latency_ms"] = pd.to_numeric(transaction_df["latency_ms"], errors="coerce")
            transaction_df["process_date"] = transaction_df["start_time"].dt.date
            process_dates = self._cache_oper_log["process_date"]
            for process_date in process_dates:
                filtered_df = transaction_df[(transaction_df["process_date"] == process_date)]

                if filtered_df.empty:
                    logger.warning(
                        f"No transaction records found for process_date {process_date} - skipping log stamping"
                    )
                    continue

                log_df = filtered_df.groupby(["process_date", "gcp_project_id"], as_index=False, dropna=False).agg(
                    total_transaction=("status_pass_failed_retry", "count"),
                    total_success_transaction=("status_pass_failed_retry", lambda x: (x == "Pass").sum()),
                    total_failed_transaction=("status_pass_failed_retry", lambda x: (x == "Failed").sum()),
                    average_response_time_sec=(
                        "latency_ms",
                        lambda x: round(x.mean() / 1000, 2) if pd.notna(x.mean()) else 0.0,
                    ),
                    min_start_time=("start_time", "min"),
                    max_end_time=("end_time", "max"),
                )
                log_df["total_runtime_sec"] = (
                    (
                        pd.to_datetime(log_df["max_end_time"], errors="coerce")
                        - pd.to_datetime(log_df["min_start_time"], errors="coerce")
                    )
                    .dt.total_seconds()
                    .round(2)
                )
                log_df = log_df.drop(columns=["min_start_time", "max_end_time"])
                env = os.environ.get("ENVIRONMENT", "").lower()
                if env == "prod":
                    log_df["environment"] = "production"
                elif env == "nprd":
                    log_df["environment"] = "non-production"
                else:
                    log_df["environment"] = env or "unknown"
                log_df = log_df.rename(columns={"gcp_project_id": "project_id"})
                log_df["project_type"] = "batch"

                log_df = log_df[
                    [
                        "process_date",
                        "environment",
                        "project_id",
                        "project_type",
                        "total_transaction",
                        "total_success_transaction",
                        "total_failed_transaction",
                        "average_response_time_sec",
                        "total_runtime_sec",
                    ]
                ]
                log_list = log_df.to_dict(orient="records")
                for log in log_list:
                    logging_ai_operation(log_instance=logger, log_obj=log, log_type="batch", message="AI-Operation-Log")
            logger.info("AI-Operation log stamping completed successfully")
        except Exception as log_err:
            logger.error(f"Failed to stamp AI-Operation logs: {log_err}", exc_info=True)

        # --- Phase 5: Send email ---

        current_dt = get_current_datetime().replace(microsecond=0)
        formatted_current_dt = current_dt.isoformat(sep=" ")
        subject = f"[AI Report] [AI-QA] [user-playground] on {self.batch_run_datetime}"
        body = f"""
            Dear All,<br>
            <br>
                The manually triggered run for user-playground is complete. <br>
                <br>
                Project: {self.project_id} <br>
                batch datetime: {self.batch_run_datetime} <br>
                <br>
            <br>
            -- Timestamp: {formatted_current_dt} <br>
            Best Regards,<br>
            [This is automatic message generated by AI - Do Not REPLY]
            """

        self.msgraph_module.send_email(
            subject=subject,
            body=body,
            sender_email=self.task_sender_email,
            receiver_email=self.task_receiver_email,
            cc_email=self.task_cc_email,
        )

        return result

    def _archive_and_cleanup(self, execution_dt: datetime) -> None:  # noqa: ARG002 -- call-site interface param
        """
        Archive prediction JSONL files from the GCS output folder to the archive folder,
        then delete working directories (input, processing, output).
        """
        output_parent = "/".join(resolve_env(self.gcs.get("output_folder")).split("/")[:-1])

        try:
            archive_folder = resolve_date(resolve_env(self.gcs.get("archive_batch_folder")), self.batch_run_datetime)
            logger.debug(f"Archive batch folder: {archive_folder}")
        except Exception as e:
            logger.error(f"Failed to resolve archive batch folder path: {e}", exc_info=True)
            raise Exception(f"Cannot determine archive batch folder path: {e}") from e

        logger.info(f"Archiving batch files from {output_parent} to {archive_folder}")

        # --- Archive prediction output files ---
        archived_count = 0
        failed_archive_count = 0

        try:
            output_files = self.gcs_module.list_files(prefix=output_parent)
            logger.info(f"Found {len(output_files)} file(s) in output folder to archive")
        except Exception as list_err:
            logger.error(f"Failed to list output files for archiving: {list_err}", exc_info=True)
            output_files = []

        for idx, file_path in enumerate(output_files, 1):
            try:
                # Strip output_parent prefix, then drop the execution-timestamp folder (index 0)
                # e.g. "output/20260304150358/prediction-model-.../predictions.jsonl"
                #   → "prediction-model-.../predictions.jsonl"
                relative_path = file_path[len(output_parent) :].lstrip("/")
                batch_path = "/".join(relative_path.split("/")[1:])
                destination = f"{archive_folder}/{batch_path}"
                logger.debug(f"Archiving [{idx}/{len(output_files)}]: {file_path} -> {destination}")
                self.gcs_module.move_file(source_path=file_path, destination_path=destination)
                archived_count += 1
            except Exception as archive_err:
                logger.error(f"Failed to archive file '{file_path}': {archive_err}", exc_info=True)
                failed_archive_count += 1
                continue

        logger.info(f"Batch archive complete. Archived: {archived_count}, Failed: {failed_archive_count}")

        # --- Cleanup working directories ---
        logger.info("Cleaning up working GCS directories")

        dirs_to_delete = []
        try:
            dirs_to_delete.append(resolve_env(self.gcs.get("input_folder")))
        except Exception as e:
            logger.warning(f"Could not resolve input_folder for cleanup: {e}")

        try:
            # Parent processing/ folder — strip "voice/%{DATA_DATE_YYYYMMDDHHMMSS}" (last 2 components)
            # e.g. "sentiment_qa/fact_check/processing/voice/%{...}" → "sentiment_qa/fact_check/processing"
            # This covers both processing/voice/ and processing/batch/ in one delete
            processing_voice = resolve_env(self.gcs.get("processing_voice_folder"))
            dirs_to_delete.append("/".join(processing_voice.split("/")[:-2]))
        except Exception as e:
            logger.warning(f"Could not resolve processing_voice_folder for cleanup: {e}")

        try:
            dirs_to_delete.append(output_parent)
        except Exception as e:
            logger.warning(f"Could not resolve output_folder parent for cleanup: {e}")

        deleted_count = 0
        failed_delete_count = 0
        for dir_path in dirs_to_delete:
            try:
                if not self.gcs_module.is_dir_exists(dir_path=dir_path):
                    logger.debug(f"Directory does not exist, skipping: {dir_path}")
                    deleted_count += 1
                    continue
                logger.info(f"Deleting directory: {dir_path}")
                self.gcs_module.delete_dir(dir_path=dir_path)
                deleted_count += 1
                logger.info(f"Deleted directory: {dir_path}")
            except Exception as delete_err:
                logger.error(f"Failed to delete directory '{dir_path}': {delete_err}", exc_info=True)
                failed_delete_count += 1

        logger.info(f"Directory cleanup complete. Deleted: {deleted_count}, Failed: {failed_delete_count}")

    def _insert_log_record(self, df: pd.DataFrame) -> None:
        """Insert transaction log records for all records in a single execution-dated file."""
        logger.info("Starting transaction log record insertion")
        default_type = "AI User-Playground"
        default_user_id = "daisyrpa"
        default_source = "SharePoint"

        execution_dt = self.get_package("execution_dt", None)

        try:
            logger.info("Creating transaction log records")
            self._transaction_log(
                default_type=default_type,
                default_user_id=default_user_id,
                default_source=default_source,
                prediction_df=df,
                execution_dt=execution_dt,
            )
            logger.info("Transaction log created successfully")
        except Exception as trans_err:
            logger.error(f"Critical error creating transaction log: {trans_err}", exc_info=True)
            raise Exception(f"Transaction log creation failed: {trans_err}") from trans_err

    def _transaction_log(
        self,
        default_type: str,
        default_user_id: str,
        default_source: str,
        prediction_df: pd.DataFrame,
        execution_dt: datetime,
    ) -> pd.DataFrame:
        """
        Create and upload transaction log entries for user-playground batch results.
        Mirrors export_output_result_task._transaction_log with type='AI User-Playground'
        and GCS URI as storage_path.
        The log file path is keyed by execution_dt (pipeline run date); data_date
        (voice call date) remains a field value inside the CSV rows.
        """
        logger.info(f"Processing {len(prediction_df)} records for transaction logging")

        model_pricing = gemini_cost(
            api_type=self.DEFAULT_COST_TYPE, model_list=prediction_df["model_version"].unique().tolist()
        )

        required_columns = [
            "file_name",
            "full_path",
            "folder",
            "record_date",
            "duration",
            "token_input",
            "token_cached",
            "token_output",
            "status",
            "message",
            "processed_time",
            "create_time",
            "model_version",
            "load_dt",
        ]
        missing_columns = [col for col in required_columns if col not in prediction_df.columns]
        if missing_columns:
            logger.error(f"Missing required columns: {missing_columns}")
            raise ValueError(f"Cannot create transaction log: missing columns {missing_columns}")

        log_payload = []
        usage_df = prediction_df[required_columns].copy()
        usage_dict = usage_df.to_dict(orient="records")

        for idx, record in enumerate(usage_dict, 1):
            try:
                if not record.get("file_name"):
                    raise ValueError(f"Record {idx}: Missing required field 'file_name'")

                control_site_name = resolve_env(get_value_by_path(self.control_access, "site_name", ""))
                if not record.get("full_path"):
                    logger.warning(f"Record {idx} ({record['file_name']}): Missing full_path, URL will be incomplete")
                    storage_path = f"https://{self.control_site}/sites/{control_site_name}/"
                else:
                    storage_path = f"https://{self.control_site}/sites/{control_site_name}/{record['full_path']}"

                usage_detail = {
                    record["file_name"]: {
                        "model": record.get("model_version"),
                        "token_input": record.get("token_input"),
                        "token_cached": record.get("token_cached"),
                        "token_output": record.get("token_output"),
                    }
                }

                token_usage_input = 0
                token_usage_output = 0
                for usage in usage_detail.values():
                    for _, token_count in usage.get("token_input", {}).items():
                        if str(token_count).isdigit():
                            token_usage_input += int(token_count)
                    for _, token_count in usage.get("token_output", {}).items():
                        if str(token_count).isdigit():
                            token_usage_output += int(token_count)

                cost_detail = GeminiBatchModule.cal_gemini_cost(
                    usage_detail=usage_detail,
                    cost_config=model_pricing,
                )
                cost_input = cost_detail[record["file_name"]].get("cost_input", 0.0) or 0.0
                cost_output = cost_detail[record["file_name"]].get("cost_output", 0.0) or 0.0
                total_cost_usd = cost_input + cost_output

                duration_val = record.get("duration")

                if duration_val is None:
                    file_metadata_sec = 0
                elif isinstance(duration_val, int):
                    file_metadata_sec = duration_val
                else:
                    file_metadata_sec = safe_cast_value(duration_val, int, 0)

                raw_start = str(record.get("processed_time", ""))
                raw_end = str(record.get("create_time", ""))
                if raw_start and raw_end:
                    try:
                        if datetime.fromisoformat(raw_start) > datetime.fromisoformat(raw_end):
                            raw_start, raw_end = raw_end, raw_start
                    except Exception:
                        pass

                transaction_log = TransactionLogSchema.from_dict(
                    TransactionPayload(
                        data_date=str(record.get("record_date", "")),
                        start_time=raw_start,
                        end_time=raw_end,
                        type=default_type,
                        gcp_project_id=resolve_env(self.gcp.get("project_id", None)),
                        gcp_project_name=resolve_env(self.gcp.get("project_name", None)),
                        user_id=default_user_id,
                        source=default_source,
                        storage_path=storage_path,
                        folder=str(record.get("folder", "")),
                        filename=str(record.get("file_name", "")),
                        file_metadata_sec=file_metadata_sec,
                        status_pass_failed_retry="Pass" if record.get("status") == "SUCCESS" else "Failed",
                        error_log_if=str(record.get("message", "")),
                        token_usage_input=token_usage_input,
                        token_usage_output=token_usage_output,
                        total_cost_usd=safe_cast_value(total_cost_usd, float, 0.0),
                        load_dt=str(record.get("load_dt", "")),
                    )
                )
                if record.get("status") != "SUCCESS":
                    transaction_log.stamp_error(
                        action="Create Transaction Log", error_message=record.get("message", "")
                    )
                else:
                    transaction_log.stamp_completion(action="Create Transaction Log")
                log_payload.append(transaction_log)

            except Exception as record_err:
                logger.error(f"CRITICAL: Record {idx}: {record_err}", exc_info=True)
                raise Exception(f"Transaction log creation failed at record {idx}: {record_err}") from record_err

        logger.info(f"Created {len(log_payload)} transaction payload record(s)")

        if not log_payload:
            raise Exception("No transaction records were created from prediction data")

        # Build DataFrame
        new_transaction_df = pd.DataFrame([log.to_dict() for log in log_payload])
        new_transaction_df = ensure_df_schema(new_transaction_df, self.TRANSACTION_LOG_SCHEMA)
        logger.info(f"Transaction log DataFrame: {len(new_transaction_df)} rows")

        # Upload all records to a single log file keyed by execution_dt (pipeline run date),
        # not by data_date (voice call date). data_date remains a field value inside the CSV.
        try:
            transaction_path = resolve_date(
                text=resolve_env(get_value_by_path(self.sharepoint, "control.transaction_log_file")),
                replace_date=execution_dt,
            )
            logger.info(f"Transaction log path (execution_dt): {transaction_path}")

            if self.sharepoint_control.is_item_exists(item_path=transaction_path):
                logger.info("Existing log found, merging")
                existing_log = self.sharepoint_control.get_item_by_path(transaction_path)
                with io.BytesIO(existing_log.content) as buf:
                    existing_df = pd.read_csv(buf)
                logger.debug(f"Existing log has {len(existing_df)} rows")
                combined_df = pd.concat([existing_df, new_transaction_df], ignore_index=True)
                logger.debug(f"Combined DataFrame has {len(combined_df)} rows")
            else:
                logger.info("No existing log found, creating new")
                combined_df = new_transaction_df

            combined_df["data_date"] = combined_df["data_date"].astype(str)
            combined_df = combined_df.fillna("").sort_values(
                by=["updated_dt", "load_dt", "data_date", "start_time", "end_time"],
                ascending=[False, False, False, False, False],
            )
            combined_df = ensure_df_schema(combined_df, list(new_transaction_df.columns))
            combined_df = replace_nan_with_default(combined_df, default_value="")

            csv_buffer = io.BytesIO()
            combined_df.to_csv(csv_buffer, index=False, encoding="utf-8-sig")
            csv_buffer.seek(0)

            self.sharepoint_control.upload_file(
                upload_path=transaction_path,
                content=csv_buffer.read(),
            )
            logger.info(f"Successfully uploaded transaction log to {transaction_path}")

            self._cache_oper_log["transaction_df"] = new_transaction_df
            self._cache_oper_log["process_date"] = (
                new_transaction_df["start_time"].dt.date.unique().tolist()
                if "start_time" in new_transaction_df.columns
                else []
            )

        except Exception as upload_err:
            logger.error(f"Failed to upload transaction log: {upload_err}", exc_info=True)
            raise Exception(f"Critical error: Cannot upload transaction log: {upload_err}") from upload_err

        return new_transaction_df

    def on_error(self, error: Exception) -> None:
        """
        Hook executed when task execution fails.
        Override this method to implement custom error handling
        (e.g., cleanup, notifications, retry logic).

        Args:
            error (Exception): The exception that occurred during execution

        Returns:
            None
        """
        try:
            current_dt = get_current_datetime().replace(microsecond=0)
            formatted_current_dt = current_dt.isoformat(sep=" ")

            subject = "[AI Failed] [AI-QA]"
            body = f"""
            Dear Operation Team,<br>
            <br>
                Project: {self.project_id} <br>
                An error occurred during execution of task '{self.task_name}' <br>
                Error details: {error} <br>
                <br>
                Please investigate. <br>
            <br>
            -- Timestamp: {formatted_current_dt} <br>
            Best Regards,<br>
            [This is automatic message generated by AI - Do Not REPLY]
            """

            self.msgraph_module.send_email(
                subject=subject,
                body=body,
                sender_email=self.msgraph_sender_email,
                receiver_email=self.msgraph_receiver_email,
                cc_email=self.msgraph_cc_email,
            )
            logger.info(f"Task error notification successfully issued: {self.task_name}")
        except Exception as email_error:
            logger.error(f"Task error notification failed to issue: {email_error}", exc_info=True)
        logger.error(f"Error in task '{self.task_name}': {error}", exc_info=True)
