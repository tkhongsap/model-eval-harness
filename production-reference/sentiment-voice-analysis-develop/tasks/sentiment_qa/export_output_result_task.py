# Library imports
import copy
import io
import os
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.core.task_interface import TaskInterface

# Source code imports
from src.core.task_registry import task_registry
from src.modules.audit_log.log_time_stamper import LogTimeStamper
from src.modules.audit_log.performance_log import (
    PerformanceLogSchema,
    PerformancePayload,
)
from src.modules.audit_log.transaction_log import (
    TransactionLogSchema,
    TransactionPayload,
)
from src.modules.google.gcs import GCSModule
from src.modules.google.gemini_batch import GeminiBatchModule
from src.modules.microsoft.msgraph import MSGraphModule
from src.modules.microsoft.sharepoint import SharePointModule
from src.utils.common import (
    get_value_by_path,
    logging_ai_operation,
    resolve_date,
    resolve_env,
    safe_cast_value,
    safe_list_get,
)
from src.utils.date_utils import (
    get_current_datetime,
)
from src.utils.file_utils import (
    load_yaml,
    read_xlsx,
)
from src.utils.logger import Logger
from src.utils.pandas_utils import clean_invalid_xml_chars, ensure_df_schema, replace_nan_with_default
from src.utils.token_utils import (
    gemini_cost,
)

logger = Logger(__name__)


@task_registry.register("QAExportOutputResultTask")
class ExportOutputResultTask(TaskInterface):
    """
    Task to export telesale sentiment analysis results to SharePoint and archive processed files.
    """

    COMMON_CONFIG_PATH = "config/common.yml"
    DEFAULT_COST_TYPE = "batch"
    DEFAULT_RECORD_DATE = "99991231"
    TRANSACTION_LOG_SCHEMA = [
        "data_date",
        "start_time",
        "end_time",
        "total_time_mins",
        "type",
        "gcp_project_id",
        "gcp_project_name",
        "user_id",
        "source",
        "storage_path",
        "folder",
        "filename",
        "file_metadata_min",
        "status_pass_failed_retry",
        "error_log_if",
        "latency_ms",
        "token_usage_input",
        "token_usage_output",
        "total_cost_usd",
        "load_dt",
        "log_id",
        "log_type",
        "action",
        "status",
        "error_message",
        "created_dt",
        "updated_dt",
        "duration_seconds",
    ]

    PERFORMANCE_LOG_SCHEMA = [
        "data_date",
        "run_date",
        "load_dt",
        "gcp_project_id",
        "gcp_project_name",
        "total_transaction",
        "total_completed",
        "total_failed",
        "success_rate",
        "total_runtime",
        "log_id",
        "log_type",
        "action",
        "status",
        "error_message",
        "created_dt",
        "updated_dt",
        "duration_seconds",
        "average_response_time_ms",
    ]

    GROUP_HEADER_SCHEMA = [
        (
            "General",
            [
                "filename",
                "agent_id",
                "call_id",
                "phone_number",
                "service_number",
                "call_duration_sec",
                "full_path",
                "department",
                "call_direction",
                "call_type",
                "call_type_confident",
                "call_month",
                "call_date",
                "full_name",
                "true_dtac",
                "updated_dt",
                "status",
                "error_code",
            ],
        ),
        (
            "Customer Insight",
            [
                "summary_story",
                "product_category",
                "product_map",
                "repeat_call",
                "fcr",
                "churn_probability",
                "churn_reason",
                "customer_insight_summary",
                "standard_gsd_name",
            ],
        ),
        (
            "Service Quality",
            [
                "greeting_standard",
                "greeting_standard_reason",
                "manners",
                "manners_reason",
                "enthusiasm",
                "enthusiasm_reason",
                "communication_skill",
                "communication_skill_reason",
                "ending_standard",
                "ending_standard_reason",
                "data_privacy",
                "data_privacy_reason",
                "legal_verification",
                "legal_verification_reason",
                "company_verification",
                "company_verification_reason",
                "sla_notification",
                "sla_notification_reason",
                "transfer_standard",
                "transfer_standard_reason",
                "problem_understanding",
                "problem_understanding_reason",
                "compensation",
                "compensation_reason",
                "hold_standard",
                "hold_standard_reason",
                "wrap_up",
                "wrap_up_reason",
                "beyond_scope_support",
                "beyond_scope_support_reason",
                "self_service",
                "self_service_reason",
                "case_ownership",
                "case_ownership_reason",
                "contact_confirm",
                "contact_confirm_reason",
                "omotenashi",
                "omotenashi_reason",
                "retention",
                "retention_reason",
                "downsell",
                "downsell_reason",
                "mnp",
                "mnp_reason",
                "upselling",
                "upselling_reason",
                "service_quality_score",
                "service_quality_performance_insight",
            ],
        ),
        (
            "Sales Opportunities",
            [
                "opportunity_recognition_in_conversation",
                "product_suggested_by_ai",
                "agent_offer_product_presentation_&_explanation",
                "product_offer_by_agent",
                "sales_outcome_&_customer_decision",
                "sales_opportunities_performance_insight",
            ],
        ),
        (
            "Customer Sentiment",
            [
                "overall_sentiment",
                "initial_sentiment",
                "final_sentiment",
                "primary_sentiment_driver",
                "csat",
                "cs_performance_insight",
            ],
        ),
        (
            "Customer Experience",
            [
                "agent_communication_&_attitude",
                "agent_communication_&_attitude_reason",
                "agent_understanding_&_resolution",
                "agent_understanding_&_resolution_reason",
                "agent_responsiveness",
                "agent_responsiveness_reason",
                "system_accessibility",
                "system_accessibility_reason",
                "ivr_usability_&_design",
                "ivr_usability_&_design_reason",
                "ces",
                "self_service_readiness",
                "cx_performance_insight",
            ],
        ),
        (
            "Network",
            [
                "issue_type",
                "problem_statement",
                "area_tag_province",
                "area_tag_district",
                "area_tag_sub_district",
                "area_tag_landmark",
            ],
        ),
    ]

    GROUP_HEADER_DAILY_SCHEMA = [
        (
            "General",
            [
                "call_date",
                "agent_id",
                "call_id",
                "phone_number",
                "service_number",
                "call_duration_sec",
                "full_path",
                "department",
                "filename",
                "call_direction",
                "call_type",
                "call_type_confident",
            ],
        ),
        (
            "Customer Insight",
            [
                "summary_story",
                "product_category",
                "product_map",
                "repeat_call",
                "fcr",
                "churn_probability",
                "churn_reason",
                "customer_insight_summary",
                "standard_gsd_name",
            ],
        ),
        ("Service Quality", []),
        (
            "Sales Opportunities",
            [
                "opportunity_recognition_in_conversation",
                "product_suggested_by_ai",
                "agent_offer_product_presentation_&_explanation",
                "product_offer_by_agent",
                "sales_outcome_&_customer_decision",
                "sales_opportunities_performance_insight",
            ],
        ),
        (
            "Customer Sentiment",
            [
                "overall_sentiment",
                "initial_sentiment",
                "final_sentiment",
                "primary_sentiment_driver",
                "csat",
                "cs_performance_insight",
            ],
        ),
        (
            "Customer Experience",
            [
                "agent_communication_&_attitude",
                "agent_communication_&_attitude_reason",
                "agent_understanding_&_resolution",
                "agent_understanding_&_resolution_reason",
                "agent_responsiveness",
                "agent_responsiveness_reason",
                "system_accessibility",
                "system_accessibility_reason",
                "ivr_usability_&_design",
                "ivr_usability_&_design_reason",
                "ces",
                "self_service_readiness",
                "cx_performance_insight",
            ],
        ),
        (
            "Network",
            [
                "issue_type",
                "problem_statement",
                "area_tag_province",
                "area_tag_district",
                "area_tag_sub_district",
                "area_tag_landmark",
            ],
        ),
    ]

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        # Load parameters from configuration
        self.gcs = self.get_config("gcs", {})
        self.gcp = self.get_config("gcp", {})
        self.sharepoint = self.get_config("sharepoint", {})
        self.framework = self.get_config("framework", {})

        try:
            common_config = load_yaml(self.COMMON_CONFIG_PATH)
            self.common_framework = common_config.get("framework", {})
            self.verint_access = common_config.get("verint", {})
            self.control_access = common_config.get("control", {})
            self.sandbox_access = common_config.get("sandbox", {})
            self.gemini_cost_path = self.control_access.get("gemini_cost_path", None)

            self.timezone = self.common_framework.get("timezone", "Asia/Bangkok")
            self.msgraph_access = common_config.get("msgraph", {})

            self.project_id = resolve_env(self.gcs.get("project_id", ""))

            self.msgraph_sender_email = resolve_env(self.msgraph_access.get("sender_email"))
            self.msgraph_receiver_email = resolve_env(self.msgraph_access.get("receiver_email"))
            self.msgraph_cc_email = resolve_env(self.msgraph_access.get("cc_email"))

            input_folder_list_inbound = resolve_env(
                get_value_by_path(self.sharepoint, "verint.input_folder_list_inbound", "")
            )
            input_folder_list_outbound = resolve_env(
                get_value_by_path(self.sharepoint, "verint.input_folder_list_outbound", "")
            )

            self.input_folder_list_inbound = [folder.strip() for folder in input_folder_list_inbound.split(",")]
            self.input_folder_list_outbound = [folder.strip() for folder in input_folder_list_outbound.split(",")]
            self.combined_folder_list = list(set(self.input_folder_list_inbound + self.input_folder_list_outbound))

            self._cache_oper_log = {
                "transaction_df": pd.DataFrame(columns=self.TRANSACTION_LOG_SCHEMA),
                "process_date": [],
            }

        except Exception as e:
            logger.error(f"Failed to load common configuration from {self.COMMON_CONFIG_PATH}: {e}", exc_info=True)
            raise

    def pre_execute(self):
        """
        Pre-execution setup: Initialize modules and connections.
        """
        logger.info("Initializing modules")

        # Initialize SharePoint Verint module
        try:
            self.verint_site = resolve_env(self.verint_access.get("site_domain"))
            self.sharepoint_verint = SharePointModule(
                client_id=resolve_env(self.verint_access.get("client_id")),
                client_secret=resolve_env(self.verint_access.get("client_secret")),
                tenant_id=resolve_env(self.verint_access.get("tenant_id")),
                site_domain=self.verint_site,
                site_path=resolve_env(self.verint_access.get("site_path")),
            )
            logger.debug(f"SharePoint Verint: {self.verint_site}")
        except Exception as e:
            logger.error(f"Failed to initialize SharePoint Verint module: {e}", exc_info=True)
            raise

        # Initialize SharePoint Control module
        try:
            control_site = resolve_env(self.control_access.get("site_domain"))
            self.sharepoint_control = SharePointModule(
                client_id=resolve_env(self.control_access.get("client_id")),
                client_secret=resolve_env(self.control_access.get("client_secret")),
                tenant_id=resolve_env(self.control_access.get("tenant_id")),
                site_domain=control_site,
                site_path=resolve_env(self.control_access.get("site_path")),
            )
            logger.debug(f"SharePoint Control module initialized: {control_site}")
        except Exception as e:
            logger.error(f"Failed to initialize SharePoint Control module: {e}", exc_info=True)
            raise

        # Initialize SharePoint Sandbox module
        try:
            sandbox_site = resolve_env(self.sandbox_access.get("site_domain"))
            self.sharepoint_sandbox = SharePointModule(
                client_id=resolve_env(self.sandbox_access.get("client_id")),
                client_secret=resolve_env(self.sandbox_access.get("client_secret")),
                tenant_id=resolve_env(self.sandbox_access.get("tenant_id")),
                site_domain=sandbox_site,
                site_path=resolve_env(self.sandbox_access.get("site_path")),
            )
            logger.debug(f"SharePoint Sandbox module initialized: {sandbox_site}")
        except Exception as e:
            logger.error(f"Failed to initialize SharePoint Sandbox module: {e}", exc_info=True)
            raise

        # Initialize GCS module
        try:
            project_id = resolve_env(self.gcs.get("project_id"))
            bucket_name = resolve_env(self.gcs.get("bucket_name"))
            self.gcs_module = GCSModule(
                project_id=project_id,
                bucket_name=bucket_name,
            )
            logger.debug(f"GCS module initialized: {project_id}/{bucket_name}")
        except Exception as e:
            logger.error(f"Failed to initialize GCS module: {e}", exc_info=True)
            raise

        # Initialize Microsoft Graph client for email notifications
        try:
            self.msgraph_module = MSGraphModule(
                tenant_id=resolve_env(self.msgraph_access.get("tenant_id")),
                client_id=resolve_env(self.msgraph_access.get("client_id")),
                client_secret=resolve_env(self.msgraph_access.get("client_secret")),
            )
            logger.debug("Microsoft Graph client initialized for email notifications")
        except Exception as e:
            logger.error(f"Failed to initialize Microsoft Graph client: {e}", exc_info=True)
            raise

    def execute_task(self):
        execution_dt = self.get_package("execution_dt", None)
        batch_results = self.pre_result.get("batch_results", [])
        list_batchs = self.pre_result.get("list_batchs", [])
        failed_batches = self.pre_result.get("failed_batches", [])

        if not batch_results:
            logger.warning("No batch results to process")
            return None

        result = self._format_output(batch_results)
        result_df = pd.DataFrame(result)
        result_df = result_df.map(clean_invalid_xml_chars)
        grouped_series = result_df.groupby("call_month")["call_date"].unique()

        for yyyymm in sorted(grouped_series.keys()):
            new_master_df = result_df[(result_df["call_month"] == yyyymm)].copy()
            master_df = self._upload_master_file(new_master_df)
            master_df["call_date"] = master_df["call_date"].astype(str)

            for yyyymmdd in sorted(grouped_series[yyyymm]):
                daily_df = master_df[master_df["call_date"] == yyyymmdd].copy()

                daily_df = daily_df.drop_duplicates(
                    subset=["agent_id", "call_id", "phone_number", "department"], keep="last"
                )

                try:
                    data_dt = daily_df["call_date"].iloc[0]
                    daily_output_path = resolve_date(
                        resolve_env(self.sharepoint.get("verint").get("daily_output_file")), data_dt
                    )
                    network_daily_output_path = resolve_date(
                        resolve_env(self.sharepoint.get("sandbox").get("daily_output_file")), data_dt
                    )
                    logger.debug(f"Daily output path: {daily_output_path}")
                    logger.debug(f"Network daily output path: {network_daily_output_path}")
                except Exception as e:
                    logger.error(f"Failed to resolve output file path: {e}", exc_info=True)
                    raise Exception(f"Cannot determine output file path: {e}") from e

                self._upload_daily_files(daily_df, daily_output_path, self.sharepoint_verint)
                self._upload_daily_files(daily_df, network_daily_output_path, self.sharepoint_sandbox)
        logger.info("Starting for achieving prediction results")
        sep_date = {}
        after_upload_list = []

        for record in batch_results:
            try:
                key = get_value_by_path(record, "file_metadata.record_date", self.DEFAULT_RECORD_DATE)
                if key not in sep_date:
                    sep_date[key] = []
                    sep_date[key].append(record)
                else:
                    sep_date[key].append(record)

                # Mock full path from GCS storage path
                gcs_full_path = get_value_by_path(record, "file_metadata.file_uri", None)
                if gcs_full_path:
                    gcs_full_path = gcs_full_path.split(
                        "/"
                    )  # gs://nprd-sentiment-qa-enquiry-bucket/sentiment_qa/processing/voice/202603/20260322/Complain/9156447740xxxxxxx_xxxxxxx_xxxxxxx_xxxxxxx_firstname_lastname_T_20260220_208.wav
                    gcs_path = f"{gcs_full_path[-2]}/{gcs_full_path[-4]}/{gcs_full_path[-3]}/{gcs_full_path[-1]}"
                    folder = os.path.dirname(gcs_path)
                else:
                    gcs_path = ""
                    folder = ""
                full_path = resolve_env("${QA_VERINT_ROOT}") + "/" + gcs_path

                # Prepare for archive and logging
                after_upload_dict = {
                    "file_name": get_value_by_path(record, "file_metadata.file_name", None)
                    + get_value_by_path(record, "file_metadata.file_ext", None),
                    "full_path": full_path,
                    "folder": folder,
                    "record_date": key,
                    "department": gcs_full_path[-2] if gcs_full_path else None,
                    "duration": get_value_by_path(record, "file_metadata.duration", None),
                    "token_input": get_value_by_path(record, "prediction.token_input", None),
                    "token_cached": get_value_by_path(record, "prediction.token_cached", None),
                    "token_output": get_value_by_path(record, "prediction.token_output", None),
                    "status": get_value_by_path(record, "prediction.status", None),
                    "message": get_value_by_path(record, "prediction.message", None),
                    "processed_time": get_value_by_path(record, "prediction.processed_time", None),
                    "create_time": get_value_by_path(record, "prediction.create_time", None),
                    "model_version": get_value_by_path(record, "prediction.model_version", None),
                    "load_dt": record.get("load_dt", None),
                }
                after_upload_list.append(after_upload_dict)
            except Exception as record_err:
                logger.error(f"Error processing record: {record_err}", exc_info=True)
                logger.warning("Skipping problematic record and continuing")
                continue

        logger.debug(f"Prepared {len(after_upload_list)} records for upload across {len(sep_date)} dates")

        try:
            tmp_output_file = resolve_env(self.sharepoint.get("verint").get("daily_output_file"))
            logger.debug(f"Output file template: {tmp_output_file}")
        except Exception as e:
            logger.error(f"Failed to resolve output file path: {e}", exc_info=True)
            raise Exception(f"Cannot determine output file path: {e}") from e

        try:
            df = pd.DataFrame(after_upload_list)
            logger.info(f"Created DataFrame with {len(df)} records for archival and logging")
            logger.debug(f"DataFrame columns: {list(df.columns)}")
        except Exception as e:
            logger.error(f"Failed to create DataFrame from upload list: {e}", exc_info=True)
            raise Exception(f"Cannot create DataFrame for archival: {e}") from e

        # Archive files
        try:
            logger.info("Starting file archival process")
            self._archive_files(execution_dt, df, list_batchs, failed_batches)
            logger.info("File archival completed successfully")
        except Exception as archive_err:
            logger.error(f"File archival failed: {archive_err}", exc_info=True)
            raise Exception(f"File archival process failed: {archive_err}") from archive_err

        # Insert transaction log records
        try:
            self._insert_log_record(df)
        except Exception as log_err:
            logger.error(f"Failed to insert transaction log records: {log_err}", exc_info=True)
            raise Exception(f"Critical error: Transaction log insertion failed: {log_err}") from log_err

        return batch_results, result_df

    def _format_output(self, raw_result: list[dict]) -> list[dict]:
        logger.info(f"Starting output formatting for {len(raw_result)} records")

        # Validate input
        if not raw_result:
            logger.warning("No raw results provided for formatting")
            return []

        result = []
        updated_dt = get_current_datetime().strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(f"Using updated_dt: {updated_dt}")

        success_count = 0
        failed_count = 0
        process_files = []

        try:
            user_config_path = resolve_env(get_value_by_path(self.sharepoint, "control.user_config_path"))
            user_config_content = self.sharepoint_control.get_item_by_path(user_config_path)
            with io.BytesIO(user_config_content.content) as user_config_file_buffer:
                weight_score_df = pd.read_excel(user_config_file_buffer, sheet_name="weight_score")
            with io.BytesIO(user_config_content.content) as user_config_file_buffer:
                product_mapping_df = pd.read_excel(user_config_file_buffer, sheet_name="product_mapping")
        except Exception as e:
            logger.error(f"Failed to fetch user config file from SharePoint: {e}", exc_info=True)
            user_config_path = resolve_env(self.framework.get("user_config_path"))
            logger.info(f"Read file from config: {user_config_path}")
            weight_score_df = read_xlsx(user_config_path, sheet_name="weight_score")
            product_mapping_df = read_xlsx(user_config_path, sheet_name="product_mapping")

        product_mapping_df["product_group"] = product_mapping_df["product_group"].ffill()
        product_mapping_df["product_group"] = product_mapping_df["product_group"].str.strip()
        product_mapping_df["product_category"] = product_mapping_df["product_category"].str.strip()

        for idx, rec in enumerate(raw_result, 1):
            try:
                filename = get_value_by_path(rec, "file_metadata.file_name", f"unknown_{idx}") + get_value_by_path(
                    rec, "file_metadata.file_ext", ""
                )
                call_month = safe_cast_value(
                    get_value_by_path(rec, "file_metadata.record_date", self.DEFAULT_RECORD_DATE)[:6], str
                )
                call_date = safe_cast_value(
                    get_value_by_path(rec, "file_metadata.record_date", self.DEFAULT_RECORD_DATE), str
                )
                agent_id = safe_cast_value(get_value_by_path(rec, "file_metadata.agent_id", None), str)
                call_id = safe_cast_value(get_value_by_path(rec, "file_metadata.call_id", None), str)
                phone_number = safe_cast_value(get_value_by_path(rec, "file_metadata.phone_number", None), str)
                file_uri = safe_cast_value(get_value_by_path(rec, "file_metadata.file_uri", None), str).split("/")
                full_path = (
                    f"Input/{safe_list_get(file_uri, 8, '')}/{safe_list_get(file_uri, 6, '')}/"
                    f"{safe_list_get(file_uri, 7, '')}/{filename}"
                )
                department = safe_list_get(file_uri, 8, "")
                process_files.append(filename)
                logger.debug(f"Processing record [{idx}/{len(raw_result)}]: {filename}")

                # Determine status
                try:
                    prediction_status = get_value_by_path(rec, "prediction.status")

                    if prediction_status != "SUCCESS":
                        status = prediction_status
                        raw_message = get_value_by_path(rec, "prediction.message")
                        # Clean message by replacing newlines for CSV compatibility
                        message = raw_message.replace("\n", " ").replace("\r", " ") if raw_message else None
                    else:
                        status = "SUCCESS"
                        message = None

                    logger.debug(f"Record [{idx}]: Status={status}, Message={message}")
                except Exception as status_err:
                    logger.error(
                        f"Record [{idx}] ({filename}): Failed to determine status: {status_err}", exc_info=True
                    )
                    raise

                # Build customer information
                try:
                    duration_val = get_value_by_path(rec, "file_metadata.duration", None)
                    if duration_val is None:
                        logger.warning(f"Record [{idx}] ({filename}): file_metadata.duration is None")

                    call_direction = safe_cast_value(get_value_by_path(rec, "file_metadata.call_direction", None), str)
                    customer_information = {
                        "filename": filename,
                        "agent_id": agent_id,
                        "call_id": call_id,
                        "phone_number": phone_number,
                        "service_number": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.service_number", None), str
                        ),
                        "call_duration_sec": safe_cast_value(duration_val, int),
                        "full_path": full_path,
                        "department": department,
                        "call_direction": call_direction,
                        "call_type": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.call_type", None), str
                        ),
                        "call_type_confident": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.call_type_confident", None), str
                        ),
                        "call_month": call_month,
                        "call_date": call_date,
                        "full_name": (
                            f"{get_value_by_path(rec, 'file_metadata.first_name', None)} "
                            f"{get_value_by_path(rec, 'file_metadata.last_name', None)}"
                        ).strip(),
                        "true_dtac": "Dtac"
                        if get_value_by_path(rec, "file_metadata.provider", None) in ["D", "d"]
                        else "True",
                    }
                    logger.debug(f"Record [{idx}]: Built customer information")
                except Exception as customer_err:
                    logger.error(
                        f"Record [{idx}] ({filename}): Failed to build customer information: {customer_err}",
                        exc_info=True,
                    )
                    raise

                # Build customer insight
                try:
                    product_category_str = safe_cast_value(
                        get_value_by_path(rec, "prediction.raw_prediction.customer_insight.product_category", None), str
                    )
                    product_list = [product.strip() for product in product_category_str.split(",")]
                    matched_rows = product_mapping_df[
                        product_mapping_df["product_category"].str.lower() == product_list[0].lower()
                    ]
                    product_map = matched_rows["product_group"].values[0] if not matched_rows.empty else "Other"

                    customer_insight = {
                        "summary_story": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.customer_insight.summary_story", None),
                            str,
                        ),
                        "product_category": product_category_str,
                        "product_map": product_map,
                        "repeat_call": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.customer_insight.repeat_call", None), str
                        ),
                        "fcr": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.customer_insight.fcr", None), str
                        ),
                        "churn_probability": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_insight.churn_probability", None
                            ),
                            str,
                        ),
                        "churn_reason": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.customer_insight.churn_reason", None), str
                        ),
                        "customer_insight_summary": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_insight.customer_insight_summary", None
                            ),
                            str,
                        ),
                        "standard_gsd_name": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_insight.standard_gsd_name", None
                            ),
                            str,
                        ),
                    }

                    if customer_insight["fcr"] == "True" or customer_insight["fcr"] == "true":
                        customer_insight["fcr"] = "Y"
                    elif customer_insight["fcr"] == "False" or customer_insight["fcr"] == "false":
                        customer_insight["fcr"] = "N"
                    else:
                        customer_insight["fcr"] = "error"
                    logger.debug(f"Record [{idx}]: Built customer insight")
                except Exception as insight_err:
                    logger.error(
                        f"Record [{idx}] ({filename}): Failed to build customer insight: {insight_err}", exc_info=True
                    )
                    raise

                # Build service quality
                try:
                    service_quality = {
                        "greeting_standard": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.greeting_standard.evaluation", None
                            ),
                            str,
                        ),
                        "greeting_standard_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.greeting_standard.reason", None
                            ),
                            str,
                        ),
                        "manners": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.manners.evaluation", None
                            ),
                            str,
                        ),
                        "manners_reason": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.service_quality.manners.reason", None),
                            str,
                        ),
                        "enthusiasm": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.enthusiasm.evaluation", None
                            ),
                            str,
                        ),
                        "enthusiasm_reason": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.service_quality.enthusiasm.reason", None),
                            str,
                        ),
                        "communication_skill": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.communication_skill.evaluation", None
                            ),
                            str,
                        ),
                        "communication_skill_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.communication_skill.reason", None
                            ),
                            str,
                        ),
                        "ending_standard": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.ending_standard.evaluation", None
                            ),
                            str,
                        ),
                        "ending_standard_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.ending_standard.reason", None
                            ),
                            str,
                        ),
                        "data_privacy": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.data_privacy.evaluation", None
                            ),
                            str,
                        ),
                        "data_privacy_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.data_privacy.reason", None
                            ),
                            str,
                        ),
                        "legal_verification": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.legal_verification.evaluation", None
                            ),
                            str,
                        ),
                        "legal_verification_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.legal_verification.reason", None
                            ),
                            str,
                        ),
                        "company_verification": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.customer_verification.evaluation", None
                            ),
                            str,
                        ),
                        "company_verification_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.customer_verification.reason", None
                            ),
                            str,
                        ),
                        "sla_notification": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.sla_notification.evaluation", None
                            ),
                            str,
                        ),
                        "sla_notification_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.sla_notification.reason", None
                            ),
                            str,
                        ),
                        "transfer_standard": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.transfer_standard.evaluation", None
                            ),
                            str,
                        ),
                        "transfer_standard_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.transfer_standard.reason", None
                            ),
                            str,
                        ),
                        "problem_understanding": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.problem_understanding.evaluation", None
                            ),
                            str,
                        ),
                        "problem_understanding_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.problem_understanding.reason", None
                            ),
                            str,
                        ),
                        "compensation": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.compensation.evaluation", None
                            ),
                            str,
                        ),
                        "compensation_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.compensation.reason", None
                            ),
                            str,
                        ),
                        "hold_standard": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.hold_standard.evaluation", None
                            ),
                            str,
                        ),
                        "hold_standard_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.hold_standard.reason", None
                            ),
                            str,
                        ),
                        "wrap_up": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.wrap_up.evaluation", None
                            ),
                            str,
                        ),
                        "wrap_up_reason": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.service_quality.wrap_up.reason", None),
                            str,
                        ),
                        "beyond_scope_support": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.beyond_scope_support.evaluation", None
                            ),
                            str,
                        ),
                        "beyond_scope_support_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.beyond_scope_support.reason", None
                            ),
                            str,
                        ),
                        "self_service": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.true_application.evaluation", None
                            ),
                            str,
                        ),
                        "self_service_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.true_application.reason", None
                            ),
                            str,
                        ),
                        "case_ownership": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.case_ownership.evaluation", None
                            ),
                            str,
                        ),
                        "case_ownership_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.case_ownership.reason", None
                            ),
                            str,
                        ),
                        "contact_confirm": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.contact_confirm.evaluation", None
                            ),
                            str,
                        ),
                        "contact_confirm_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.contact_confirm.reason", None
                            ),
                            str,
                        ),
                        "omotenashi": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.omotenashi.evaluation", None
                            ),
                            str,
                        ),
                        "omotenashi_reason": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.service_quality.omotenashi.reason", None),
                            str,
                        ),
                        "retention": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.retention.evaluation", None
                            ),
                            str,
                        ),
                        "retention_reason": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.service_quality.retention.reason", None),
                            str,
                        ),
                        "downsell": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.downsell.evaluation", None
                            ),
                            str,
                        ),
                        "downsell_reason": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.service_quality.downsell.reason", None),
                            str,
                        ),
                        "mnp": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.service_quality.mnp.evaluation", None),
                            str,
                        ),
                        "mnp_reason": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.service_quality.mnp.reason", None), str
                        ),
                        "upselling": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.service_quality.upselling.evaluation", None
                            ),
                            str,
                        ),
                        "upselling_reason": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.service_quality.upselling.reason", None),
                            str,
                        ),
                        "service_quality_performance_insight": safe_cast_value(
                            get_value_by_path(
                                rec,
                                "prediction.raw_prediction.service_quality.service_quality_performance_insight",
                                None,
                            ),
                            str,
                        ),
                    }

                    total_weight = 0
                    total_score = 0
                    for key, value in service_quality.items():
                        if key in weight_score_df["item"].values:
                            if call_direction == "IN":
                                weight = weight_score_df.loc[weight_score_df["item"] == key, "inbound_weight"].values[0]
                            elif call_direction == "OUT":
                                weight = weight_score_df.loc[weight_score_df["item"] == key, "outbound_weight"].values[
                                    0
                                ]
                            else:
                                logger.error(
                                    f"Invalid call direction '{call_direction}' extracted from filename "
                                    f"'{filename}'. Expected 'IN' or 'OUT'."
                                )
                                weight = 0

                            if value.lower() == "meet":
                                total_score += weight
                                total_weight += weight
                            elif value.lower() == "below":
                                total_weight += weight
                            else:
                                continue
                    service_quality["service_quality_score"] = (
                        f"{(total_score / total_weight) * 100:.2f}" if total_weight > 0 else None
                    )

                    logger.debug(f"Record [{idx}]: Built service quality")
                except Exception as service_err:
                    logger.error(
                        f"Record [{idx}] ({filename}): Failed to build service quality: {service_err}", exc_info=True
                    )
                    raise

                # Build sale opportunity
                try:
                    sale_opportunity = {
                        "opportunity_recognition_in_conversation": safe_cast_value(
                            get_value_by_path(
                                rec,
                                "prediction.raw_prediction.sale_opportunity.opportunity_recognition_in_conversation",
                                None,
                            ),
                            str,
                        ),
                        "product_suggested_by_ai": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.sale_opportunity.product_suggested_by_ai", None
                            ),
                            str,
                        ),
                        "agent_offer_product_presentation_&_explanation": safe_cast_value(
                            get_value_by_path(
                                rec,
                                "prediction.raw_prediction.sale_opportunity.agent_offer_product_presentation_&_explanation",
                                None,
                            ),
                            str,
                        ),
                        "product_offer_by_agent": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.sale_opportunity.product_offer_by_agent", None
                            ),
                            str,
                        ),
                        "sales_outcome_&_customer_decision": safe_cast_value(
                            get_value_by_path(
                                rec,
                                "prediction.raw_prediction.sale_opportunity.sales_outcome_&_customer_decision",
                                None,
                            ),
                            str,
                        ),
                        "sales_opportunities_performance_insight": safe_cast_value(
                            get_value_by_path(
                                rec,
                                "prediction.raw_prediction.sale_opportunity.sales_opportunities_performance_insight",
                                None,
                            ),
                            str,
                        ),
                    }
                    if (
                        sale_opportunity["opportunity_recognition_in_conversation"] == "True"
                        or sale_opportunity["opportunity_recognition_in_conversation"] == "true"
                    ):
                        sale_opportunity["opportunity_recognition_in_conversation"] = "Y"
                    elif (
                        sale_opportunity["opportunity_recognition_in_conversation"] == "False"
                        or sale_opportunity["opportunity_recognition_in_conversation"] == "false"
                    ):
                        sale_opportunity["opportunity_recognition_in_conversation"] = "N"
                    else:
                        sale_opportunity["opportunity_recognition_in_conversation"] = "error"

                    if (
                        sale_opportunity["agent_offer_product_presentation_&_explanation"] == "True"
                        or sale_opportunity["agent_offer_product_presentation_&_explanation"] == "true"
                    ):
                        sale_opportunity["agent_offer_product_presentation_&_explanation"] = "Y"
                    elif (
                        sale_opportunity["agent_offer_product_presentation_&_explanation"] == "False"
                        or sale_opportunity["agent_offer_product_presentation_&_explanation"] == "false"
                    ):
                        sale_opportunity["agent_offer_product_presentation_&_explanation"] = "N"
                    else:
                        sale_opportunity["agent_offer_product_presentation_&_explanation"] = "error"

                    if (
                        sale_opportunity["sales_outcome_&_customer_decision"] == "True"
                        or sale_opportunity["sales_outcome_&_customer_decision"] == "true"
                    ):
                        sale_opportunity["sales_outcome_&_customer_decision"] = "Y"
                    elif (
                        sale_opportunity["sales_outcome_&_customer_decision"] == "False"
                        or sale_opportunity["sales_outcome_&_customer_decision"] == "false"
                    ):
                        sale_opportunity["sales_outcome_&_customer_decision"] = "N"
                    else:
                        sale_opportunity["sales_outcome_&_customer_decision"] = "error"

                    logger.debug(f"Record [{idx}]: Built sale opportunity")
                except Exception as sale_err:
                    logger.error(
                        f"Record [{idx}] ({filename}): Failed to build sale opportunity: {sale_err}", exc_info=True
                    )
                    raise

                # Build customer sentiment
                try:
                    customer_sentiment = {
                        "overall_sentiment": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_sentiment.overall_sentiment", None
                            ),
                            str,
                        ),
                        "initial_sentiment": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_sentiment.initial_sentiment", None
                            ),
                            str,
                        ),
                        "final_sentiment": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_sentiment.final_sentiment", None
                            ),
                            str,
                        ),
                        "primary_sentiment_driver": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_sentiment.primary_sentiment_driver", None
                            ),
                            str,
                        ),
                        "csat": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.customer_sentiment.csat", None), str
                        ),
                        "cs_performance_insight": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_sentiment.cs_performance_insight", None
                            ),
                            str,
                        ),
                    }
                    logger.debug(f"Record [{idx}]: Built customer sentiment")
                except Exception as sentiment_err:
                    logger.error(
                        f"Record [{idx}] ({filename}): Failed to build customer sentiment: {sentiment_err}",
                        exc_info=True,
                    )
                    raise

                # Build customer experience
                try:
                    customer_experience = {
                        "agent_communication_&_attitude": safe_cast_value(
                            get_value_by_path(
                                rec,
                                "prediction.raw_prediction.customer_experience.agent_communication_&_attitude",
                                None,
                            ),
                            str,
                        ),
                        "agent_communication_&_attitude_reason": safe_cast_value(
                            get_value_by_path(
                                rec,
                                "prediction.raw_prediction.customer_experience.agent_communication_&_attitude_reason",
                                None,
                            ),
                            str,
                        ),
                        "agent_understanding_&_resolution": safe_cast_value(
                            get_value_by_path(
                                rec,
                                "prediction.raw_prediction.customer_experience.agent_understanding_&_resolution",
                                None,
                            ),
                            str,
                        ),
                        "agent_understanding_&_resolution_reason": safe_cast_value(
                            get_value_by_path(
                                rec,
                                "prediction.raw_prediction.customer_experience.agent_understanding_&_resolution_reason",
                                None,
                            ),
                            str,
                        ),
                        "agent_responsiveness": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_experience.agent_responsiveness", None
                            ),
                            str,
                        ),
                        "agent_responsiveness_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_experience.agent_responsiveness_reason", None
                            ),
                            str,
                        ),
                        "system_accessibility": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_experience.system_accessibility", None
                            ),
                            str,
                        ),
                        "system_accessibility_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_experience.system_accessibility_reason", None
                            ),
                            str,
                        ),
                        "ivr_usability_&_design": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_experience.ivr_usability_&_design", None
                            ),
                            str,
                        ),
                        "ivr_usability_&_design_reason": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_experience.ivr_usability_&_design_reason", None
                            ),
                            str,
                        ),
                        "ces": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.customer_experience.ces", None), str
                        ),
                        "self_service_readiness": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_experience.self_service_readiness", None
                            ),
                            str,
                        ),
                        "cx_performance_insight": safe_cast_value(
                            get_value_by_path(
                                rec, "prediction.raw_prediction.customer_experience.cx_performance_insight", None
                            ),
                            str,
                        ),
                    }

                    if (
                        customer_experience["system_accessibility"] == "True"
                        or customer_experience["system_accessibility"] == "true"
                    ):
                        customer_experience["system_accessibility"] = "Y"
                    elif (
                        customer_experience["system_accessibility"] == "False"
                        or customer_experience["system_accessibility"] == "false"
                    ):
                        customer_experience["system_accessibility"] = "N"
                    else:
                        customer_experience["system_accessibility"] = "error"

                    if (
                        customer_experience["ivr_usability_&_design"] == "True"
                        or customer_experience["ivr_usability_&_design"] == "true"
                    ):
                        customer_experience["ivr_usability_&_design"] = "Y"
                    elif (
                        customer_experience["ivr_usability_&_design"] == "False"
                        or customer_experience["ivr_usability_&_design"] == "false"
                    ):
                        customer_experience["ivr_usability_&_design"] = "N"
                    else:
                        customer_experience["ivr_usability_&_design"] = "error"

                    logger.debug(f"Record [{idx}]: Built customer experience")
                except Exception as experience_err:
                    logger.error(
                        f"Record [{idx}] ({filename}): Failed to build customer experience: {experience_err}",
                        exc_info=True,
                    )
                    raise

                # Build network
                try:
                    network = {
                        "issue_type": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.network.issue_type", None), str
                        ),
                        "problem_statement": ", ".join(
                            get_value_by_path(rec, "prediction.raw_prediction.network.problem_statement", None)
                        ),  # safe_cast_value(get_value_by_path(rec, "prediction.raw_prediction.network.problem_statement", None), str),  # noqa: E501
                        "area_tag_province": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.network.area_tag_province", None), str
                        ),
                        "area_tag_district": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.network.area_tag_district", None), str
                        ),
                        "area_tag_sub_district": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.network.area_tag_sub_district", None), str
                        ),
                        "area_tag_landmark": safe_cast_value(
                            get_value_by_path(rec, "prediction.raw_prediction.network.area_tag_landmark", None), str
                        ),
                    }
                    logger.debug(f"Record [{idx}]: Built network")
                except Exception as network_err:
                    logger.error(f"Record [{idx}] ({filename}): Failed to build network: {network_err}", exc_info=True)
                    raise

                # Combine all results
                try:
                    result.append(
                        {
                            **customer_information,
                            **customer_insight,
                            **service_quality,
                            **sale_opportunity,
                            **customer_sentiment,
                            **customer_experience,
                            **network,
                            "updated_dt": updated_dt,
                            "status": status,
                            "error_code": message,
                        }
                    )
                    success_count += 1
                    logger.debug(f"Record [{idx}] ({filename}): Successfully formatted")
                except Exception as combine_err:
                    logger.error(
                        f"Record [{idx}] ({filename}): Failed to combine results: {combine_err}", exc_info=True
                    )
                    raise

            except Exception as record_err:
                # Clean error message by replacing newlines with spaces for better CSV compatibility
                error_msg = str(record_err).replace("\n", " ").replace("\r", " ")

                result.append(
                    {
                        "filename": filename,
                        "agent_id": agent_id,
                        "call_id": call_id,
                        "phone_number": phone_number,
                        "full_path": full_path,
                        "department": department,
                        "call_month": call_month,
                        "call_date": call_date,
                        "updated_dt": updated_dt,
                        "status": "FAILED",
                        "error_code": error_msg,
                    }
                )
                failed_count += 1
                logger.error(f"Record [{idx}]: Failed to format output: {record_err}", exc_info=True)
                logger.warning(f"Skipping problematic record [{idx}] and continuing")
                continue

        # Process batch processing log for failed files
        try:
            log_path = resolve_env(get_value_by_path(self.sharepoint, "control.batch_processing_log_file"))
            logger.debug(f"Checking batch processing log at: {log_path}")

            if self.sharepoint_control.is_item_exists(item_path=log_path):
                logger.info("Found batch processing log, checking for failed files")
                try:
                    batch_processing_log = self.sharepoint_control.get_item_by_path(log_path)
                    with io.BytesIO(batch_processing_log.content) as log_buffer:
                        batch_processing_log_df = pd.read_csv(log_buffer)

                    logger.debug(f"Batch processing log has {len(batch_processing_log_df)} total records")

                    batch_processing_log_df = (
                        batch_processing_log_df.sort_values(
                            by=["filename", "updated_dt"], ascending=[True, False]
                        ).drop_duplicates(subset=["filename"], keep="first")
                    )[["batch_job_id", "batch_job_display_name", "filename", "status", "error_message", "updated_dt"]]
                    batch_job_detail_df = batch_processing_log_df[
                        batch_processing_log_df["filename"].isin(process_files)
                    ][["batch_job_id", "batch_job_display_name"]].drop_duplicates()
                    logger.debug(f"Found {len(batch_job_detail_df)} unique batch jobs for processed files")

                    failed_files_df = batch_processing_log_df.merge(
                        batch_job_detail_df, on=["batch_job_id", "batch_job_display_name"], how="inner"
                    )
                    failed_files_df = failed_files_df[failed_files_df["status"] == "FAILED"]

                    logger.info(f"Found {len(failed_files_df)} failed files in batch processing log")

                    appended_count = 0
                    for failed_file in failed_files_df.itertuples():
                        try:
                            # Clean error message by replacing newlines for CSV compatibility
                            clean_error = (
                                failed_file.error_message.replace("\n", " ").replace("\r", " ")
                                if failed_file.error_message
                                else None
                            )

                            result.append(
                                {
                                    "filename": failed_file.filename,
                                    "summary": None,
                                    "updated_dt": updated_dt,
                                    "status": "FAILED",
                                    "error_code": clean_error,
                                }
                            )
                            appended_count += 1
                            failed_count += 1
                        except Exception as append_err:
                            # Clean error message
                            clean_append_err = str(append_err).replace("\n", " ").replace("\r", " ")

                            result.append(
                                {
                                    "filename": None,
                                    "summary": None,
                                    "updated_dt": updated_dt,
                                    "status": "FAILED",
                                    "error_code": f"Unknown error during log append {clean_append_err}",
                                }
                            )
                            logger.error(f"Failed to append failed file from log: {append_err}", exc_info=True)
                            failed_count += 1
                            continue

                    logger.info(f"Appended {appended_count} failed files from batch processing log")

                except Exception as log_err:
                    logger.error(f"Failed to process batch processing log for failed files: {log_err}", exc_info=True)
                    logger.warning("Skipping appending failed files from batch processing log")
            else:
                logger.warning(f"Batch processing log not found at: {log_path}")
        except Exception as log_path_err:
            logger.error(f"Failed to resolve batch processing log path: {log_path_err}", exc_info=True)
            logger.warning("Skipping batch processing log check")

        logger.info(
            f"Output formatting complete: {success_count} successful, {failed_count} failed, "
            f"total results: {len(result)}"
        )

        if not result:
            logger.error("No records were successfully formatted - all records failed")
            raise Exception("Failed to format any output records")

        if failed_count > 0:
            logger.warning(
                f"{failed_count} record(s) failed to format but continuing with {success_count} successful records"
            )
            logger.warning(f"Success rate: {(success_count / (success_count + failed_count) * 100):.2f}%")
        else:
            logger.info("All records formatted successfully")

        return result

    def _archive_files(
        self, execution_dt: datetime, df: pd.DataFrame, list_batchs: list[str], failed_batches: list[str]
    ):
        logger.info("Starting file archival process")
        logger.debug(f"Processing {len(df['record_date'].unique())} unique record dates")

        for record_date in df["record_date"].unique():
            # Archive successfully processed voice files
            logger.info(f"Archiving successfully processed voice files for record date: {record_date}")

            for product in self.combined_folder_list:
                archived_success_count = 0
                archived_skipped_count = 0
                deleted_success_count = 0
                deleted_skipped_count = 0

                try:
                    partition_df = df[df["record_date"] == record_date]
                    success_records = partition_df[
                        (partition_df["status"] == "SUCCESS") & (partition_df["department"] == product)
                    ].drop_duplicates(subset=["file_name"])
                    logger.info(f"Found {len(success_records)} successfully processed files to archive")
                except Exception as filter_err:
                    logger.error(f"Failed to filter success records: {filter_err}", exc_info=True)
                    logger.warning("Skipping voice file archival for this partition")
                    continue

                # Resolve folder paths once (non-critical errors - continue with empty lists)
                try:
                    processing_voice_folder = resolve_date(
                        text=self.gcs["processing_voice_folder"],
                        replace_date=record_date,
                    )
                    archive_voice_folder = resolve_date(
                        text=self.gcs["archive_voice_folder"],
                        replace_date=record_date,
                    )
                    input_voice_folder = resolve_date(
                        text=self.gcs["input_folder"],
                        replace_date=record_date,
                    )

                    processing_voice_folder = processing_voice_folder.replace("{PRODUCT}", product)
                    archive_voice_folder = archive_voice_folder.replace("{PRODUCT}", product)
                    input_voice_folder = input_voice_folder.replace("{PRODUCT}", product)

                    logger.debug(f"Processing folder: {processing_voice_folder}")
                    logger.debug(f"Archive folder: {archive_voice_folder}")
                    logger.debug(f"Input folder: {input_voice_folder}")
                except Exception as path_err:
                    logger.error(f"Failed to resolve folder paths: {path_err}", exc_info=True)
                    logger.warning("Skipping voice file archival due to path resolution error")
                    continue

                # List files in folders (non-critical - continue with empty lists)
                try:
                    if self.gcs_module.is_dir_exists(dir_path=processing_voice_folder):
                        files_in_processing = self.gcs_module.list_files(prefix=processing_voice_folder)
                        files_in_processing = [os.path.basename(f) for f in files_in_processing]
                        logger.info(f"Found {len(files_in_processing)} files in processing folder")
                    else:
                        files_in_processing = []
                        logger.warning(f"Processing voice folder does not exist: {processing_voice_folder}")
                except Exception as list_err:
                    logger.error(f"Error listing processing folder: {list_err}", exc_info=True)
                    files_in_processing = []

                try:
                    if self.gcs_module.is_dir_exists(dir_path=input_voice_folder):
                        files_in_input = self.gcs_module.list_files(prefix=input_voice_folder)
                        files_in_input = [os.path.basename(f) for f in files_in_input]
                        logger.info(f"Found {len(files_in_input)} files in input folder")
                    else:
                        files_in_input = []
                        logger.warning(f"Input voice folder does not exist: {input_voice_folder}")
                except Exception as list_err:
                    logger.error(f"Error listing input folder: {list_err}", exc_info=True)
                    files_in_input = []

                for success_record in success_records.itertuples():
                    # Archive from processing folder
                    try:
                        logger.debug(f"Archiving voice file: {success_record.file_name}")
                        logger.debug(f"Files in processing folder: {files_in_processing}")
                        if success_record.file_name in files_in_processing:
                            self.gcs_module.move_file(
                                source_path=f"{processing_voice_folder}/{success_record.file_name}",
                                destination_path=f"{archive_voice_folder}/{success_record.file_name}",
                            )
                            archived_success_count += 1
                            logger.debug(f"Archived voice file: {success_record.file_name}")
                        else:
                            logger.warning(f"Voice file not found in processing folder: {success_record.file_name}")
                            archived_skipped_count += 1
                    except Exception as archive_err:
                        logger.error(
                            f"Failed to archive voice file '{success_record.file_name}': {archive_err}", exc_info=True
                        )
                        archived_skipped_count += 1

                    # Delete from input folder
                    try:
                        if success_record.file_name in files_in_input:
                            self.gcs_module.delete_file(file_path=f"{input_voice_folder}/{success_record.file_name}")
                            deleted_success_count += 1
                            logger.debug(f"Deleted voice file from input folder: {success_record.file_name}")
                        else:
                            logger.debug(f"Voice file not found in input folder: {success_record.file_name}")
                            deleted_skipped_count += 1
                    except Exception as delete_err:
                        logger.error(
                            f"Failed to delete voice file '{success_record.file_name}' from input folder: {delete_err}",
                            exc_info=True,
                        )
                        deleted_skipped_count += 1

                logger.info(
                    f"Voice file '{product}' archival complete. Archived: {archived_success_count}, "
                    f"Skipped: {archived_skipped_count}"
                )
                logger.info(
                    f"Voice file '{product}' deletion from input complete. Deleted: {deleted_success_count}, "
                    f"Skipped: {deleted_skipped_count}"
                )

        # Archive batch files
        logger.info("Archiving processed batch files")
        if failed_batches:
            logger.warning(f"Excluding {len(failed_batches)} failed batches from archival")
            for batch in failed_batches:
                logger.warning(f"Failed batch: {batch}")

        # Filter out failed batches from archive list
        archive_batch_list = [batch for batch in list_batchs if batch not in failed_batches]
        logger.info(f"Archiving {len(archive_batch_list)} processed batch files...")

        archived_batch_count = 0
        failed_archive_count = 0
        output_folders = set()  # Collect unique output folders

        for idx, batch in enumerate(archive_batch_list, 1):
            try:
                try:
                    archive_batch_folder = resolve_date(
                        text=self.gcs["archive_batch_folder"],
                        replace_date=execution_dt.strftime("%Y-%m-%d"),
                    )
                except Exception as path_err:
                    logger.error(f"Failed to resolve archive path for batch {batch}: {path_err}")
                    failed_archive_count += 1
                    continue

                # Extract last two path components for archive structure
                try:
                    parts = Path(batch).parts
                    batch_path = "/".join(parts[-2:])
                    destination = f"{archive_batch_folder}/{batch_path}"
                except Exception as path_parse_err:
                    logger.error(f"Failed to parse batch path '{batch}': {path_parse_err}")
                    failed_archive_count += 1
                    continue

                logger.debug(f"Archiving batch [{idx}/{len(archive_batch_list)}]: {batch} -> {destination}")
                self.gcs_module.move_file(
                    source_path=batch,
                    destination_path=destination,
                )
                archived_batch_count += 1
                logger.info(f"Archived batch [{idx}/{len(archive_batch_list)}]: {batch_path}")

                # Collect output folder for later cleanup
                try:
                    output_folders.add(os.path.dirname(batch))
                except Exception as folder_err:
                    logger.debug(f"Could not extract folder from batch path: {folder_err}")

            except Exception as batch_archive_err:
                logger.error(f"Failed to archive batch file '{batch}': {batch_archive_err}", exc_info=True)
                failed_archive_count += 1
                continue

        logger.info(f"Batch archival complete. Archived: {archived_batch_count}, Failed: {failed_archive_count}")

        logger.info("Cleaning up processed output directories")

        deleted_output_count = 0
        failed_delete_count = 0
        logger.info(f"Checking {len(output_folders)} output folders for cleanup...")

        for output_folder in output_folders:
            try:
                if not self.gcs_module.is_dir_exists(dir_path=output_folder):
                    logger.debug(f"Output folder does not exist, skipping: {output_folder}")
                    deleted_output_count += 1  # Count as success since it's already gone
                    continue

                logger.info(f"Cleaning up output folder: {output_folder}")
                self.gcs_module.delete_dir(dir_path=output_folder)
                deleted_output_count += 1
                logger.info(f"Cleaned up output folder: {output_folder}")

            except Exception as delete_folder_err:
                logger.error(f"Failed to cleanup output folder '{output_folder}': {delete_folder_err}", exc_info=True)
                failed_delete_count += 1

        logger.info(
            f"Output directory cleanup complete. Cleaned: {deleted_output_count}, Skipped: {failed_delete_count}"
        )

        # Final summary
        if failed_batches:
            logger.warning(f"Retrieve Batch Task completed with {len(failed_batches)} failed batches")
        else:
            logger.info("Retrieve Batch Task completed successfully")

    def _insert_log_record(self, df: pd.DataFrame):
        logger.info("Starting log record insertion")
        default_type = "AI Classification"
        default_user_id = "daisyrpa"
        default_source = "SharePoint"

        try:
            unique_dates = df["record_date"].unique()
            logger.info(f"Processing log records for {len(unique_dates)} unique dates")
        except Exception as e:
            logger.error(f"Failed to extract unique record dates: {e}", exc_info=True)
            raise Exception(f"Cannot process log records: {e}") from e

        unique_dates.sort()
        for record_date in unique_dates:
            try:
                prediction_df = df[df["record_date"] == record_date]
                logger.debug(f"Processing {len(prediction_df)} records for date {record_date}")
            except Exception as e:
                logger.error(f"Failed to filter records for date {record_date}: {e}", exc_info=True)
                continue

            # Validate input DataFrame
            if prediction_df is None or prediction_df.empty:
                logger.warning(f"No records found for date {record_date}. Skipping log insertion.")
                continue

            # Transaction log is critical - raise error if it fails
            try:
                logger.info("Creating transaction log records")
                transaction_log_df = self._transaction_log(
                    default_type=default_type,
                    default_user_id=default_user_id,
                    default_source=default_source,
                    prediction_df=prediction_df,
                )
                logger.info("Transaction log creation completed successfully")
            except Exception as trans_err:
                logger.error(f"Critical error creating transaction log: {trans_err}", exc_info=True)
                raise Exception(f"Transaction log creation failed: {trans_err}") from trans_err

            # Performance log is non-critical - skip if it fails but continue execution
            if transaction_log_df is not None and not transaction_log_df.empty:
                try:
                    logger.info("Creating performance log records")
                    self._performance_log(transaction_log_df=transaction_log_df)
                    logger.info("Performance log creation completed successfully")
                except Exception as perf_err:
                    logger.error(f"Non-critical error creating performance log: {perf_err}", exc_info=True)
                    logger.warning("Skipping performance log creation due to error, continuing execution")
            else:
                logger.warning("Transaction log DataFrame is None or empty. Skipping performance log creation.")

    def _transaction_log(
        self, default_type: str, default_user_id: str, default_source: str, prediction_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        Create transaction log entries based on prediction DataFrame.
        Parameters:
            default_type (str): The default transaction type.
            default_user_id (str): The default user ID for the transaction.
            default_source (str): The default source of the transaction.
            prediction_df (pd.DataFrame): The DataFrame containing prediction results.
        Returns:
            pd.DataFrame: The updated DataFrame after logging transactions.
        Raises:
            ValueError: If required columns are missing in the prediction DataFrame.
            Exception: For any other unexpected errors during transaction logging.
        """
        logger.info(f"Processing {len(prediction_df)} prediction records for transaction logging")
        logger.debug(f"DataFrame columns: {list(prediction_df.columns)}")

        model_pricing = gemini_cost(
            api_type=self.DEFAULT_COST_TYPE, model_list=prediction_df["model_version"].unique().tolist()
        )

        # Validate required columns exist
        required_columns = [
            "file_name",
            "full_path",
            "folder",
            "record_date",
            "duration",
            "token_input",
            "token_cached",
            "token_output",
            "status",
            "message",
            "processed_time",
            "create_time",
            "model_version",
            "load_dt",
        ]
        missing_columns = [col for col in required_columns if col not in prediction_df.columns]
        if missing_columns:
            logger.error(f"Missing required columns in prediction_df: {missing_columns}")
            raise ValueError(f"Cannot create transaction log: missing columns {missing_columns}")

        logger.debug("Extracting usage data from prediction DataFrame")
        logger.info("Transaction log processing in STRICT MODE - will fail on any record error")
        log_payload = []
        usage_df = prediction_df[required_columns].copy()
        usage_dict = usage_df.to_dict(orient="records")

        # Pre-resolve Verint site name
        verint_site_name = resolve_env(get_value_by_path(self.verint_access, "site_name", ""))

        for idx, record in enumerate(usage_dict, 1):
            try:
                # Validate critical fields - STRICT MODE
                if not record.get("file_name"):
                    error_msg = f"Record {idx}: Missing required field 'file_name'"
                    logger.error(error_msg)
                    raise ValueError(error_msg)

                if not record.get("full_path"):
                    logger.warning(f"Record {idx} ({record['file_name']}): Missing full_path, URL will be incomplete")
                    source_files_url = f"https://{self.verint_site}/sites/{verint_site_name}/"
                else:
                    source_files_url = f"https://{self.verint_site}/sites/{verint_site_name}/{record['full_path']}"

                usage_detail = {
                    record["file_name"]: {
                        "model": record.get("model_version"),
                        "token_input": record.get("token_input"),
                        "token_cached": record.get("token_cached"),
                        "token_output": record.get("token_output"),
                    }
                }

                token_usage_input = 0
                token_usage_output = 0
                for usage in usage_detail.values():
                    for _token_type, token_count in usage.get("token_input", {}).items():
                        if str(token_count).isdigit():
                            token_usage_input += int(token_count)
                    for _token_type, token_count in usage.get("token_output", {}).items():
                        if str(token_count).isdigit():
                            token_usage_output += int(token_count)

                cost_detail = GeminiBatchModule.cal_gemini_cost(
                    usage_detail=usage_detail,
                    cost_config=model_pricing,
                )
                cost_input = cost_detail[record["file_name"]].get("cost_input", 0.0) or 0.0
                cost_output = cost_detail[record["file_name"]].get("cost_output", 0.0) or 0.0
                total_cost_usd = cost_input + cost_output

                # Create transaction payload with safe defaults
                try:
                    # Pre-process None values to avoid safe_cast_value warnings
                    duration_val = record.get("duration")
                    if duration_val is None:
                        file_metadata_sec = 0
                    elif isinstance(duration_val, int):
                        file_metadata_sec = duration_val
                    else:
                        file_metadata_sec = safe_cast_value(duration_val, int, 0)

                    # Ensure cost is float
                    cost_val = safe_cast_value(total_cost_usd, float, 0.0)

                    raw_start = str(record.get("processed_time", ""))
                    raw_end = str(record.get("create_time", ""))
                    if raw_start and raw_end:
                        try:
                            if datetime.fromisoformat(raw_start) > datetime.fromisoformat(raw_end):
                                raw_start, raw_end = raw_end, raw_start
                        except Exception:
                            pass

                    transaction_log = TransactionLogSchema.from_dict(
                        TransactionPayload(
                            data_date=str(record.get("record_date", "")),
                            start_time=raw_start,
                            end_time=raw_end,
                            type=default_type,
                            gcp_project_id=resolve_env(self.gcp.get("project_id", None)),
                            gcp_project_name=resolve_env(self.gcp.get("project_name", None)),
                            user_id=default_user_id,
                            source=default_source,
                            storage_path=source_files_url,
                            folder=str(record.get("folder", "")),
                            filename=str(record.get("file_name", "")),
                            file_metadata_sec=file_metadata_sec,
                            status_pass_failed_retry="Pass" if record.get("status") == "SUCCESS" else "Failed",
                            error_log_if=str(record.get("message", "")),
                            token_usage_input=token_usage_input,
                            token_usage_output=token_usage_output,
                            total_cost_usd=cost_val,
                            load_dt=str(record.get("load_dt", "")),
                        )
                    )
                    if record.get("status") != "SUCCESS":
                        transaction_log.stamp_error(
                            action="Create Transaction Log", error_message=record.get("message", "")
                        )
                    else:
                        transaction_log.stamp_completion(action="Create Transaction Log")
                    log_payload.append(transaction_log)
                except Exception as payload_err:
                    logger.error(
                        f"CRITICAL: Record {idx} ({record['file_name']}): Failed to create "
                        f"TransactionPayload: {payload_err}",
                        exc_info=True,
                    )
                    raise Exception(f"Transaction log creation failed at record {idx}: {payload_err}") from payload_err

            except Exception as record_err:
                logger.error(f"CRITICAL: Record {idx}: Unexpected error processing record: {record_err}", exc_info=True)
                logger.debug(f"Problematic record data: {record}")
                raise Exception(f"Transaction log creation failed at record {idx}: {record_err}") from record_err

        logger.info(f"All {len(log_payload)} transaction payload records created successfully (STRICT MODE)")

        if not log_payload:
            error_msg = "No transaction records were created from prediction data"
            logger.error(error_msg)
            raise Exception(error_msg)

        # Insert transaction logs
        try:
            logger.debug("Calling audit_log_module.log_transaction...")
            new_transaction_df = pd.DataFrame([log.to_dict() for log in log_payload])
            new_transaction_df = ensure_df_schema(new_transaction_df, self.TRANSACTION_LOG_SCHEMA)
            logger.info(
                f"Transaction log DataFrame created with {len(new_transaction_df)} rows and "
                f"{len(new_transaction_df.columns)} columns"
            )
        except Exception as log_err:
            logger.error(f"Failed to create transaction log DataFrame: {log_err}", exc_info=True)
            raise Exception(f"Critical error: Cannot create transaction log: {log_err}") from log_err

        # Filter out empty data_date values
        try:
            data_dates = [d for d in new_transaction_df["data_date"].unique().tolist() if d and str(d).strip()]
            logger.info(f"Processing transaction logs for {len(data_dates)} unique dates: {data_dates}")
        except Exception as date_err:
            logger.error(f"Failed to extract unique data_dates: {date_err}", exc_info=True)
            raise Exception(f"Cannot process transaction logs: {date_err}") from date_err

        # Upload logs to SharePoint by date
        upload_success_count = 0
        upload_failed_count = 0

        for data_date in data_dates:
            try:
                transaction_path = resolve_date(
                    text=resolve_env(get_value_by_path(self.sharepoint, "control.transaction_log_file")),
                    replace_date=data_date,
                )
                logger.debug(f"Processing transaction log for date {data_date}: {transaction_path}")

                # Get date partition
                date_partition_df = new_transaction_df[new_transaction_df["data_date"] == data_date]
                logger.debug(f"Date partition has {len(date_partition_df)} records")

                if self.sharepoint_control.is_item_exists(item_path=transaction_path):
                    logger.info(f"Existing transaction log found for date {data_date}, merging with new data")
                    try:
                        existing_log = self.sharepoint_control.get_item_by_path(transaction_path)
                        with io.BytesIO(existing_log.content) as existing_buffer:
                            existing_df = pd.read_csv(existing_buffer)

                        logger.debug(
                            f"Existing log has {len(existing_df)} rows with {len(existing_df.columns)} columns"
                        )

                        combined_df = pd.concat([existing_df, date_partition_df], ignore_index=True)
                        logger.debug(f"Combined DataFrame has {len(combined_df)} rows before deduplication")
                    except Exception as merge_err:
                        logger.error(f"Error merging with existing transaction log: {merge_err}", exc_info=True)
                        raise Exception(
                            f"Cannot merge transaction log for date {data_date}: {merge_err}"
                        ) from merge_err
                else:
                    logger.info(f"No existing transaction log found for date {data_date}, creating new log")
                    combined_df = date_partition_df

                # Prepare and upload CSV
                try:
                    combined_df["data_date"] = combined_df["data_date"].astype(str)
                    combined_df = combined_df.fillna("").sort_values(
                        by=["updated_dt", "load_dt", "data_date", "start_time", "end_time"],
                        ascending=[False, False, False, False, False],
                    )
                    combined_df = ensure_df_schema(combined_df, list(new_transaction_df.columns))
                    combined_df = replace_nan_with_default(combined_df, default_value="")

                    csv_buffer = io.BytesIO()
                    combined_df.to_csv(csv_buffer, index=False, encoding="utf-8-sig")
                    csv_buffer.seek(0)

                    self.sharepoint_control.upload_file(
                        upload_path=transaction_path,
                        content=csv_buffer.read(),
                    )
                    logger.info(f"Successfully uploaded transaction log to {transaction_path}")
                    upload_success_count += 1
                except Exception as upload_err:
                    logger.error(f"Failed to upload transaction log for date {data_date}: {upload_err}", exc_info=True)
                    upload_failed_count += 1

            except Exception as date_err:
                logger.error(f"Failed to process transaction log for date {data_date}: {date_err}", exc_info=True)
                upload_failed_count += 1
                continue

        logger.info(f"Transaction log upload complete: {upload_success_count} successful, {upload_failed_count} failed")

        if upload_failed_count > 0:
            logger.warning(f"{upload_failed_count} transaction log(s) failed to upload to SharePoint")
        else:
            logger.info("All transaction logs uploaded/updated successfully")

        self._cache_oper_log["transaction_df"] = combined_df[combined_df["type"] == default_type]
        self._cache_oper_log["process_date"] = (
            new_transaction_df["start_time"].dt.date.unique().tolist()
            if "start_time" in new_transaction_df.columns
            else []
        )

        return new_transaction_df

    def _performance_log(self, transaction_log_df: pd.DataFrame) -> None:
        """
        Create performance log entries based on transaction log DataFrame.
        Parameters:
            transaction_log_df (pd.DataFrame): The DataFrame containing transaction log records.
        Returns:
            None
        Raises:
            ValueError: If required columns are missing in the transaction log DataFrame.
        """
        logger.info(f"Processing {len(transaction_log_df)} transaction log records for performance logging")
        logger.debug(f"DataFrame columns: {list(transaction_log_df.columns)}")

        # Validate required columns exist
        required_columns = [
            "data_date",
            "start_time",
            "load_dt",
            "gcp_project_id",
            "gcp_project_name",
            "status_pass_failed_retry",
            "latency_ms",
        ]
        missing_columns = [col for col in required_columns if col not in transaction_log_df.columns]
        if missing_columns:
            logger.error(f"Missing required columns in transaction_log_df: {missing_columns}")
            raise ValueError(f"Cannot create performance log: missing columns {missing_columns}")

        logger.debug("Extracting and aggregating data from transaction log DataFrame")
        pre_df = transaction_log_df[required_columns].copy()

        # Convert start_time to datetime, setting problematic values to None
        pre_df["start_time"] = pd.to_datetime(pre_df["start_time"], errors="coerce")
        invalid_count = pre_df["start_time"].isna().sum()
        if invalid_count > 0:
            logger.warning(f"{invalid_count} records have invalid start_time values, setting them to None")

        pre_df["load_dt"] = pd.to_datetime(pre_df["load_dt"], errors="coerce")
        invalid_count = pre_df["load_dt"].isna().sum()
        if invalid_count > 0:
            logger.warning(f"{invalid_count} records have invalid load_dt values, setting them to None")

        # Format valid datetime values, keep None for invalid ones (don't filter out)
        pre_df["start_time"] = pre_df["start_time"].apply(lambda x: x.strftime("%Y%m%d") if pd.notna(x) else None)

        pre_df["load_dt"] = pre_df["load_dt"].apply(lambda x: x.strftime("%Y%m%d") if pd.notna(x) else None)

        performance_df = (
            pre_df.groupby(
                ["data_date", "start_time", "load_dt", "gcp_project_id", "gcp_project_name"],
                as_index=False,
                dropna=False,  # Keep groups with None values
            )
            .agg(
                total_transactions=("status_pass_failed_retry", "count"),
                total_completed=("status_pass_failed_retry", lambda x: (x == "Pass").sum()),
                total_failed=("status_pass_failed_retry", lambda x: (x == "Failed").sum()),
                total_runtime=("latency_ms", "sum"),
                average_response_time_ms=("latency_ms", lambda x: round(x.mean(), 2) if pd.notna(x.mean()) else 0.0),
            )
            .sort_values(by=["start_time"], ascending=False)
        )

        logger.debug(f"Aggregated performance data: {len(performance_df)} records")
        logger.info("Performance log processing in ERROR-TOLERANT MODE - will skip failed records")
        performance_dict = performance_df.to_dict(orient="records")
        log_payload = []
        skipped_records = 0

        for idx, record in enumerate(performance_dict, 1):
            try:
                performance_log = PerformanceLogSchema.from_dict(
                    PerformancePayload(
                        data_date=str(record.get("data_date", "")),
                        run_date=str(record.get("start_time", "")),
                        load_dt=str(record.get("load_dt", "")),
                        gcp_project_id=str(record.get("gcp_project_id", "")),
                        gcp_project_name=str(record.get("gcp_project_name", "")),
                        total_transaction=int(record.get("total_transactions", 0)),
                        total_completed=int(record.get("total_completed", 0)),
                        total_failed=int(record.get("total_failed", 0)),
                        total_runtime=str(record.get("total_runtime", "0.0")),
                        average_response_time_ms=str(record.get("average_response_time_ms", "0.0")),
                    )
                )
                performance_log.stamp_completion(action="Create Performance Log")
                log_payload.append(performance_log)
            except Exception as payload_err:
                logger.error(
                    f"Performance record {idx}: Failed to create PerformancePayload: {payload_err}", exc_info=True
                )
                logger.warning(f"Skipping performance record {idx} and continuing...")
                skipped_records += 1
                continue

        logger.info(
            f"Performance payload created: {len(log_payload)} succeeded, {skipped_records} skipped "
            f"(ERROR-TOLERANT MODE)"
        )

        if skipped_records > 0:
            logger.warning(f"Note: {skipped_records} performance records were skipped due to errors")

        if not log_payload:
            logger.warning(f"All {len(performance_dict)} performance records failed to process - returning None")
            return

        # Insert performance logs
        try:
            logger.debug("Calling audit_log_module.log_performance...")
            df = pd.DataFrame(log.to_dict() for log in log_payload)
            df = ensure_df_schema(df, self.PERFORMANCE_LOG_SCHEMA)
            logger.info(f"Performance log DataFrame created with {len(df)} rows and {len(df.columns)} columns")
        except Exception as log_err:
            logger.error(f"Failed to create performance log DataFrame: {log_err}", exc_info=True)
            raise Exception(f"Cannot create performance log: {log_err}") from log_err

        # Filter out empty data_date values
        try:
            data_dates = [d for d in df["data_date"].unique().tolist() if d and str(d).strip()]
            logger.info(f"Processing performance logs for {len(data_dates)} unique dates: {data_dates}")
        except Exception as date_err:
            logger.error(f"Failed to extract unique data_dates: {date_err}", exc_info=True)
            raise Exception(f"Cannot process performance logs: {date_err}") from date_err

        # Upload logs to SharePoint by date
        upload_success_count = 0
        upload_failed_count = 0

        for data_date in data_dates:
            try:
                performance_path = resolve_date(
                    text=resolve_env(get_value_by_path(self.sharepoint, "control.performance_log_file")),
                    replace_date=data_date,
                )
                logger.debug(f"Processing performance log for date {data_date}: {performance_path}")

                # Get date partition
                date_partition_df = df[df["data_date"] == data_date]
                logger.debug(f"Date partition has {len(date_partition_df)} records")

                if self.sharepoint_control.is_item_exists(item_path=performance_path):
                    logger.info(f"Existing performance log found for date {data_date}, merging with new data")
                    try:
                        existing_log = self.sharepoint_control.get_item_by_path(performance_path)
                        with io.BytesIO(existing_log.content) as existing_buffer:
                            existing_df = pd.read_csv(existing_buffer)

                        logger.debug(
                            f"Existing log has {len(existing_df)} rows with {len(existing_df.columns)} columns"
                        )

                        combined_df = pd.concat([existing_df, date_partition_df], ignore_index=True)
                        logger.debug(f"Combined DataFrame has {len(combined_df)} rows before deduplication")
                    except Exception as merge_err:
                        logger.error(f"Error merging with existing performance log: {merge_err}", exc_info=True)
                        raise Exception(
                            f"Cannot merge performance log for date {data_date}: {merge_err}"
                        ) from merge_err
                else:
                    logger.info(f"No existing performance log found for date {data_date}, creating new log")
                    combined_df = date_partition_df

                # Prepare and upload CSV - use combined_df to preserve all records
                try:
                    combined_df["data_date"] = combined_df["data_date"].astype(str)
                    combined_df = combined_df.fillna("").sort_values(
                        by=["load_dt", "data_date"], ascending=[False, False]
                    )
                    combined_df = ensure_df_schema(combined_df, list(df.columns))
                    combined_df = replace_nan_with_default(combined_df, default_value="")

                    csv_buffer = io.BytesIO()
                    combined_df.to_csv(csv_buffer, index=False, encoding="utf-8-sig")
                    csv_buffer.seek(0)

                    self.sharepoint_control.upload_file(
                        upload_path=performance_path,
                        content=csv_buffer.read(),
                    )
                    logger.info(
                        f"Successfully uploaded performance log to {performance_path} with "
                        f"{len(combined_df)} total records"
                    )
                    upload_success_count += 1
                except Exception as upload_err:
                    logger.error(f"Failed to upload performance log for date {data_date}: {upload_err}", exc_info=True)
                    upload_failed_count += 1

            except Exception as date_err:
                logger.error(f"Failed to process performance log for date {data_date}: {date_err}", exc_info=True)
                upload_failed_count += 1
                continue

        logger.info(f"Performance log upload complete: {upload_success_count} successful, {upload_failed_count} failed")

        if upload_failed_count > 0:
            logger.warning(f"{upload_failed_count} performance log(s) failed to upload to SharePoint")
        else:
            logger.info("All performance logs uploaded/updated successfully")

    def _upload_master_file(self, single_month_master_df: pd.DataFrame) -> pd.DataFrame:
        """Appends a list of data dicts to XLSX master files based on input folder structure."""
        logger.info("Update Master file")

        single_month_master_df = single_month_master_df.replace(["None", "none", "null", "Null"], "").fillna("")
        # Flatten all expected column names
        all_columns = [col for _, cols in self.GROUP_HEADER_SCHEMA for col in cols]

        # ✅ Ensure consistent column order
        for col in all_columns:
            if col not in single_month_master_df.columns:
                single_month_master_df[col] = ""  # fill missing columns
        single_month_master_df = single_month_master_df[all_columns]  # reorder columns

        try:
            data_dt = single_month_master_df["call_date"].iloc[0]
            master_output_path = resolve_date(
                resolve_env(self.sharepoint.get("verint").get("master_output_file")), data_dt
            )
            logger.debug(f"Master output path: {master_output_path}")
        except Exception as e:
            logger.error(f"Failed to resolve output file path: {e}", exc_info=True)
            raise Exception(f"Cannot determine output file path: {e}") from e

        try:
            # Try downloading existing XLSX
            master_file = self.sharepoint_verint.get_item_by_path(item_path=master_output_path)
            master_file_bytes = master_file.content
            with io.BytesIO(master_file_bytes) as existing_file:
                existing_df = pd.read_excel(
                    existing_file,
                    header=1,
                    dtype={"call_date": str, "agent_id": str, "call_id": str, "phone_number": str},
                    keep_default_na=False,
                )

                combined_df = pd.concat([existing_df, single_month_master_df], ignore_index=True)

                existing_file.seek(0)
                existing_wb = load_workbook(existing_file)
                ws = existing_wb.active
                self._sync_excel_schema(ws, self.GROUP_HEADER_SCHEMA, all_columns)
                start_row = ws.max_row + 1
                for r_idx, row in enumerate(single_month_master_df.itertuples(index=False, name=None), start=start_row):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
        except Exception as e:
            if hasattr(e, "response") and e.response.status_code == 404:
                logger.info("No existing master file found, creating new one")

                combined_df = single_month_master_df.copy()
                # Create a new workbook
                existing_wb = Workbook()
                ws = existing_wb.active

                # First row (group headers)
                col_idx = 1
                for group_name, cols in self.GROUP_HEADER_SCHEMA:
                    start_col = col_idx
                    col_idx += len(cols)
                    end_col = col_idx - 1
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                    cell = ws.cell(row=1, column=start_col, value=group_name)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.font = Font(bold=True)

                # Second row (actual headers)
                col_idx = 1
                for _, cols in self.GROUP_HEADER_SCHEMA:
                    for col_name in cols:
                        cell = ws.cell(row=2, column=col_idx, value=col_name)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        cell.font = Font(bold=True)
                        col_idx += 1

                # Data rows start from row 3
                for r_idx, row in enumerate(single_month_master_df.itertuples(index=False, name=None), start=3):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
            else:
                logger.error(f"Failed to process existing Excel file: {e}", exc_info=True)
                raise Exception(f"Cannot process existing master file: {e}") from e

        # Auto-fit column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)

        # ✅ MOVED INSIDE LOOP - Save and upload for EACH folder
        output_stream = io.BytesIO()
        existing_wb.save(output_stream)
        output_stream.seek(0)

        self.sharepoint_verint.upload_file(
            upload_path=master_output_path,
            content=output_stream.read(),
        )

        return combined_df

    def _upload_daily_files(self, daily_df: pd.DataFrame, daily_output_path: str, sharepoint_site) -> None:
        """Appends a list of data dicts to XLSX daily files based on input folder structure."""

        group_header_daily_schema = copy.deepcopy(self.GROUP_HEADER_DAILY_SCHEMA)
        logger.info("Update Daily file")

        # ====== Grouping ======
        try:
            user_config_path = resolve_env(get_value_by_path(self.sharepoint, "control.user_config_path"))
            user_config_content = self.sharepoint_control.get_item_by_path(user_config_path)
            with io.BytesIO(user_config_content.content) as user_config_file_buffer:
                service_quality_group_df = pd.read_excel(user_config_file_buffer, sheet_name="service_quality_group")
        except Exception as e:
            logger.error(f"Failed to fetch user config file from SharePoint: {e}", exc_info=True)
            user_config_path = resolve_env(self.framework.get("user_config_path"))
            logger.info(f"Read file from config: {user_config_path}")
            service_quality_group_df = read_xlsx(user_config_path, sheet_name="service_quality_group")

        service_quality_group_df["sub_category"] = service_quality_group_df["sub_category"].str.strip()

        group_names = service_quality_group_df["sub_category"].dropna().loc[lambda s: s != ""].unique()

        service_quality_schema = []

        # Group
        for group_name in group_names:
            matching_items = (
                service_quality_group_df[service_quality_group_df["sub_category"] == group_name]["item"]
                .str.strip()
                .tolist()
            )
            existing_columns = [col for col in matching_items if col in daily_df.columns]

            service_quality_schema += matching_items
            group_reason = f"{group_name}_reason"

            if existing_columns:
                daily_df[group_name] = daily_df.apply(
                    lambda row, _cols=existing_columns: self._calculate_category([str(row[col]) for col in _cols]),
                    axis=1,
                )

                daily_df[group_reason] = daily_df.apply(
                    lambda row, _cols=existing_columns: ", \n".join(
                        [f"{col}: {row[f'{col}_reason']}" for col in _cols if pd.notna(row[f"{col}_reason"])]
                    ),
                    axis=1,
                )
            else:
                daily_df[group_name] = ""
                daily_df[group_reason] = ""

            service_quality_schema.append(group_name)
            service_quality_schema.append(group_reason)
        # Non Group
        na_items = (
            service_quality_group_df[service_quality_group_df["sub_category"].isna()]["item"].str.strip().tolist()
        )
        service_quality_schema += na_items

        service_quality_schema.append("service_quality_score")
        service_quality_schema.append("service_quality_performance_insight")
        next(cols for name, cols in group_header_daily_schema if name == "Service Quality").extend(
            service_quality_schema
        )
        # ========================

        # Flatten all expected column names
        all_columns = [col for _, cols in group_header_daily_schema for col in cols]

        daily_df = daily_df.reindex(columns=all_columns)

        # ====== NEW FEATURE: Schema & Existence Validation ======
        try:
            existing_file_data = sharepoint_site.get_item_by_path(daily_output_path)
            logger.info(f"File already exists at {daily_output_path}. Checking schema...")

            # Read existing schema (row 2 contains actual headers)
            with io.BytesIO(existing_file_data.content) as existing_buf:
                existing_wb = load_workbook(existing_buf, read_only=True)
                ws_existing = existing_wb.active
                # Extract headers from Row 2
                existing_headers = [
                    ws_existing.cell(row=2, column=c).value for c in range(1, ws_existing.max_column + 1)
                ]
                existing_wb.close()

            # Filter out any None values if columns are empty
            existing_headers = [h for h in existing_headers if h is not None]

            # 3. Check if schemas are equal
            if existing_headers != all_columns:
                logger.warning("Schema mismatch detected! Preparing to upload as a new file.")

                # 1. Find columns that are in the new schema but missing from the existing file
                added_columns = [col for col in all_columns if col not in existing_headers]

                # 2. Find columns that were in the existing file but are removed in the new schema
                removed_columns = [col for col in existing_headers if col not in all_columns]

                # 3. Log the differences cleanly
                if added_columns:
                    logger.info(f"Added new columns: {added_columns}")
                if removed_columns:
                    logger.info(f"Removed columns from existing: {removed_columns}")

                # 4. If the items are the same but just out of order, let yourself know!
                if not added_columns and not removed_columns:
                    logger.info("The columns equal, but their structural ORDER has changed.")

                base, ext = os.path.splitext(daily_output_path)
                timestamp = datetime.now(tz=ZoneInfo(self.timezone)).strftime("%Y%m%d_%H%M%S")
                archived_path = f"{base}_old_grouping_schema_{timestamp}{ext}"
                sharepoint_site.rename_file(current_path=daily_output_path, new_path=archived_path)

                logger.info(
                    f"Old file renamed to '{archived_path}'. Proceeding to write new layout to '{daily_output_path}'"
                )
            else:
                logger.info("Schemas match. Continuing with regular upload.")

        except Exception as e:
            logger.info(f"File does not exist or could not be verified ({e}). Proceeding with fresh upload.")
        # ========================================================

        existing_wb = Workbook()
        ws = existing_wb.active

        # First row (group headers)
        col_idx = 1
        for group_name, cols in group_header_daily_schema:
            start_col = col_idx
            col_idx += len(cols)
            end_col = col_idx - 1
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=group_name)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(bold=True)

        # Second row (actual headers)
        col_idx = 1
        for _, cols in group_header_daily_schema:
            for col_name in cols:
                cell = ws.cell(row=2, column=col_idx, value=col_name)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = Font(bold=True)
                col_idx += 1

        # Data rows start from row 3
        for r_idx, row in enumerate(daily_df.itertuples(index=False, name=None), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        logger.info(f"Write file with {len(daily_df)} records: {daily_output_path}")

        # Auto-fit column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)

        # Save and upload
        output_stream = io.BytesIO()
        existing_wb.save(output_stream)
        output_stream.seek(0)

        sharepoint_site.upload_file(
            upload_path=daily_output_path,
            content=output_stream.read(),
        )

    def _sync_excel_schema(self, ws, group_headers, all_columns):
        """
        Ensure Excel columns match latest schema.
        Add missing columns and rebuild header rows if needed.
        """
        # Read existing header row (row 2)
        existing_headers = [cell.value for cell in ws[2]]

        if existing_headers == all_columns:
            return  # schema already correct

        # Read existing data into memory (skip first 2 rows)
        data = list(ws.iter_rows(min_row=3, values_only=True))
        existing_df = pd.DataFrame(data, columns=existing_headers)

        # Add missing columns
        for col in all_columns:
            if col not in existing_df.columns:
                existing_df[col] = ""

        # Reorder columns to match new schema
        existing_df = existing_df[all_columns]

        # 1. Unmerge all merged cells first (IMPORTANT)
        if ws.merged_cells.ranges:
            for merged_range in list(ws.merged_cells.ranges):
                ws.unmerge_cells(str(merged_range))

        # Clear worksheet
        ws.delete_rows(1, ws.max_row)

        # --- Recreate headers ---
        col_idx = 1
        for group_name, cols in group_headers:
            start_col = col_idx
            col_idx += len(cols)
            end_col = col_idx - 1
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=group_name)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(bold=True)

        col_idx = 1
        for _, cols in group_headers:
            for col_name in cols:
                cell = ws.cell(row=2, column=col_idx, value=col_name)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = Font(bold=True)
                col_idx += 1

        # Rewrite old data
        for r_idx, row in enumerate(existing_df.itertuples(index=False, name=None), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    def _calculate_category(self, sub_cate: list) -> str:
        """Calculate category based on sub-category list."""

        sub_cate = [item.lower() for item in sub_cate]
        sub_cate = set(sub_cate)

        if "below" in sub_cate:
            return "Below"
        if "meet" in sub_cate:
            return "Meet"
        return "N/A"

    def on_error(self, error: Exception) -> None:
        """
        Hook executed when task execution fails.
        Override this method to implement custom error handling
        (e.g., cleanup, notifications, retry logic).

        Args:
            error (Exception): The exception that occurred during execution

        Returns:
            None
        """
        try:
            current_dt = get_current_datetime().replace(microsecond=0)
            formatted_current_dt = current_dt.isoformat(sep=" ")

            subject = "[AI Failed] [AI-QA]"
            body = f"""
            Dear Operation Team,<br>
            <br>
                Project: {self.project_id} <br>
                An error occurred during execution of task '{self.task_name}' <br>
                Error details: {error} <br>
                <br>
                Please investigate. <br>
            <br>
            -- Timestamp: {formatted_current_dt} <br>
            Best Regards,<br>
            [This is automatic message generated by AI - Do Not REPLY]
            """

            self.msgraph_module.send_email(
                subject=subject,
                body=body,
                sender_email=self.msgraph_sender_email,
                receiver_email=self.msgraph_receiver_email,
                cc_email=self.msgraph_cc_email,
            )
            logger.info(f"Task error notification successfully issued: {self.task_name}")
        except Exception as email_error:
            logger.error(f"Task error notification failed to issue: {email_error}", exc_info=True)
        logger.error(f"Error in task '{self.task_name}': {error}", exc_info=True)

    def post_execute(self, result):
        """
        Post-execution hook for the task. This method is called after the main execution logic is completed.
        It is responsible for stamping AI-Operation logs based on the transaction log data.
        """
        logger.info("Stamping AI-Operation logs")
        if not self._cache_oper_log.get("process_date"):
            logger.warning("No process_date found for stamping logs - skipping logging")
            return result

        transaction_df = self._cache_oper_log.get("transaction_df")
        if transaction_df is None or transaction_df.empty:
            logger.warning("No transaction log DataFrame found for stamping logs - skipping logging")
            return result
        try:
            required_columns = ["start_time", "end_time", "gcp_project_id", "status_pass_failed_retry", "latency_ms"]
            transaction_df = ensure_df_schema(transaction_df, required_columns).copy()
            transaction_df["start_time"] = pd.to_datetime(
                transaction_df["start_time"], errors="coerce", utc=True
            ).dt.tz_convert(LogTimeStamper.CONFIG_TIMEZONE)
            transaction_df["end_time"] = pd.to_datetime(
                transaction_df["end_time"], errors="coerce", utc=True
            ).dt.tz_convert(LogTimeStamper.CONFIG_TIMEZONE)
            transaction_df["latency_ms"] = pd.to_numeric(transaction_df["latency_ms"], errors="coerce")
            transaction_df["process_date"] = transaction_df["start_time"].dt.date
            process_dates = self._cache_oper_log["process_date"]
            for process_date in process_dates:
                filtered_df = transaction_df[(transaction_df["process_date"] == process_date)]

                if filtered_df.empty:
                    logger.warning(
                        f"No transaction records found for process_date {process_date} - skipping log stamping"
                    )
                    continue

                log_df = filtered_df.groupby(["process_date", "gcp_project_id"], as_index=False, dropna=False).agg(
                    total_transaction=("status_pass_failed_retry", "count"),
                    total_success_transaction=("status_pass_failed_retry", lambda x: (x == "Pass").sum()),
                    total_failed_transaction=("status_pass_failed_retry", lambda x: (x == "Failed").sum()),
                    average_response_time_sec=(
                        "latency_ms",
                        lambda x: round(x.mean() / 1000, 2) if pd.notna(x.mean()) else 0.0,
                    ),
                    min_start_time=("start_time", "min"),
                    max_end_time=("end_time", "max"),
                )
                log_df["total_runtime_sec"] = (
                    (
                        pd.to_datetime(log_df["max_end_time"], errors="coerce")
                        - pd.to_datetime(log_df["min_start_time"], errors="coerce")
                    )
                    .dt.total_seconds()
                    .round(2)
                )
                log_df = log_df.drop(columns=["min_start_time", "max_end_time"])
                env = os.environ.get("ENVIRONMENT", "").lower()
                if env == "prod":
                    log_df["environment"] = "production"
                elif env == "nprd":
                    log_df["environment"] = "non-production"
                else:
                    log_df["environment"] = env or "unknown"
                log_df = log_df.rename(columns={"gcp_project_id": "project_id"})
                log_df["project_type"] = "batch"

                log_df = log_df[
                    [
                        "process_date",
                        "environment",
                        "project_id",
                        "project_type",
                        "total_transaction",
                        "total_success_transaction",
                        "total_failed_transaction",
                        "average_response_time_sec",
                        "total_runtime_sec",
                    ]
                ]
                log_list = log_df.to_dict(orient="records")
                for log in log_list:
                    logging_ai_operation(log_instance=logger, log_obj=log, log_type="batch", message="AI-Operation-Log")
            logger.info("AI-Operation log stamping completed successfully")
            return result
        except Exception as log_err:
            logger.error(f"Failed to stamp AI-Operation logs: {log_err}", exc_info=True)
            return result
