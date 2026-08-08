"""The command line: the only place the pieces are wired together.

Every other module in `evalgen` refuses to know about its neighbours. `outcomes`
decides `parse_ok` and knows nothing about runs; `runner` calls a model and judges
nothing; `flatten` changes the grain and scores nothing; `report` renders and computes
no metric. That separation is what makes each of them testable, and it leaves exactly
one place where a mistake can be made about how they fit: here.

So the wiring is written down rather than implied.

    testset ──> prompts.build_messages ──> runner.run ──> outcomes.classify
                                                              │
                                       run.jsonl <────────────┤
                                                              ▼
                        flatten.to_rows (one row per PRODUCT, not per call)
                                                              │
                                              records.from_row (normalisation)
                                                              │
                              metrics.score_* ────────────────┴──> report.render
                              report.mechanism_table / n_flip

Four decisions this file makes that no other module can, each of which would otherwise
be made silently by whoever wrote the first script:

1. **The aggregate metrics table is scored on replicate 1, and says so.** Pooling every
   replicate into one prediction list looks obvious and is wrong: `metrics.outer_join`
   builds `{r.key: r for r in pred}` (`metrics.py:122`), so three replicates of the same
   item collapse to whichever one happened to be last in the list, and the resulting
   table describes a run nobody made. The mechanism verdicts and `n_flip` are the
   things that read every replicate, which is the division of labour `report.py`
   already argues for.

2. **A dry run costs nothing and proves the request.** `--dry-run` drives the REAL
   runner with a client that records instead of calling, so the bodies written out are
   the bodies the run loop would have produced -- built by `request.build_request`,
   the same function `client.complete` sends. A dry run that rebuilt the request
   separately would review a request nobody makes.

3. **Two arms with the same name are refused.** `report.render` keys its mechanism
   tables by arm name; two runs both called `incumbent` collapse into one dict entry
   and the paired comparison silently becomes a comparison of an arm with itself.

4. **A comparison across different testsets, ground truths or harness versions is
   refused**, via `manifest.assert_comparable` plus an explicit sha check. Those are
   the mismatches that still produce a full, plausible-looking table.

`configure_stdout()` is the first statement in `main()`. It has to be: this CLI prints
Thai -- the testset's `expected_failure` prose reaches stdout through
`MechanismRow.detail` -- and a Thai-locale Windows console is cp874, which raises
`UnicodeEncodeError` on the report itself, after the run has been paid for.

Exit codes are three, and the middle one matters: `1` means the harness ran and found
problems in the inputs, `2` means it refused to run at all. A caller that cannot tell
those apart will eventually treat a refusal as a clean sheet.
"""

from __future__ import annotations

import argparse
import json
import statistics
import sys
import time
from collections.abc import Mapping, Sequence
from dataclasses import asdict, dataclass, replace
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from evalgen.artifacts import (
    ArtifactError,
    RunJournal,
    append_jsonl,
    assert_shareable_payload,
    atomic_write_bytes,
    atomic_write_text,
    file_sha256,
    require_private_destination,
)
from evalgen.contracts import ApplicationSpec, RETENTION_APPLICATION
from evalgen.console import configure_stdout
from evalgen.config import ENV_FILES, find_api_key, find_runtime_api_key, load_env_file
from evalgen.decoding import decoding_schema
from evalgen.flatten import named_no_product, to_rows
from evalgen.judge import find_disagreements, run_judge, shareable_report
from evalgen.experiments import (
    PlanError,
    arm_by_id,
    canonical_sha,
    decision as experiment_decision,
    load_plan,
    logical_call_budget,
    operational_summary,
    projected_execution_budget,
    qualification as assess_qualification,
    qualification_contract_sha,
    runtime_gate,
    stability_disagreement,
    validate_plan,
)
from evalgen.prompts import Prompt, PromptError
from evalgen.prompts import get as get_prompt
from evalgen.prompts import validate_manifest as validate_prompt_manifest
from evalgen.report import ArmSummary, ReportError, mechanism_table, n_flip, render
from evalgen.request import REASONING_EFFORTS
from evalgen.runtime import (
    OPENROUTER_RUNTIME,
    RuntimeBackend,
    RuntimeSpec,
    build_runtime_request,
    execution_provenance,
    local_gpu_runtime,
)
from evalgen.runner import (
    ItemResult,
    RunConfig,
    RunError,
    RunResult,
    item_result_to_row,
    run,
    write_run_log,
)
from evalgen.testsets import (
    TestItem,
    TestSet,
    TestsetError,
    load_testset,
    testset_sha,
    validate,
)
from evalharness import manifest as manifest_mod
from evalharness.adapters.retention import load_csv
from evalharness.compare import (
    CoverageMismatch,
    check_coverage,
    disagreement,
    paired_verdict,
    regressions,
)
from evalharness.keys import hmac_key
from evalharness.labelspaces import RETENTION
from evalharness.metrics import (
    DimensionResult,
    score_call_result,
    score_product,
    score_reason,
)
from evalharness.records import Record, from_row

__all__ = [
    "COST_PER_CORRECT_DIMENSION",
    "EXIT_OK",
    "EXIT_PROBLEMS",
    "EXIT_REFUSED",
    "CliError",
    "LoadedRun",
    "arm_summary",
    "build_client",
    "build_parser",
    "load_run",
    "main",
    "new_run_dir",
    "replicate_records",
    "response_format",
]

PACKAGE_DIR = Path(__file__).resolve().parent
REPO_ROOT = PACKAGE_DIR.parent.parent

DEFAULT_TESTSET = REPO_ROOT / "tests" / "fixtures" / "testsets" / "retention_v1.jsonl"
DEFAULT_GT = REPO_ROOT / "tests" / "fixtures" / "testsets" / "retention_v1.gt.csv"
SCHEMA_PATH = PACKAGE_DIR / "schemas" / "retention.json"
DEFAULT_EXPERIMENT_PLAN = REPO_ROOT / "experiments" / "retention-e5.plan.json"
SYNTHETIC_FIXTURE_ROOT = (REPO_ROOT / "tests" / "fixtures").resolve()
_CLASSIFICATION_RANK = {"synthetic": 0, "internal": 1, "customer": 2}

APPLICATIONS: dict[str, ApplicationSpec] = {
    RETENTION_APPLICATION.application_id: RETENTION_APPLICATION,
}


def _application_spec(application_id: str) -> ApplicationSpec:
    """Return the reviewed application contract wired into this executable.

    Adding a label space alone is not enough to make a new application runnable: its
    adapter, prompt, schema and decision units must move together.  The generic
    ``ApplicationSpec`` is the extension seam; this registry is the explicit point at
    which an implementation becomes executable.
    """
    try:
        return APPLICATIONS[application_id]
    except KeyError as exc:
        raise CliError(
            f"application {application_id!r} has no executable application contract. "
            "Define and review its adapter, prompt, schema, testset reference and "
            "decision units before running models."
        ) from exc


def _inside(path: Path, root: Path) -> bool:
    resolved = Path(path).expanduser().resolve(strict=False)
    return resolved == root or root in resolved.parents


def _validate_declared_classification(
    classification: str, *, testset_path: Path, gt_path: Path
) -> None:
    """Prevent the convenient synthetic default from becoming a privacy bypass."""
    if classification not in _CLASSIFICATION_RANK:
        raise CliError(f"unknown data classification {classification!r}")
    if classification != "synthetic":
        return
    outside = [
        str(path)
        for path in (testset_path, gt_path)
        if not _inside(path, SYNTHETIC_FIXTURE_ROOT)
    ]
    if outside:
        raise CliError(
            "--data-classification synthetic is allowed only for committed synthetic "
            f"fixtures below {SYNTHETIC_FIXTURE_ROOT}; outside path(s): {outside}. "
            "Classify custom or real inputs as internal/customer and store outputs "
            "below EVAL_HARNESS_DATA_DIR."
        )

# `out/` is ignored by .gitignore as a directory, which is the second line of defence
# described in that file's header. Run artifacts carry model output verbatim, so they
# go somewhere already denied rather than somewhere that has to be remembered.
DEFAULT_OUT = REPO_ROOT / "out" / "runs"

EXIT_OK = 0
EXIT_PROBLEMS = 1  # the harness ran; the inputs or the arms have problems
EXIT_REFUSED = 2  # the harness declined to run, or to compare

# The order dimensions are scored and printed in. `report._DIMENSION_ORDER` prints in
# the same order; this list is what decides which scorers actually run.
_SCORERS = (
    ("call_result", score_call_result, RETENTION.call_result),
    ("reason", score_reason, RETENTION.reason),
    ("product", score_product, RETENTION.product),
)

# How many lines of the assembled prompt a dry run echoes. Enough to show the Role
# line, the transcript/audio repair and the opening of the product rules, which is the
# part a reviewer can actually check against production.
DRY_RUN_PROMPT_LINES = 40

# Which scored dimension defines "a correct answer" for the cost-per-correct figure.
#
# `call_result` and NOT a new notion of correctness invented for a cost table. It is the
# label the Retention app acts on, it is first in `_SCORERS`, and its one-vs-rest
# true-positive count is already the arithmetic section 5's weighted recall is built
# from -- so the ratio divides a number the report prints two sections earlier rather
# than one only this line knows about.
#
# Named as a constant so `report._performance_section` can print WHICH dimension it
# divided by. A cost-per-correct whose definition of correct is not stated is a ratio
# nobody can reproduce, and the three dimensions here would give three different answers
# (`metrics.py:11-13`: three dimensions, three denominators).
#
# It must stay a SINGLE-LABEL dimension. `_correct_answers` counts true positives, and
# that count is a count of correct ROWS only while a row can produce at most one -- which
# holds for `call_result` and `product` and does NOT hold for `reason`, where one row
# carries up to three labels and a row half right would contribute a true positive
# anyway. Pointed at `reason` this constant would silently start counting something
# else and the ratio would keep printing.
COST_PER_CORRECT_DIMENSION = "call_result"


# ======================================================================== errors


class CliError(RuntimeError):
    """A refusal, phrased for someone at a terminal.

    Raised where continuing would produce output that looks like a result. `main`
    turns it into exit code 2 and a one-line message, never a traceback: a stack trace
    for "the API key is not set" teaches the reader to ignore stack traces.
    """


# ======================================================================== loading


def response_format(schema_path: Path = SCHEMA_PATH) -> dict[str, Any]:
    """The retention schema, transformed for decoding and wrapped as `json_schema`.

    Wrapped rather than sent bare because `runner._required_keys` unwraps exactly this
    shape to learn which top-level keys `outcomes.classify` must find. The keys are
    therefore read out of the request that was sent, not from a second list kept in
    step by hand -- for this schema they are `product`, `call_event_detection` and
    `recommendation`.

    Production does not do this. It calls Vertex with forced function calling
    (`main.py:1109-1114`, `"functionCallingConfig": {"mode": "ANY"}`), which this
    harness cannot reproduce over OpenRouter chat. `response_format` is a CONSTRAINED
    DECODER rather than an encouragement, and `decoding.decoding_schema` is where the
    three consequences of that difference are named, counted and measured. The file on
    disk stays the port; only what is sent is transformed.

    **No `"strict": true`, deliberately, and that is a measurement rather than an
    omission.** It was tried on 2026-08-04 against both arms. Both endpoints accept it,
    and it changed nothing observable: RET-01, RET-02 and RET-03 came back with the
    same content and the same completion-token counts with and without it. What
    actually guarantees the schema was enforced is `provider: {"require_parameters":
    true}`, which `request.build_request` sends -- it makes OpenRouter refuse an
    endpoint that cannot honour `response_format` rather than drop the parameter and
    serve prompt-coaxed output that looks identical. Shipping an unmeasurable flag as
    though it were the guard is how a pin comes to be trusted for something it does not
    do.
    """
    schema = decoding_schema(json.loads(schema_path.read_text(encoding="utf-8")))
    return {
        "type": "json_schema",
        "json_schema": {"name": schema.get("title", "analyze_call"), "schema": schema},
    }


def _load_testset(path: Path, *, app: str = "retention") -> TestSet:
    try:
        return load_testset(path, app=app)
    except FileNotFoundError as exc:
        raise CliError(f"testset not found: {path}") from exc
    except TestsetError as exc:
        raise CliError(str(exc)) from exc


def _load_gt(path: Path) -> list[Record]:
    try:
        return load_csv(path)
    except FileNotFoundError as exc:
        raise CliError(f"ground truth not found: {path}") from exc


def _prompt(prompt_id: str) -> Prompt:
    try:
        return get_prompt(prompt_id)
    except PromptError as exc:
        raise CliError(str(exc)) from exc


def _sig(record: Record) -> tuple:
    """One ground-truth row reduced to what a comparison of two GT sources can check."""
    return (
        record.call_id,
        record.phone,
        record.product,
        record.call_result,
        tuple(sorted(record.reasons)),
    )


def _gt_disagreements(testset: TestSet, gt: Sequence[Record]) -> list[str]:
    """Where the testset's own `gt` and the ground-truth CSV disagree.

    THE ITEM AND THE CSV ARE TWO GROUND TRUTHS, and nothing else in the pipeline ever
    puts them side by side. `flatten._gt_skeleton` builds a failure row per product in
    `item.gt`; `metrics.*` scores against the CSV. If the two drift, a failed item
    emits rows keyed on products the scorer is not expecting, those rows join with
    nothing, and the arm's `Coverage.items_joined` quietly falls -- on the arm that
    failed, which is exactly where nobody is surprised by a lower number.

    Checked here, in `check`, because it costs nothing and because after a run it is
    too late to be worth knowing.
    """
    problems: list[str] = []
    by_call: dict[str, list[Record]] = {}
    for record in gt:
        by_call.setdefault(record.call_id, []).append(record)

    for item in testset.items:
        rows = by_call.pop(item.call_id, [])
        if not rows:
            problems.append(
                f"{item.item_id}: call_id {item.call_id} has no row in the ground-truth "
                "CSV. report.mechanism_table refuses this item rather than scoring it, "
                "because an empty truth matches an empty prediction and reports PASS."
            )
            continue
        from_item = sorted(
            _sig(from_row({"call_id": item.call_id, "phone_number": item.phone_number, **row}))
            for row in item.gt
        )
        from_csv = sorted(_sig(record) for record in rows)
        if from_item != from_csv:
            problems.append(
                f"{item.item_id}: the item's own `gt` and the ground-truth CSV disagree.\n"
                f"    testset: {from_item}\n"
                f"    csv    : {from_csv}\n"
                "    flatten builds failure rows from the item and metrics scores against "
                "the CSV, so a disagreement moves failure rows off the merge key."
            )

    for call_id, rows in sorted(by_call.items()):
        problems.append(
            f"ground-truth CSV has {len(rows)} row(s) for call_id {call_id}, which no "
            "testset item claims. Every dimension's denominator is ground-truth driven "
            "(metrics.py:11-13), so an unclaimed row scores as a miss on every arm."
        )
    return problems


# ======================================================================== run I/O


@dataclass(frozen=True)
class LoadedRun:
    """One completed arm, read back off disk.

    Deliberately reconstructs `runner.RunResult` rather than inventing a second result
    type: `write_run_log` writes every `ItemResult` field, so the round trip is
    lossless and every consumer downstream -- scoring, the mechanism table, the report
    -- runs against the same objects whether the rows came from memory or from a file.
    A parallel "logged row" type would be a second place for the flatten step to be
    called slightly differently.
    """

    directory: Path
    meta: dict[str, Any]
    result: RunResult

    @property
    def arm(self) -> str:
        return self.result.config.arm

    @property
    def decoding(self) -> dict[str, Any]:
        return dict(self.meta["decoding"])


def new_run_dir(base: Path, arm: str, *, now: datetime | None = None) -> Path:
    """Create a fresh run directory. Never reuses one that already holds a run.

    The timestamp is second-resolution, and two arms started inside the same second is
    not hypothetical: it is what a script comparing an incumbent against a candidate
    does, and it is what the test suite does. With `exist_ok=True` and a shared name
    the second run's `run.jsonl` overwrites the first, both `compare` arguments resolve
    to the same directory, and the report prints a model compared against itself with
    every mechanism in perfect agreement.

    So the name is disambiguated rather than reused, and a directory that already
    contains a run is never written into. Suffixes are `-2`, `-3`, ... rather than a
    random token so that two runs a second apart still sort into the order they
    happened.
    """
    stamp = (now or datetime.now(timezone.utc)).strftime("%Y%m%d-%H%M%SZ")
    base = Path(base)
    for suffix in range(1, 1000):
        name = f"{stamp}-{arm}" if suffix == 1 else f"{stamp}-{arm}-{suffix}"
        directory = base / name
        try:
            directory.mkdir(parents=True, exist_ok=False)
        except FileExistsError:
            continue
        return directory
    raise CliError(
        f"could not create a run directory under {base}: 999 runs of arm {arm!r} already "
        "exist for this second. Something is looping."
    )


def _output_base(args: argparse.Namespace) -> Path:
    """Resolve an output root, enforcing the private-data policy before any write."""
    base = Path(args.out)
    classification = getattr(args, "data_classification", "synthetic")
    if classification != "synthetic":
        try:
            return require_private_destination(base)
        except ArtifactError as exc:
            raise CliError(str(exc)) from exc
    return base.resolve(strict=False)


def _run_directory(args: argparse.Namespace, name: str) -> Path:
    resume = getattr(args, "resume_run", None)
    if not resume:
        return new_run_dir(_output_base(args), name)
    directory = Path(resume).expanduser().resolve(strict=False)
    if getattr(args, "data_classification", "synthetic") != "synthetic":
        try:
            require_private_destination(directory)
        except ArtifactError as exc:
            raise CliError(str(exc)) from exc
    if not directory.is_dir():
        raise CliError(f"--resume-run is not an existing run directory: {directory}")
    meta_path = directory / "run.json"
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise CliError(f"cannot read resume manifest {meta_path}: {exc}") from exc
        if not isinstance(meta, dict):
            raise CliError(f"cannot resume {directory}: run.json is not an object")
        if meta.get("status") == "COMPLETE":
            raise CliError(
                f"{directory} is already COMPLETE according to run.json. Finalized "
                "evidence is immutable even if run.state.json was removed; start a "
                "new run instead."
            )
    state_path = directory / "run.state.json"
    if state_path.exists():
        try:
            state = json.loads(state_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise CliError(f"cannot read resume state {state_path}: {exc}") from exc
        if isinstance(state, dict) and state.get("status") == "COMPLETE":
            raise CliError(
                f"{directory} is already COMPLETE. Finalized evidence is immutable; "
                "start a new run instead of resuming and rewriting it."
            )
    return directory


def _resume_snapshot(directory: Path, name: str) -> Path:
    """Return one immutable input for a resumed run, independent of source mounts."""
    path = directory / "inputs" / name
    if not path.is_file():
        raise CliError(
            f"cannot resume {directory}: immutable snapshot {path} is missing. "
            "Restoring the original source path is not enough because resume is bound "
            "to the exact bytes captured before the first paid call."
        )
    return path


def _snapshot_file(source: Path, target: Path) -> str:
    """Copy one immutable input once; resume refuses changed source bytes."""
    try:
        content = Path(source).read_bytes()
    except OSError as exc:
        raise CliError(f"cannot snapshot input {source}: {exc}") from exc
    if target.exists():
        if target.read_bytes() != content:
            raise CliError(
                f"resume source {source} differs from immutable snapshot {target}. "
                "Start a new run instead of mixing input versions."
            )
    else:
        atomic_write_bytes(target, content)
    return file_sha256(target)


def _snapshot_run_inputs(
    directory: Path,
    *,
    testset: TestSet,
    gt_path: Path,
    prompt: Prompt,
) -> dict[str, Any]:
    inputs = directory / "inputs"
    testset_target = inputs / "testset.jsonl"
    gt_target = inputs / "ground_truth.csv"
    schema_target = inputs / "response_schema.json"
    prompt_target = inputs / "prompt.txt"
    hashes = {
        "testset_sha": _snapshot_file(testset.path, testset_target),
        "gt_sha": _snapshot_file(gt_path, gt_target),
        "schema_sha": _snapshot_file(SCHEMA_PATH, schema_target),
    }
    prompt_bytes = prompt.system_text.encode("utf-8")
    if prompt_target.exists():
        if prompt_target.read_bytes() != prompt_bytes:
            raise CliError(
                f"assembled prompt differs from immutable snapshot {prompt_target}; "
                "start a new run"
            )
    else:
        atomic_write_bytes(prompt_target, prompt_bytes)
    hashes["prompt_sha"] = prompt.sha
    return {
        **hashes,
        "testset_path": testset_target,
        "gt_path": gt_target,
        "schema_path": schema_target,
        "prompt_path": prompt_target,
        "source_testset_path": testset.path.resolve(),
        "source_gt_path": Path(gt_path).resolve(),
    }


def _write_run_state(directory: Path, payload: Mapping[str, Any]) -> None:
    atomic_write_text(
        directory / "run.state.json",
        json.dumps(dict(payload), ensure_ascii=False, indent=2, sort_keys=True) + "\n",
    )


def _write_meta(directory: Path, meta: dict[str, Any]) -> Path:
    path = directory / "run.json"
    atomic_write_text(
        path,
        json.dumps(meta, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
    )
    return path


def _item_result_from_log(row: dict[str, Any]) -> ItemResult:
    """One JSONL line back into the dataclass that wrote it."""
    return ItemResult(
        item_id=row["item_id"],
        call_id=row["call_id"],
        replicate=row["replicate"],
        outcome=row["outcome"],
        parse_ok=row["parse_ok"],
        truncated=row["truncated"],
        payload=row["payload"],
        repairs=tuple(row["repairs"]),
        observed_model=row["observed_model"],
        # `.get` on the four fields added after the 2026-08-04 runs, and only on those
        # four. A run log written before they existed is still a real run, and refusing
        # to load it would mean the fix for "the violations were undiagnosable" started
        # by making the runs that contain them unreadable. Everything above is indexed,
        # so a log missing a field that has always been written is still a KeyError and
        # still refused.
        provider=row.get("provider"),
        generation_id=row.get("generation_id"),
        finish_reason=row.get("finish_reason"),
        raw_content=row.get("raw_content"),
        latency_s=row["latency_s"],
        prompt_tokens=row["prompt_tokens"],
        completion_tokens=row["completion_tokens"],
        reasoning_tokens=row["reasoning_tokens"],
        cost=row["cost"],
        error=row["error"],
        http_status=row.get("http_status"),
        attempt_count=row.get("attempt_count", 1),
        runtime_id=row.get("runtime_id"),
        runtime_backend=row.get("runtime_backend"),
        runtime_fingerprint=row.get("runtime_fingerprint"),
        system_fingerprint=row.get("system_fingerprint"),
    )


def _resolve_run_dir(path: Path) -> Path:
    """Accept the run directory, its `run.json`, or its `run.jsonl`."""
    path = Path(path)
    if path.is_dir():
        return path
    if path.is_file():
        return path.parent
    raise CliError(f"run not found: {path}")


def _recorded_artifact_path(directory: Path, value: object) -> Path:
    """Resolve portable run-relative paths while retaining legacy absolute paths."""
    path = Path(str(value))
    if path.is_absolute():
        return path
    root = Path(directory).resolve()
    resolved = (root / path).resolve(strict=False)
    if resolved != root and root not in resolved.parents:
        raise CliError(f"recorded artifact path {value!r} escapes run directory {root}")
    return resolved


def _run_relative_path(directory: Path, path: Path) -> str:
    """Record a self-contained artifact path relative to its run bundle."""
    root = Path(directory).resolve()
    resolved = Path(path).resolve()
    try:
        return resolved.relative_to(root).as_posix()
    except ValueError as exc:
        raise CliError(
            f"artifact {resolved} is outside run directory {root}; evidence-grade "
            "inputs must be copied into the portable run bundle"
        ) from exc


def load_run(path: Path) -> LoadedRun:
    """Read a run directory back into a `RunResult` plus its recorded provenance.

    Every parse failure here is turned into a `CliError` naming the file. A run
    directory is something a person may have edited, moved or truncated, and a
    `KeyError: 'decoding'` traceback teaches the reader that this tool crashes rather
    than that their run.json is incomplete.
    """
    directory = _resolve_run_dir(path)
    meta_path = directory / "run.json"
    log_path = directory / "run.jsonl"
    for required in (meta_path, log_path):
        if not required.exists():
            raise CliError(
                f"{directory} is not a run directory: {required.name} is missing. A run "
                "directory holds run.json (provenance) and run.jsonl (one row per item "
                "per replicate)."
            )

    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        rows = [
            json.loads(line)
            for line in log_path.read_text(encoding="utf-8").split("\n")
            if line.strip()
        ]
    except (OSError, UnicodeError, ValueError, TypeError) as exc:
        raise CliError(f"{directory} holds unreadable JSON: {exc}") from exc

    if not isinstance(meta, dict):
        raise CliError(f"{meta_path} must contain a JSON object")
    if not all(isinstance(row, dict) for row in rows):
        raise CliError(f"{log_path} must contain one JSON object per non-blank line")

    if not rows:
        raise CliError(f"{log_path} has no rows. There is nothing to compare.")

    _validate_run_artifacts(directory, meta, rows, log_path)

    try:
        config = RunConfig(
            model=meta["model_requested"],
            arm=meta["arm"],
            prompt_id=meta["prompt_id"],
            temperature=meta["decoding"]["temperature"],
            top_p=meta["decoding"]["top_p"],
            seed=meta["decoding"]["seed"],
            max_tokens=meta["decoding"]["max_tokens"],
            repeats=meta["repeats"],
            concurrency=meta["concurrency"],
            provider=meta.get("provider_requested"),
            reasoning_effort=meta.get("decoding", {}).get(
                "reasoning_effort", meta.get("reasoning_effort", "provider-default")
            ),
            max_attempts=meta.get("max_attempts", 3),
        )
        result = RunResult(
            config=config,
            prompt_sha=meta["prompt_sha"],
            testset_sha=meta["testset_sha"],
            results=tuple(_item_result_from_log(row) for row in rows),
        )
    except (KeyError, TypeError, ValueError) as exc:
        raise CliError(
            f"{directory} is missing {exc} in run.json or run.jsonl. Both are written "
            "together by a run; an incomplete pair cannot be scored, and guessing the "
            "missing field would put an invented decoding parameter in the report."
        ) from exc
    if meta.get("outcome_counts") is not None and result.outcome_counts() != meta["outcome_counts"]:
        raise CliError(
            f"{directory} outcome_counts do not match run.jsonl; refusing a summary "
            "that was not computed from the rows being scored"
        )
    return LoadedRun(directory=directory, meta=meta, result=result)


def _validate_run_artifacts(
    directory: Path,
    meta: Mapping[str, Any],
    rows: Sequence[Mapping[str, Any]],
    log_path: Path,
) -> None:
    """Validate integrity, row identity, and the exact item/replicate matrix."""
    artifact_version = meta.get("artifact_schema_version")
    v2_markers = {
        "run_contract",
        "run_contract_sha",
        "run_log_sha256",
        "journal_sha256",
        "application_contract_sha",
    }
    if artifact_version not in (None, 2):
        raise CliError(
            f"{directory} uses unsupported artifact_schema_version "
            f"{artifact_version!r}; refusing to interpret it as a legacy run"
        )
    sibling_v2_artifacts = any(
        (directory / name).exists()
        for name in ("run.state.json", "run.journal.jsonl", "inputs")
    )
    if artifact_version is None and (
        v2_markers.intersection(meta) or sibling_v2_artifacts
    ):
        raise CliError(
            f"{directory} has evidence-grade fields but no artifact_schema_version. "
            "Refusing a metadata downgrade that would bypass integrity checks."
        )
    evidence_grade = artifact_version == 2
    if evidence_grade:
        required_v2 = {
            "status",
            "run_contract",
            "run_contract_sha",
            "run_log_sha256",
            "run_log_bytes",
            "journal_sha256",
            "journal_bytes",
            "application_contract",
            "application_contract_sha",
            "runtime",
            "runtime_fingerprint",
            "execution_contract",
            "execution_sha",
            "arm",
            "model_requested",
            "provider_requested",
            "prompt_id",
            "decoding",
            "testset_path",
            "gt_path",
            "schema_path",
            "prompt_path",
            "testset_sha",
            "gt_sha",
            "schema_sha",
            "prompt_sha",
            "items",
            "repeats",
            "rows",
        }
        missing_v2 = sorted(required_v2 - set(meta))
        if missing_v2:
            raise CliError(
                f"{directory} is missing evidence-grade field(s): {missing_v2}"
            )
        if meta.get("status") != "COMPLETE":
            raise CliError(f"{directory} is not finalized COMPLETE")
        for field in ("items", "repeats", "rows"):
            value = meta.get(field)
            if isinstance(value, bool) or not isinstance(value, int) or value < 1:
                raise CliError(
                    f"{directory} evidence-grade {field} must be a positive integer, "
                    f"found {value!r}"
                )
        if meta["rows"] != meta["items"] * meta["repeats"]:
            raise CliError(
                f"{directory} declares rows={meta['rows']}, but items x repeats is "
                f"{meta['items'] * meta['repeats']}"
            )
        state_path = directory / "run.state.json"
        journal_path = directory / "run.journal.jsonl"
        if not state_path.exists() or not journal_path.exists():
            raise CliError(
                f"{directory} is an evidence-grade run but its state or journal is missing"
            )
        try:
            state = json.loads(state_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, ValueError, TypeError) as exc:
            raise CliError(f"{state_path} is unreadable: {exc}") from exc
        if not isinstance(state, dict) or state.get("status") != "COMPLETE":
            raise CliError(f"{state_path} does not mark the run COMPLETE")
        if canonical_sha(meta.get("run_contract")) != meta.get("run_contract_sha"):
            raise CliError(f"{directory} run_contract_sha does not match run.json")
        if state.get("run_contract_sha") != meta.get("run_contract_sha"):
            raise CliError(f"{state_path} belongs to a different run contract")
        if state.get("run_log_sha256") != meta.get("run_log_sha256"):
            raise CliError(f"{state_path} names a different finalized run log")
        if file_sha256(journal_path) != meta.get("journal_sha256"):
            raise CliError(f"{journal_path} does not match journal_sha256 in run.json")
        if journal_path.stat().st_size != meta.get("journal_bytes"):
            raise CliError(f"{journal_path} size does not match journal_bytes in run.json")
        try:
            application = ApplicationSpec.from_manifest(meta["application_contract"])
            runtime_block = meta["runtime"]
            if not isinstance(runtime_block, Mapping):
                raise ValueError("runtime provenance must be an object")
            runtime_manifest = runtime_block.get("runtime")
            if not isinstance(runtime_manifest, Mapping):
                raise ValueError("runtime provenance has no runtime manifest")
            runtime = RuntimeSpec.from_manifest(runtime_manifest)
        except (TypeError, ValueError) as exc:
            raise CliError(f"{directory} has invalid contract provenance: {exc}") from exc
        if application.fingerprint() != meta.get("application_contract_sha"):
            raise CliError(f"{directory} application_contract_sha does not match its manifest")
        if runtime.fingerprint() != meta.get("runtime_fingerprint"):
            raise CliError(f"{directory} runtime_fingerprint does not match its manifest")
        if runtime_block.get("runtime_fingerprint") != meta.get("runtime_fingerprint"):
            raise CliError(f"{directory} runtime provenance names a different fingerprint")
        execution_contract = meta.get("execution_contract")
        if not isinstance(execution_contract, Mapping):
            raise CliError(f"{directory} execution_contract must be an object")
        if canonical_sha(execution_contract) != meta.get("execution_sha"):
            raise CliError(f"{directory} execution_sha does not match execution_contract")
        if execution_contract.get("runtime_fingerprint") != meta.get(
            "runtime_fingerprint"
        ):
            raise CliError(
                f"{directory} execution_contract names a different runtime fingerprint"
            )
        contract = meta.get("run_contract")
        if not isinstance(contract, Mapping):
            raise CliError(f"{directory} run_contract must be an object")
        for field in (
            "application_contract_sha",
            "runtime_fingerprint",
            "testset_sha",
            "gt_sha",
            "schema_sha",
            "prompt_sha",
        ):
            if contract.get(field) != meta.get(field):
                raise CliError(
                    f"{directory} run_contract {field} does not match run.json"
                )
        for path_field, sha_field in (
            ("testset_path", "testset_sha"),
            ("gt_path", "gt_sha"),
            ("schema_path", "schema_sha"),
            ("prompt_path", "prompt_sha"),
        ):
            snapshot_path = _recorded_artifact_path(directory, meta[path_field])
            if not snapshot_path.is_file():
                raise CliError(f"{directory} immutable snapshot is missing: {snapshot_path}")
            if file_sha256(snapshot_path) != meta[sha_field]:
                raise CliError(
                    f"{snapshot_path} does not match {sha_field} in run.json"
                )
        try:
            journal = RunJournal.open(journal_path, contract)
        except ArtifactError as exc:
            raise CliError(f"{journal_path} is not a valid run journal: {exc}") from exc
        if journal.trailing_torn_record:
            raise CliError(
                f"{journal_path} ends in a torn record. Prior paid results remain "
                "recoverable, but a finalized evidence bundle cannot contain an "
                "ambiguous tail."
            )
        unresolved = journal.unresolved_cells()
        if unresolved:
            raise CliError(
                f"{journal_path} has {len(unresolved)} durably started cell(s) with no "
                f"durable result; first unresolved cell is {unresolved[0]}"
            )
    recorded_sha = meta.get("run_log_sha256")
    if recorded_sha is not None and file_sha256(log_path) != recorded_sha:
        raise CliError(f"{log_path} does not match run_log_sha256 in run.json")
    recorded_bytes = meta.get("run_log_bytes")
    if recorded_bytes is not None and log_path.stat().st_size != recorded_bytes:
        raise CliError(f"{log_path} size does not match run_log_bytes in run.json")
    if meta.get("rows") is not None and len(rows) != meta["rows"]:
        raise CliError(
            f"{directory} declares {meta['rows']} rows but run.jsonl contains {len(rows)}"
        )

    identity_fields = {
        "arm": meta.get("arm"),
        "model_requested": meta.get("model_requested"),
        "provider_requested": meta.get("provider_requested"),
        "prompt_id": meta.get("prompt_id"),
        "prompt_sha": meta.get("prompt_sha"),
        "testset_sha": meta.get("testset_sha"),
        "reasoning_effort": (
            meta.get("decoding", {}).get("reasoning_effort")
            if isinstance(meta.get("decoding"), Mapping)
            else None
        ),
    }
    cells: set[tuple[str, int]] = set()
    log_rows_by_cell: dict[tuple[str, int], Mapping[str, Any]] = {}
    calls: dict[str, str] = {}
    replicates_by_item: dict[str, set[int]] = {}
    for index, row in enumerate(rows, start=1):
        if evidence_grade:
            required_row_fields = {
                *identity_fields,
                "item_id",
                "call_id",
                "replicate",
                "outcome",
                "parse_ok",
                "runtime_id",
                "runtime_backend",
                "runtime_fingerprint",
            }
            missing_row = sorted(required_row_fields - set(row))
            if missing_row:
                raise CliError(
                    f"{log_path}:{index} is missing evidence identity field(s): "
                    f"{missing_row}"
                )
        for field, expected in identity_fields.items():
            if field in row and row[field] != expected:
                raise CliError(
                    f"{log_path}:{index} records {field}={row[field]!r}, expected "
                    f"{expected!r} from run.json"
                )
        try:
            raw_item_id = row["item_id"]
            raw_replicate = row["replicate"]
            raw_call_id = row["call_id"]
        except KeyError as exc:
            raise CliError(f"{log_path}:{index} has no valid logical-cell identity") from exc
        if evidence_grade:
            if not isinstance(raw_item_id, str) or not raw_item_id:
                raise CliError(f"{log_path}:{index} item_id must be a non-empty string")
            if not isinstance(raw_call_id, str) or not raw_call_id:
                raise CliError(f"{log_path}:{index} call_id must be a non-empty string")
            if (
                isinstance(raw_replicate, bool)
                or not isinstance(raw_replicate, int)
                or raw_replicate < 1
            ):
                raise CliError(
                    f"{log_path}:{index} replicate must be a positive integer, "
                    f"found {raw_replicate!r}"
                )
            if not isinstance(row.get("outcome"), str) or not row["outcome"]:
                raise CliError(f"{log_path}:{index} outcome must be a non-empty string")
            if not isinstance(row.get("parse_ok"), bool):
                raise CliError(f"{log_path}:{index} parse_ok must be boolean")
            item_id = raw_item_id
            replicate = raw_replicate
            call_id = raw_call_id
        else:
            try:
                item_id = str(raw_item_id)
                replicate = int(raw_replicate)
                call_id = str(raw_call_id)
            except (TypeError, ValueError) as exc:
                raise CliError(
                    f"{log_path}:{index} has no valid logical-cell identity"
                ) from exc
        if replicate < 1:
            raise CliError(f"{log_path}:{index} has invalid replicate {replicate}")
        row_runtime = row.get("runtime_fingerprint")
        if row_runtime is not None and row_runtime != meta.get("runtime_fingerprint"):
            raise CliError(
                f"{log_path}:{index} records runtime fingerprint {row_runtime!r}, "
                f"expected {meta.get('runtime_fingerprint')!r}"
            )
        if evidence_grade:
            expected_runtime = runtime.fingerprint()
            if row_runtime != expected_runtime:
                raise CliError(
                    f"{log_path}:{index} records runtime fingerprint {row_runtime!r}, "
                    f"expected {expected_runtime!r}"
                )
            if row.get("runtime_id") != runtime.runtime_id:
                raise CliError(
                    f"{log_path}:{index} records runtime_id {row.get('runtime_id')!r}, "
                    f"expected {runtime.runtime_id!r}"
                )
            if row.get("runtime_backend") != runtime.backend.value:
                raise CliError(
                    f"{log_path}:{index} records runtime_backend "
                    f"{row.get('runtime_backend')!r}, expected {runtime.backend.value!r}"
                )
        cell = (item_id, replicate)
        if cell in cells:
            raise CliError(f"{log_path} repeats logical cell {cell}")
        cells.add(cell)
        log_rows_by_cell[cell] = row
        previous_call = calls.setdefault(item_id, call_id)
        if previous_call != call_id:
            raise CliError(
                f"{log_path} assigns item {item_id!r} to multiple call ids: "
                f"{previous_call!r} and {call_id!r}"
            )
        replicates_by_item.setdefault(item_id, set()).add(replicate)

    expected_repeats = meta.get("repeats")
    expected_items = meta.get("items")
    if evidence_grade and isinstance(expected_repeats, int) and expected_repeats > 0:
        wanted = set(range(1, expected_repeats + 1))
        bad = {
            item_id: sorted(values)
            for item_id, values in replicates_by_item.items()
            if values != wanted
        }
        if bad:
            first = next(iter(bad.items()))
            raise CliError(
                f"{log_path} does not contain the exact replicate matrix; first "
                f"mismatch is {first[0]} -> {first[1]}, expected {sorted(wanted)}"
            )
    if isinstance(expected_items, int) and len(replicates_by_item) != expected_items:
        raise CliError(
            f"{directory} declares {expected_items} items but run.jsonl contains "
            f"{len(replicates_by_item)} distinct item ids"
        )
    if evidence_grade:
        journal_rows_by_cell = {
            (row["item_id"], row["replicate"]): row
            for row in journal.completed_rows()
        }
        if set(journal_rows_by_cell) != set(log_rows_by_cell):
            missing = sorted(set(journal_rows_by_cell) - set(log_rows_by_cell))[:3]
            extra = sorted(set(log_rows_by_cell) - set(journal_rows_by_cell))[:3]
            raise CliError(
                f"{directory} finalized log and paid-call journal contain different "
                f"logical cells (missing_from_log={missing}, extra_in_log={extra})"
            )
        for cell, row in log_rows_by_cell.items():
            if canonical_sha(row) != canonical_sha(journal_rows_by_cell[cell]):
                raise CliError(
                    f"{directory} finalized log differs from the paid-call journal "
                    f"for logical cell {cell}"
                )


# ==================================================================== the bridge


def replicate_records(
    result: RunResult, items: Sequence[TestItem]
) -> list[list[Record]]:
    """Model payloads -> scorable records, one list per replicate, in replicate order.

    This is the grain change, and it is the step to distrust: `flatten.to_rows` turns a
    CALL-shaped payload into PRODUCT-shaped rows, because the scorer merges on
    `(call_id, phone_number, product)` (`fact_checker.py:1075`). RET-16 is one call with
    three products and therefore three rows. An implementation that emitted one row per
    call would still score, still print, and silently lose two thirds of that item.

    Failures are included, never skipped. `to_rows` is documented to return a
    ground-truth-shaped skeleton for a payload that could not be read, so every item
    contributes its rows on every replicate -- which is what keeps the two arms'
    denominators equal (`flatten.py`, "Failure rows keep the grain").

    **The grain is checked here, because here is where both numbers exist.** `to_rows`
    deliberately does not lowercase -- `records.from_row` -> `norm_text` folds case
    once, and `flatten.py` argues at length that two normalisation sites is one too
    many. The cost of that correct decision is that a payload naming both `Postpaid`
    and `POSTPAID` (or `" Postpaid "`) emits two rows that collapse to one merge key,
    and `metrics.outer_join` builds `{r.key: r for r in pred}` (`metrics.py:122`) and
    silently keeps whichever came LAST. Measured: the surviving `call_result` is the
    second one, and the discarded row is counted nowhere in `Coverage`. This is the one
    seam where `len(rows)` and `len({record.key})` are both in scope, so it is the only
    place the loss can be turned into a number.
    """
    by_id = {item.item_id: item for item in items}
    per_replicate: dict[int, list[Record]] = {}
    for row in result.results:
        item = by_id.get(row.item_id)
        if item is None:
            raise CliError(
                f"the run log contains item_id {row.item_id!r}, which is not in the "
                "testset it names. The run and the testset have drifted apart; scoring "
                "would compare a model's answers against labels for other items."
            )
        flat_rows = to_rows(row.payload, item, parse_ok=row.parse_ok)
        records = [from_row(flat) for flat in flat_rows]
        _refuse_colliding_rows(row, flat_rows, records)
        per_replicate.setdefault(row.replicate, []).extend(records)
    return [per_replicate[key] for key in sorted(per_replicate)]


def _require_run_matrix(loaded: LoadedRun, testset: TestSet) -> None:
    """Bind a run log to the exact item/call matrix in the hashed testset."""
    expected_calls: dict[str, str] = {}
    for item in testset.items:
        if item.item_id in expected_calls:
            raise CliError(f"testset repeats item id {item.item_id!r}")
        expected_calls[item.item_id] = item.call_id
    observed: dict[str, set[int]] = {}
    wrong_calls = 0
    for row in loaded.result.results:
        observed.setdefault(row.item_id, set()).add(row.replicate)
        if expected_calls.get(row.item_id) != row.call_id:
            wrong_calls += 1
    missing = set(expected_calls) - set(observed)
    extra = set(observed) - set(expected_calls)
    wanted_replicates = set(range(1, loaded.result.config.repeats + 1))
    bad_replicates = sum(values != wanted_replicates for values in observed.values())
    if missing or extra or wrong_calls or bad_replicates:
        raise CliError(
            f"{loaded.directory} does not match the exact testset task matrix: "
            f"{len(missing)} missing item id(s), {len(extra)} extra item id(s), "
            f"{wrong_calls} item/call mismatch(es), {bad_replicates} replicate "
            "matrix mismatch(es)."
        )


def _refuse_colliding_rows(
    row: ItemResult, flat_rows: Sequence[dict], records: Sequence[Record]
) -> None:
    """One item's rows must have one merge key each. See `replicate_records`.

    A refusal rather than a warning, and it is affordable: the run log is already on
    disk, so nothing paid for is lost and the payload that caused it can be read. The
    alternative is a scorer that drops a prediction with no number moving anywhere,
    which is the failure mode this whole repository is arranged against.

    **Both product spellings are named, the raw one and the folded one.** The message
    is the only thing a reader gets before they open `run.jsonl`, and the merge key is
    already normalised by the time `records` exist -- reporting it alone prints
    `products ['postpaid', 'postpaid']`, which reads as an impossible duplicate rather
    than as the case collision it is, and does not say which key to look for. The raw
    rows are `to_rows` output, where the model's own spelling survives.

    It should be unreachable in a run made by this CLI. `schemas/retention.json` sets
    `additionalProperties: false` over `Postpaid`/`TOL`/`TVS`/`unknown`, and
    `request.build_request` sends `provider: {"require_parameters": true}` so a
    provider that cannot enforce that is refused rather than silently routed to. That
    is exactly why it is worth asserting: it is the check that tells you the
    enforcement stopped working, on the run where it stopped.
    """
    keys = {record.key for record in records}
    if len(keys) == len(records):
        return
    raw = [flat.get("product") for flat in flat_rows]
    folded = [record.product for record in records]
    raise CliError(
        f"{row.item_id} replicate {row.replicate} flattened to {len(records)} row(s) "
        f"but only {len(keys)} distinct merge key(s).\n"
        f"    the payload named : {raw}\n"
        f"    which folded into : {folded}\n"
        "Two product names that differ only in case or padding fold together in "
        "records.norm_text, and metrics.outer_join keeps whichever came last "
        "(metrics.py:122) -- the other prediction is discarded with no number moving in "
        "Coverage. Refusing to score rather than reporting the survivor as the arm's "
        "answer. The payload is in run.jsonl."
    )


def _score(gt: Sequence[Record], pred: Sequence[Record]) -> dict[str, object]:
    """The three dimensions, three denominators. See `metrics.py:11-13`."""
    return {
        name: scorer(list(gt), list(pred), classes) for name, scorer, classes in _SCORERS
    }


# ================================================================== the cost accounting
#
# `run.jsonl` has carried tokens, cost and latency per call since the runner was
# written, and the xlsx export prints them on its "Per call" sheet -- but the TEXT
# report showed none of it, so a reader of `out/reports/compare-*.txt` could see which
# mechanisms an arm passed and nothing at all about what asking it cost. These four
# helpers are the whole of the addition; `report.py` renders them and computes none of
# them, which is the same division of labour every other number in that file follows.


def _token_totals(rows: Sequence[ItemResult]) -> tuple[int, int, int]:
    """(prompt, completion, reasoning) summed over every row, a missing count as zero.

    Zero for a missing value is safe HERE and is not safe for the cost, and the
    difference is worth stating because the two lines look identical. These are sums
    with no denominator attached: a row that reported no usage contributes nothing
    whether it is skipped or added as zero, and the total is a floor either way. The
    cost cannot be treated that way, because a zero there is a positive claim that the
    call was free -- so `_cost_totals` returns the count of the rows it could not see.

    `reasoning_tokens` is the one most often absent: `client.py:271` reads it off
    `usage.completion_tokens_details`, which a backend without a reasoning trace does
    not send. It is a BREAKDOWN of the completion tokens and is therefore already
    inside the second number, never to be added to it -- `report._performance_section`
    prints that in the report, where the person adding two columns together is.
    """
    return (
        sum(row.prompt_tokens or 0 for row in rows),
        sum(row.completion_tokens or 0 for row in rows),
        sum(row.reasoning_tokens or 0 for row in rows),
    )


def _cost_totals(rows: Sequence[ItemResult]) -> tuple[float, int]:
    """(cost in USD, how many rows reported no cost at all). The pair, never the sum.

    The sum alone is `runner.total_cost()` restricted to these rows, and it is a LOWER
    BOUND for the reason that method gives (`runner.py:411-421`): OpenRouter reports
    `usage.cost` only for the providers that supply it, `client.py:272` keeps a missing
    value as None rather than zero, and a None is skipped rather than counted free.

    A floor whose coverage is invisible is worse than no number, because it is quotable:
    0.02 USD over 40 calls and 0.02 USD over the 3 calls that happened to report both
    print as 0.02. So the count travels with the total, all the way onto the page.
    """
    reported = [row.cost for row in rows if row.cost is not None]
    return (sum(reported), len(rows) - len(reported))


def _latency_stats(rows: Sequence[ItemResult]) -> tuple[float | None, float | None]:
    """(median, max) request seconds over EVERY call, including the ones that died.

    Every call, not the parsed ones. A median computed over successes alone gets better
    as the failures get worse, which is the wrong direction for the one number whose job
    is to say what this arm cost in wall clock. A 120s timeout is 120 seconds that were
    spent, and `outcome_counts` in section 3 is what says how many rows were timeouts.

    `ItemResult.latency_s` is the LAST attempt's request time with backoff sleeps
    excluded (`runner.py:236-238`), so a retried item reports the attempt that answered
    rather than the wall clock it occupied. That is the runner's definition and this
    function does not second-guess it; the report states it so the number is read as
    what it is.

    (None, None) for no rows, rather than (0.0, 0.0). A zero-second median is a claim
    about a run that did not happen.
    """
    values = [row.latency_s for row in rows]
    if not values:
        return (None, None)
    return (statistics.median(values), max(values))


def _correct_answers(dimensions: Mapping[str, object], dimension: str) -> int:
    """How many rows the named dimension got RIGHT, off the scorer's own result object.

    `sum(c.tp for c in DimensionResult.classes)` -- and for a one-vs-rest split over a
    single-label dimension that sum IS the count of correct rows, because a row can
    contribute a true positive to at most one class: `tp` for class `c` needs the ground
    truth to be `c` and the prediction to be `c`, and a row has one of each. Rows whose
    ground-truth label sits outside the label space contribute nothing, which is right:
    a label the scorer cannot score is not one an arm can get right.

    Zero for a dimension that was not scored, rather than a guess. `report` prints "0
    correct rows, and a cost divided by zero correct answers is not a number" rather
    than a ratio, so an unscored dimension cannot become a cost per correct answer that
    happens to look plausible.
    """
    result = dimensions.get(dimension)
    if not isinstance(result, DimensionResult):
        return 0
    true_positives, _fp, _fn, _tn = result.totals()
    return true_positives


def arm_summary(
    loaded: LoadedRun,
    gt: Sequence[Record],
    per_replicate: Sequence[Sequence[Record]],
) -> ArmSummary:
    """Everything the report prints about one arm.

    `dimensions` is scored on replicate 1 alone. See this module's docstring: pooling
    replicates would let `outer_join`'s dict silently keep the last one, and the table
    would describe a run that never happened. `n_flip` and the mechanism verdicts read
    every replicate, which is where the replicates earn their cost.

    `answered_nothing` counts every ROW of the run, not replicate 1, because
    `outcome_counts` does and the two print beside each other. It is the count
    `Coverage.parse_failures` cannot see: a response that parsed, satisfied every
    required key and still named no product takes the ground-truth skeleton with
    `parse_ok=True`, scores a product true positive per ground-truth product, and moves
    nothing in coverage (`flatten.py`, KNOWN CONSEQUENCE). Counted from the payloads
    rather than from the flattened rows so it cannot be confused with a parse failure.

    **Tokens, cost and latency are run totals; cost per correct answer is not.** The
    first three describe what was spent, so they cover every row the run made -- the
    replicates were paid for whether or not they were scored. The ratio cannot: its
    denominator is `dimensions`, which is one replicate, so its numerator has to be that
    same replicate's cost or the figure is `repeats` times the number an app owner would
    act on. Production makes one call per item.

    **Which replicate, read off the log rather than assumed to be 1.** `dimensions` is
    scored on `per_replicate[0]`, and `replicate_records` returns
    `[per_replicate[key] for key in sorted(per_replicate)]` -- index 0 is the LOWEST
    replicate number present, which is 1 in every log this CLI writes and need not be in
    a log that was filtered by hand. Taking the minimum makes the numerator select the
    same calls the denominator scored by construction, instead of by a convention that
    is true today. A hardcoded 1 against a log starting at replicate 2 would divide an
    empty cost by a real hit count and print 0.000000 USD per correct answer.
    """
    rows = loaded.result.results
    dimensions = _score(gt, per_replicate[0])

    prompt_tokens, completion_tokens, reasoning_tokens = _token_totals(rows)
    cost_usd, calls_without_cost = _cost_totals(rows)
    latency_median, latency_max = _latency_stats(rows)

    scored_replicate = min((row.replicate for row in rows), default=None)
    scored_rows = [row for row in rows if row.replicate == scored_replicate]
    scored_cost, scored_without_cost = _cost_totals(scored_rows)
    correct = _correct_answers(dimensions, COST_PER_CORRECT_DIMENSION)

    # Two ways this ratio must refuse to exist, and they are different findings, so the
    # report is handed None and the counts rather than a number and a footnote:
    #   * no correct answer -- a division by zero, and "infinite" is not a cost;
    #   * no row of the scored replicate reported a cost at all -- the numerator is
    #     UNKNOWN, and 0.000000 USD per correct answer reads as a free model.
    # A partially reported numerator is allowed through and is a lower bound, which is
    # what the whole section says of every number in it.
    cost_per_correct = None
    if correct and scored_without_cost < len(scored_rows):
        cost_per_correct = scored_cost / correct

    return ArmSummary(
        arm=loaded.arm,
        model=loaded.result.config.model,
        prompt_sha=loaded.result.prompt_sha,
        observed_models=loaded.result.observed_models(),
        outcome_counts=loaded.result.outcome_counts(),
        n_flip=n_flip(per_replicate),
        dimensions=dimensions,
        replicates=loaded.result.config.repeats,
        decoding=loaded.decoding,
        answered_nothing=sum(
            1
            for row in loaded.result.results
            if named_no_product(row.payload, parse_ok=row.parse_ok)
        ),
        calls=len(rows),
        prompt_tokens=prompt_tokens,
        completion_tokens=completion_tokens,
        reasoning_tokens=reasoning_tokens,
        cost_usd_lower_bound=cost_usd,
        calls_without_cost=calls_without_cost,
        latency_median_s=latency_median,
        latency_max_s=latency_max,
        cost_per_correct_usd=cost_per_correct,
        cost_per_correct_dimension=COST_PER_CORRECT_DIMENSION,
        correct_answers=correct,
        scored_replicate_cost_usd=scored_cost,
    )


# ================================================================ the dry client


@dataclass(frozen=True)
class _NoCall:
    """What the dry-run client returns in place of a completion.

    Empty content on purpose. `outcomes.classify` files it as `empty_other`, which is
    never `parse_ok`, so nothing a dry run produces can be mistaken for a measurement
    if it is ever read by accident. The alternative -- returning a well-formed payload
    -- would make a dry run's rows score, and a scored dry run is a fabricated result.

    It also cannot be retried: `empty_other` is a classified response, not a transport
    failure, and `runner._is_retryable` never sees it.
    """

    content: str | None = None
    finish_reason: str | None = None
    observed_model: str | None = None
    generation_id: str | None = None
    provider: str | None = None
    prompt_tokens: int | None = None
    completion_tokens: int | None = None
    reasoning_tokens: int | None = None
    cost: float | None = None
    latency_s: float = 0.0


class _DryRunClient:
    """Records the request bodies the runner would send, and sends none of them.

    Substituted for `OpenRouterClient` at the seam the runner already has (`ClientLike`
    is a Protocol precisely so this is possible), so the dry run exercises the real run
    loop: the same prompt assembly, the same decoding parameters, the same per-item
    ordering. The bodies are built by `request.build_request`, which is what
    `client.complete` sends -- so what is written out is what would have gone over the
    wire, not a reconstruction of it.
    """

    def __init__(self, testset: TestSet, runtime: RuntimeSpec = OPENROUTER_RUNTIME) -> None:
        self._by_transcript = {item.transcript_th: item for item in testset.items}
        self.requests: list[dict[str, Any]] = []
        self.runtime = runtime

    def complete(
        self,
        *,
        model: str,
        messages: Sequence[dict[str, Any]],
        max_tokens: int,
        temperature: float,
        top_p: float | None = None,
        seed: int | None = None,
        response_format: dict[str, Any] | None = None,
        provider: str | None = None,
        reasoning_effort: str = "provider-default",
    ) -> _NoCall:
        item = self._by_transcript[messages[1]["content"]]
        request = build_runtime_request(
            self.runtime,
            model=model,
            messages=messages,
            max_tokens=max_tokens,
            temperature=temperature,
            top_p=top_p,
            seed=seed,
            response_format=response_format,
            provider=provider,
            reasoning_effort=reasoning_effort,
        )
        self.requests.append(
            {
                "item_id": item.item_id,
                "call_id": item.call_id,
                "runtime_fingerprint": self.runtime.fingerprint(),
                "request": request,
            }
        )
        return _NoCall()


def build_client(
    api_key: str,
    *,
    timeout: float = 120.0,
    runtime: RuntimeSpec = OPENROUTER_RUNTIME,
):
    """Construct the real OpenAI-compatible client for a reviewed runtime.

    A named module function rather than an inline constructor so a test can substitute
    a fake without a network, and so `openai` is imported only when a real run is
    actually about to happen: `check` and `--dry-run` must stay runnable in an
    environment with no SDK installed, which is the environment CI builds (the root
    `requirements.txt` deliberately omits `openai`).
    """
    from evalgen.client import OpenAICompatibleClient

    return OpenAICompatibleClient(api_key, runtime=runtime, timeout=timeout)


def _api_key(runtime: RuntimeSpec = OPENROUTER_RUNTIME) -> str:
    for path in ENV_FILES:
        load_env_file(path)
    if runtime == OPENROUTER_RUNTIME:
        value, name = find_api_key()
    else:
        value, name = find_runtime_api_key(runtime)
    if not value:
        if runtime != OPENROUTER_RUNTIME:
            raise CliError(
                f"no API key found in {runtime.api_key_env} for runtime "
                f"{runtime.runtime_id!r}. For a local endpoint with authentication "
                "disabled, set an explicit documented placeholder value; the SDK "
                "still requires a non-empty key."
            )
        raise CliError(
            "no OpenRouter API key found. Set one of "
            f"{', '.join(('OPENROUTER_API_KEY', 'OPEN_ROUTER_API', 'OPENROUTER_KEY'))} "
            f"in the environment or in one of {[str(p) for p in ENV_FILES]}. Use "
            "--dry-run to review the request bodies without a key."
        )
    print(f"  api key      loaded from {name}")
    return value


def _metadata_pairs(values: Sequence[str] | None) -> dict[str, str]:
    result: dict[str, str] = {}
    for value in values or ():
        if "=" not in value:
            raise CliError(f"runtime metadata must be KEY=VALUE, found {value!r}")
        key, raw = value.split("=", 1)
        if not key.strip() or not raw.strip() or key.strip() in result:
            raise CliError(f"runtime metadata has blank or duplicate key: {value!r}")
        result[key.strip()] = raw.strip()
    return result


def _runtime_from_args(args: argparse.Namespace) -> RuntimeSpec:
    manifest_path = getattr(args, "runtime_manifest", None)
    backend = getattr(args, "runtime_backend", None) or "openrouter"
    metadata = _metadata_pairs(getattr(args, "runtime_metadata", None))
    try:
        if manifest_path:
            conflicting = [
                name
                for name, value in (
                    ("--runtime-backend", getattr(args, "runtime_backend", None)),
                    ("--runtime-id", getattr(args, "runtime_id", None)),
                    ("--base-url", getattr(args, "base_url", None)),
                    ("--api-key-env", getattr(args, "api_key_env", None)),
                    ("--runtime-metadata", metadata),
                    (
                        "--allow-insecure-http",
                        bool(getattr(args, "allow_insecure_http", False)),
                    ),
                )
                if value
            ]
            if conflicting:
                raise CliError(
                    "--runtime-manifest is the complete runtime contract; do not "
                    f"combine it with {', '.join(conflicting)}"
                )
            path = Path(manifest_path)
            try:
                raw_manifest = json.loads(path.read_text(encoding="utf-8"))
            except (OSError, UnicodeError, json.JSONDecodeError) as exc:
                raise CliError(f"cannot read runtime manifest {path}: {exc}") from exc
            if not isinstance(raw_manifest, dict):
                raise CliError(f"runtime manifest {path} must contain a JSON object")
            runtime = RuntimeSpec.from_manifest(raw_manifest)
        elif backend == "openrouter":
            runtime = RuntimeSpec.openrouter(
                runtime_id=getattr(args, "runtime_id", None) or "openrouter",
                api_key_env=getattr(args, "api_key_env", None) or "OPENROUTER_API_KEY",
                base_url=getattr(args, "base_url", None) or OPENROUTER_RUNTIME.base_url,
                build_metadata=metadata,
            )
        else:
            runtime = local_gpu_runtime(
                runtime_id=getattr(args, "runtime_id", None) or "local-gpu",
                base_url=getattr(args, "base_url", None) or "http://127.0.0.1:8000/v1",
                api_key_env=getattr(args, "api_key_env", None) or "EVALGEN_GPU_API_KEY",
                build_metadata=metadata,
                allow_insecure_http=bool(getattr(args, "allow_insecure_http", False)),
            )
    except CliError:
        raise
    except (TypeError, ValueError) as exc:
        raise CliError(f"invalid runtime configuration: {exc}") from exc
    if runtime.backend is RuntimeBackend.OPENAI_COMPATIBLE and getattr(
        args, "provider", None
    ) is not None:
        raise CliError("--provider is OpenRouter-only; omit it for an internal GPU runtime")
    if runtime.backend is RuntimeBackend.OPENAI_COMPATIBLE and getattr(
        args, "reasoning_effort", "provider-default"
    ) != "provider-default":
        raise CliError(
            "--reasoning-effort is OpenRouter-only for now; identify a GPU reasoning "
            "variant through --model and immutable runtime metadata"
        )
    return runtime


# ==================================================================== subcommands


def cmd_check(args: argparse.Namespace) -> int:
    """Validate the pack. No network, no key, no cost.

    Three things are checked and each has a distinct failure: the prompt assembles from
    its committed assets, the testset's evidence spans still appear verbatim in their
    own transcripts, and the two ground truths agree with each other. The first two
    have owners (`prompts.PromptError`, `testsets.validate`); the third has none, which
    is why `_gt_disagreements` exists.
    """
    testset = _load_testset(Path(args.testset), app=args.app)
    gt = _load_gt(Path(args.gt))
    prompt = _prompt(args.prompt_id)

    print(f"testset      {args.testset}")
    print(f"  sha        {testset_sha(args.testset)}")
    print(f"  items      {len(testset.items)}")
    families: dict[str, int] = {}
    for item in testset.items:
        families[item.family] = families.get(item.family, 0) + 1
    print(f"  families   {', '.join(f'{k}={v}' for k, v in families.items())}")
    print(f"ground truth {args.gt}")
    print(f"  rows       {len(gt)}  (one per product, not one per call)")
    print(f"prompt       {prompt.id}")
    print(f"  sha        {prompt.sha}")
    print(f"  chars      {len(prompt.system_text)}")
    schema = response_format()
    required = tuple(schema["json_schema"]["schema"].get("required", ()))
    print(f"schema       {SCHEMA_PATH.name}  required keys: {', '.join(required)}")

    problems = validate(testset) + _gt_disagreements(testset, gt)
    if problems:
        print(f"\n{len(problems)} PROBLEM(S):")
        for problem in problems:
            print(f"  - {problem}")
        return EXIT_PROBLEMS

    print("\nOK. No problem found. This says the pack is internally consistent;")
    print("it says nothing about any model.")
    return EXIT_OK


def _subset_testset(testset: TestSet, item_ids: Sequence[str], directory: Path) -> TestSet:
    """Write the named items to their own file and load it back.

    Written to a real file rather than filtered in memory because `RunResult` records
    `testset_sha(testset.path)`, and a sha describing 20 items over a run of 3 is
    provenance that is worse than none: it reads as though the whole pack was run.
    The subset file stays in the run directory beside the log it describes.
    """
    wanted = {item_id.strip() for item_id in item_ids if item_id.strip()}
    known = {item.item_id for item in testset.items}
    missing = sorted(wanted - known)
    if missing:
        raise CliError(
            f"--items names {missing}, which are not in {testset.path}. Known ids: "
            f"{', '.join(sorted(known))}."
        )

    # Testset order, not command-line order: the run log's order is the testset's
    # (runner.run), and a subset that reordered them would produce logs that cannot be
    # diffed against the full run.
    lines = [
        line
        for line in testset.path.read_text(encoding="utf-8").split("\n")
        if line.strip() and json.loads(line)["item_id"] in wanted
    ]
    target = directory / "testset.jsonl"
    content = ("\n".join(lines) + "\n").encode("utf-8")
    if target.exists() and target.read_bytes() != content:
        raise CliError(
            f"existing subset {target} differs from the requested items; start a new "
            "run instead of changing an incomplete run's workload"
        )
    if not target.exists():
        atomic_write_bytes(target, content)
    return _load_testset(target, app=testset.app)


def _preflight(testset: TestSet, gt_path: Path) -> None:
    """Everything that must hold BEFORE a single token is bought.

    Every one of these fails late and expensively otherwise:

      * The ground truth is not read until the run is over -- `manifest.file_hash` is
        called while writing `run.json`. A missing or moved CSV therefore raises after
        every call has been paid for, with the results already in memory and nothing
        written yet.
      * A pack whose labels have drifted from its transcripts still runs, still
        returns, and still scores. The numbers are simply about labels nobody can
        defend. `validate` costs milliseconds; the run costs money and an hour.
      * The item's own `gt` and the ground-truth CSV are TWO ground truths, and
        `_gt_disagreements` is the only thing that ever puts them side by side.
        `flatten._gt_skeleton` builds failure rows from the item while `metrics.*`
        scores against the CSV, so a disagreement moves failure rows off the merge key
        -- on the arm that failed, which is where a lower number surprises nobody. It
        used to be checked in `check` alone, which is a command a caller can skip. It
        costs one pass over 22 rows, and it is the same argument the bullet above it
        already makes for checking the CSV's existence here.

    The ground truth is SCOPED to the call ids this run will actually request, the way
    `cmd_stability` scopes it before scoring. `_gt_disagreements` also reports CSV rows
    no testset item claims, which is a real problem for the whole pack and a normal
    state for a `--items` subset; unscoped, every stability run would refuse on 17
    rows it was never asked about. Whether the full CSV is claimed is `check`'s
    question, and `check` still asks it.

    `check` reports all of these in full and reads them as a list. Here they are a
    refusal, because at this point the next thing that happens is spending.
    """
    if not gt_path.exists():
        raise CliError(
            f"ground truth not found: {gt_path}. Checked before the run rather than "
            "after it: run.json hashes this file when the run finishes, so a missing "
            "path would otherwise raise once every call had already been paid for."
        )
    call_ids = {item.call_id for item in testset.items}
    scoped_gt = [record for record in _load_gt(gt_path) if record.call_id in call_ids]
    problems = validate(testset) + _gt_disagreements(testset, scoped_gt)
    if problems:
        raise CliError(
            f"{testset.path} has {len(problems)} validation problem(s), the first being: "
            f"{problems[0]} Run `evalgen check` for the full list. Refusing before the "
            "run rather than scoring a model against labels that have drifted away from "
            "the transcripts they describe."
        )


def _execute_run(
    args: argparse.Namespace,
    *,
    testset: TestSet,
    gt_path: Path,
    directory: Path,
) -> tuple[int, LoadedRun | None]:
    """Run one arm, or dry-run it. Returns (exit code, the loaded run or None)."""
    application = _application_spec(args.app)
    prompt = _prompt(args.prompt_id)
    runtime = _runtime_from_args(args)
    config = RunConfig(
        model=args.model,
        arm=args.arm,
        prompt_id=prompt.id,
        temperature=args.temperature,
        top_p=args.top_p,
        seed=args.seed,
        max_tokens=args.max_tokens,
        repeats=args.repeats,
        concurrency=args.concurrency,
        provider=args.provider,
        reasoning_effort=args.reasoning_effort,
        max_attempts=args.max_attempts,
    )
    classification_testset_path = Path(
        getattr(args, "classification_testset_path", testset.path)
    )
    classification_gt_path = Path(
        getattr(args, "classification_gt_path", gt_path)
    )

    print(f"arm          {config.arm}")
    print(f"model        {config.model}")
    print(
        f"runtime      {runtime.runtime_id} ({runtime.backend.value})\n"
        f"  endpoint   {runtime.base_url}\n"
        f"  sha        {runtime.fingerprint()}"
    )
    if config.provider:
        print(f"provider     {config.provider}  (pinned, allow_fallbacks=false)")
    elif runtime.backend is RuntimeBackend.OPENROUTER:
        print(
            "provider     UNPINNED -- the router may serve this arm from more than one "
            "backend.\n             Check prompt_token_spread afterwards; a 60-call run "
            "was split this way\n             on 2026-08-04 and observed_models showed "
            "one id throughout."
        )
    print(f"testset      {testset.path}  ({len(testset.items)} items)")
    print(f"  sha        {testset_sha(testset.path)}")
    print(f"prompt       {prompt.id}")
    print(f"  sha        {prompt.sha}")
    print(
        f"decoding     temperature={config.temperature} top_p={config.top_p} "
        f"seed={config.seed} max_tokens={config.max_tokens} "
        f"reasoning={config.reasoning_effort}"
    )
    print(f"repeats      {config.repeats}   concurrency {config.concurrency}")
    print(f"out          {directory}")

    if args.dry_run:
        # A dry run spends nothing and produces no score, so it is not gated on the
        # pack validating. Reviewing the prompt is exactly what you want to be able to
        # do while a label is still being argued about.
        _validate_declared_classification(
            getattr(args, "data_classification", "synthetic"),
            testset_path=classification_testset_path,
            gt_path=classification_gt_path,
        )
        schema = response_format()
        return _dry_run(testset, prompt, config, schema, directory, runtime=runtime), None

    _preflight(testset, Path(gt_path))
    _validate_declared_classification(
        getattr(args, "data_classification", "synthetic"),
        testset_path=classification_testset_path,
        gt_path=classification_gt_path,
    )

    # Snapshot every mutable input before credentials are read. The model executes
    # against these bytes, not against paths that can change during a paid run.
    snapshot = _snapshot_run_inputs(
        directory,
        testset=testset,
        gt_path=Path(gt_path),
        prompt=prompt,
    )
    testset = _load_testset(Path(snapshot["testset_path"]), app=args.app)
    gt_path = Path(snapshot["gt_path"])
    schema = response_format(Path(snapshot["schema_path"]))
    _preflight(testset, gt_path)

    experiment_plan_sha = getattr(args, "experiment_plan_sha", None)
    run_contract = {
        "schema_version": 2,
        "app": args.app,
        "application_contract": application.manifest(),
        "application_contract_sha": application.fingerprint(),
        "arm": config.arm,
        "model": config.model,
        "prompt_id": config.prompt_id,
        "prompt_sha": prompt.sha,
        "testset_sha": snapshot["testset_sha"],
        "gt_sha": snapshot["gt_sha"],
        "schema_sha": snapshot["schema_sha"],
        "decoding": {
            "temperature": config.temperature,
            "top_p": config.top_p,
            "seed": config.seed,
            "max_tokens": config.max_tokens,
            "reasoning_effort": config.reasoning_effort,
        },
        "repeats": config.repeats,
        "concurrency": config.concurrency,
        "max_attempts": config.max_attempts,
        "timeout": args.timeout,
        "provider": config.provider,
        "runtime": runtime.manifest(),
        "runtime_fingerprint": runtime.fingerprint(),
        "experiment_plan_sha": experiment_plan_sha,
        "experiment_mode": getattr(args, "experiment_mode", None),
        "data_classification": getattr(args, "data_classification", "synthetic"),
    }
    journal_path = directory / "run.journal.jsonl"
    try:
        journal = (
            RunJournal.open(journal_path, run_contract)
            if journal_path.exists()
            else RunJournal.create(journal_path, run_contract)
        )
        completed = tuple(_item_result_from_log(row) for row in journal.completed_rows())
    except (ArtifactError, KeyError, TypeError, ValueError) as exc:
        raise CliError(f"cannot open resumable run journal: {exc}") from exc
    unresolved = journal.unresolved_cells()
    if journal.trailing_torn_record:
        raise CliError(
            f"run journal {journal_path} ends in a torn record. Prior completed rows "
            "remain readable, but automatic continuation cannot prove what the final "
            "append contained; preserve this run for review and start a new run."
        )
    if unresolved:
        raise CliError(
            f"run journal has {len(unresolved)} cell(s) that were durably dispatched "
            "but have no durable result. They may have completed server-side, so "
            "automatic replay would select a second draw. Preserve this run for "
            "review and start a new run; the first unresolved cell is "
            f"{unresolved[0]}."
        )

    _write_run_state(
        directory,
        {
            "schema_version": 1,
            "status": "INCOMPLETE",
            "run_contract": run_contract,
            "run_contract_sha": canonical_sha(run_contract),
            "completed_cells": len(completed),
        },
    )

    total = len(testset.items) * config.repeats
    remaining = total - len(completed)
    if remaining < 0:
        raise CliError(
            f"journal contains {len(completed)} cells for a {total}-cell run contract"
        )
    print(
        f"\ncalling {remaining} remaining logical cell(s) ({len(completed)}/{total} "
        "already journaled). Transport retries remain visible in attempt_count; parse "
        "failures are never re-rolled."
    )

    if remaining:
        api_key = _api_key(runtime)
        if runtime == OPENROUTER_RUNTIME:
            # Preserve the historical monkeypatch/client seam for existing callers.
            client = build_client(api_key, timeout=args.timeout)
        else:
            client = build_client(api_key, timeout=args.timeout, runtime=runtime)
    else:
        # `run` validates the recovered exact matrix and submits no request.
        client = _DryRunClient(testset, runtime)

    def progress(done: int, count: int, row: ItemResult) -> None:
        print(
            f"  [{done:>3}/{count}] {row.item_id} rep{row.replicate} "
            f"{row.outcome:<16} {row.latency_s:6.2f}s"
        )

    run_identity = RunResult(
        config=config,
        prompt_sha=prompt.sha,
        testset_sha=str(snapshot["testset_sha"]),
        results=(),
    )

    def with_runtime_identity(row: ItemResult) -> ItemResult:
        if row.runtime_fingerprint not in (None, runtime.fingerprint()):
            raise CliError(
                f"{row.item_id} rep{row.replicate} reported runtime fingerprint "
                f"{row.runtime_fingerprint!r}, expected {runtime.fingerprint()!r}"
            )
        if row.runtime_id not in (None, runtime.runtime_id):
            raise CliError(
                f"{row.item_id} rep{row.replicate} reported runtime id "
                f"{row.runtime_id!r}, expected {runtime.runtime_id!r}"
            )
        if row.runtime_backend not in (None, runtime.backend.value):
            raise CliError(
                f"{row.item_id} rep{row.replicate} reported runtime backend "
                f"{row.runtime_backend!r}, expected {runtime.backend.value!r}"
            )
        return replace(
            row,
            runtime_id=row.runtime_id or runtime.runtime_id,
            runtime_backend=row.runtime_backend or runtime.backend.value,
            runtime_fingerprint=row.runtime_fingerprint or runtime.fingerprint(),
        )

    def checkpoint(done: int, count: int, row: ItemResult) -> None:
        try:
            journal.append(item_result_to_row(run_identity, with_runtime_identity(row)))
        except ArtifactError as exc:
            raise CliError(f"cannot journal paid result {row.item_id} rep{row.replicate}: {exc}") from exc

    started = time.perf_counter()
    try:
        result = run(
            testset,
            client=client,
            prompt=prompt,
            config=config,
            response_format=schema,
            progress=progress,
            started=journal.mark_started,
            checkpoint=checkpoint,
            completed=completed,
        )
    except (RunError, ArtifactError) as exc:
        raise CliError(str(exc)) from exc
    result = replace(
        result,
        results=tuple(with_runtime_identity(row) for row in result.results),
    )
    wall_time_s = time.perf_counter() - started

    final_rows = [item_result_to_row(result, row) for row in result.results]
    final_by_cell = {
        (row["item_id"], row["replicate"]): row for row in final_rows
    }
    journal_by_cell = {
        (row["item_id"], row["replicate"]): row
        for row in journal.completed_rows()
    }
    if set(final_by_cell) != set(journal_by_cell):
        raise CliError(
            "refusing to finalize: the completed result matrix differs from the "
            "fsynced paid-call journal"
        )
    for cell, row in final_by_cell.items():
        if canonical_sha(row) != canonical_sha(journal_by_cell[cell]):
            raise CliError(
                f"refusing to finalize: result for logical cell {cell} differs from "
                "the fsynced paid-call journal"
            )

    log_path = directory / "run.jsonl"
    write_run_log(result, log_path)
    gt_sha = str(snapshot["gt_sha"])
    schema_sha = str(snapshot["schema_sha"])
    # The workload is the common evaluation question. Runtime and operational knobs
    # are deliberately recorded in a separate per-arm contract: OpenRouter and an
    # internal GPU endpoint need not expose the same reasoning controls, retry policy,
    # concurrency, or timeout to remain valid arms over the same scored workload.
    workload_contract = {
        "app": args.app,
        "testset_sha": result.testset_sha,
        "gt_sha": gt_sha,
        "prompt_sha": prompt.sha,
        "schema_sha": schema_sha,
        "repeats": config.repeats,
        "application_contract_sha": application.fingerprint(),
    }
    if experiment_plan_sha:
        workload_contract["experiment_plan_sha"] = experiment_plan_sha
    execution_contract = {
        "runtime_fingerprint": runtime.fingerprint(),
        "provider": config.provider,
        "decoding": {
            "temperature": config.temperature,
            "top_p": config.top_p,
            "seed": config.seed,
            "max_tokens": config.max_tokens,
            "reasoning_effort": config.reasoning_effort,
        },
        "max_attempts": config.max_attempts,
        "concurrency": config.concurrency,
        "timeout": args.timeout,
    }
    system_fingerprints: dict[str, int] = {}
    for row in result.results:
        key = row.system_fingerprint or "<not reported>"
        system_fingerprints[key] = system_fingerprints.get(key, 0) + 1
    meta = {
        "artifact_schema_version": 2,
        "status": "COMPLETE",
        "run_id": directory.name,
        "created_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "arm": config.arm,
        "app": args.app,
        "application_contract": application.manifest(),
        "application_contract_sha": application.fingerprint(),
        "backend": runtime.backend.value,
        "runtime": execution_provenance(runtime),
        "runtime_fingerprint": runtime.fingerprint(),
        "model_requested": config.model,
        # The pin as requested, recorded even when it is null. A run.json with no
        # provider key at all would leave "was this pinned?" answerable only by the
        # file's age, and the answer decides whether the arm is one system or a blend.
        "provider_requested": config.provider,
        "prompt_id": prompt.id,
        "prompt_sha": prompt.sha,
        "testset_path": _run_relative_path(directory, testset.path),
        "source_testset_path": str(snapshot["source_testset_path"]),
        "testset_sha": result.testset_sha,
        "gt_path": _run_relative_path(directory, gt_path),
        "source_gt_path": str(snapshot["source_gt_path"]),
        "gt_sha": gt_sha,
        "schema_path": _run_relative_path(directory, Path(snapshot["schema_path"])),
        "prompt_path": _run_relative_path(directory, Path(snapshot["prompt_path"])),
        # New runs separate the code that classified responses from the code that
        # scores them. `scorer_sha` remains as a compatibility alias for readers of
        # historical run.json files.
        "outcome_contract_sha": manifest_mod.outcome_contract_sha(),
        "generation_contract_sha": manifest_mod.generation_contract_sha(),
        "decision_policy_sha": manifest_mod.decision_policy_sha(),
        "scoring_code_sha": manifest_mod.scoring_code_sha(),
        "scorer_sha": manifest_mod.scoring_code_sha(),
        "schema_sha": schema_sha,
        "workload_contract": workload_contract,
        "workload_sha": manifest_mod.workload_sha(workload_contract),
        "execution_contract": execution_contract,
        "execution_sha": canonical_sha(execution_contract),
        "experiment_plan_sha": experiment_plan_sha,
        "experiment_mode": getattr(args, "experiment_mode", None),
        "qualification_sha": getattr(args, "qualification_sha", None),
        "decoding": {
            "temperature": config.temperature,
            "top_p": config.top_p,
            "seed": config.seed,
            "max_tokens": config.max_tokens,
            "reasoning_effort": config.reasoning_effort,
        },
        "max_attempts": config.max_attempts,
        "repeats": config.repeats,
        "concurrency": config.concurrency,
        "items": len(testset.items),
        "rows": len(result.results),
        "run_contract": run_contract,
        "run_contract_sha": canonical_sha(run_contract),
        "run_log_sha256": file_sha256(log_path),
        "run_log_bytes": log_path.stat().st_size,
        "journal_sha256": file_sha256(journal_path),
        "journal_bytes": journal_path.stat().st_size,
        "resumed_cells": len(completed),
        "data_classification": getattr(args, "data_classification", "synthetic"),
        "outcome_counts": result.outcome_counts(),
        "observed_models": result.observed_models(),
        "observed_providers": result.observed_providers(),
        "system_fingerprints": system_fingerprints,
        # The two facts that decide whether this arm is one system. `split_items` is
        # the measurement (a backend's own tokenizer cannot be echoed); the spread is
        # kept in full beside it so the claim can be checked rather than believed.
        "prompt_token_spread": {k: list(v) for k, v in result.prompt_token_spread().items()},
        "split_items": {k: list(v) for k, v in result.split_items().items()},
        "calls_without_prompt_usage": result.calls_without_prompt_usage(),
        "wall_time_s": wall_time_s,
        "throughput_calls_per_s": remaining / wall_time_s if wall_time_s else None,
        "truncated_rate": result.truncated_rate(),
        # A LOWER BOUND, not a total: OpenRouter reports usage.cost only for providers
        # that supply it, and a missing value stays None rather than becoming zero.
        "total_cost_usd_lower_bound": result.total_cost(),
    }
    _write_meta(directory, meta)
    _write_run_state(
        directory,
        {
            "schema_version": 1,
            "status": "COMPLETE",
            "run_contract_sha": canonical_sha(run_contract),
            "run_log_sha256": meta["run_log_sha256"],
            "completed_cells": len(result.results),
        },
    )

    print("\noutcome counts (every row, failures included):")
    for name, count in result.outcome_counts().items():
        print(f"  {name:<18} {count}")
    print("observed models (what the runtime actually served):")
    for name, count in result.observed_models().items():
        print(f"  {count:>5}  {name}")
    if len(result.observed_models()) > 1:
        print(
            "  WARNING: more than one entry. This arm is not one model; its numbers are "
            "a blend of two systems and no downstream statistic can unblend them."
        )
    _print_backend_identity(result, config.provider)
    print(f"truncated rate  {result.truncated_rate():.3f}")
    print(f"cost (lower bound, USD)  {result.total_cost():.6f}")
    print(f"\nwrote {directory / 'run.jsonl'} and {directory / 'run.json'}")

    return EXIT_OK, LoadedRun(directory=directory, meta=meta, result=result)


def _print_backend_identity(result: RunResult, requested: str | None) -> None:
    """Which backend answered, and the one check that cannot be answered by the router.

    Printed after every run, pinned or not, because "was this arm one system?" is a
    question about the run that happened rather than about the flags it was given.

    Two signals, deliberately not merged into one verdict:

      * `observed_providers` -- what OpenRouter says served the call. Useful, and it is
        the router describing its own routing. A pinned run whose provider field echoes
        the pin has proved that the field echoes the pin.
      * `split_items` -- items whose byte-identical replicates came back with more than
        one `prompt_tokens` value. This one is produced by the backend's tokenizer, so
        it is the router being measured rather than asked. It is the PASS/FAIL line.

    A run with fewer than two replicates cannot split, and says so instead of printing
    a zero that would read as a clean sheet.
    """
    providers = result.observed_providers()
    print("observed providers (which backend, not which model id):")
    if not providers:
        print("      -  none reported. This backend does not expose `provider`;")
        print("         prompt_tokens below is the identity signal, and it is the stronger one.")
    for name, count in providers.items():
        print(f"  {count:>5}  {name}")
    if len(providers) > 1:
        print(
            "  WARNING: more than one backend served this arm. Different quantisation, "
            "chat template\n           and tokenizer, reported under one model id."
        )
    if requested and providers and set(providers) != {requested}:
        print(
            f"  WARNING: {requested!r} was pinned with allow_fallbacks=false, but the "
            f"served set is {sorted(providers)}."
        )

    spread = result.prompt_token_spread()
    split = result.split_items()
    checked = sum(1 for values in spread.values() if values)
    print(
        f"prompt_tokens fingerprint  {checked - len(split)}/{checked} items returned "
        "exactly one value"
    )
    if result.config.repeats < 2:
        print(
            "  repeats=1, so no item could disagree with itself. This number is "
            "arithmetic, not evidence."
        )
    elif split:
        print(
            "  FAILED. A byte-identical request returned two token counts, which is two "
            "tokenizers\n  and therefore two builds. The arm is a blend regardless of "
            "what `provider` reported:"
        )
        for item_id, values in list(split.items())[:10]:
            print(f"    {item_id}  {list(values)}")
    else:
        print("  every item returned one value: one tokenizer answered, so one build did.")


def _dry_run(
    testset: TestSet,
    prompt: Prompt,
    config: RunConfig,
    schema: dict[str, Any],
    directory: Path,
    *,
    runtime: RuntimeSpec = OPENROUTER_RUNTIME,
) -> int:
    """Write the exact request bodies and the prompt. Zero API calls, zero key needed.

    `repeats` is forced to 1 and `concurrency` to 1. Not a shortcut: `RunConfig`
    documents that every replicate sends the SAME request, seed included, so one body
    per item is the complete set of distinct bodies, and a single worker keeps the file
    in testset order without a sort that would have to be trusted.
    """
    dry_config = replace(config, repeats=1, concurrency=1)
    client = _DryRunClient(testset, runtime)
    run(
        testset,
        client=client,
        prompt=prompt,
        config=dry_config,
        response_format=schema,
    )

    bodies = directory / "requests.jsonl"
    atomic_write_text(
        bodies,
        "".join(json.dumps(entry, ensure_ascii=False) + "\n" for entry in client.requests),
    )
    prompt_file = directory / "prompt.txt"
    atomic_write_text(prompt_file, prompt.system_text)

    print("\nDRY RUN. No API call was made and no key was read.")
    print(f"  requests   {len(client.requests)} bodies -> {bodies}")
    print(f"  prompt     {prompt_file}")
    print(f"  runtime    {runtime.runtime_id} sha={runtime.fingerprint()}")
    print(
        f"  repeats={config.repeats} is configured; every replicate sends this same "
        "body, seed included (RunConfig), so one body per item is the whole set."
    )

    lines = prompt.system_text.split("\n")
    shown = min(DRY_RUN_PROMPT_LINES, len(lines))
    print(f"\nSYSTEM PROMPT, first {shown} of {len(lines)} lines (sha {prompt.sha[:16]}):")
    print("-" * 78)
    for number, line in enumerate(lines[:shown], start=1):
        print(f"{number:>3} | {line}")
    print("-" * 78)
    print("The user turn is the transcript, verbatim, with no framing added.")
    return EXIT_OK


def cmd_baseline(args: argparse.Namespace) -> int:
    """Run one arm over the whole pack."""
    if getattr(args, "resume_run", None):
        directory = _run_directory(args, args.arm)
        args.classification_testset_path = Path(args.testset)
        args.classification_gt_path = Path(args.gt)
        testset_path = _resume_snapshot(directory, "testset.jsonl")
        gt_path = _resume_snapshot(directory, "ground_truth.csv")
    else:
        testset_path = Path(args.testset)
        gt_path = Path(args.gt)
        testset = _load_testset(testset_path, app=args.app)
        directory = _run_directory(args, args.arm)
    testset = _load_testset(testset_path, app=args.app)
    code, _ = _execute_run(args, testset=testset, gt_path=gt_path, directory=directory)
    return code


def cmd_stability(args: argparse.Namespace) -> int:
    """The N_flip probe: a few items, many replicates, identical requests every time.

    What it measures is provider variance at fixed decoding parameters, and it is real
    rather than theoretical: `qwen3.6-27b` returned an empty response on run 2 of three
    identical Thai round-trips (`outcomes._classify_empty`). Run this before a full
    comparison; a FLAKY verdict argued about afterwards is an argument about the
    replicate count, not about the model.
    """
    if getattr(args, "resume_run", None):
        directory = _run_directory(args, args.arm)
        args.classification_testset_path = Path(args.testset)
        args.classification_gt_path = Path(args.gt)
        testset = _load_testset(
            _resume_snapshot(directory, "testset.jsonl"), app=args.app
        )
        gt_path = _resume_snapshot(directory, "ground_truth.csv")
    else:
        full = _load_testset(Path(args.testset), app=args.app)
        directory = _run_directory(args, args.arm)
        testset = _subset_testset(full, args.items.split(","), directory)
        args.classification_testset_path = full.path
        gt_path = Path(args.gt)

    code, loaded = _execute_run(
        args, testset=testset, gt_path=gt_path, directory=directory
    )
    if code != EXIT_OK or loaded is None:
        return code

    gt = _load_gt(
        _recorded_artifact_path(loaded.directory, loaded.meta["gt_path"])
    )
    call_ids = {item.call_id for item in testset.items}
    scoped_gt = [record for record in gt if record.call_id in call_ids]
    per_replicate = replicate_records(loaded.result, testset.items)

    flips = n_flip(per_replicate)
    print(f"\nN_flip = {flips} over {len(per_replicate)} replicates")
    print(
        "  A cell is (scored row x dimension), not a row. N_flip = 0 at one replicate "
        "is arithmetic, not stability."
    )
    if len(per_replicate) < 2:
        print("  Only one replicate ran, so this number cannot mean anything yet.")

    try:
        rows = mechanism_table(scoped_gt, per_replicate, list(testset.items),
                               group_by=lambda item: item.item_id)
    except ReportError as exc:
        raise CliError(str(exc)) from exc

    print("\nper item:")
    for row in rows:
        print(f"  {row.mechanism:<10} {row.verdict:<6} {row.detail.splitlines()[0]}")
    return EXIT_OK


# =========================================================== enterprise experiment


def _checked_plan(path: Path) -> tuple[dict[str, Any], str]:
    """Load and validate a plan against the current checkout, returning its SHA."""
    try:
        plan = load_plan(path)
        problems = validate_plan(plan, root=REPO_ROOT)
    except PlanError as exc:
        raise CliError(str(exc)) from exc
    if problems:
        raise CliError(
            f"{path} has {len(problems)} problem(s); first: {problems[0]}. "
            "Run `evalgen experiment-check --plan ...` for the complete list."
        )
    return plan, canonical_sha(plan)


def cmd_experiment_check(args: argparse.Namespace) -> int:
    """Validate the preregistration, assets and prompt registry without a key."""
    path = Path(args.plan)
    try:
        plan = load_plan(path)
        problems = validate_plan(plan, root=REPO_ROOT)
    except PlanError as exc:
        print(f"1 PROBLEM:\n  - {exc}")
        return EXIT_PROBLEMS
    problems.extend(validate_prompt_manifest())
    print(f"experiment   {plan.get('experiment_id', '<missing>')}")
    print(f"status       {plan.get('status', '<missing>')}")
    print(f"plan sha     {canonical_sha(plan)}")
    try:
        budget = logical_call_budget(plan)
    except (KeyError, TypeError, ValueError) as exc:
        problems.append(f"cannot compute call budget: {exc}")
    else:
        print(
            f"call budget  qualification<={budget['qualification_max']} "
            f"full={budget['full']} load={budget['load']} total<="
            f"{budget['grand_total_max']} for current inventory"
        )
    if problems:
        print(f"\n{len(problems)} PROBLEM(S):")
        for problem in problems:
            print(f"  - {problem}")
        return EXIT_PROBLEMS
    print("\nOK. The plan and its committed assets are internally consistent.")
    if plan.get("status") == "locked":
        print(
            "No model call was made; committed provider qualification artifacts "
            "were verified."
        )
    else:
        print("No model call was made; provider availability has not been asserted.")
    return EXIT_OK


def cmd_experiment_budget(args: argparse.Namespace) -> int:
    """Print the current no-network approval ceiling from snapshotted prices."""
    plan, plan_sha = _checked_plan(Path(args.plan))
    try:
        budget = projected_execution_budget(plan, root=REPO_ROOT)
    except (KeyError, TypeError, ValueError, OSError, PlanError) as exc:
        raise CliError(f"cannot calculate experiment budget: {exc}") from exc
    print(f"experiment   {plan['experiment_id']}")
    print(f"plan sha     {plan_sha}")
    print(f"prices       snapshot {budget['price_snapshot_utc']}")
    print("\nfull + load ceiling (every call spends all 8,000 output tokens):")
    for arm_id, facts in budget["arms"].items():
        if facts["availability"] == "UNAVAILABLE":
            print(f"  {arm_id:<22} UNAVAILABLE  $0.00")
            continue
        print(
            f"  {arm_id:<22} ${facts['maximum_usd']:.2f}  "
            f"{facts['provider']} ({facts['provider_selection']})"
        )
    print(f"  {'full + load':<22} ${budget['full_load_maximum_usd']:.2f}")
    print(f"  {'qualification':<22} ${budget['qualification_maximum_usd']:.2f}")
    print(f"  {'grand maximum':<22} ${budget['grand_maximum_usd']:.2f}")
    print(
        "\nThis is a conservative approval ceiling, not likely spend. Refresh inventory "
        "and prices before approval; a change amends the plan and its SHA."
    )
    return EXIT_OK


def _plan_asset(plan: Mapping[str, Any], name: str) -> Path:
    return REPO_ROOT / str(plan["assets"][name]["path"])


def _planned_args(
    source: argparse.Namespace,
    *,
    plan: Mapping[str, Any],
    plan_sha: str,
    arm: Mapping[str, Any],
    provider: str,
    repeats: int,
    concurrency: int,
    max_attempts: int,
    reasoning_effort: str,
    mode: str,
    qualification_sha: str | None = None,
) -> argparse.Namespace:
    workload = plan["workload"]
    return argparse.Namespace(
        model=arm["model"],
        arm=arm["id"],
        app=plan["app"],
        prompt_id=plan["assets"]["prompt"]["id"],
        provider=provider,
        repeats=repeats,
        concurrency=concurrency,
        timeout=source.timeout,
        temperature=workload["temperature"],
        top_p=workload["top_p"],
        seed=workload["seed"],
        max_tokens=workload["max_tokens"],
        reasoning_effort=reasoning_effort,
        max_attempts=max_attempts,
        dry_run=source.dry_run,
        experiment_plan_sha=plan_sha,
        experiment_mode=mode,
        qualification_sha=qualification_sha,
        data_classification=getattr(source, "data_classification", "synthetic"),
        resume_run=getattr(source, "resume_run", None),
        runtime_manifest=getattr(source, "runtime_manifest", None),
        runtime_backend=getattr(source, "runtime_backend", None),
        runtime_id=getattr(source, "runtime_id", None),
        base_url=getattr(source, "base_url", None),
        api_key_env=getattr(source, "api_key_env", None),
        runtime_metadata=getattr(source, "runtime_metadata", []),
        allow_insecure_http=getattr(source, "allow_insecure_http", False),
        classification_testset_path=_plan_asset(plan, "testset"),
    )


def cmd_qualify(args: argparse.Namespace) -> int:
    """Run one bounded, pinned provider qualification probe (six logical calls)."""
    plan, plan_sha = _checked_plan(Path(args.plan))
    if plan.get("status") == "locked":
        raise CliError(
            "the plan is locked. Additional provider evidence requires returning it "
            "to draft/qualified status and a new review, not changing the approved arm."
        )
    try:
        arm = arm_by_id(plan, args.arm)
    except PlanError as exc:
        raise CliError(str(exc)) from exc
    candidates = arm.get("provider_candidates", [])
    if args.provider not in candidates:
        raise CliError(
            f"provider {args.provider!r} is not preregistered for {args.arm}; "
            f"candidates are {candidates}. Refresh the inventory and amend the draft "
            "plan before probing a newly listed endpoint."
        )
    qualification_plan = plan["qualification"]
    full = _load_testset(_plan_asset(plan, "testset"), app=plan["app"])
    directory = _run_directory(args, f"qual-{args.arm}-{args.provider}")
    testset = _subset_testset(full, qualification_plan["item_ids"], directory)
    run_args = _planned_args(
        args,
        plan=plan,
        plan_sha=plan_sha,
        arm=arm,
        provider=args.provider,
        repeats=qualification_plan["replicates"],
        concurrency=1,
        max_attempts=qualification_plan["max_attempts"],
        reasoning_effort=qualification_plan["reasoning_effort"],
        mode="qualification",
    )
    code, loaded = _execute_run(
        run_args,
        testset=testset,
        gt_path=_plan_asset(plan, "ground_truth"),
        directory=directory,
    )
    if code != EXIT_OK or loaded is None:
        return code
    return _write_qualification_artifact(
        plan=plan,
        plan_sha=plan_sha,
        arm=arm,
        provider=args.provider,
        loaded=loaded,
        target=directory / "qualification.json",
    )


def _write_qualification_artifact(
    *,
    plan: Mapping[str, Any],
    plan_sha: str,
    arm: Mapping[str, Any],
    provider: str,
    loaded: LoadedRun,
    target: Path,
) -> int:
    """Deterministically classify and write one safe qualification artifact."""
    result = assess_qualification(
        loaded.result,
        expected_model=str(arm["model"]),
        expected_provider=provider,
    )
    artifact = {
        "schema_version": 1,
        "experiment_id": plan["experiment_id"],
        "plan_sha": plan_sha,
        "arm": arm["id"],
        "model": arm["model"],
        "provider": provider,
        "qualification_contract_sha": qualification_contract_sha(
            plan, arm_id=str(arm["id"]), provider=provider
        ),
        "run_id": loaded.directory.name,
        "qualification": result.to_dict(),
    }
    artifact["qualification_sha"] = canonical_sha(artifact)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(
        json.dumps(artifact, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    print(f"\nqualification  {result.status}")
    for reason in result.reasons:
        print(f"  - {reason}")
    print(f"qualification sha  {artifact['qualification_sha']}")
    print(f"wrote {target}")
    return EXIT_OK if result.passed else EXIT_PROBLEMS


def cmd_qualification_report(args: argparse.Namespace) -> int:
    """Reclassify an existing qualification run without making a model call."""
    plan, plan_sha = _checked_plan(Path(args.plan))
    loaded = load_run(Path(args.run))
    meta = loaded.meta
    if meta.get("experiment_mode") != "qualification":
        raise CliError(
            f"run mode is {meta.get('experiment_mode')!r}, not 'qualification'"
        )
    if meta.get("experiment_plan_sha") != plan_sha:
        raise CliError(
            "qualification run plan SHA does not match the current plan; reclassify "
            "against the exact draft that produced the paid rows"
        )
    arm_id = str(meta.get("arm") or "")
    try:
        arm = arm_by_id(plan, arm_id)
    except PlanError as exc:
        raise CliError(str(exc)) from exc
    provider = str(meta.get("provider_requested") or "")
    if provider not in arm.get("provider_candidates", []):
        raise CliError(
            f"recorded provider {provider!r} is not preregistered for {arm_id}"
        )
    if meta.get("model_requested") != arm.get("model"):
        raise CliError(
            f"recorded model {meta.get('model_requested')!r} does not match "
            f"plan model {arm.get('model')!r}"
        )
    target = Path(args.out) if args.out else loaded.directory / "qualification.json"
    print("No model call was made and no key was read.")
    return _write_qualification_artifact(
        plan=plan,
        plan_sha=plan_sha,
        arm=arm,
        provider=provider,
        loaded=loaded,
        target=target,
    )


def cmd_experiment_run(args: argparse.Namespace) -> int:
    """Execute one locked full or load arm; never infer provider selection."""
    plan, plan_sha = _checked_plan(Path(args.plan))
    if plan.get("status") != "locked":
        raise CliError(
            f"plan status is {plan.get('status')!r}, not 'locked'. Qualification does "
            "not authorize the full run; record selected providers and qualification "
            "hashes, review cost, then lock the plan."
        )
    if args.confirm_plan_sha != plan_sha:
        raise CliError(
            f"--confirm-plan-sha is {args.confirm_plan_sha!r}, but the current plan is "
            f"{plan_sha}. Review the changed plan rather than executing an old approval."
        )
    try:
        arm = arm_by_id(plan, args.arm)
    except PlanError as exc:
        raise CliError(str(exc)) from exc
    if arm.get("availability") == "UNAVAILABLE":
        raise CliError(
            f"{args.arm} is locked as UNAVAILABLE in the production-like regime; "
            "there is deliberately no provider to run."
        )
    provider = str(arm["selected_provider"])
    qualification_sha = str(arm["qualification_sha"])
    full = _load_testset(_plan_asset(plan, "testset"), app=plan["app"])
    workload = plan["workload"]
    if args.mode == "full":
        testset = full
        repeats = workload["replicates"]
        concurrency = workload["concurrency"]
        max_attempts = workload["max_attempts"]
        directory = _run_directory(args, f"{plan['experiment_id']}-full-{args.arm}")
    else:
        allowed = plan["operations"]["concurrency_levels"]
        if args.concurrency_level not in allowed:
            raise CliError(
                f"load concurrency must be one of {allowed}, found "
                f"{args.concurrency_level!r}"
            )
        concurrency = args.concurrency_level
        repeats = plan["operations"]["replicates_per_level"]
        max_attempts = plan["operations"]["max_attempts"]
        directory = _run_directory(args, f"e5-load-c{concurrency}-{args.arm}")
        testset = _subset_testset(full, plan["operations"]["item_ids"], directory)
    run_args = _planned_args(
        args,
        plan=plan,
        plan_sha=plan_sha,
        arm=arm,
        provider=provider,
        repeats=repeats,
        concurrency=concurrency,
        max_attempts=max_attempts,
        reasoning_effort=workload["reasoning_effort"],
        mode=args.mode,
        qualification_sha=qualification_sha,
    )
    code, _ = _execute_run(
        run_args,
        testset=testset,
        gt_path=_plan_asset(plan, "ground_truth"),
        directory=directory,
    )
    return code


def _parse_run_assignments(values: Sequence[str], *, separator: str = "=") -> dict[str, Path]:
    parsed: dict[str, Path] = {}
    for value in values:
        if separator not in value:
            raise CliError(f"run assignment must be NAME{separator}PATH, found {value!r}")
        name, raw_path = value.split(separator, 1)
        if not name or name in parsed:
            raise CliError(f"run assignment has blank or duplicate name: {value!r}")
        parsed[name] = Path(raw_path)
    return parsed


def _slice_disagreements(
    gt: Sequence[Record],
    incumbent: Sequence[Record],
    candidate: Sequence[Record],
    *,
    call_ids: set[str],
) -> list[dict[str, Any]]:
    scoped_gt = [row for row in gt if row.call_id in call_ids]
    scoped_inc = [row for row in incumbent if row.call_id in call_ids]
    scoped_cand = [row for row in candidate if row.call_id in call_ids]
    return [
        asdict(paired_verdict(disagreement(scoped_gt, scoped_inc, scoped_cand, name)))
        for name in ("call_result", "reason", "product")
    ]


def cmd_experiment_report(args: argparse.Namespace) -> int:
    """Assemble machine-readable and Markdown quality-first experiment reports."""
    plan, plan_sha = _checked_plan(Path(args.plan))
    if plan.get("status") != "locked":
        raise CliError("experiment-report requires the same locked plan that ran the arms")
    assignments = _parse_run_assignments(args.run)
    expected_arms = {
        arm["id"] for arm in plan["arms"] if arm.get("availability") == "QUALIFIED"
    }
    if set(assignments) != expected_arms:
        raise CliError(
            f"full run assignments must equal qualified arms; expected "
            f"{sorted(expected_arms)}, found {sorted(assignments)}"
        )
    incumbent_id = next(
        arm["id"] for arm in plan["arms"] if arm.get("role") == "incumbent"
    )
    if incumbent_id not in assignments:
        raise CliError(f"--run must include incumbent assignment {incumbent_id}=PATH")
    loaded = {name: load_run(path) for name, path in assignments.items()}
    load_assignments = _parse_run_assignments(args.load_run)
    for name, run_data in loaded.items():
        if run_data.meta.get("experiment_plan_sha") != plan_sha:
            raise CliError(
                f"{name} records plan {run_data.meta.get('experiment_plan_sha')!r}, "
                f"not current {plan_sha}"
            )
        if run_data.meta.get("experiment_mode") != "full":
            raise CliError(f"{name} is not a full run for {plan['experiment_id']}")
    allowed_levels = set(plan["operations"]["concurrency_levels"])
    expected_load_names = {
        f"{arm_id}@{level}" for arm_id in assignments for level in allowed_levels
    }
    # Load probes are an operational diagnostic, not an input to the quality
    # decision.  A full-arm-only experiment may omit them; once any load run is
    # supplied, however, require the complete matrix so a partial benchmark cannot
    # masquerade as a cross-arm operations comparison.
    if load_assignments and set(load_assignments) != expected_load_names:
        missing = sorted(expected_load_names - set(load_assignments))
        extra = sorted(set(load_assignments) - expected_load_names)
        raise CliError(
            "a load report requires every full arm at every preregistered concurrency; "
            f"missing={missing}, extra={extra}"
        )
    loaded_load = {name: load_run(path) for name, path in load_assignments.items()}
    for name, run_data in loaded_load.items():
        arm_id, raw_level = name.rsplit("@", 1)
        level = int(raw_level)
        if run_data.meta.get("experiment_plan_sha") != plan_sha:
            raise CliError(f"load run {name} records a different experiment plan")
        if run_data.meta.get("experiment_mode") != "load":
            raise CliError(f"{name} is not an Experiment 5 load run")
        if run_data.arm != arm_id or run_data.result.config.concurrency != level:
            raise CliError(
                f"{name} identity does not match its run: arm={run_data.arm}, "
                f"concurrency={run_data.result.config.concurrency}"
            )
    incumbent = loaded[incumbent_id]
    gt = _load_gt(_plan_asset(plan, "ground_truth"))
    testset = _load_testset(_plan_asset(plan, "testset"), app=plan["app"])
    per_rep = {
        name: replicate_records(run_data.result, testset.items)
        for name, run_data in loaded.items()
    }
    item_call = {item.item_id: item.call_id for item in testset.items}
    planned_arms = {arm["id"]: arm for arm in plan["arms"]}
    runtime_problems = {
        name: runtime_gate(
            run_data.result,
            expected_model=str(planned_arms[name]["model"]),
            expected_provider=str(planned_arms[name]["selected_provider"]),
        )
        for name, run_data in loaded.items()
    }
    arm_facts = {
        name: {
            **operational_summary(run_data.result),
            "wall_time_s": run_data.meta.get("wall_time_s"),
            "throughput_calls_per_s": run_data.meta.get("throughput_calls_per_s"),
            "runtime_gate_problems": runtime_problems[name],
        }
        for name, run_data in loaded.items()
    }
    load_facts: dict[str, dict[str, Any]] = {name: {} for name in assignments}
    for assignment, run_data in loaded_load.items():
        arm_id, level = assignment.rsplit("@", 1)
        load_facts[arm_id][level] = {
            **operational_summary(run_data.result),
            "wall_time_s": run_data.meta.get("wall_time_s"),
            "throughput_calls_per_s": run_data.meta.get("throughput_calls_per_s"),
        }
    summary: dict[str, Any] = {
        "schema_version": 1,
        "experiment_id": plan["experiment_id"],
        "decision_policy": (
            "legacy_v1"
            if plan["experiment_id"] == "retention-e5"
            else "decision_grade_v2"
        ),
        # Policy is bound when the report is produced, not when model outputs are
        # generated. This permits the same immutable runs to be audited under a fixed
        # legacy rule or the current decision-grade rule without pretending the
        # runtime endpoint determines statistical policy.
        "decision_policy_sha": manifest_mod.decision_policy_sha(),
        "plan_sha": plan_sha,
        "reconciled": False,
        "unavailable_arms": [
            arm["id"] for arm in plan["arms"] if arm.get("availability") == "UNAVAILABLE"
        ],
        "arms": arm_facts,
        "load": load_facts,
        "comparisons": {},
    }
    try:
        regression_key = hmac_key()
    except RuntimeError as exc:
        raise CliError(
            f"{exc} Experiment reports retain hashed item-level regressions; refusing "
            "to replace that deliverable with an empty list."
        ) from exc
    incumbent_ok = sum(row.outcome == "ok" for row in incumbent.result.results)
    for name, candidate in loaded.items():
        if name == incumbent_id:
            continue
        _refuse_incomparable(incumbent, candidate)
        slices: dict[str, Any] = {}
        for slice_name, slice_plan in plan["slices"].items():
            call_ids = {item_call[item_id] for item_id in slice_plan["item_ids"]}
            slices[slice_name] = _slice_disagreements(
                gt,
                per_rep[incumbent_id][0],
                per_rep[name][0],
                call_ids=call_ids,
            )
        stability = paired_verdict(
            stability_disagreement(incumbent.result, candidate.result)
        )
        quality = [
            paired_verdict(
                disagreement(
                    gt,
                    per_rep[incumbent_id][0],
                    per_rep[name][0],
                    dimension,
                )
            )
            for dimension in ("call_result", "reason", "product")
        ]
        ok = sum(row.outcome == "ok" for row in candidate.result.results)
        verdict = experiment_decision(
            qualification_status="QUALIFIED",
            parse_ok=ok,
            total=len(candidate.result.results),
            quality_verdicts=quality,
            stability_verdict=stability,
            runtime_problems=runtime_problems[name],
            reference_parse_ok=incumbent_ok,
            reference_total=len(incumbent.result.results),
            reference_runtime_problems=runtime_problems[incumbent_id],
            policy=summary["decision_policy"],
        )
        regression_rows = [
            row
            for dimension in ("call_result", "reason", "product")
            for row in regressions(
                gt,
                per_rep[incumbent_id][0],
                per_rep[name][0],
                dimension,
                regression_key,
            )
        ]
        summary["comparisons"][name] = {
            "incumbent": incumbent_id,
            "slices": slices,
            "stability": asdict(stability),
            "decision": asdict(verdict),
            "regressions": [asdict(row) for row in regression_rows],
        }

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    for name, facts in summary["arms"].items():
        (out / f"arm-{name}.json").write_text(
            json.dumps(facts, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
            newline="\n",
        )
        (out / f"arm-{name}.md").write_text(
            "\n".join([
                f"# {name}",
                "",
                "**RECONCILED: NO.**",
                "",
                f"Parse valid: {facts['parse_ok']}/{facts['calls']} "
                f"({facts['parse_valid_rate']:.3%})",
                f"Latency p50/p95/p99/max: {facts['latency_s']['p50']:.3f}s / "
                f"{facts['latency_s']['p95']:.3f}s / {facts['latency_s']['p99']:.3f}s / "
                f"{facts['latency_s']['max']:.3f}s",
                f"Cost lower bound: ${facts['cost_usd']['lower_bound']:.6f}; "
                f"missing on {facts['cost_usd']['calls_missing']} calls.",
                "Runtime gate: " + (
                    "PASS" if not facts["runtime_gate_problems"]
                    else "; ".join(facts["runtime_gate_problems"])
                ),
                "",
            ]),
            encoding="utf-8",
            newline="\n",
        )
    for name, facts in summary["comparisons"].items():
        (out / f"comparison-{incumbent_id}-vs-{name}.json").write_text(
            json.dumps(facts, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
            newline="\n",
        )
        comparison_lines = [
            f"# {incumbent_id} vs {name}",
            "",
            "**RECONCILED: NO.**",
            "",
            f"Decision: **{facts['decision']['status']}** — "
            + "; ".join(facts["decision"]["reasons"]),
            "",
            "| Slice | Dimension | d | Net | Band | Verdict |",
            "|---|---|---:|---:|---:|---|",
        ]
        for slice_name, rows in facts["slices"].items():
            for row in rows:
                comparison_lines.append(
                    f"| {slice_name} | {row['dimension']} | {row['discordant']} | "
                    f"{row['net']:+d} | {row['band'] if row['band'] is not None else '—'} | "
                    f"{row['verdict']} |"
                )
        comparison_lines += [
            "",
            f"Item-level regressions: {len(facts['regressions'])}. See the paired JSON "
            "artifact for every hashed regression row.",
            "",
        ]
        (out / f"comparison-{incumbent_id}-vs-{name}.md").write_text(
            "\n".join(comparison_lines), encoding="utf-8", newline="\n"
        )
    json_path = out / "summary.json"
    json_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    lines = [
        f"# {plan['experiment_id']} summary",
        "",
        "**RECONCILED: NO. This is harness output, not a migration verdict.**",
        "",
        f"Plan: `{plan_sha}`",
        "",
        "| Arm | Parse valid | p50 latency | p95 latency | Cost lower bound |",
        "|---|---:|---:|---:|---:|",
    ]
    for name, facts in summary["arms"].items():
        latency = facts["latency_s"]
        lines.append(
            f"| {name} | {facts['parse_ok']}/{facts['calls']} | "
            f"{latency['p50']:.3f}s | {latency['p95']:.3f}s | "
            f"${facts['cost_usd']['lower_bound']:.6f} |"
        )
    for name, comparison in summary["comparisons"].items():
        lines += [
            "",
            f"## {name} vs {incumbent_id}",
            "",
            f"Decision: **{comparison['decision']['status']}** — "
            + "; ".join(comparison["decision"]["reasons"]),
            "",
            "| Slice | Dimension | d | Net | Band | Verdict |",
            "|---|---|---:|---:|---:|---|",
        ]
        for slice_name, rows in comparison["slices"].items():
            for row in rows:
                lines.append(
                    f"| {slice_name} | {row['dimension']} | {row['discordant']} | "
                    f"{row['net']:+d} | {row['band'] if row['band'] is not None else '—'} | "
                    f"{row['verdict']} |"
                )
        lines += [
            "",
            f"Item-level regressions: {len(comparison['regressions'])}",
            "",
        ]
        if comparison["regressions"]:
            lines += [
                "| Item key | Dimension | Ground truth | Incumbent | Candidate | Error |",
                "|---|---|---|---|---|---|",
            ]
            for row in comparison["regressions"]:
                lines.append(
                    f"| {row['item_key']} | {row['dimension']} | {row['gt_label']} | "
                    f"{row['incumbent_label']} | {row['candidate_label']} | "
                    f"{row['error_type']} |"
                )
    lines += [
        "",
        "## Load",
        "",
        "| Arm | Concurrency | Parse valid | Calls/s | p50 | p95 | p99 |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    for name, levels in summary["load"].items():
        for level in sorted(levels, key=int):
            facts = levels[level]
            latency = facts["latency_s"]
            throughput = facts["throughput_calls_per_s"]
            lines.append(
                f"| {name} | {level} | {facts['parse_ok']}/{facts['calls']} | "
                f"{throughput:.3f} | {latency['p50']:.3f}s | "
                f"{latency['p95']:.3f}s | {latency['p99']:.3f}s |"
            )
    md_path = out / "summary.md"
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")
    xlsx_path = out / "summary.xlsx"
    _write_experiment_xlsx(summary, xlsx_path)
    print(f"wrote {json_path}, {md_path}, and {xlsx_path}")
    return EXIT_OK


def _write_experiment_xlsx(summary: Mapping[str, Any], path: Path) -> None:
    """Write a compact workbook; import openpyxl only on the reporting path."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:  # pragma: no cover - generation environment defect
        raise CliError(
            "experiment-report requires the root pinned openpyxl dependency"
        ) from exc

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["RECONCILED", "NO"])
    overview.append(["Experiment", summary["experiment_id"]])
    overview.append(["Plan SHA", summary["plan_sha"]])
    overview.append([])
    overview.append([
        "Arm", "Parse valid", "Calls", "Rate", "p50 s", "p95 s",
        "Cost lower bound", "Missing cost",
    ])
    for name, facts in summary["arms"].items():
        overview.append([
            name, facts["parse_ok"], facts["calls"], facts["parse_valid_rate"],
            facts["latency_s"]["p50"], facts["latency_s"]["p95"],
            facts["cost_usd"]["lower_bound"], facts["cost_usd"]["calls_missing"],
        ])

    quality = workbook.create_sheet("Quality")
    quality.append([
        "Candidate", "Slice", "Dimension", "Discordant", "Net", "Band",
        "Verdict", "Decision",
    ])
    for candidate, comparison in summary["comparisons"].items():
        for slice_name, rows in comparison["slices"].items():
            for row in rows:
                quality.append([
                    candidate, slice_name, row["dimension"], row["discordant"],
                    row["net"], row["band"], row["verdict"],
                    comparison["decision"]["status"],
                ])

    regression_sheet = workbook.create_sheet("Regressions")
    regression_sheet.append([
        "Candidate", "Item key", "Dimension", "Ground truth", "Incumbent",
        "Candidate output", "Error type",
    ])
    for candidate, comparison in summary["comparisons"].items():
        for row in comparison["regressions"]:
            regression_sheet.append([
                candidate, row["item_key"], row["dimension"], row["gt_label"],
                row["incumbent_label"], row["candidate_label"], row["error_type"],
            ])

    load_sheet = workbook.create_sheet("Load")
    load_sheet.append([
        "Arm", "Concurrency", "Parse valid", "Calls", "Calls/s", "p50 s",
        "p95 s", "p99 s", "Max s", "Cost lower bound", "Missing cost",
    ])
    for name, levels in summary["load"].items():
        for level in sorted(levels, key=int):
            facts = levels[level]
            load_sheet.append([
                name, int(level), facts["parse_ok"], facts["calls"],
                facts["throughput_calls_per_s"], facts["latency_s"]["p50"],
                facts["latency_s"]["p95"], facts["latency_s"]["p99"],
                facts["latency_s"]["max"], facts["cost_usd"]["lower_bound"],
                facts["cost_usd"]["calls_missing"],
            ])
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def _refuse_incomparable(incumbent: LoadedRun, candidate: LoadedRun) -> None:
    """Every mismatch that would still produce a full, plausible-looking table."""
    if incumbent.arm == candidate.arm:
        raise CliError(
            f"both runs are named {incumbent.arm!r}. report.render keys its mechanism "
            "tables by arm name, so two arms sharing one name collapse into a single "
            "entry and the paired comparison silently becomes an arm compared with "
            "itself. Re-run one of them with a different --arm."
        )
    for loaded in (incumbent, candidate):
        current = loaded.meta.get("scoring_code_sha")
        legacy = loaded.meta.get("scorer_sha")
        if current is not None and legacy is not None and current != legacy:
            raise CliError(
                f"{loaded.arm} records conflicting scoring hashes: "
                f"scoring_code_sha={current} but scorer_sha={legacy}. The compatibility "
                "alias must identify the same code; refusing a tampered or partial "
                "manifest."
            )
    for field, label in (
        ("testset_sha", "testset"),
        ("gt_sha", "ground truth"),
    ):
        left, right = incumbent.meta.get(field), candidate.meta.get(field)
        if left != right:
            raise CliError(
                f"the arms ran against different {label} files ({field}: {left} vs "
                f"{right}). A paired comparison requires the same items and the same "
                "labels; the arm that ran the easier set would look better for it."
            )

    # New-contract runs must have been classified by the same outcome code and must
    # represent the same common evaluation workload. Generation/runtime code and final
    # decision policy are recorded per arm/report rather than blockers: those are the
    # very layers an OpenRouter-to-GPU migration is expected to change.
    new_fields = (
        "application_contract_sha",
        "outcome_contract_sha",
        "workload_sha",
    )
    for field in new_fields:
        left, right = incumbent.meta.get(field), candidate.meta.get(field)
        if (left is None) != (right is None):
            raise CliError(
                f"one arm predates {field} provenance ({left!r} vs {right!r}). Refusing "
                "to mix a legacy run with a content-addressed enterprise run."
            )
        if left is not None and left != right:
            raise CliError(
                f"the arms differ on {field} ({left} vs {right}). A common-prompt "
                "comparison requires identical classification code and workload."
            )


def cmd_compare(args: argparse.Namespace) -> int:
    """The mechanism report: which mechanisms each arm passes.

    Not a percentage. Aggregate production metrics operate on normalized product rows;
    paired inference operates on independent call clusters. `report.py` states the
    grain map in full; this function's job is to hand it two arms it may compare.
    """
    incumbent = load_run(Path(args.incumbent))
    candidate = load_run(Path(args.candidate))
    _refuse_incomparable(incumbent, candidate)

    testset_path = (
        Path(args.testset)
        if args.testset
        else _recorded_artifact_path(incumbent.directory, incumbent.meta["testset_path"])
    )
    gt_path = (
        Path(args.gt)
        if args.gt
        else _recorded_artifact_path(incumbent.directory, incumbent.meta["gt_path"])
    )
    if not testset_path.exists():
        raise CliError(
            f"the run names testset {testset_path}, which no longer exists. Pass "
            "--testset to point at it; the sha recorded in run.json is then checked."
        )
    testset = _load_testset(testset_path, app=incumbent.meta.get("app", "retention"))
    actual = testset_sha(testset_path)
    if actual != incumbent.meta["testset_sha"]:
        raise CliError(
            f"{testset_path} has changed since the run: sha is {actual}, the run "
            f"recorded {incumbent.meta['testset_sha']}. Scoring against an edited "
            "testset would compare answers with labels they never saw."
        )
    _require_run_matrix(incumbent, testset)
    _require_run_matrix(candidate, testset)

    gt = _load_gt(gt_path)
    if manifest_mod.file_hash(gt_path) != incumbent.meta["gt_sha"]:
        raise CliError(
            f"{gt_path} has changed since the run. Refusing rather than scoring against "
            "labels that moved."
        )

    manifests = {
        loaded.arm: manifest_mod.Manifest(
            items_sha=manifest_mod.items_hash(gt),
            labels_sha=loaded.meta["gt_sha"],
            item_count=len(gt),
            scorer_sha=loaded.meta.get("scoring_code_sha", loaded.meta["scorer_sha"]),
            arm=loaded.arm,
            backend=loaded.meta.get("backend", "openrouter"),
            model_id=_observed_model_id(loaded),
            output_mechanism="guided_json",
            prompt_sha=loaded.result.prompt_sha,
            generation_config=loaded.decoding,
        )
        for loaded in (incumbent, candidate)
    }
    try:
        deltas = manifest_mod.assert_comparable(
            manifests[incumbent.arm], manifests[candidate.arm]
        )
    except manifest_mod.ManifestMismatch as exc:
        raise CliError(str(exc)) from exc

    per_replicate = {
        loaded.arm: replicate_records(loaded.result, testset.items)
        for loaded in (incumbent, candidate)
    }
    try:
        summaries = {
            loaded.arm: arm_summary(loaded, gt, per_replicate[loaded.arm])
            for loaded in (incumbent, candidate)
        }
        mechanisms = {
            loaded.arm: mechanism_table(gt, per_replicate[loaded.arm], list(testset.items))
            for loaded in (incumbent, candidate)
        }
    except ReportError as exc:
        raise CliError(str(exc)) from exc

    # The coverage gate, per dimension. It cannot fire while `flatten.to_rows` keeps
    # emitting a ground-truth skeleton for every failure -- which is the point of that
    # guarantee -- so this is here to fail loudly if that ever stops being true.
    coverage_warnings: list[str] = []
    for name in ("call_result", "reason", "product"):
        try:
            check_coverage(
                summaries[incumbent.arm].dimensions[name],  # type: ignore[arg-type]
                summaries[candidate.arm].dimensions[name],  # type: ignore[arg-type]
            )
        except CoverageMismatch as exc:
            coverage_warnings.append(str(exc))

    inc_rows, cand_rows = per_replicate[incumbent.arm][0], per_replicate[candidate.arm][0]
    disagreements = [
        disagreement(gt, inc_rows, cand_rows, name)
        for name in ("call_result", "reason", "product")
    ]
    try:
        key = hmac_key()
    except RuntimeError as exc:
        raise CliError(
            f"{exc} Without it the per-item regression list cannot be written, and an "
            "empty list would print as 'none.' -- which reads as 'the candidate lost "
            "nothing'. Set a value (any string; True's team holds the real one) and "
            "re-run."
        ) from exc
    regression_rows = [
        row
        for name in ("call_result", "reason", "product")
        for row in regressions(gt, inc_rows, cand_rows, name, key)
    ]

    text = render(
        summaries[incumbent.arm],
        summaries[candidate.arm],
        mechanisms,
        disagreements,
        regression_rows,
    )
    text += _footer(incumbent, candidate, deltas, coverage_warnings)

    print()
    print(text)
    if args.report:
        report_path = Path(args.report)
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(text, encoding="utf-8", newline="\n")
        print(f"wrote {report_path}")

    verdicts = {row.verdict for rows in mechanisms.values() for row in rows}
    return EXIT_PROBLEMS if verdicts & {"FAIL", "FLAKY"} else EXIT_OK


def cmd_judge(args: argparse.Namespace) -> int:
    """Independent-model adjudication of scorer disagreements. Diagnostic only.

    Reuses `_refuse_incomparable` -- the same gate `compare` applies -- because a judge
    run over two arms that were not comparable in the first place (different testset,
    ground truth or scorer version) would adjudicate disagreements that are artifacts of
    the mismatch, not of either model. This command never prints a verdict, a net, or
    anything `report.render`'s mechanism table would recognise: see `evalgen.judge`'s
    module docstring for how that separation is enforced rather than only claimed.
    """
    incumbent = load_run(Path(args.incumbent))
    candidate = load_run(Path(args.candidate))
    _refuse_incomparable(incumbent, candidate)

    runtime = _runtime_from_args(args)
    if runtime.backend is RuntimeBackend.OPENROUTER and not args.provider:
        raise CliError(
            "the advisory judge must pin one OpenRouter provider; pass --provider, "
            "or select an openai-compatible internal runtime where provider routing "
            "does not apply"
        )

    testset_path = (
        Path(args.testset)
        if args.testset
        else _recorded_artifact_path(incumbent.directory, incumbent.meta["testset_path"])
    )
    gt_path = (
        Path(args.gt)
        if args.gt
        else _recorded_artifact_path(incumbent.directory, incumbent.meta["gt_path"])
    )
    if not testset_path.is_file():
        raise CliError(f"judge testset not found: {testset_path}")
    if not gt_path.is_file():
        raise CliError(f"judge ground truth not found: {gt_path}")
    actual_testset_sha = testset_sha(testset_path)
    if actual_testset_sha != incumbent.meta.get("testset_sha"):
        raise CliError(
            f"{testset_path} has changed since the source runs: sha is "
            f"{actual_testset_sha}, expected {incumbent.meta.get('testset_sha')}"
        )
    actual_gt_sha = manifest_mod.file_hash(gt_path)
    if actual_gt_sha != incumbent.meta.get("gt_sha"):
        raise CliError(
            f"{gt_path} has changed since the source runs; refusing to ask a judge "
            "about labels the arms were not scored against"
        )
    testset = _load_testset(testset_path, app=incumbent.meta.get("app", "retention"))
    gt = _load_gt(gt_path)
    _require_run_matrix(incumbent, testset)
    _require_run_matrix(candidate, testset)

    source_models = {
        str(value)
        for value in (
            incumbent.meta.get("model_requested"),
            candidate.meta.get("model_requested"),
            *incumbent.result.observed_models().keys(),
            *candidate.result.observed_models().keys(),
        )
        if value
    }
    if args.model in source_models:
        raise CliError(
            f"judge model {args.model!r} is one of the evaluated arm models. The "
            "diagnostic requires a distinct model, not self-review."
        )

    def recorded_classification(loaded: LoadedRun) -> str:
        value = loaded.meta.get("data_classification")
        if value in _CLASSIFICATION_RANK:
            return str(value)
        path = _recorded_artifact_path(
            loaded.directory, loaded.meta.get("testset_path", "")
        )
        return "synthetic" if _inside(path, SYNTHETIC_FIXTURE_ROOT) else "customer"

    required_classification = max(
        (recorded_classification(incumbent), recorded_classification(candidate)),
        key=_CLASSIFICATION_RANK.__getitem__,
    )
    requested_classification = getattr(args, "data_classification", "synthetic")
    if _CLASSIFICATION_RANK[requested_classification] < _CLASSIFICATION_RANK[required_classification]:
        raise CliError(
            f"judge output classification {requested_classification!r} would downgrade "
            f"source evidence classified {required_classification!r}"
        )
    effective_classification = max(
        (requested_classification, required_classification),
        key=_CLASSIFICATION_RANK.__getitem__,
    )

    per_replicate = {
        loaded.arm: replicate_records(loaded.result, testset.items)
        for loaded in (incumbent, candidate)
    }
    inc_rows = per_replicate[incumbent.arm][0]
    cand_rows = per_replicate[candidate.arm][0]

    dimensions = args.dimension if args.dimension else ["call_result", "reason", "product"]

    if args.dry_run:
        total = sum(len(find_disagreements(gt, inc_rows, cand_rows, d)) for d in dimensions)
        print(
            f"dry run: {total} disagreement item(s) across {dimensions} would be sent "
            f"to {args.model} through {runtime.runtime_id}. No call made."
        )
        return EXIT_OK

    private_arg = getattr(args, "private_out", None)
    legacy_out = getattr(args, "out", None)
    if private_arg and legacy_out and Path(private_arg) != Path(legacy_out):
        raise CliError("--out is an alias for --private-out; pass only one destination")
    private_value = private_arg or legacy_out
    if not private_value:
        raise CliError(
            "judge calls require --private-out so exact request/raw-response evidence "
            "is fsynced after every paid call"
        )
    private_path = Path(private_value).expanduser().resolve(strict=False)
    if effective_classification != "synthetic":
        try:
            private_path = require_private_destination(private_path)
        except ArtifactError as exc:
            raise CliError(str(exc)) from exc
    raw_path = private_path.with_name(private_path.name + ".raw.jsonl")
    for target in (private_path, raw_path):
        if target.exists():
            raise CliError(
                f"refusing to mix or overwrite judge evidence at {target}; choose a "
                "new --private-out"
            )

    api_key = _api_key(runtime)
    append_jsonl(
        raw_path,
        {
            "type": "judge_private_journal",
            "schema_version": 1,
            "created_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "judge_model": args.model,
            "judge_runtime_fingerprint": runtime.fingerprint(),
            "incumbent_run_id": incumbent.meta.get("run_id", incumbent.directory.name),
            "candidate_run_id": candidate.meta.get("run_id", candidate.directory.name),
        },
    )
    if runtime == OPENROUTER_RUNTIME:
        client = build_client(api_key, timeout=args.timeout)
    else:
        client = build_client(api_key, timeout=args.timeout, runtime=runtime)

    source_provenance = {
        "incumbent_manifest_sha": canonical_sha(incumbent.meta),
        "candidate_manifest_sha": canonical_sha(candidate.meta),
        "selected_replicate": 1,
        "testset_sha": actual_testset_sha,
        "gt_sha": actual_gt_sha,
        "application_contract_sha": incumbent.meta.get("application_contract_sha"),
        "outcome_contract_sha": incumbent.meta.get("outcome_contract_sha"),
        "generation_contract_sha": incumbent.meta.get("generation_contract_sha"),
        "scoring_code_sha": incumbent.meta.get(
            "scoring_code_sha", incumbent.meta.get("scorer_sha")
        ),
        "incumbent_run_contract_sha": incumbent.meta.get("run_contract_sha"),
        "candidate_run_contract_sha": candidate.meta.get("run_contract_sha"),
    }

    report = run_judge(
        testset=testset,
        gt=gt,
        incumbent=inc_rows,
        candidate=cand_rows,
        dimensions=dimensions,
        client=client,
        model=args.model,
        provider=args.provider,
        reasoning_effort=(
            "none"
            if runtime.backend is RuntimeBackend.OPENROUTER
            else "provider-default"
        ),
        max_items=args.max_items,
        source_provenance=source_provenance,
        private_record_sink=lambda record: append_jsonl(raw_path, record),
    )
    report["incumbent_arm"] = incumbent.arm
    report["candidate_arm"] = candidate.arm
    report["restricted_source_run_ids"] = {
        "incumbent": incumbent.meta.get("run_id", incumbent.directory.name),
        "candidate": candidate.meta.get("run_id", candidate.directory.name),
    }

    text = _render_judge_report(report)
    print()
    print(text)

    atomic_write_text(
        private_path,
        json.dumps(report, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
    )
    print(f"wrote restricted review report {private_path}")
    print(f"wrote restricted raw call journal {raw_path}")
    safe_report = shareable_report(report)
    assert_shareable_payload(safe_report)
    if getattr(args, "shareable_out", None):
        shareable_path = Path(args.shareable_out)
        atomic_write_text(
            shareable_path,
            json.dumps(safe_report, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
        )
        print(f"wrote shareable judge report {shareable_path}")
    if args.report:
        report_path = Path(args.report)
        atomic_write_text(report_path, text + "\n")
        print(f"wrote {report_path}")
    return EXIT_OK


def _render_judge_report(report: dict[str, Any]) -> str:
    """A short, honest summary. No net, no verdict, no headline -- see the module
    docstring on `evalgen.judge` for why that is a structural constraint here, not a
    style choice."""
    summary = report["summary"]
    items = report.get("items", [])
    identity_counts: dict[str, int] = {}
    evidence_counts: dict[str, int] = {}
    for row in items:
        identity = str(row.get("identity_status", "unknown"))
        evidence = str(row.get("evidence_status", "unknown"))
        identity_counts[identity] = identity_counts.get(identity, 0) + 1
        evidence_counts[evidence] = evidence_counts.get(evidence, 0) + 1
    lines = [
        "=" * 78,
        "INDEPENDENT JUDGE -- DIAGNOSTIC ONLY, NOT A SCORED DIMENSION",
        "=" * 78,
        f"incumbent   {report['incumbent_arm']}",
        f"candidate   {report['candidate_arm']}",
        f"judge model {report['model_requested']} on "
        f"{report['provider_requested'] or 'internal runtime'}",
        f"dimensions  {', '.join(report['dimensions'])}",
        "",
        f"disagreement items found: {report['candidates_found']}"
        + ("  (TRUNCATED by --max-items)" if report["truncated"] else ""),
        f"items adjudicated:        {summary['total']}",
        f"usable opinions:           {summary.get('usable_count', 0)}",
        f"transport errors:          {summary.get('transport_error_count', 0)}",
        f"identity/execution errors: {summary.get('execution_error_count', 0)}",
        f"identity status:           {identity_counts}",
        f"evidence status:           {evidence_counts}",
        "",
    ]
    for verdict in ("ground_truth_correct", "defensible_disagreement", "ground_truth_error", "unclear"):
        lines.append(
            f"  {verdict:<26} {summary['counts'][verdict]:>4}  "
            f"({summary[f'{verdict}_rate']:.1%})"
        )
    lines.append(
        f"  of which parse_error=True {summary['parse_error_count']:>4}  "
        f"({summary['parse_error_rate']:.1%})"
    )
    if summary["counts"]["ground_truth_error"]:
        lines += [
            "",
            "Items the judge flagged as a likely GROUND TRUTH ERROR (review these by hand,",
            "the same way RET-11 was found -- this is an opinion from one model, not a",
            "correction applied to any fixture):",
        ]
        for row in report["items"]:
            if (
                row["verdict"] == "ground_truth_error"
                and row.get("execution_status") == "completed"
                and not row.get("parse_error")
            ):
                lines.append(
                    f"    {row['judgment_unit_id']} [{row['dimension']}] "
                    f"evidence={row.get('evidence_status')} "
                    f"identity={row.get('identity_status')}"
                )
    lines += [
        "",
        "This is one model's opinion about a disputed label, recorded once per item at",
        "temperature 0. It does not change any score, does not join any verdict, and is",
        "not itself RECONCILED against anything. Read it the way you would read a second",
        "reviewer's comment, not a second scorer's number. Transcript spans, rationale,",
        "and raw responses exist only in the restricted review bundle.",
    ]
    return "\n".join(lines)


def _observed_model_id(loaded: LoadedRun) -> str:
    """The model that answered, or a visible statement that more than one did."""
    observed = loaded.result.observed_models()
    if not observed:
        return "unknown"
    if len(observed) == 1:
        return next(iter(observed))
    return "MIXED: " + ", ".join(f"{name}x{count}" for name, count in observed.items())


def _footer(
    incumbent: LoadedRun,
    candidate: LoadedRun,
    deltas: Sequence[str],
    coverage_warnings: Sequence[str],
) -> str:
    """What the CLI knows and `render` cannot: which replicate was scored, and where.

    `render` builds its report from two `ArmSummary` objects and has no way to know
    that `dimensions` came from replicate 1 of several. Stating it here rather than
    leaving it in a docstring is the difference between a documented limitation and a
    number a reader will assume covers the whole run.
    """
    lines = [
        "",
        "=" * 78,
        "HOW THIS REPORT WAS PRODUCED",
        "=" * 78,
        f"incumbent run  {incumbent.directory}",
        f"candidate run  {candidate.directory}",
        "Section 5's aggregate metrics are scored on REPLICATE 1 of each arm. Pooling",
        "  every replicate into one prediction list would be wrong, not merely",
        "  approximate: metrics.outer_join keys predictions by (call_id, phone,",
        "  product), so replicates of one item overwrite each other and only the last",
        "  survives. Sections 1 and 4 -- the mechanism verdicts and N_flip -- are the",
        "  ones that read every replicate.",
        "Section 6 is split on the same line and deliberately. Its token, cost and latency",
        "  totals cover EVERY replicate, because every replicate was paid for whether or",
        "  not it was scored. Its cost per correct answer covers REPLICATE 1 alone on both",
        "  sides of the division, because the denominator is section 5's and a numerator",
        "  holding the whole bill would report repeats-times the cost of an answer.",
    ]
    if deltas:
        lines.append("Recorded arm differences (expected, not blocking):")
        lines.extend(f"  - {delta}" for delta in deltas)
    if coverage_warnings:
        lines.append("COVERAGE REFUSALS (the arms did not score comparable item sets):")
        lines.extend(f"  - {warning}" for warning in coverage_warnings)
    return "\n".join(lines) + "\n"


# ========================================================================= parser


def _add_runtime_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--runtime-manifest",
        default=None,
        metavar="JSON",
        help=(
            "complete non-secret RuntimeSpec manifest; mutually exclusive with the "
            "individual runtime flags"
        ),
    )
    parser.add_argument(
        "--runtime-backend",
        choices=("openrouter", "openai-compatible"),
        default=None,
        help=(
            "model gateway: historical OpenRouter behavior, or an internal "
            "OpenAI-compatible vLLM/TGI/SGLang endpoint"
        ),
    )
    parser.add_argument("--runtime-id", default=None, help="stable deployment/runtime id")
    parser.add_argument(
        "--base-url",
        default=None,
        help="runtime API root; local GPU default is http://127.0.0.1:8000/v1",
    )
    parser.add_argument(
        "--api-key-env",
        default=None,
        help="name of the environment variable holding the runtime key; never serialized",
    )
    parser.add_argument(
        "--runtime-metadata",
        action="append",
        default=[],
        metavar="KEY=VALUE",
        help=(
            "repeatable immutable non-secret deployment fact, e.g. "
            "image_digest=sha256:... or gpu_type=H100"
        ),
    )
    parser.add_argument(
        "--allow-insecure-http",
        action="store_true",
        help="allow plain HTTP to a non-loopback endpoint on an explicitly trusted network",
    )


def _add_run_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--model", required=True, help="model id understood by the selected runtime")
    parser.add_argument("--testset", default=str(DEFAULT_TESTSET))
    parser.add_argument("--gt", default=str(DEFAULT_GT))
    parser.add_argument("--app", default="retention")
    parser.add_argument("--prompt-id", default="v9_16_base")
    parser.add_argument("--out", default=str(DEFAULT_OUT))
    parser.add_argument(
        "--data-classification",
        choices=("synthetic", "internal", "customer"),
        default="synthetic",
        help=(
            "internal/customer artifacts must live below EVAL_HARNESS_DATA_DIR; "
            "synthetic preserves the committed fixture workflow"
        ),
    )
    parser.add_argument(
        "--resume-run",
        default=None,
        metavar="DIRECTORY",
        help="resume an incomplete journaled run with the exact same contract",
    )
    _add_runtime_arguments(parser)
    # No default, and that is the decision. A default provider would be this file
    # quietly choosing a backend for every model id anyone ever passes in, including
    # ones where the choice is wrong. The pin has to be argued for per model -- which
    # regime production runs, which parameters the endpoint can honour -- so it is
    # named on the command line and recorded in run.json.
    parser.add_argument(
        "--provider",
        default=None,
        metavar="NAME",
        help=(
            "pin the OpenRouter backend by provider name (the `provider_name` field of "
            "GET /api/v1/models/<id>/endpoints, e.g. 'DeepInfra', 'CoreWeave'). Sends "
            "order=[NAME] with allow_fallbacks=false, so a busy or unavailable endpoint "
            "fails the call instead of silently serving the arm from a second build. "
            "Unpinned by default; check the prompt_tokens fingerprint afterwards."
        ),
    )
    parser.add_argument("--concurrency", type=int, default=4)
    parser.add_argument("--timeout", type=float, default=120.0)
    # Production's decoding, transcribed from main.py:1116-1119. top_p=0.0 is kept as
    # production wrote it rather than "corrected" to 1.0: this harness reproduces the
    # arm it is comparing against, not the arm it would have configured.
    parser.add_argument("--temperature", type=float, default=0.0)
    parser.add_argument("--top-p", type=float, default=0.0)
    parser.add_argument("--seed", type=int, default=0)
    parser.add_argument("--max-tokens", type=int, default=8000)
    parser.add_argument(
        "--reasoning-effort",
        choices=sorted(REASONING_EFFORTS),
        default="provider-default",
        help=(
            "OpenRouter normalized reasoning effort. 'none' explicitly disables "
            "reasoning; 'provider-default' omits the field for historical compatibility."
        ),
    )
    parser.add_argument(
        "--max-attempts",
        type=int,
        default=3,
        help=(
            "maximum API attempts for failures that emitted no generation; enterprise "
            "experiments use 1 so provider reliability is measured directly"
        ),
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="write the exact request bodies and the prompt sha, making ZERO API calls",
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="evalgen",
        description=(
            "Run a Thai retention eval set through two models and report which "
            "mechanisms each one passes."
        ),
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    check = subparsers.add_parser(
        "check", help="validate the pack. No network, no key, no cost."
    )
    check.add_argument("--testset", default=str(DEFAULT_TESTSET))
    check.add_argument("--gt", default=str(DEFAULT_GT))
    check.add_argument("--app", default="retention")
    check.add_argument("--prompt-id", default="v9_16_base")
    check.set_defaults(handler=cmd_check)

    baseline = subparsers.add_parser("baseline", help="run one arm over the whole pack")
    baseline.add_argument("--arm", required=True)
    baseline.add_argument("--repeats", type=int, default=3)
    _add_run_arguments(baseline)
    baseline.set_defaults(handler=cmd_baseline)

    stability = subparsers.add_parser(
        "stability", help="the N_flip probe: a few items, many identical replicates"
    )
    stability.add_argument("--arm", default="stability")
    stability.add_argument(
        "--items", required=True, help="comma-separated item ids, e.g. RET-01,RET-11,RET-16"
    )
    stability.add_argument("--repeats", type=int, default=5)
    _add_run_arguments(stability)
    stability.set_defaults(handler=cmd_stability)

    compare = subparsers.add_parser("compare", help="the mechanism report")
    compare.add_argument("--incumbent", required=True, help="a run directory")
    compare.add_argument("--candidate", required=True, help="a run directory")
    compare.add_argument("--testset", default=None, help="override the recorded path")
    compare.add_argument("--gt", default=None, help="override the recorded path")
    compare.add_argument("--report", default=None, help="also write the report here")
    compare.set_defaults(handler=cmd_compare)

    judge = subparsers.add_parser(
        "judge",
        help="independent-model adjudication of scorer disagreements (diagnostic only)",
    )
    judge.add_argument("--incumbent", required=True, help="a run directory")
    judge.add_argument("--candidate", required=True, help="a run directory")
    judge.add_argument("--testset", default=None, help="override the recorded path")
    judge.add_argument("--gt", default=None, help="override the recorded path")
    judge.add_argument(
        "--dimension",
        action="append",
        choices=("call_result", "reason", "product"),
        default=None,
        help="repeatable; default is all three",
    )
    judge.add_argument("--model", required=True, help="runtime model id for the judge")
    judge.add_argument(
        "--provider",
        default=None,
        help="required OpenRouter provider pin; omit for an internal GPU runtime",
    )
    judge.add_argument("--max-items", type=int, default=None)
    judge.add_argument("--timeout", type=float, default=120.0)
    judge.add_argument(
        "--data-classification",
        choices=("synthetic", "internal", "customer"),
        default="synthetic",
    )
    _add_runtime_arguments(judge)
    judge.add_argument(
        "--private-out",
        default=None,
        help="restricted JSON review report; a sibling raw JSONL journal is checkpointed",
    )
    judge.add_argument(
        "--shareable-out",
        default=None,
        help="sanitized diagnostic JSON with no transcript spans, rationale, raw text, or item ids",
    )
    judge.add_argument(
        "--out",
        default=None,
        help="deprecated alias for --private-out",
    )
    judge.add_argument("--report", default=None, help="write a shareable text summary here")
    judge.add_argument(
        "--dry-run",
        action="store_true",
        help="count disagreement items and exit. No call, no key required.",
    )
    judge.set_defaults(handler=cmd_judge)

    experiment_check = subparsers.add_parser(
        "experiment-check",
        help="validate a preregistered enterprise experiment; no network or key",
    )
    experiment_check.add_argument("--plan", default=str(DEFAULT_EXPERIMENT_PLAN))
    experiment_check.set_defaults(handler=cmd_experiment_check)

    experiment_budget = subparsers.add_parser(
        "experiment-budget",
        help="print the current no-network qualification/full/load cost ceiling",
    )
    experiment_budget.add_argument("--plan", default=str(DEFAULT_EXPERIMENT_PLAN))
    experiment_budget.set_defaults(handler=cmd_experiment_budget)

    qualify = subparsers.add_parser(
        "qualify",
        help="run one six-call explicit-regime provider qualification probe",
    )
    qualify.add_argument("--plan", default=str(DEFAULT_EXPERIMENT_PLAN))
    qualify.add_argument("--arm", required=True)
    qualify.add_argument("--provider", required=True)
    qualify.add_argument("--out", default=str(DEFAULT_OUT))
    qualify.add_argument("--timeout", type=float, default=120.0)
    qualify.add_argument("--data-classification", choices=("synthetic", "internal", "customer"), default="synthetic")
    qualify.add_argument("--resume-run", default=None)
    _add_runtime_arguments(qualify)
    qualify.add_argument("--dry-run", action="store_true")
    qualify.set_defaults(handler=cmd_qualify)

    qualification_report = subparsers.add_parser(
        "qualification-report",
        help="reclassify one recorded qualification run; no network or key",
    )
    qualification_report.add_argument(
        "--plan", default=str(DEFAULT_EXPERIMENT_PLAN)
    )
    qualification_report.add_argument("--run", required=True)
    qualification_report.add_argument(
        "--out", default=None, help="default: overwrite RUN/qualification.json"
    )
    qualification_report.set_defaults(handler=cmd_qualification_report)

    experiment_run = subparsers.add_parser(
        "experiment-run",
        help="run one arm from a locked experiment plan",
    )
    experiment_run.add_argument("--plan", default=str(DEFAULT_EXPERIMENT_PLAN))
    experiment_run.add_argument("--arm", required=True)
    experiment_run.add_argument("--mode", choices=("full", "load"), default="full")
    experiment_run.add_argument("--concurrency-level", type=int, default=None)
    experiment_run.add_argument("--confirm-plan-sha", required=True)
    experiment_run.add_argument("--out", default=str(DEFAULT_OUT))
    experiment_run.add_argument("--timeout", type=float, default=120.0)
    experiment_run.add_argument("--data-classification", choices=("synthetic", "internal", "customer"), default="synthetic")
    experiment_run.add_argument("--resume-run", default=None)
    _add_runtime_arguments(experiment_run)
    experiment_run.add_argument("--dry-run", action="store_true")
    experiment_run.set_defaults(handler=cmd_experiment_run)

    experiment_report = subparsers.add_parser(
        "experiment-report",
        help="assemble quality-first JSON and Markdown reports from full runs",
    )
    experiment_report.add_argument("--plan", default=str(DEFAULT_EXPERIMENT_PLAN))
    experiment_report.add_argument(
        "--run",
        action="append",
        required=True,
        metavar="ARM=PATH",
        help="repeat once per full arm",
    )
    experiment_report.add_argument(
        "--load-run",
        action="append",
        default=[],
        metavar="ARM@CONCURRENCY=PATH",
        help="repeat for each completed load arm; all three levels are required per arm",
    )
    experiment_report.add_argument("--out", required=True)
    experiment_report.set_defaults(handler=cmd_experiment_report)

    return parser


def main(argv: Sequence[str] | None = None) -> int:
    """Entry point. `configure_stdout()` first, before anything can print.

    First is not a style preference. This CLI prints Thai -- `MechanismRow.detail`
    carries the testset's `expected_failure` prose verbatim -- and on a Thai-locale
    Windows console `sys.stdout` is cp874, which raises `UnicodeEncodeError` rather
    than mangling a character. Anything printed before this call is printed on a stream
    that cannot say ยกเลิก, and it would fail after the model calls had been paid for.
    """
    configure_stdout()

    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return int(args.handler(args))
    except CliError as exc:
        print(f"\nREFUSED: {exc}", file=sys.stderr)
        return EXIT_REFUSED


if __name__ == "__main__":  # pragma: no cover - exercised through scripts/evalgen.py
    raise SystemExit(main())
