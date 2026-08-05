"""Unit tests for app/processors/report_builder.py."""
from __future__ import annotations

import io

import openpyxl
import pandas as pd
import pytest

from app.processors.report_builder import ReportBuilder

# ---------------------------------------------------------------------------
# protect_excel_value
# ---------------------------------------------------------------------------

def test_protect_date_string() -> None:
    assert ReportBuilder.protect_excel_value("2024-01-15") == "'2024-01-15"


def test_protect_fraction_string() -> None:
    assert ReportBuilder.protect_excel_value("3/3") == "'3/3"


def test_protect_single_digit_fraction() -> None:
    assert ReportBuilder.protect_excel_value("0/2") == "'0/2"


def test_protect_passthrough_plain_string() -> None:
    assert ReportBuilder.protect_excel_value("hello") == "hello"


def test_protect_passthrough_integer() -> None:
    assert ReportBuilder.protect_excel_value(42) == 42


def test_protect_passthrough_float() -> None:
    assert ReportBuilder.protect_excel_value(3.14) == 3.14


def test_protect_passthrough_none() -> None:
    assert ReportBuilder.protect_excel_value(None) is None


def test_protect_does_not_quote_datetime() -> None:
    # "2024-01-15 10:30:00" is not a pure date → should NOT be quoted
    val = "2024-01-15 10:30:00"
    assert ReportBuilder.protect_excel_value(val) == val


# ---------------------------------------------------------------------------
# build_user_excel — sheet presence and column headers
# ---------------------------------------------------------------------------

def _make_df(cols: list[str], n_rows: int = 2) -> pd.DataFrame:
    data = {c: [f"{c}_{i}" for i in range(n_rows)] for c in cols}
    return pd.DataFrame(data)


def test_build_user_excel_returns_bytesio() -> None:
    schema1 = ["Col_A", "Col_B"]
    schema2 = ["Col_X", "Col_Y"]
    builder = ReportBuilder(schema1, schema2)
    df1 = _make_df(schema1)
    df2 = _make_df(schema2)
    buf = builder.build_user_excel(df1, df2)
    assert isinstance(buf, io.BytesIO)
    assert buf.tell() == 0


def test_build_user_excel_sheet_names() -> None:
    schema1 = ["Col_A", "Col_B"]
    schema2 = ["Col_X", "Col_Y"]
    builder = ReportBuilder(schema1, schema2)
    df1 = _make_df(schema1)
    df2 = _make_df(schema2)
    buf = builder.build_user_excel(df1, df2)
    wb = openpyxl.load_workbook(buf)
    assert "Incompliant Photo Retailer" in wb.sheetnames
    assert "Suspicious Retailer (Active)" in wb.sheetnames


def test_build_user_excel_header_row() -> None:
    schema1 = ["Name", "Status", "Zone"]
    builder = ReportBuilder(schema1, [])
    df1 = _make_df(schema1)
    buf = builder.build_user_excel(df1)
    wb = openpyxl.load_workbook(buf)
    ws = wb["Incompliant Photo Retailer"]
    headers = [ws.cell(1, c).value for c in range(1, len(schema1) + 1)]
    assert headers == schema1


def test_build_user_excel_no_suspicious_sheet() -> None:
    schema1 = ["Col_A"]
    builder = ReportBuilder(schema1, [])
    df1 = _make_df(schema1)
    buf = builder.build_user_excel(df1)
    wb = openpyxl.load_workbook(buf)
    assert "Suspicious Retailer (Active)" not in wb.sheetnames


def test_build_user_excel_row_count() -> None:
    schema1 = ["A", "B"]
    builder = ReportBuilder(schema1, [])
    df1 = _make_df(schema1, n_rows=5)
    buf = builder.build_user_excel(df1)
    wb = openpyxl.load_workbook(buf)
    ws = wb["Incompliant Photo Retailer"]
    # Row 1 = header, rows 2-6 = data
    assert ws.max_row == 6


# ---------------------------------------------------------------------------
# build_transaction_row
# ---------------------------------------------------------------------------

def _make_log(**overrides: object) -> dict:
    base: dict = {
        "status": "success",
        "start_time": "2024-01-15 10:00:00",
        "end_time": "2024-01-15 10:01:30",
        "process_time": 90.0,
        "message": "success process : RTR001-Shop",
        "rtr_code": "RTR001",
        "rtr_name": "Shop One",
        "image_parts": ["bucket/shop/p1.jpg", "bucket/shop/p2.jpg"],
        "meta_data": {
            "text_input_tokens": 100,
            "image_input_tokens": 200,
            "text_cache_tokens": 0,
            "image_cache_tokens": 0,
            "output_tokens": 50,
        },
    }
    base.update(overrides)
    return base


def test_build_transaction_row_required_keys() -> None:
    builder = ReportBuilder([], [])
    row = builder.build_transaction_row(_make_log(), "20240115", "proj-id", "proj-name")
    required = {
        "data_date", "start_time", "end_time", "total_time_mins",
        "type", "gcp_project_id", "gcp_project_name",
        "status_pass_failed_retry", "latency_ms",
        "token_usage_input", "token_usage_output", "total_cost_usd",
    }
    assert required.issubset(set(row.keys()))


def test_build_transaction_row_data_date() -> None:
    builder = ReportBuilder([], [])
    row = builder.build_transaction_row(_make_log(), "20240115", "pid", "pname")
    assert row["data_date"] == "20240115"


def test_build_transaction_row_token_input() -> None:
    builder = ReportBuilder([], [])
    row = builder.build_transaction_row(_make_log(), "20240115", "pid", "pname")
    # text_input (100) + image_input (200) = 300
    assert row["token_usage_input"] == 300


def test_build_transaction_row_cost_positive() -> None:
    builder = ReportBuilder([], [])
    row = builder.build_transaction_row(_make_log(), "20240115", "pid", "pname")
    assert row["total_cost_usd"] > 0


def test_build_transaction_row_latency_ms() -> None:
    builder = ReportBuilder([], [])
    log = _make_log(process_time=1.5)
    row = builder.build_transaction_row(log, "20240115", "pid", "pname")
    assert row["latency_ms"] == pytest.approx(1500.0, rel=1e-3)


def test_build_transaction_row_error_log_only_on_failure() -> None:
    builder = ReportBuilder([], [])
    success_row = builder.build_transaction_row(_make_log(status="success"), "20240115", "p", "p")
    fail_row = builder.build_transaction_row(_make_log(status="fail", message="oops"), "20240115", "p", "p")
    assert success_row["error_log_if"] == ""
    assert fail_row["error_log_if"] == "oops"


def test_build_transaction_row_folder_from_image_parts() -> None:
    builder = ReportBuilder([], [])
    row = builder.build_transaction_row(
        _make_log(image_parts=["bucket/shop/p1.jpg"]), "20240115", "p", "p"
    )
    assert row["folder"] == "bucket/shop/p1.jpg"


def test_build_transaction_row_empty_image_parts() -> None:
    builder = ReportBuilder([], [])
    row = builder.build_transaction_row(_make_log(image_parts=[]), "20240115", "p", "p")
    assert row["folder"] == ""
