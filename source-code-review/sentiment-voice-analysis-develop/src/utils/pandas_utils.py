import io
from pathlib import Path

import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter

from src.utils.logger import Logger

logger = Logger(__name__)


def export_pandas_df_to_markdown(df: pd.DataFrame, file_path: str | Path, encoding: str = "utf-8") -> None:
    """
    Export a pandas DataFrame to a markdown file.
    Parameters:
        df (pd.DataFrame): The DataFrame to export.
        file_path (str|Path): The path to the output markdown file.
        encoding (str): The file encoding, default is 'utf-8'.
    Returns:
        None
    """
    logger.debug(f"Exporting DataFrame with {len(df)} rows to markdown file: {file_path}")
    if isinstance(file_path, str):
        file_path = Path(file_path)
    if not file_path.parent.exists():
        logger.info(f"Creating directory: {file_path.parent}")
        file_path.parent.mkdir(parents=True, exist_ok=True)
    with open(file_path, "w", encoding=encoding) as f:
        f.write(df.to_markdown(index=False))
    logger.info(f"Successfully exported DataFrame to markdown: {file_path}")


def convert_df_to_markdown(df: pd.DataFrame) -> str:
    """
    Convert a pandas DataFrame to a markdown string.
    Parameters:
        df (pd.DataFrame): The DataFrame to convert.
    Returns:
        str: The markdown representation of the DataFrame.
    """
    logger.debug(f"Converting DataFrame with {len(df)} rows to markdown string")
    markdown_str = df.to_markdown(index=False)
    logger.info("Successfully converted DataFrame to markdown string")
    return markdown_str


def convert_df_to_excel(df: pd.DataFrame) -> bytes:
    """
    Convert a pandas DataFrame to Excel format and return as bytes.
    Parameters:
        df (pd.DataFrame): The DataFrame to convert.
    Returns:
        bytes: The Excel file content in bytes.
    """
    logger.debug(f"Converting DataFrame with {len(df)} rows and {len(df.columns)} columns to Excel bytes")
    with io.BytesIO() as output:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
        data = output.getvalue()
    logger.info(f"Successfully converted DataFrame to Excel bytes ({len(data)} bytes)")
    return data


def ensure_df_schema(df: pd.DataFrame, schemas: list) -> pd.DataFrame:
    """
    Ensure that the DataFrame contains all required schema columns.

    Parameters:
        df (pd.DataFrame): Input DataFrame to validate.
        schemas (list): List of required schema field names.
    Returns:
        pd.DataFrame: DataFrame with all schema columns.
    """
    logger.debug(f"Validating DataFrame schema with {len(schemas)} required columns")
    # Get all schema field names in order
    schema_columns = list(schemas)

    # Add missing columns with None values
    missing_columns = []
    for col in schema_columns:
        if col not in df.columns:
            logger.debug(f"Adding missing column: {col}")
            df[col] = pd.NA
            missing_columns.append(col)

    if missing_columns:
        logger.warning(f"Added {len(missing_columns)} missing columns: {missing_columns}")
    else:
        logger.debug("All required columns are present")

    # Reorder columns to match schema
    df = df[schema_columns]
    logger.info(f"DataFrame schema validated and reordered to match {len(schema_columns)} columns")

    return df


def replace_nan_with_default(df: pd.DataFrame, default_value: any = "", columns: list = None) -> pd.DataFrame:
    """
    Replace NaN values (both actual NaN and string 'nan'/'NaN') with a default value.

    Parameters:
        df (pd.DataFrame): Input DataFrame to clean.
        default_value (any): The value to replace NaN with. Default is empty string ''.
        columns (list): Specific columns to process. If None, processes all columns.
    Returns:
        pd.DataFrame: DataFrame with NaN values replaced.
    """
    logger.debug(f"Replacing NaN values with default value: '{default_value}'")
    df_copy = df.copy()

    # Determine which columns to process
    target_columns = columns if columns is not None else df_copy.columns

    # Opt into the non-downcasting future behavior so ``replace`` does not emit pandas'
    # downcasting FutureWarning (pandas 2.2). The dtype is left as-is; values are blank-filled
    # for the text/CSV exports this helper feeds.
    with pd.option_context("future.no_silent_downcasting", True):
        for col in target_columns:
            if col in df_copy.columns:
                # Replace string 'nan'/'NaN'/'None' with actual NaN first, then fill.
                df_copy[col] = df_copy[col].replace(["nan", "NaN", "None", "none"], pd.NA)
                df_copy[col] = df_copy[col].fillna(default_value)
                logger.debug(f"Replaced NaN values in column '{col}'")

    logger.info(f"Successfully replaced NaN values in {len(target_columns)} columns")
    return df_copy


_EXCEL_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _escape_excel_formula(value):
    """Prefix string cells starting with a formula trigger with a single quote.

    Mitigates CSV/Excel injection: an analyst opening the file with an
    unsanitized ``=cmd|'/C calc'!A1`` cell would trigger formula evaluation.
    """
    if isinstance(value, str) and value.startswith(_EXCEL_FORMULA_PREFIXES):
        return "'" + value
    return value


def df_to_excel_bytes(
    df: pd.DataFrame, sheet_name: str = "Sheet1", freeze_panes: str = None, text_columns: list = None
) -> bytes:
    """
    Convert a pandas DataFrame to a formatted Excel file in bytes.
    Parameters:
        df (pd.DataFrame): The DataFrame to be formatted.
        sheet_name (str): The name of the Excel sheet. Default is 'Sheet1'.
        freeze_panes (str): Cell reference to freeze panes (e.g., 'A2' to freeze top row). Default is None.
        text_columns (list | str): List of column names to format as text, or "ALL" for all columns.
            Default is None.
    Returns:
        bytes: The formatted Excel file in bytes.
    """
    logger.debug(f"Converting DataFrame ({len(df)} rows, {len(df.columns)} columns) to formatted Excel bytes")
    try:
        # Sanitize formula-injection vectors before writing (mutates a copy only).
        _sanitize = df.map if hasattr(df, "map") else df.applymap
        df = _sanitize(_escape_excel_formula)

        with io.BytesIO() as output:
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]

                # Format specified columns as text using Excel '@' format (prevents scientific notation)
                if text_columns and isinstance(text_columns, list):
                    logger.debug(f"Applying text format to columns: {text_columns}")

                    for col_name in text_columns:
                        if col_name in df.columns:
                            col_idx = df.columns.get_loc(col_name) + 1  # openpyxl is 1-indexed
                            col_letter = get_column_letter(col_idx)

                            # Apply '@' number format to all cells in column (Excel text format)
                            for cell in worksheet[col_letter]:
                                cell.number_format = "@"

                            logger.debug(f"Applied text format '@' to column '{col_name}' ({col_letter})")
                elif text_columns and isinstance(text_columns, str) and text_columns == "ALL":
                    logger.debug("Applying text format '@' to all columns")
                    for col_idx in range(1, len(df.columns) + 1):
                        col_letter = get_column_letter(col_idx)
                        for cell in worksheet[col_letter]:
                            cell.number_format = "@"

                # Add auto-filter
                logger.debug(f"Adding auto-filter to sheet '{sheet_name}'")
                worksheet.auto_filter.ref = worksheet.dimensions

                # Freeze top row
                if freeze_panes:
                    logger.debug(f"Freezing panes at: {freeze_panes}")
                    worksheet.freeze_panes = worksheet[freeze_panes]

                # Adjust column width
                logger.debug("Adjusting column widths")
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass

                    # Set a reasonable max width to avoid extremely wide columns
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            data = output.getvalue()
        logger.info(
            f"Successfully converted DataFrame to formatted Excel bytes ({len(data)} bytes) with sheet '{sheet_name}'"
        )
        return data
    except Exception as e:
        logger.error(f"Error converting DataFrame to Excel bytes: {e}")
        raise Exception(f"Error converting DataFrame to Excel bytes: {e}") from e


def trim_all_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Trim whitespace from all string columns in a DataFrame.
    Parameters:
        df (pd.DataFrame): The DataFrame to be trimmed.
    Returns:
        pd.DataFrame: The trimmed DataFrame.
    """
    if not isinstance(df, pd.DataFrame):
        raise ValueError("Input must be a pandas DataFrame")
    try:
        for col in df.select_dtypes(include=["object"]).columns:
            df[col] = df[col].str.strip()
        return df
    except Exception as e:
        raise ValueError(f"Error trimming DataFrame columns: {e}") from e


def clean_invalid_xml_chars(val):
    if isinstance(val, str):
        # Removes control characters that crash openpyxl
        return ILLEGAL_CHARACTERS_RE.sub("", val)
    return val
