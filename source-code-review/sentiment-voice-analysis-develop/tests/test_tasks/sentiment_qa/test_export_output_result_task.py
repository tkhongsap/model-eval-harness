import io
from datetime import date, datetime
from types import SimpleNamespace
from unittest.mock import MagicMock, Mock, patch

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

import tasks.sentiment_qa.export_output_result_task as qa_export_module
from tasks.sentiment_qa.export_output_result_task import ExportOutputResultTask

COMMON_YAML = {
    "verint": {
        "site_domain": "verint.sharepoint.com",
        "client_id": "verint-client",
        "client_secret": "verint-secret",
        "tenant_id": "tenant-id",
        "site_path": "/sites/verint",
    },
    "control": {
        "site_domain": "control.sharepoint.com",
        "client_id": "control-client",
        "client_secret": "control-secret",
        "tenant_id": "tenant-id",
        "site_path": "/sites/control",
        "site_name": "control-site",
        "gemini_cost_path": "config/gemini_cost.yml",
    },
    "msgraph": {
        "tenant_id": "tenant-id",
        "client_id": "graph-client",
        "client_secret": "graph-secret",
        "sender_email": "sender@example.com",
        "receiver_email": "receiver@example.com",
        "cc_email": "cc@example.com",
    },
    "sandbox": {
        "site_domain": "sandbox.sharepoint.com",
        "client_id": "sandbox-client",
        "client_secret": "sandbox-secret",
        "tenant_id": "tenant-id",
        "site_path": "/sites/sandbox",
    },
}


def _excel_bytes(df: pd.DataFrame, sheet_name: str) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    return buf.getvalue()


def _mock_config_excel() -> bytes:
    user_prompt_inbound_df = pd.DataFrame(
        [
            {
                "no_cate": 2,
                "category": "Greeting",
                "sub_category": "opening",
                "item": "salutation",
                "rule_and_logic": "Say hello politely.",
            },
            {
                "no_cate": pd.NA,
                "category": pd.NA,
                "sub_category": "issue_type",
                "item": pd.NA,
                "rule_and_logic": "Describe issue type only for network calls.",
            },
        ]
    )

    user_prompt_outbound_df = pd.DataFrame(
        [
            {
                "no_cate": 2,
                "category": "Greeting",
                "sub_category": "opening",
                "item": "salutation",
                "rule_and_logic": "Say hello politely.",
            },
            {
                "no_cate": pd.NA,
                "category": pd.NA,
                "sub_category": "issue_type",
                "item": pd.NA,
                "rule_and_logic": "Describe issue type only for network calls.",
            },
        ]
    )

    weight_df = pd.DataFrame(
        [
            {"item": "greeting_standard", "inbound_weight": 10, "outbound_weight": 10},
            {"item": "manners", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "enthusiasm", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "communication_skill", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "ending_standard", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "data_privacy", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "legal_verification", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "company_verification", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "sla_notification", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "transfer_standard", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "problem_understanding", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "compensation", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "hold_standard", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "wrap_up", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "beyond_scope_support", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "self_service", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "case_ownership", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "contact_confirm", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "omotenashi", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "retention", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "downsell", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "mnp", "inbound_weight": 5, "outbound_weight": 5},
            {"item": "upselling", "inbound_weight": 5, "outbound_weight": 5},
        ]
    )

    product_mapping_df = pd.DataFrame(
        [
            {"product_group": "TOL", "product_category": "CCTV"},
            {"product_group": None, "product_category": "IoT Device"},
            {"product_group": None, "product_category": "IoT Device and Service"},
            {"product_group": None, "product_category": "True Online"},
            {"product_group": None, "product_category": "True ID Box"},
            {"product_group": None, "product_category": "Trueonline mesh wi-fi"},
            {"product_group": None, "product_category": "TrueID TV"},
            {"product_group": None, "product_category": "TrueX"},
            {"product_group": "Postpay", "product_category": "Dtac Mobile Postpaid"},
            {"product_group": None, "product_category": "Mobile data topping package"},
            {"product_group": None, "product_category": "Mobile Device"},
            {"product_group": None, "product_category": "Mobile Postpaid"},
            {"product_group": None, "product_category": "True Mobile Postpaid"},
            {"product_group": None, "product_category": "Mobile Voice topping package"},
            {"product_group": "Prepay", "product_category": "Dtac Mobile Prepaid"},
            {"product_group": None, "product_category": "True Mobile Prepaid"},
            {"product_group": None, "product_category": "Mobile Prepaid"},
            {"product_group": "TVS", "product_category": "Entertainment Content"},
            {"product_group": None, "product_category": "TrueVisions"},
            {"product_group": "Other", "product_category": "General Service (not telecom related)"},
            {"product_group": None, "product_category": "Other"},
            {"product_group": None, "product_category": "True Agent Dealer Commission"},
            {"product_group": None, "product_category": "True Space Service"},
            {"product_group": None, "product_category": "Netflix"},
            {"product_group": None, "product_category": "True Money"},
            {"product_group": "Application", "product_category": "True App"},
            {"product_group": None, "product_category": "TrueID App"},
        ]
    )

    service_quality_df = pd.DataFrame(
        [
            {"sub_category": "communication", "item": "greeting_standard"},
            {"sub_category": "communication", "item": "manners"},
            {"sub_category": "communication", "item": "enthusiasm"},
            {"sub_category": "communication", "item": "communication_skill"},
            {"sub_category": "communication", "item": "ending_standard"},
            {"sub_category": "communication", "item": "transfer_standard"},
            {"sub_category": "communication", "item": "hold_standard"},
            {"sub_category": "communication", "item": "contact_confirm"},
            {"sub_category": "communication", "item": "upselling"},
            {"sub_category": "compliance", "item": "data_privacy"},
            {"sub_category": "compliance", "item": "legal_verification"},
            {"sub_category": "compliance", "item": "company_verification"},
            {"sub_category": "compliance", "item": "wrap_up"},
            {"sub_category": "compliance", "item": "self_service"},
            {"sub_category": "compliance", "item": "retention"},
            {"sub_category": "compliance", "item": "downsell"},
            {"sub_category": "compliance", "item": "mnp"},
            {"sub_category": "solving", "item": "sla_notification"},
            {"sub_category": "solving", "item": "problem_understanding"},
            {"sub_category": "solving", "item": "compensation"},
            {"sub_category": "solving", "item": "beyond_scope_support"},
            {"sub_category": "solving", "item": "case_ownership"},
            {"sub_category": None, "item": "omotenashi"},
        ]
    )

    standard_gsd_df = pd.DataFrame(
        [
            {
                "standard_gsd_name": "Account Login / Self-service Access",
                "definition": "ใช้กับเคสที่ลูกค้าเข้าแอปหรือระบบ Self-service ไม่ได้ เช่น ล็อกอินไม่สำเร็จ หน้าแอป error หรือไม่สามารถจัดการบริการผ่านแอปได้",  # noqa: E501
            },
            {
                "standard_gsd_name": "Account Login / Self-service Access",
                "definition": "ใช้กับเคสที่มีปัญหาการล็อกอินหรือการผูกบัญชีกับหมายเลข/บริการ ทำให้ลูกค้าเข้าใช้งานหรือจัดการข้อมูลในระบบดิจิทัลไม่ได้",  # noqa: E501
            },
            {
                "standard_gsd_name": "OTP / Verification Issue",
                "definition": "ใช้กับเคสที่ลูกค้าไม่ได้รับ OTP หรือรหัสยืนยันอื่นไม่ได้ ส่งผลให้ยืนยันตัวตน เปลี่ยนแปลงข้อมูล หรือทำรายการต่อไม่ได้",
            },
            {
                "standard_gsd_name": "Account Linkage / Registration",
                "definition": "ใช้กับเคสลงทะเบียนหรือเชื่อมบัญชีที่ต้องยืนยันตัวตนผ่านวิดีโอคอล/ส่งเอกสาร แต่ลูกค้าติดขั้นตอนหรือสถานะยังไม่ผ่าน",
            },
            {
                "standard_gsd_name": "PUK / SIM Unlock Support",
                "definition": "ใช้กับเคสที่ซิมถูกล็อกหรือต้องใช้รหัส PUK เพื่อปลดล็อก เช่น ใส่ PIN ผิดหลายครั้งจนใช้งานซิมไม่ได้",
            },
        ]
    )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        user_prompt_inbound_df.to_excel(writer, sheet_name="user_prompt_inbound", index=False)
        user_prompt_outbound_df.to_excel(writer, sheet_name="user_prompt_outbound", index=False)
        weight_df.to_excel(writer, sheet_name="weight_score", index=False)
        product_mapping_df.to_excel(writer, sheet_name="product_mapping", index=False)
        service_quality_df.to_excel(writer, sheet_name="service_quality_group", index=False)
        standard_gsd_df.to_excel(writer, sheet_name="standard_gsd", index=False)

    return output.getvalue()


def _legacy_master_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A1:C1")
    ws["A1"] = "Legacy"
    headers = ["filename", "agent_id", "call_date"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=idx, value=header)
    ws.cell(row=3, column=1, value="existing.wav")
    ws.cell(row=3, column=2, value="OLD001")
    ws.cell(row=3, column=3, value="20241231")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def mock_deps():
    with (
        patch("tasks.sentiment_qa.export_output_result_task.SharePointModule") as sp_cls,
        patch("tasks.sentiment_qa.export_output_result_task.GCSModule") as gcs_cls,
        patch("tasks.sentiment_qa.export_output_result_task.MSGraphModule") as graph_cls,
        patch("tasks.sentiment_qa.export_output_result_task.load_yaml") as load_yaml,
        patch("tasks.sentiment_qa.export_output_result_task.resolve_env") as resolve_env,
        patch("tasks.sentiment_qa.export_output_result_task.resolve_date") as resolve_date,
    ):
        load_yaml.return_value = COMMON_YAML
        resolve_env.side_effect = lambda value: value
        resolve_date.side_effect = lambda text, replace_date=None: text
        yield {
            "sharepoint": sp_cls,
            "gcs": gcs_cls,
            "msgraph": graph_cls,
            "resolve_env": resolve_env,
            "resolve_date": resolve_date,
        }


@pytest.fixture
def task(mock_deps):
    instance = ExportOutputResultTask()
    instance.sharepoint = {
        "verint": {
            "master_output_file": "Reports/master.xlsx",
            "daily_output_file": "Reports/daily.xlsx",
            "input_folder_list_inbound": "",
            "input_folder_list_outbound": "",
        },
        "sandbox": {
            "daily_output_file": "Reports/network_daily.xlsx",
        },
        "control": {
            "weight_score_file": "Control/weights.xlsx",
            "batch_processing_log_file": "Control/batch_log.csv",
            "transaction_log_file": "Control/transaction.csv",
        },
    }
    instance.framework = {"weight_score_file": "fallback_weights.xlsx"}
    instance.gcs = {"project_id": "qa-proj", "bucket_name": "qa-bucket"}
    instance.gcp = {"project_id": "qa-proj", "project_name": "QA Project"}
    instance.pre_result = {"batch_results": [], "list_batchs": [], "failed_batches": []}
    instance.packages = {"execution_dt": datetime(2025, 1, 2, 3, 4, 5)}
    instance.get_package = lambda key, default=None: instance.packages.get(key, default)
    instance.sharepoint_verint = Mock()
    instance.sharepoint_control = Mock()
    instance.sharepoint_sandbox = Mock()
    instance.gcs_module = Mock()
    instance.msgraph_module = Mock()
    # Initialize combined_folder_list for tests
    instance.combined_folder_list = []
    return instance


def _qa_record(*, prediction_status: str = "SUCCESS", network_problem_statement=None):
    return {
        "file_metadata": {
            "file_name": "123_0891111111_120000_A001_jane_doe_D_20250101_123_OUT",
            "file_ext": ".wav",
            "record_date": "20250101",
            "duration": 123,
            "call_id": "123",
            "phone_number": "0891111111",
            "agent_id": "A001",
            "first_name": "Jane",
            "last_name": "Doe",
            "provider": "D",
            "call_direction": "OUT",
            "file_uri": "gs://qa-bucket/sentiment_qa/processing/voice/202501/20250101/Complain/123_0891111111_120000_A001_jane_doe_D_20250101_123_OUT.wav",
        },
        "prediction": {
            "status": prediction_status,
            "message": "Original failure\nmessage" if prediction_status != "SUCCESS" else None,
            "model_version": "gemini-2.5-flash",
            "token_input": {"text": 100, "audio": 20},
            "token_cached": 10,
            "token_output": {"text": 30},
            "raw_prediction": {
                "service_number": "1678",
                "call_type": "Complaint",
                "call_type_confident": "0.98",
                "customer_insight": {
                    "summary_story": "Customer called about signal",
                    "product_category": "Mobile",
                    "repeat_call": "No",
                    "fcr": "True",
                    "churn_probability": "Low",
                    "churn_reason": "None",
                    "customer_insight_summary": "Resolved in one call",
                },
                "service_quality": {
                    "greeting_standard": {"evaluation": "Meet", "reason": "hello agent"},
                    "manners": {"evaluation": "Below", "reason": "hello agent"},
                    "enthusiasm": {"evaluation": "Meet", "reason": "hello agent"},
                    "communication_skill": {"evaluation": "Meet", "reason": "hello agent"},
                    "ending_standard": {"evaluation": "Meet", "reason": "hello agent"},
                    "data_privacy": {"evaluation": "Meet", "reason": "hello agent"},
                    "legal_verification": {"evaluation": "Meet", "reason": "hello agent"},
                    "customer_verification": {"evaluation": "Meet", "reason": "hello agent"},
                    "sla_notification": {"evaluation": "Meet", "reason": "hello agent"},
                    "transfer_standard": {"evaluation": "Meet", "reason": "hello agent"},
                    "problem_understanding": {"evaluation": "Meet", "reason": "hello agent"},
                    "compensation": {"evaluation": "Meet", "reason": "hello agent"},
                    "hold_standard": {"evaluation": "Meet", "reason": "hello agent"},
                    "wrap_up": {"evaluation": "Meet", "reason": "hello agent"},
                    "beyond_scope_support": {"evaluation": "Meet", "reason": "hello agent"},
                    "true_application": {"evaluation": "Meet", "reason": "hello agent"},
                    "case_ownership": {"evaluation": "Meet", "reason": "hello agent"},
                    "contact_confirm": {"evaluation": "Meet", "reason": "hello agent"},
                    "omotenashi": {"evaluation": "Meet", "reason": "hello agent"},
                    "retention": {"evaluation": "Meet", "reason": "hello agent"},
                    "downsell": {"evaluation": "N/A", "reason": "hello agent"},
                    "mnp": {"evaluation": "N/A", "reason": "hello agent"},
                    "upselling": {"evaluation": "Meet", "reason": "hello agent"},
                    "service_quality_performance_insight": "Good overall quality",
                },
                "sale_opportunity": {
                    "opportunity_recognition_in_conversation": "True",
                    "product_suggested_by_ai": "Power Pack",
                    "agent_offer_product_presentation_&_explanation": "False",
                    "product_offer_by_agent": "Handset insurance",
                    "sales_outcome_&_customer_decision": "True",
                    "sales_opportunities_performance_insight": "Moderate opportunity",
                },
                "customer_sentiment": {
                    "overall_sentiment": "Neutral",
                    "initial_sentiment": "Negative",
                    "final_sentiment": "Positive",
                    "primary_sentiment_driver": "Issue resolved",
                    "csat": "4",
                    "cs_performance_insight": "Customer ended satisfied",
                },
                "customer_experience": {
                    "agent_communication_&_attitude": "Meet",
                    "agent_communication_&_attitude_reason": "Polite",
                    "agent_understanding_&_resolution": "Meet",
                    "agent_understanding_&_resolution_reason": "Understood quickly",
                    "agent_responsiveness": "Meet",
                    "agent_responsiveness_reason": "Fast responses",
                    "system_accessibility": "True",
                    "system_accessibility_reason": "System available",
                    "ivr_usability_&_design": "False",
                    "ivr_usability_&_design_reason": "Confusing menu",
                    "ces": "3",
                    "self_service_readiness": "Medium",
                    "cx_performance_insight": "Average CX",
                },
                "network": {
                    "issue_type": "Signal",
                    "problem_statement": network_problem_statement
                    if network_problem_statement is not None
                    else ["Signal drop", "Slow internet"],
                    "area_tag_province": "Bangkok",
                    "area_tag_district": "Chatuchak",
                    "area_tag_sub_district": "Lat Yao",
                    "area_tag_landmark": "Office tower",
                },
            },
        },
    }


def test_pre_execute_initializes_modules(mock_deps):
    instance = ExportOutputResultTask()
    instance.sharepoint_verint = None
    instance.sharepoint_control = None
    instance.gcs_module = None
    instance.msgraph_module = None
    instance.gcs = {"project_id": "qa-proj", "bucket_name": "qa-bucket"}

    instance.pre_execute()

    assert mock_deps["sharepoint"].call_count == 3
    mock_deps["gcs"].assert_called_once_with(project_id="qa-proj", bucket_name="qa-bucket")
    mock_deps["msgraph"].assert_called_once()


def test_execute_task_empty_returns_none(task):
    task.pre_result["batch_results"] = []

    assert task.execute_task() is None


def test_execute_task_happy_path(task):
    task.pre_result = {
        "batch_results": [
            {
                "file_metadata": {
                    "file_name": "file1",
                    "file_ext": ".wav",
                    "record_date": "20250101",
                    "file_uri": "gs://qa-bucket/sentiment_qa/processing/voice/202501/20250101/Support/file1.wav",
                    "duration": 60,
                },
                "prediction": {
                    "status": "SUCCESS",
                    "token_input": {"text": 1},
                    "token_cached": 0,
                    "token_output": {"text": 1},
                    "model_version": "gemini-2.5-flash",
                },
                "load_dt": "2025-01-02 03:04:05",
            },
            {
                "file_metadata": {
                    "file_name": "file2",
                    "file_ext": ".wav",
                    "record_date": "20250102",
                    "file_uri": "gs://qa-bucket/sentiment_qa/processing/voice/202501/20250102/Billing/file2.wav",
                    "duration": 65,
                },
                "prediction": {
                    "status": "FAILED",
                    "message": "bad file",
                    "token_input": {"text": 2},
                    "token_cached": 0,
                    "token_output": {"text": 1},
                    "model_version": "gemini-2.5-flash",
                },
                "load_dt": "2025-01-02 03:04:05",
            },
        ],
        "list_batchs": ["output/batch1.jsonl"],
        "failed_batches": [],
    }

    formatted_rows = [
        {
            "filename": "file1.wav",
            "call_month": "202501",
            "call_date": "20250101",
            "agent_id": "A001",
            "call_id": "1",
            "phone_number": "0891",
            "department": "Support",
        },
        {
            "filename": "file2.wav",
            "call_month": "202501",
            "call_date": "20250102",
            "agent_id": "A002",
            "call_id": "2",
            "phone_number": "0892",
            "department": "Billing",
        },
    ]

    with (
        patch.object(task, "_format_output", return_value=formatted_rows),
        patch.object(task, "_upload_master_file", side_effect=lambda df: df.copy()) as upload_master,
        patch.object(task, "_upload_daily_files") as upload_daily,
        patch.object(task, "_archive_files") as archive_files,
        patch.object(task, "_insert_log_record") as insert_log,
    ):
        batch_results, result_df = task.execute_task()

    assert batch_results == task.pre_result["batch_results"]
    assert len(result_df) == 2
    upload_master.assert_called_once()
    assert upload_daily.call_count == 4
    archive_files.assert_called_once()
    insert_log.assert_called_once()


def test_format_output_realistic_payload_and_failed_log(task):
    weight_item = SimpleNamespace(content=_mock_config_excel())
    log_csv = (
        "batch_job_id,batch_job_display_name,filename,status,error_message,updated_dt\n"
        "job-1,QA batch,123_0891111111_120000_A001_jane_doe_D_20250101_123_OUT.wav,SUCCESS,,2025-01-02 01:00:00\n"
        "job-1,QA batch,missing_audio.wav,FAILED,Audio decode error\ncontinued,2025-01-02 01:01:00\n"
    )
    log_item = SimpleNamespace(content=log_csv.encode("utf-8"))

    def get_item_side(item_path):
        if item_path == "Control/weights.xlsx":
            return weight_item
        if item_path == "Control/batch_log.csv":
            return log_item
        raise AssertionError(f"Unexpected path: {item_path}")

    task.sharepoint_control.get_item_by_path.side_effect = get_item_side
    task.sharepoint_control.is_item_exists.side_effect = lambda item_path: item_path == "Control/batch_log.csv"

    def resolve_env_side_effect(value):
        if not value or value == "None":
            return "Control/weights.xlsx"
        return value

    if hasattr(task, "config_path"):
        task.config_path = "Control/weights.xlsx"
    if hasattr(task, "user_config_path"):
        task.user_config_path = "Control/weights.xlsx"

    with (
        patch("tasks.sentiment_qa.export_output_result_task.resolve_env", side_effect=resolve_env_side_effect),
        patch(
            "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
            return_value=datetime(2025, 1, 2, 3, 4, 5),
        ),
    ):
        result = task._format_output([_qa_record()])

    assert len(result) == 2
    formatted = result[0]
    failed_from_log = result[1]

    assert formatted["filename"] == "123_0891111111_120000_A001_jane_doe_D_20250101_123_OUT.wav"
    assert (
        formatted["full_path"]
        == "Input/Complain/202501/20250101/123_0891111111_120000_A001_jane_doe_D_20250101_123_OUT.wav"
    )
    assert formatted["true_dtac"] == "Dtac"
    assert formatted["fcr"] == "Y"
    assert formatted["system_accessibility"] == "Y"
    assert formatted["ivr_usability_&_design"] == "N"
    assert formatted["opportunity_recognition_in_conversation"] == "Y"
    assert formatted["agent_offer_product_presentation_&_explanation"] == "N"
    assert formatted["sales_outcome_&_customer_decision"] == "Y"
    assert formatted["problem_statement"] == "Signal drop, Slow internet"
    assert formatted["service_quality_score"] is not None
    assert formatted["status"] == "SUCCESS"

    assert failed_from_log["filename"] == "missing_audio.wav"
    assert failed_from_log["status"] == "FAILED"
    assert failed_from_log["error_code"] == "Audio decode error"


def test_format_output_broken_record_returns_failed_row(task):
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(content=_mock_config_excel())
    task.sharepoint_control.is_item_exists.return_value = False

    with patch(
        "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
        return_value=datetime(2025, 1, 2, 3, 4, 5),
    ):
        result = task._format_output([_qa_record(network_problem_statement=None)])

    # success branch uses default problem list; explicit None forces join failure
    assert len(result) == 1
    assert result[0]["status"] == "SUCCESS"

    with patch(
        "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
        return_value=datetime(2025, 1, 2, 3, 4, 5),
    ):
        broken = task._format_output([_qa_record(network_problem_statement=123)])

    assert len(broken) == 1
    assert broken[0]["status"] == "FAILED"
    assert "join" in broken[0]["error_code"]


@pytest.mark.parametrize(
    ("sub_categories", "expected"),
    [(["Meet", "N/A"], "Meet"), (["Meet", "Below"], "Below"), (["N/A"], "N/A")],
)
def test_calculate_category(task, sub_categories, expected):
    assert task._calculate_category(sub_categories) == expected


def test_sync_excel_schema_rebuilds_headers(task):
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A1:B1")
    ws["A1"] = "Old Group"
    ws["A2"] = "alpha"
    ws["B2"] = "beta"
    ws["A3"] = "row-a"
    ws["B3"] = "row-b"

    group_headers = [("Group One", ["alpha", "beta", "gamma"]), ("Group Two", ["delta"])]
    all_columns = ["alpha", "beta", "gamma", "delta"]

    task._sync_excel_schema(ws, group_headers, all_columns)

    assert [ws.cell(2, col).value for col in range(1, 5)] == all_columns
    assert ws.cell(3, 1).value == "row-a"
    assert ws.cell(3, 2).value == "row-b"
    assert ws.cell(3, 3).value == ""
    assert ws.cell(3, 4).value == ""
    assert {str(rng) for rng in ws.merged_cells.ranges} == {"A1:C1", "D1"}


def test_upload_master_file_updates_existing_workbook(task):
    task.sharepoint_verint.get_item_by_path.return_value = SimpleNamespace(content=_legacy_master_workbook())
    master_df = pd.DataFrame(
        [
            {
                "filename": "new.wav",
                "agent_id": "A100",
                "call_id": "C100",
                "phone_number": "0811111111",
                "department": "Support",
                "call_date": "20250101",
                "call_month": "202501",
            }
        ]
    )

    combined = task._upload_master_file(master_df)

    assert len(combined) == 2
    upload_kwargs = task.sharepoint_verint.upload_file.call_args.kwargs
    uploaded_wb = load_workbook(io.BytesIO(upload_kwargs["content"]))
    ws = uploaded_wb.active

    assert ws.cell(2, 1).value == "filename"
    assert ws.cell(3, 1).value == "existing.wav"
    assert ws.cell(4, 1).value == "new.wav"
    assert upload_kwargs["upload_path"] == "Reports/master.xlsx"


def test_upload_master_file_creates_new_workbook_on_404(task):
    not_found = Exception("missing")
    not_found.response = SimpleNamespace(status_code=404)
    task.sharepoint_verint.get_item_by_path.side_effect = not_found

    master_df = pd.DataFrame(
        [
            {
                "filename": "fresh.wav",
                "agent_id": "A200",
                "call_id": "C200",
                "phone_number": "0822222222",
                "department": "Billing",
                "call_date": "20250102",
                "call_month": "202501",
            }
        ]
    )

    combined = task._upload_master_file(master_df)

    assert len(combined) == 1
    uploaded_wb = load_workbook(io.BytesIO(task.sharepoint_verint.upload_file.call_args.kwargs["content"]))
    ws = uploaded_wb.active
    assert ws.cell(1, 1).value == "General"
    assert ws.cell(2, 1).value == "filename"
    assert ws.cell(3, 1).value == "fresh.wav"


def test_upload_daily_files_creates_schema_workbook(task):
    daily_df = pd.DataFrame(
        [
            {
                "call_date": "20250101",
                "agent_id": "A001",
                "call_id": "C001",
                "phone_number": "0891111111",
                "service_number": "1678",
                "call_duration_sec": 123,
                "full_path": "Input/Complain/202501/20250101/test.wav",
                "department": "Complain",
                "filename": "test.wav",
                "call_direction": "OUTBOUND",
                "call_type": "Complaint",
                "call_type_confident": "0.98",
            }
        ]
    )

    config_item = SimpleNamespace(content=_mock_config_excel())

    task.sharepoint_control.get_item_by_path.return_value = config_item

    def resolve_env_side_effect(value):
        if not value or value == "None":
            return "Control/weights.xlsx"
        return value

    if hasattr(task, "config_path"):
        task.config_path = "Control/weights.xlsx"
    if hasattr(task, "user_config_path"):
        task.user_config_path = "Control/weights.xlsx"

    with patch("tasks.sentiment_qa.export_output_result_task.resolve_env", side_effect=resolve_env_side_effect):
        task._upload_daily_files(daily_df, "Reports/daily.xlsx", task.sharepoint_verint)

    uploaded_wb = load_workbook(io.BytesIO(task.sharepoint_verint.upload_file.call_args.kwargs["content"]))
    ws = uploaded_wb.active
    headers = [ws.cell(2, col).value for col in range(1, 13)]

    assert headers[:4] == ["call_date", "agent_id", "call_id", "phone_number"]
    assert ws.cell(3, 1).value == "20250101"
    assert ws.cell(3, 9).value == "test.wav"
    assert task.sharepoint_verint.upload_file.call_args.kwargs["upload_path"] == "Reports/daily.xlsx"


def test_post_execute_skips_without_process_date(task):
    task._cache_oper_log = {"transaction_df": pd.DataFrame(), "process_date": []}

    assert task.post_execute("sentinel") == "sentinel"


def test_post_execute_stamps_ai_operation_logs(task, monkeypatch):
    task._cache_oper_log = {
        "process_date": [date(2025, 1, 1)],
        "transaction_df": pd.DataFrame(
            [
                {
                    "start_time": "2025-01-01T00:00:00Z",
                    "end_time": "2025-01-01T00:00:05Z",
                    "gcp_project_id": "qa-proj",
                    "status_pass_failed_retry": "Pass",
                    "latency_ms": 5000,
                },
                {
                    "start_time": "2025-01-01T00:01:00Z",
                    "end_time": "2025-01-01T00:01:02Z",
                    "gcp_project_id": "qa-proj",
                    "status_pass_failed_retry": "Failed",
                    "latency_ms": 2000,
                },
            ]
        ),
    }
    monkeypatch.setenv("ENVIRONMENT", "nprd")

    with patch("tasks.sentiment_qa.export_output_result_task.logging_ai_operation") as log_ai_operation:
        result = task.post_execute(["done"])

    assert result == ["done"]
    log_ai_operation.assert_called_once()
    log_obj = log_ai_operation.call_args.kwargs["log_obj"]
    assert log_obj["environment"] == "non-production"
    assert log_obj["total_transaction"] == 2
    assert log_obj["total_success_transaction"] == 1
    assert log_obj["total_failed_transaction"] == 1


def _export_prediction_df():
    return pd.DataFrame(
        [
            {
                "file_name": "call-1.wav",
                "full_path": "Input/Support/202501/20250101/call-1.wav",
                "folder": "Support/20250101",
                "record_date": "20250101",
                "duration": 60,
                "token_input": {"text": 100, "audio": 20},
                "token_cached": 10,
                "token_output": {"text": 30},
                "status": "SUCCESS",
                "message": "",
                "processed_time": "2025-01-01T00:01:00Z",
                "create_time": "2025-01-01T00:00:00Z",
                "model_version": "gemini-2.5-flash",
                "load_dt": "2025-01-02 03:04:05",
            },
            {
                "file_name": "call-2.wav",
                "full_path": "Input/Billing/202501/20250101/call-2.wav",
                "folder": "Billing/20250101",
                "record_date": "20250101",
                "duration": 120,
                "token_input": {"text": 80, "audio": 10},
                "token_cached": 0,
                "token_output": {"text": 40},
                "status": "FAILED",
                "message": "decoder error",
                "processed_time": "2025-01-01T00:05:00Z",
                "create_time": "2025-01-01T00:04:00Z",
                "model_version": "gemini-2.5-flash",
                "load_dt": "2025-01-02 03:04:05",
            },
        ]
    )


def test_archive_files_moves_voice_and_batch_outputs(task):
    task.combined_folder_list = ["Support", "Billing"]
    task.sharepoint["verint"]["input_folder_list"] = "Support,Billing"
    task.gcs.update(
        {
            "processing_voice_folder": "processing/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_voice_folder": "archive/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "input_folder": "input/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_batch_folder": "archive_batch/%{DATA_DATE_YYYYMMDD}",
        }
    )
    df = pd.DataFrame(
        [
            {"record_date": "20250101", "status": "SUCCESS", "department": "Support", "file_name": "call-1.wav"},
            {"record_date": "20250101", "status": "SUCCESS", "department": "Billing", "file_name": "call-2.wav"},
            {"record_date": "20250101", "status": "FAILED", "department": "Support", "file_name": "call-3.wav"},
        ]
    )

    def dir_exists(dir_path):
        return dir_path in {
            "processing/Support/%{DATA_DATE_YYYYMMDD}",
            "processing/Billing/%{DATA_DATE_YYYYMMDD}",
            "input/Support/%{DATA_DATE_YYYYMMDD}",
            "input/Billing/%{DATA_DATE_YYYYMMDD}",
            "output/run1",
        }

    def list_files(prefix):
        mapping = {
            "processing/Support/%{DATA_DATE_YYYYMMDD}": ["processing/Support/%{DATA_DATE_YYYYMMDD}/call-1.wav"],
            "processing/Billing/%{DATA_DATE_YYYYMMDD}": ["processing/Billing/%{DATA_DATE_YYYYMMDD}/call-2.wav"],
            "input/Support/%{DATA_DATE_YYYYMMDD}": ["input/Support/%{DATA_DATE_YYYYMMDD}/call-1.wav"],
            "input/Billing/%{DATA_DATE_YYYYMMDD}": ["input/Billing/%{DATA_DATE_YYYYMMDD}/call-2.wav"],
        }
        return mapping[prefix]

    task.gcs_module.is_dir_exists.side_effect = dir_exists
    task.gcs_module.list_files.side_effect = list_files

    task._archive_files(
        datetime(2025, 1, 2, 3, 4, 5),
        df,
        ["output/run1/batch1.jsonl", "output/run2/batch2.jsonl"],
        ["output/run2/batch2.jsonl"],
    )

    moved_paths = [
        (call.kwargs["source_path"], call.kwargs["destination_path"])
        for call in task.gcs_module.move_file.call_args_list
    ]
    assert (
        "processing/Support/%{DATA_DATE_YYYYMMDD}/call-1.wav",
        "archive/Support/%{DATA_DATE_YYYYMMDD}/call-1.wav",
    ) in moved_paths
    assert (
        "processing/Billing/%{DATA_DATE_YYYYMMDD}/call-2.wav",
        "archive/Billing/%{DATA_DATE_YYYYMMDD}/call-2.wav",
    ) in moved_paths
    assert ("output/run1/batch1.jsonl", "archive_batch/%{DATA_DATE_YYYYMMDD}/run1/batch1.jsonl") in moved_paths

    deleted_files = [call.kwargs["file_path"] for call in task.gcs_module.delete_file.call_args_list]
    assert "input/Support/%{DATA_DATE_YYYYMMDD}/call-1.wav" in deleted_files
    assert "input/Billing/%{DATA_DATE_YYYYMMDD}/call-2.wav" in deleted_files
    task.gcs_module.delete_dir.assert_called_once_with(dir_path="output/run1")


def test_transaction_log_creates_upload_and_updates_cache(task):
    task.sharepoint["control"]["transaction_log_file"] = "Control/transaction_%{DATA_DATE_YYYYMMDD}.csv"
    task.verint_access["site_name"] = "verint-site"
    task.verint_site = "verint.sharepoint.com"
    prediction_df = _export_prediction_df()
    existing_df = pd.DataFrame(
        [
            {
                "data_date": "20250101",
                "start_time": "2025-01-01 00:00:00+00:00",
                "end_time": "2025-01-01 00:00:10+00:00",
                "type": "AI Classification",
                "updated_dt": "2025-01-01 01:00:00",
                "load_dt": "2025-01-01 01:00:00",
                "filename": "existing.wav",
                "status_pass_failed_retry": "Pass",
            }
        ]
    )
    task.sharepoint_control.is_item_exists.side_effect = lambda item_path: (
        item_path == "Control/transaction_%{DATA_DATE_YYYYMMDD}.csv"
    )
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(
        content=existing_df.to_csv(index=False).encode("utf-8")
    )

    with (
        patch("tasks.sentiment_qa.export_output_result_task.gemini_cost", return_value={}),
        patch(
            "tasks.sentiment_qa.export_output_result_task.GeminiBatchModule.cal_gemini_cost",
            side_effect=lambda usage_detail, cost_config: {
                name: {"cost_input": 0.1, "cost_output": 0.2} for name in usage_detail
            },
        ),
    ):
        new_df = task._transaction_log("AI Classification", "daisyrpa", "SharePoint", prediction_df)

    assert len(new_df) == 2
    upload_kwargs = task.sharepoint_control.upload_file.call_args.kwargs
    uploaded_df = pd.read_csv(io.BytesIO(upload_kwargs["content"]))
    assert upload_kwargs["upload_path"] == "Control/transaction_%{DATA_DATE_YYYYMMDD}.csv"
    assert set(uploaded_df["filename"]) >= {"existing.wav", "call-1.wav", "call-2.wav"}
    assert not task._cache_oper_log["transaction_df"].empty
    assert task._cache_oper_log["process_date"]


def test_performance_log_aggregates_and_uploads(task):
    task.sharepoint["control"]["performance_log_file"] = "Control/performance_%{DATA_DATE_YYYYMMDD}.csv"
    transaction_log_df = pd.DataFrame(
        [
            {
                "data_date": "20250101",
                "start_time": "2025-01-01 00:00:00+00:00",
                "load_dt": "2025-01-02 03:04:05",
                "gcp_project_id": "qa-proj",
                "gcp_project_name": "QA Project",
                "status_pass_failed_retry": "Pass",
                "latency_ms": 5000,
            },
            {
                "data_date": "20250101",
                "start_time": "2025-01-01 00:05:00+00:00",
                "load_dt": "2025-01-02 03:04:05",
                "gcp_project_id": "qa-proj",
                "gcp_project_name": "QA Project",
                "status_pass_failed_retry": "Failed",
                "latency_ms": 2000,
            },
        ]
    )
    task.sharepoint_control.is_item_exists.side_effect = lambda item_path: (
        item_path == "Control/performance_%{DATA_DATE_YYYYMMDD}.csv"
    )
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(
        content=pd.DataFrame([{"data_date": "20241231", "load_dt": "20241231", "updated_dt": "2024-12-31 00:00:00"}])
        .to_csv(index=False)
        .encode("utf-8")
    )

    task._performance_log(transaction_log_df)

    upload_kwargs = task.sharepoint_control.upload_file.call_args.kwargs
    uploaded_df = pd.read_csv(io.BytesIO(upload_kwargs["content"]))
    assert upload_kwargs["upload_path"] == "Control/performance_%{DATA_DATE_YYYYMMDD}.csv"
    assert "success_rate" in uploaded_df.columns
    assert (uploaded_df["gcp_project_id"] == "qa-proj").any()


def test_insert_log_record_calls_transaction_and_performance(task):
    df = pd.DataFrame([{"record_date": "20250101", "file_name": "x.wav"}])
    transaction_df = pd.DataFrame([{"data_date": "20250101"}])

    with (
        patch.object(task, "_transaction_log", return_value=transaction_df) as transaction_log,
        patch.object(task, "_performance_log") as performance_log,
    ):
        task._insert_log_record(df)

    transaction_log.assert_called_once()
    performance_log.assert_called_once_with(transaction_log_df=transaction_df)


def test_insert_log_record_skips_performance_when_transaction_empty(task):
    df = pd.DataFrame([{"record_date": "20250101", "file_name": "x.wav"}])

    with (
        patch.object(task, "_transaction_log", return_value=pd.DataFrame()) as transaction_log,
        patch.object(task, "_performance_log") as performance_log,
    ):
        task._insert_log_record(df)

    transaction_log.assert_called_once()
    performance_log.assert_not_called()


def test_on_error_sends_email(task):
    task.task_name = "QAExportOutputResultTask"
    with patch(
        "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
        return_value=datetime(2025, 1, 2, 3, 4, 5),
    ):
        task.on_error(RuntimeError("boom"))

    task.msgraph_module.send_email.assert_called_once()


def test_post_execute_skips_when_transaction_df_empty(task):
    task._cache_oper_log = {"process_date": [date(2025, 1, 1)], "transaction_df": pd.DataFrame()}

    assert task.post_execute("ok") == "ok"


def test_pre_execute_raises_for_dependency_failures(task, mock_deps):
    scenarios = [
        (Exception("verint boom"), None, None, None, None, "verint boom"),
        (Mock(), Exception("control boom"), None, None, None, "control boom"),
        (Mock(), Mock(), Exception("sandbox boom"), None, None, "sandbox boom"),
        (Mock(), Mock(), Mock(), Exception("gcs boom"), None, "gcs boom"),
        (Mock(), Mock(), Mock(), None, Exception("graph boom"), "graph boom"),
    ]

    for sp_verint, sp_control, sp_sandbox, gcs_side, graph_side, expected in scenarios:
        mock_deps["sharepoint"].reset_mock(side_effect=True)
        mock_deps["gcs"].reset_mock(side_effect=True)
        mock_deps["msgraph"].reset_mock(side_effect=True)

        mock_deps["sharepoint"].side_effect = [sp_verint, sp_control, sp_sandbox]
        mock_deps["gcs"].side_effect = gcs_side
        mock_deps["msgraph"].side_effect = graph_side

        with pytest.raises(Exception, match=expected):
            task.pre_execute()

    mock_deps["sharepoint"].side_effect = None
    mock_deps["gcs"].side_effect = None
    mock_deps["msgraph"].side_effect = None


def test_execute_task_wraps_archive_and_log_failures(task):
    batch_record = {
        "file_metadata": {
            "file_name": "file1",
            "file_ext": ".wav",
            "record_date": "20250101",
            "file_uri": None,
            "duration": 60,
        },
        "prediction": {"status": "SUCCESS", "model_version": "gemini-2.5-flash"},
        "load_dt": "2025-01-02 03:04:05",
    }
    task.pre_result = {"batch_results": [batch_record], "list_batchs": ["batch1"], "failed_batches": []}
    formatted_rows = [
        {
            "filename": "file1.wav",
            "call_month": "202501",
            "call_date": "20250101",
            "agent_id": "A1",
            "call_id": "1",
            "phone_number": "0891",
            "department": "Support",
        }
    ]

    with (
        patch.object(task, "_format_output", return_value=formatted_rows),
        patch.object(task, "_upload_master_file", side_effect=lambda df: df.copy()),
        patch.object(task, "_upload_daily_files"),
        patch.object(task, "_archive_files", side_effect=RuntimeError("archive broke")),
        patch.object(task, "_insert_log_record"),
        pytest.raises(Exception, match="File archival process failed"),
    ):
        task.execute_task()

    with (
        patch.object(task, "_format_output", return_value=formatted_rows),
        patch.object(task, "_upload_master_file", side_effect=lambda df: df.copy()),
        patch.object(task, "_upload_daily_files"),
        patch.object(task, "_archive_files"),
        patch.object(task, "_insert_log_record", side_effect=RuntimeError("log broke")),
        pytest.raises(Exception, match="Transaction log insertion failed"),
    ):
        task.execute_task()


def test_format_output_fallback_weights_and_non_boolean_values(task):
    fallback_weights_df = pd.DataFrame([{"item": "greeting_standard", "inbound_weight": 10, "outbound_weight": 10}])
    fallback_mapping_df = pd.DataFrame([{"product_group": "TOL", "product_category": "CCTV"}])
    fallback_quality_df = pd.DataFrame([{"sub_category": "communication", "item": "greeting_standard"}])
    fallback_gsd_df = pd.DataFrame([{"standard_gsd_name": "Account Login", "definition": "Test"}])
    record = _qa_record(prediction_status="FAILED")
    record["file_metadata"]["duration"] = None
    record["prediction"]["raw_prediction"]["customer_insight"]["fcr"] = "maybe"
    record["prediction"]["raw_prediction"]["sale_opportunity"]["opportunity_recognition_in_conversation"] = "maybe"
    record["prediction"]["raw_prediction"]["sale_opportunity"]["agent_offer_product_presentation_&_explanation"] = (
        "maybe"
    )
    record["prediction"]["raw_prediction"]["sale_opportunity"]["sales_outcome_&_customer_decision"] = "maybe"
    record["prediction"]["raw_prediction"]["customer_experience"]["system_accessibility"] = "maybe"
    record["prediction"]["raw_prediction"]["customer_experience"]["ivr_usability_&_design"] = "maybe"
    task.sharepoint_control.get_item_by_path.side_effect = Exception("missing sharepoint weights")
    task.sharepoint_control.is_item_exists.return_value = False

    def read_xlsx_side_effect(file_path, sheet_name=None, **kwargs):
        if sheet_name == "product_mapping":
            return fallback_mapping_df
        if sheet_name == "service_quality_group":
            return fallback_quality_df
        if sheet_name == "standard_gsd":
            return fallback_gsd_df
        return fallback_weights_df

    with (
        patch("tasks.sentiment_qa.export_output_result_task.read_xlsx", side_effect=read_xlsx_side_effect),
        patch(
            "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
            return_value=datetime(2025, 1, 2, 3, 4, 5),
        ),
    ):
        result = task._format_output([record])

    assert result[0]["status"] == "FAILED"
    assert result[0]["error_code"] == "Original failure message"
    assert result[0]["fcr"] == "error"
    assert result[0]["opportunity_recognition_in_conversation"] == "error"
    assert result[0]["agent_offer_product_presentation_&_explanation"] == "error"
    assert result[0]["sales_outcome_&_customer_decision"] == "error"
    assert result[0]["system_accessibility"] == "error"
    assert result[0]["ivr_usability_&_design"] == "error"
    assert result[0]["call_duration_sec"] is None


def test_format_output_handles_failed_log_append_error(task):
    task.sharepoint_control.get_item_by_path.side_effect = [
        SimpleNamespace(content=_mock_config_excel()),
        SimpleNamespace(content=_mock_config_excel()),
        SimpleNamespace(content=b"ignored"),
    ]
    task.sharepoint_control.is_item_exists.side_effect = lambda item_path: item_path == "Control/batch_log.csv"
    broken_log_df = pd.DataFrame(
        [
            {
                "batch_job_id": "job-1",
                "batch_job_display_name": "QA batch",
                "filename": "123_0891111111_120000_A001_jane_doe_D_20250101_123_OUT.wav",
                "status": "SUCCESS",
                "error_message": "ok",
                "updated_dt": "2025-01-02 01:00:00",
            },
            {
                "batch_job_id": "job-1",
                "batch_job_display_name": "QA batch",
                "filename": "bad_failed.wav",
                "status": "FAILED",
                "error_message": object(),
                "updated_dt": "2025-01-02 01:01:00",
            },
        ]
    )

    with (
        patch("tasks.sentiment_qa.export_output_result_task.pd.read_csv", return_value=broken_log_df),
        patch(
            "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
            return_value=datetime(2025, 1, 2, 3, 4, 5),
        ),
    ):
        result = task._format_output([_qa_record()])

    failed_rows = [row for row in result if row["status"] == "FAILED"]
    assert any("Unknown error during log append" in row.get("error_code", "") for row in failed_rows)


def test_transaction_log_without_existing_file_and_missing_full_path(task):
    task.sharepoint["control"]["transaction_log_file"] = "Control/transaction_%{DATA_DATE_YYYYMMDD}.csv"
    task.verint_access["site_name"] = "verint-site"
    task.verint_site = "verint.sharepoint.com"
    df = _export_prediction_df().iloc[[0]].copy()
    df.loc[df.index[0], "full_path"] = ""
    task.sharepoint_control.is_item_exists.return_value = False

    with (
        patch("tasks.sentiment_qa.export_output_result_task.gemini_cost", return_value={}),
        patch(
            "tasks.sentiment_qa.export_output_result_task.GeminiBatchModule.cal_gemini_cost",
            return_value={"call-1.wav": {"cost_input": 0.1, "cost_output": 0.2}},
        ),
    ):
        new_df = task._transaction_log("AI Classification", "daisyrpa", "SharePoint", df)

    assert len(new_df) == 1
    assert (
        task.sharepoint_control.upload_file.call_args.kwargs["upload_path"]
        == "Control/transaction_%{DATA_DATE_YYYYMMDD}.csv"
    )


def test_transaction_log_missing_columns_raises(task):
    bad_df = pd.DataFrame([{"file_name": "x", "model_version": "gemini-2.5-flash"}])
    with patch("tasks.sentiment_qa.export_output_result_task.gemini_cost", return_value={}):
        with pytest.raises(ValueError, match="missing columns"):
            task._transaction_log("AI Classification", "daisyrpa", "SharePoint", bad_df)


def test_performance_log_skips_invalid_payloads_and_returns_none(task):
    task.sharepoint["control"]["performance_log_file"] = "Control/performance_%{DATA_DATE_YYYYMMDD}.csv"
    transaction_log_df = pd.DataFrame(
        [
            {
                "data_date": "20250101",
                "start_time": "bad-date",
                "load_dt": "bad-date",
                "gcp_project_id": "qa-proj",
                "gcp_project_name": "QA Project",
                "status_pass_failed_retry": "Pass",
                "latency_ms": 5000,
            }
        ]
    )

    with patch(
        "tasks.sentiment_qa.export_output_result_task.PerformanceLogSchema.from_dict",
        side_effect=Exception("invalid payload"),
    ):
        assert task._performance_log(transaction_log_df) is None


def test_upload_master_raise_on_invalid_paths(task):
    with patch("tasks.sentiment_qa.export_output_result_task.resolve_date", side_effect=RuntimeError("bad path")):
        with pytest.raises(Exception, match="Cannot determine output file path"):
            task._upload_master_file(pd.DataFrame([{"call_date": "20250101"}]))


def test_upload_master_file_raises_for_non_404_error(task):
    task.sharepoint_verint.get_item_by_path.side_effect = RuntimeError("broken workbook")
    with pytest.raises(Exception, match="Cannot process existing master file"):
        task._upload_master_file(pd.DataFrame([{"filename": "x.wav", "call_date": "20250101"}]))


def test_on_error_handles_email_failure(task):
    task.task_name = "QAExportOutputResultTask"
    task.msgraph_module.send_email.side_effect = RuntimeError("mail down")
    with patch(
        "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
        return_value=datetime(2025, 1, 2, 3, 4, 5),
    ):
        task.on_error(RuntimeError("boom"))

    task.msgraph_module.send_email.assert_called_once()


def test_post_execute_uses_unknown_environment_and_handles_logging_error(task, monkeypatch):
    task._cache_oper_log = {
        "process_date": [date(2025, 1, 2)],
        "transaction_df": pd.DataFrame(
            [
                {
                    "start_time": "2025-01-02T00:00:00Z",
                    "end_time": "2025-01-02T00:00:02Z",
                    "gcp_project_id": "qa-proj",
                    "status_pass_failed_retry": "Pass",
                    "latency_ms": 2000,
                }
            ]
        ),
    }
    monkeypatch.setenv("ENVIRONMENT", "")
    with patch(
        "tasks.sentiment_qa.export_output_result_task.logging_ai_operation",
        side_effect=RuntimeError("log sink down"),
    ):
        assert task.post_execute("ok") == "ok"


def test_archive_files_handles_missing_voice_dirs_and_batch_archive_failure(task):
    task.sharepoint["verint"]["input_folder_list"] = "Support"
    task.gcs.update(
        {
            "processing_voice_folder": "processing/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_voice_folder": "archive/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "input_folder": "input/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_batch_folder": "archive_batch/%{DATA_DATE_YYYYMMDD}",
        }
    )
    df = pd.DataFrame(
        [{"record_date": "20250101", "status": "SUCCESS", "department": "Support", "file_name": "missing.wav"}]
    )
    task.gcs_module.is_dir_exists.side_effect = lambda dir_path: False
    task.gcs_module.move_file.side_effect = RuntimeError("batch archive failure")

    task._archive_files(datetime(2025, 1, 2, 3, 4, 5), df, ["output/run1/batch1.jsonl"], [])

    task.gcs_module.move_file.assert_called_once()
    task.gcs_module.delete_file.assert_not_called()
    task.gcs_module.delete_dir.assert_not_called()


def test_archive_files_handles_output_cleanup_delete_failure(task):
    task.sharepoint["verint"]["input_folder_list"] = "Support"
    task.gcs.update(
        {
            "processing_voice_folder": "processing/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_voice_folder": "archive/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "input_folder": "input/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_batch_folder": "archive_batch/%{DATA_DATE_YYYYMMDD}",
        }
    )
    df = pd.DataFrame(
        [{"record_date": "20250101", "status": "SUCCESS", "department": "Support", "file_name": "call-1.wav"}]
    )

    def dir_exists(dir_path):
        return dir_path in {
            "processing/Support/%{DATA_DATE_YYYYMMDD}",
            "input/Support/%{DATA_DATE_YYYYMMDD}",
            "output/run1",
        }

    def list_files(prefix):
        mapping = {
            "processing/Support/%{DATA_DATE_YYYYMMDD}": ["processing/Support/%{DATA_DATE_YYYYMMDD}/call-1.wav"],
            "input/Support/%{DATA_DATE_YYYYMMDD}": ["input/Support/%{DATA_DATE_YYYYMMDD}/call-1.wav"],
        }
        return mapping[prefix]

    task.gcs_module.is_dir_exists.side_effect = dir_exists
    task.gcs_module.list_files.side_effect = list_files

    def move_side_effect(source_path=None, destination_path=None):
        if source_path == "output/run1/batch1.jsonl":
            return
        return

    task.gcs_module.move_file.side_effect = move_side_effect
    task.gcs_module.delete_dir.side_effect = RuntimeError("cleanup failed")

    task._archive_files(datetime(2025, 1, 2, 3, 4, 5), df, ["output/run1/batch1.jsonl"], [])

    task.gcs_module.delete_dir.assert_called_once_with(dir_path="output/run1")


def test_performance_log_missing_columns_raises(task):
    with pytest.raises(ValueError, match="missing columns"):
        task._performance_log(pd.DataFrame([{"data_date": "20250101"}]))


def test_transaction_log_raises_for_missing_file_name(task):
    broken_df = _export_prediction_df().iloc[[0]].copy()
    broken_df.loc[broken_df.index[0], "file_name"] = None
    with patch("tasks.sentiment_qa.export_output_result_task.gemini_cost", return_value={}):
        with pytest.raises(Exception, match="Record 1"):
            task._transaction_log("AI Classification", "daisyrpa", "SharePoint", broken_df)


class _FakeSeries:
    def __init__(self, values, mask_key="mask"):
        self._values = list(values)
        self._mask_key = mask_key

    def unique(self):
        return list(self._values)

    def __eq__(self, _other):
        return self._mask_key


class _MaskDrivenFrame:
    def __init__(self, values, on_mask, mask_key="mask"):
        self._series = _FakeSeries(values, mask_key=mask_key)
        self._on_mask = on_mask
        self._mask_key = mask_key

    def __getitem__(self, key):
        if key == "record_date":
            return self._series
        if key == self._mask_key:
            return self._on_mask()
        raise KeyError(key)


class _BrokenDataDateFrame:
    columns = ["data_date"]

    def __len__(self):
        return 1

    def __getitem__(self, key):
        if key == "data_date":
            raise RuntimeError("bad data dates")
        raise KeyError(key)


class _BrokenUniqueDatesFrame:
    def __getitem__(self, _key):
        raise RuntimeError("date key boom")


def test_init_raises_when_common_config_load_fails():
    with patch("tasks.sentiment_qa.export_output_result_task.load_yaml", side_effect=RuntimeError("config boom")):
        with pytest.raises(RuntimeError, match="config boom"):
            ExportOutputResultTask()


def test_execute_task_groups_same_day_records_and_skips_broken_after_upload_entry(task):
    batch_results = [
        {
            "file_metadata": {
                "file_name": "file1",
                "file_ext": ".wav",
                "record_date": "20250101",
                "file_uri": "gs://qa-bucket/sentiment_qa/processing/voice/202501/20250101/Support/file1.wav",
                "duration": 60,
            },
            "prediction": {"status": "SUCCESS", "model_version": "gemini-2.5-flash"},
            "load_dt": "2025-01-02 03:04:05",
        },
        {
            "file_metadata": {
                "file_name": "file2",
                "file_ext": ".wav",
                "record_date": "20250101",
                "file_uri": "gs://qa-bucket/sentiment_qa/processing/voice/202501/20250101/Support/file2.wav",
                "duration": 65,
            },
            "prediction": {"status": "SUCCESS", "model_version": "gemini-2.5-flash"},
            "load_dt": "2025-01-02 03:04:05",
        },
        {
            "file_metadata": {
                "file_name": None,
                "file_ext": ".wav",
                "record_date": "20250101",
                "file_uri": "gs://qa-bucket/sentiment_qa/processing/voice/202501/20250101/Support/file3.wav",
                "duration": 70,
            },
            "prediction": {"status": "SUCCESS", "model_version": "gemini-2.5-flash"},
            "load_dt": "2025-01-02 03:04:05",
        },
    ]
    task.pre_result = {"batch_results": batch_results, "list_batchs": ["batch1"], "failed_batches": []}
    formatted_rows = [
        {
            "filename": "file1.wav",
            "call_month": "202501",
            "call_date": "20250101",
            "agent_id": "A1",
            "call_id": "1",
            "phone_number": "0891",
            "department": "Support",
        },
        {
            "filename": "file2.wav",
            "call_month": "202501",
            "call_date": "20250101",
            "agent_id": "A2",
            "call_id": "2",
            "phone_number": "0892",
            "department": "Support",
        },
    ]

    with (
        patch.object(task, "_format_output", return_value=formatted_rows),
        patch.object(task, "_upload_master_file", side_effect=lambda df: df.copy()),
        patch.object(task, "_upload_daily_files") as upload_daily,
        patch.object(task, "_archive_files") as archive_files,
        patch.object(task, "_insert_log_record") as insert_log,
    ):
        task.execute_task()

    assert upload_daily.call_count == 2

    first_call_path = upload_daily.call_args_list[0].args[1]
    second_call_path = upload_daily.call_args_list[1].args[1]
    assert first_call_path == "Reports/daily.xlsx"
    assert second_call_path == "Reports/network_daily.xlsx"

    archived_df = archive_files.call_args.args[1]
    assert list(archived_df["file_name"]) == ["file1.wav", "file2.wav"]
    insert_df = insert_log.call_args.args[0]
    assert list(insert_df["record_date"]) == ["20250101", "20250101"]


def test_execute_task_raises_when_daily_path_resolution_fails(task):
    task.pre_result = {
        "batch_results": [
            {
                "file_metadata": {
                    "file_name": "file1",
                    "file_ext": ".wav",
                    "record_date": "20250101",
                    "file_uri": "gs://qa-bucket/sentiment_qa/processing/voice/202501/20250101/Support/file1.wav",
                    "duration": 60,
                },
                "prediction": {"status": "SUCCESS", "model_version": "gemini-2.5-flash"},
                "load_dt": "2025-01-02 03:04:05",
            }
        ],
        "list_batchs": [],
        "failed_batches": [],
    }
    formatted_rows = [
        {
            "filename": "file1.wav",
            "call_month": "202501",
            "call_date": "20250101",
            "agent_id": "A1",
            "call_id": "1",
            "phone_number": "0891",
            "department": "Support",
        }
    ]

    with (
        patch.object(task, "_format_output", return_value=formatted_rows),
        patch.object(task, "_upload_master_file", side_effect=lambda df: df.copy()),
        patch.object(task, "_upload_daily_files"),
        patch("tasks.sentiment_qa.export_output_result_task.resolve_env", side_effect=RuntimeError("bad daily path")),
        pytest.raises(Exception, match="Cannot determine output file path"),
    ):
        task.execute_task()


def test_execute_task_raises_when_archive_dataframe_creation_fails(task):
    task.pre_result = {
        "batch_results": [
            {
                "file_metadata": {
                    "file_name": "file1",
                    "file_ext": ".wav",
                    "record_date": "20250101",
                    "file_uri": "gs://qa-bucket/sentiment_qa/processing/voice/202501/20250101/Support/file1.wav",
                    "duration": 60,
                },
                "prediction": {"status": "SUCCESS", "model_version": "gemini-2.5-flash"},
                "load_dt": "2025-01-02 03:04:05",
            }
        ],
        "list_batchs": [],
        "failed_batches": [],
    }
    formatted_rows = [
        {
            "filename": "file1.wav",
            "call_month": "202501",
            "call_date": "20250101",
            "agent_id": "A1",
            "call_id": "1",
            "phone_number": "0891",
            "department": "Support",
        }
    ]
    real_dataframe = pd.DataFrame
    calls = {"count": 0}

    def dataframe_side_effect(data=None, *args, **kwargs):
        calls["count"] += 1
        if calls["count"] == 1:
            return real_dataframe(data, *args, **kwargs)
        raise RuntimeError("df boom")

    with (
        patch.object(task, "_format_output", return_value=formatted_rows),
        patch.object(task, "_upload_master_file", side_effect=lambda df: df.copy()),
        patch.object(task, "_upload_daily_files"),
        patch("tasks.sentiment_qa.export_output_result_task.pd.DataFrame", side_effect=dataframe_side_effect),
        pytest.raises(Exception, match="Cannot create DataFrame for archival"),
    ):
        task.execute_task()


def test_format_output_returns_empty_for_no_records(task):
    assert task._format_output([]) == []


def test_format_output_handles_status_failure(task):
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(content=_mock_config_excel())
    task.sharepoint_control.is_item_exists.return_value = False
    record = _qa_record(prediction_status="FAILED")
    record["prediction"]["message"] = object()

    with patch(
        "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
        return_value=datetime(2025, 1, 2, 3, 4, 5),
    ):
        result = task._format_output([record])

    assert result[0]["status"] == "FAILED"
    assert "replace" in result[0]["error_code"]


def test_format_output_handles_customer_information_failure(task):
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(content=_mock_config_excel())
    task.sharepoint_control.is_item_exists.return_value = False
    record = _qa_record()
    record["prediction"]["raw_prediction"]["service_number"] = "__customer_boom__"
    original_safe_cast = qa_export_module.safe_cast_value

    def safe_cast_side_effect(value, cast_type, default=None):
        if value == "__customer_boom__":
            raise RuntimeError("customer info boom")
        return original_safe_cast(value, cast_type, default)

    with (
        patch("tasks.sentiment_qa.export_output_result_task.safe_cast_value", side_effect=safe_cast_side_effect),
        patch(
            "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
            return_value=datetime(2025, 1, 2, 3, 4, 5),
        ),
    ):
        result = task._format_output([record])

    assert result[0]["status"] == "FAILED"
    assert "customer info boom" in result[0]["error_code"]


def test_format_output_handles_false_fcr_and_customer_insight_failure(task):
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(content=_mock_config_excel())
    task.sharepoint_control.is_item_exists.return_value = False
    ok_record = _qa_record()
    ok_record["prediction"]["raw_prediction"]["customer_insight"]["fcr"] = "False"
    broken_record = _qa_record()
    broken_record["file_metadata"]["file_name"] = "other_call"
    broken_record["prediction"]["raw_prediction"]["customer_insight"]["product_category"] = "__insight_boom__"
    original_safe_cast = qa_export_module.safe_cast_value

    def safe_cast_side_effect(value, cast_type, default=None):
        if value == "__insight_boom__":
            raise RuntimeError("insight boom")
        return original_safe_cast(value, cast_type, default)

    with (
        patch("tasks.sentiment_qa.export_output_result_task.safe_cast_value", side_effect=safe_cast_side_effect),
        patch(
            "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
            return_value=datetime(2025, 1, 2, 3, 4, 5),
        ),
    ):
        result = task._format_output([ok_record, broken_record])

    assert result[0]["fcr"] == "N"
    assert result[1]["status"] == "FAILED"
    assert "insight boom" in result[1]["error_code"]


def test_format_output_maps_sale_and_experience_boolean_variants(task):
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(content=_mock_config_excel())
    task.sharepoint_control.is_item_exists.return_value = False
    record = _qa_record()
    sale = record["prediction"]["raw_prediction"]["sale_opportunity"]
    sale["opportunity_recognition_in_conversation"] = "False"
    sale["agent_offer_product_presentation_&_explanation"] = "True"
    sale["sales_outcome_&_customer_decision"] = "False"
    experience = record["prediction"]["raw_prediction"]["customer_experience"]
    experience["system_accessibility"] = "False"
    experience["ivr_usability_&_design"] = "True"

    with patch(
        "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
        return_value=datetime(2025, 1, 2, 3, 4, 5),
    ):
        result = task._format_output([record])

    assert result[0]["opportunity_recognition_in_conversation"] == "N"
    assert result[0]["agent_offer_product_presentation_&_explanation"] == "Y"
    assert result[0]["sales_outcome_&_customer_decision"] == "N"
    assert result[0]["system_accessibility"] == "N"
    assert result[0]["ivr_usability_&_design"] == "Y"


def test_upload_daily_files_handles_service_quality_build_failure(task):
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(content=_mock_config_excel())
    task.sharepoint_control.is_item_exists.return_value = False

    raw_record = _qa_record()
    flat_record = {
        **raw_record["file_metadata"],
        **raw_record["prediction"]["raw_prediction"]["service_quality"],
        "status": "SUCCESS",
    }

    for col in flat_record.copy():
        if col not in ["file_metadata", "prediction", "status"]:
            flat_record[f"{col}_reason"] = "Some reason"

    daily_df = pd.DataFrame([flat_record])

    with patch.object(task, "_calculate_category", side_effect=RuntimeError("service quality boom")):
        with pytest.raises(RuntimeError, match="service quality boom"):
            task._upload_daily_files(
                daily_df=daily_df, daily_output_path="mock/path/output.xlsx", sharepoint_site=MagicMock()
            )


def test_format_output_handles_sale_opportunity_failure(task):
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(content=_mock_config_excel())
    task.sharepoint_control.is_item_exists.return_value = False
    record = _qa_record()
    record["prediction"]["raw_prediction"]["sale_opportunity"]["product_suggested_by_ai"] = "__sale_boom__"
    original_safe_cast = qa_export_module.safe_cast_value

    def safe_cast_side_effect(value, cast_type, default=None):
        if value == "__sale_boom__":
            raise RuntimeError("sale boom")
        return original_safe_cast(value, cast_type, default)

    with (
        patch("tasks.sentiment_qa.export_output_result_task.safe_cast_value", side_effect=safe_cast_side_effect),
        patch(
            "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
            return_value=datetime(2025, 1, 2, 3, 4, 5),
        ),
    ):
        result = task._format_output([record])

    assert result[0]["status"] == "FAILED"
    assert "sale boom" in result[0]["error_code"]


def test_format_output_handles_customer_sentiment_failure(task):
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(content=_mock_config_excel())
    task.sharepoint_control.is_item_exists.return_value = False
    record = _qa_record()
    record["prediction"]["raw_prediction"]["customer_sentiment"]["overall_sentiment"] = "__sentiment_boom__"
    original_safe_cast = qa_export_module.safe_cast_value

    def safe_cast_side_effect(value, cast_type, default=None):
        if value == "__sentiment_boom__":
            raise RuntimeError("sentiment boom")
        return original_safe_cast(value, cast_type, default)

    with (
        patch("tasks.sentiment_qa.export_output_result_task.safe_cast_value", side_effect=safe_cast_side_effect),
        patch(
            "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
            return_value=datetime(2025, 1, 2, 3, 4, 5),
        ),
    ):
        result = task._format_output([record])

    assert result[0]["status"] == "FAILED"
    assert "sentiment boom" in result[0]["error_code"]


def test_format_output_handles_customer_experience_failure(task):
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(content=_mock_config_excel())
    task.sharepoint_control.is_item_exists.return_value = False
    record = _qa_record()
    record["prediction"]["raw_prediction"]["customer_experience"]["agent_communication_&_attitude"] = (
        "__experience_boom__"
    )
    original_safe_cast = qa_export_module.safe_cast_value

    def safe_cast_side_effect(value, cast_type, default=None):
        if value == "__experience_boom__":
            raise RuntimeError("experience boom")
        return original_safe_cast(value, cast_type, default)

    with (
        patch("tasks.sentiment_qa.export_output_result_task.safe_cast_value", side_effect=safe_cast_side_effect),
        patch(
            "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
            return_value=datetime(2025, 1, 2, 3, 4, 5),
        ),
    ):
        result = task._format_output([record])

    assert result[0]["status"] == "FAILED"
    assert "experience boom" in result[0]["error_code"]


def test_format_output_skips_failed_log_processing_errors(task):
    task.sharepoint_control.get_item_by_path.side_effect = [
        SimpleNamespace(content=_mock_config_excel()),
        SimpleNamespace(content=_mock_config_excel()),
        SimpleNamespace(content=b"bad log"),
    ]
    task.sharepoint_control.is_item_exists.side_effect = lambda item_path: item_path == "Control/batch_log.csv"

    with (
        patch("tasks.sentiment_qa.export_output_result_task.pd.read_csv", side_effect=RuntimeError("bad log csv")),
        patch(
            "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
            return_value=datetime(2025, 1, 2, 3, 4, 5),
        ),
    ):
        result = task._format_output([_qa_record()])

    assert len(result) == 1
    assert result[0]["status"] == "SUCCESS"


def test_format_output_skips_failed_log_when_path_resolution_fails(task):
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(content=_mock_config_excel())
    task.sharepoint_control.is_item_exists.return_value = False

    def resolve_env_side_effect(value):
        if value == "Control/batch_log.csv":
            raise RuntimeError("bad log path")
        return value

    with (
        patch("tasks.sentiment_qa.export_output_result_task.resolve_env", side_effect=resolve_env_side_effect),
        patch(
            "tasks.sentiment_qa.export_output_result_task.get_current_datetime",
            return_value=datetime(2025, 1, 2, 3, 4, 5),
        ),
    ):
        result = task._format_output([_qa_record()])

    assert result[0]["status"] == "SUCCESS"


def test_archive_files_skips_partition_when_record_filtering_fails(task):
    task.sharepoint["verint"]["input_folder_list"] = "Support"
    frame = _MaskDrivenFrame(["20250101"], lambda: (_ for _ in ()).throw(RuntimeError("filter boom")))

    task._archive_files(datetime(2025, 1, 2, 3, 4, 5), frame, [], [])

    task.gcs_module.move_file.assert_not_called()


def test_archive_files_skips_product_when_folder_path_resolution_fails(task):
    task.sharepoint["verint"]["input_folder_list"] = "Support"
    task.gcs.update(
        {
            "processing_voice_folder": "processing/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_voice_folder": "archive/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "input_folder": "input/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_batch_folder": "archive_batch/%{DATA_DATE_YYYYMMDD}",
        }
    )
    df = pd.DataFrame(
        [{"record_date": "20250101", "status": "SUCCESS", "department": "Support", "file_name": "call-1.wav"}]
    )

    with patch(
        "tasks.sentiment_qa.export_output_result_task.resolve_date", side_effect=RuntimeError("bad folder path")
    ):
        task._archive_files(datetime(2025, 1, 2, 3, 4, 5), df, [], [])

    task.gcs_module.list_files.assert_not_called()


def test_archive_files_handles_processing_and_input_list_errors(task):
    task.sharepoint["verint"]["input_folder_list"] = "Support"
    task.gcs.update(
        {
            "processing_voice_folder": "processing/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_voice_folder": "archive/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "input_folder": "input/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_batch_folder": "archive_batch/%{DATA_DATE_YYYYMMDD}",
        }
    )
    df = pd.DataFrame(
        [{"record_date": "20250101", "status": "SUCCESS", "department": "Support", "file_name": "call-1.wav"}]
    )

    def is_dir_exists(dir_path):
        if dir_path.startswith("processing/"):
            raise RuntimeError("processing list boom")
        if dir_path.startswith("input/"):
            raise RuntimeError("input list boom")
        return False

    task.gcs_module.is_dir_exists.side_effect = is_dir_exists
    task._archive_files(datetime(2025, 1, 2, 3, 4, 5), df, [], [])

    task.gcs_module.move_file.assert_not_called()
    task.gcs_module.delete_file.assert_not_called()


def test_archive_files_handles_voice_archive_and_delete_failures(task):
    task.sharepoint["verint"]["input_folder_list"] = "Support"
    task.combined_folder_list = ["Support"]
    task.gcs.update(
        {
            "processing_voice_folder": "processing/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_voice_folder": "archive/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "input_folder": "input/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_batch_folder": "archive_batch/%{DATA_DATE_YYYYMMDD}",
        }
    )
    df = pd.DataFrame(
        [{"record_date": "20250101", "status": "SUCCESS", "department": "Support", "file_name": "call-1.wav"}]
    )
    task.gcs_module.is_dir_exists.return_value = True
    task.gcs_module.list_files.side_effect = [
        ["processing/Support/%{DATA_DATE_YYYYMMDD}/call-1.wav"],
        ["input/Support/%{DATA_DATE_YYYYMMDD}/call-1.wav"],
    ]

    def move_side_effect(source_path=None, destination_path=None):
        if source_path and source_path.startswith("processing/"):
            raise RuntimeError("archive voice boom")
        return

    task.gcs_module.move_file.side_effect = move_side_effect
    task.gcs_module.delete_file.side_effect = RuntimeError("delete input boom")

    task._archive_files(datetime(2025, 1, 2, 3, 4, 5), df, [], [])

    task.gcs_module.delete_file.assert_called_once()


def test_archive_files_handles_batch_path_resolution_and_parse_failures(task):
    task.sharepoint["verint"]["input_folder_list"] = ""
    df = pd.DataFrame(
        [{"record_date": "20250101", "status": "SUCCESS", "department": "Support", "file_name": "call-1.wav"}]
    )

    with patch(
        "tasks.sentiment_qa.export_output_result_task.resolve_date",
        side_effect=["processing", "archive", "input", RuntimeError("bad batch path")],
    ):
        task._archive_files(datetime(2025, 1, 2, 3, 4, 5), df, ["output/run1/batch1.jsonl"], [])

    with patch("tasks.sentiment_qa.export_output_result_task.Path") as path_cls:
        task.gcs.update({"archive_batch_folder": "archive_batch/%{DATA_DATE_YYYYMMDD}"})
        path_cls.return_value.parts = property(lambda self: (_ for _ in ()).throw(RuntimeError("bad parts")))
        task._archive_files(datetime(2025, 1, 2, 3, 4, 5), df, ["output/run2/batch2.jsonl"], [])


def test_archive_files_handles_output_folder_extraction_and_missing_cleanup_dir(task):
    task.sharepoint["verint"]["input_folder_list"] = "Support"
    task.gcs.update(
        {
            "processing_voice_folder": "processing/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_voice_folder": "archive/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "input_folder": "input/{PRODUCT}/%{DATA_DATE_YYYYMMDD}",
            "archive_batch_folder": "archive_batch/%{DATA_DATE_YYYYMMDD}",
        }
    )
    df = pd.DataFrame(
        [{"record_date": "20250101", "status": "SUCCESS", "department": "Support", "file_name": "call-1.wav"}]
    )
    task.gcs_module.is_dir_exists.side_effect = lambda dir_path: False

    with patch(
        "tasks.sentiment_qa.export_output_result_task.os.path.dirname", side_effect=RuntimeError("dirname boom")
    ):
        task._archive_files(datetime(2025, 1, 2, 3, 4, 5), df, ["output/run1/batch1.jsonl"], [])

    task.gcs_module.move_file.reset_mock()
    task.gcs_module.is_dir_exists.side_effect = lambda dir_path: False
    task._archive_files(datetime(2025, 1, 2, 3, 4, 5), df, ["output/run1/batch1.jsonl"], [])

    task.gcs_module.delete_dir.assert_not_called()


def test_insert_log_record_raises_when_unique_dates_cannot_be_read(task):
    with pytest.raises(Exception, match="Cannot process log records"):
        task._insert_log_record(_BrokenUniqueDatesFrame())


def test_insert_log_record_skips_filter_failures_and_empty_partitions(task):
    filter_fail_frame = _MaskDrivenFrame(["20250101"], lambda: (_ for _ in ()).throw(RuntimeError("filter date boom")))
    task._insert_log_record(filter_fail_frame)

    empty_frame = _MaskDrivenFrame(["20250101"], lambda: pd.DataFrame())
    with patch.object(task, "_transaction_log") as transaction_log:
        task._insert_log_record(empty_frame)

    transaction_log.assert_not_called()


def test_insert_log_record_wraps_transaction_log_failures(task):
    df = pd.DataFrame([{"record_date": "20250101", "file_name": "x.wav"}])

    with patch.object(task, "_transaction_log", side_effect=RuntimeError("transaction boom")):
        with pytest.raises(Exception, match="Transaction log creation failed"):
            task._insert_log_record(df)


def test_insert_log_record_ignores_performance_log_failures(task):
    df = pd.DataFrame([{"record_date": "20250101", "file_name": "x.wav"}])
    transaction_df = pd.DataFrame([{"data_date": "20250101"}])

    with (
        patch.object(task, "_transaction_log", return_value=transaction_df),
        patch.object(task, "_performance_log", side_effect=RuntimeError("perf boom")),
    ):
        task._insert_log_record(df)


def test_transaction_log_handles_none_and_string_duration_values(task):
    task.sharepoint["control"]["transaction_log_file"] = "Control/transaction_%{DATA_DATE_YYYYMMDD}.csv"
    task.verint_access["site_name"] = "verint-site"
    task.verint_site = "verint.sharepoint.com"
    df = _export_prediction_df().copy()
    df.loc[df.index[0], "duration"] = None
    df.loc[df.index[1], "duration"] = "120"
    task.sharepoint_control.is_item_exists.return_value = False

    with (
        patch("tasks.sentiment_qa.export_output_result_task.gemini_cost", return_value={}),
        patch(
            "tasks.sentiment_qa.export_output_result_task.GeminiBatchModule.cal_gemini_cost",
            side_effect=lambda usage_detail, cost_config: {
                name: {"cost_input": 0.1, "cost_output": 0.2} for name in usage_detail
            },
        ),
    ):
        result = task._transaction_log("AI Classification", "daisyrpa", "SharePoint", df)

    assert len(result) == 2
    assert task.sharepoint_control.upload_file.called


def test_transaction_log_raises_when_payload_creation_fails(task):
    task.verint_access["site_name"] = "verint-site"
    task.verint_site = "verint.sharepoint.com"
    df = _export_prediction_df().iloc[[0]].copy()
    task.sharepoint_control.is_item_exists.return_value = False

    with (
        patch("tasks.sentiment_qa.export_output_result_task.gemini_cost", return_value={}),
        patch(
            "tasks.sentiment_qa.export_output_result_task.GeminiBatchModule.cal_gemini_cost",
            return_value={"call-1.wav": {"cost_input": 0.1, "cost_output": 0.2}},
        ),
        patch(
            "tasks.sentiment_qa.export_output_result_task.TransactionLogSchema.from_dict",
            side_effect=RuntimeError("payload boom"),
        ),
        pytest.raises(Exception, match="Transaction log creation failed at record 1"),
    ):
        task._transaction_log("AI Classification", "daisyrpa", "SharePoint", df)


def test_transaction_log_raises_when_no_payload_records_are_created(task):
    empty_df = pd.DataFrame(
        columns=[
            "file_name",
            "full_path",
            "folder",
            "record_date",
            "duration",
            "token_input",
            "token_cached",
            "token_output",
            "status",
            "message",
            "processed_time",
            "create_time",
            "model_version",
            "load_dt",
        ]
    )
    with patch("tasks.sentiment_qa.export_output_result_task.gemini_cost", return_value={}):
        with pytest.raises(Exception, match="No transaction records were created"):
            task._transaction_log("AI Classification", "daisyrpa", "SharePoint", empty_df)


def test_transaction_log_raises_when_schema_enforcement_fails(task):
    task.verint_access["site_name"] = "verint-site"
    task.verint_site = "verint.sharepoint.com"
    df = _export_prediction_df().iloc[[0]].copy()
    task.sharepoint_control.is_item_exists.return_value = False

    with (
        patch("tasks.sentiment_qa.export_output_result_task.gemini_cost", return_value={}),
        patch(
            "tasks.sentiment_qa.export_output_result_task.GeminiBatchModule.cal_gemini_cost",
            return_value={"call-1.wav": {"cost_input": 0.1, "cost_output": 0.2}},
        ),
        patch("tasks.sentiment_qa.export_output_result_task.ensure_df_schema", side_effect=RuntimeError("schema boom")),
    ):
        with pytest.raises(Exception, match="Cannot create transaction log"):
            task._transaction_log("AI Classification", "daisyrpa", "SharePoint", df)


def test_transaction_log_raises_when_data_date_extraction_fails(task):
    task.verint_access["site_name"] = "verint-site"
    task.verint_site = "verint.sharepoint.com"
    df = _export_prediction_df().iloc[[0]].copy()

    with (
        patch("tasks.sentiment_qa.export_output_result_task.gemini_cost", return_value={}),
        patch(
            "tasks.sentiment_qa.export_output_result_task.GeminiBatchModule.cal_gemini_cost",
            return_value={"call-1.wav": {"cost_input": 0.1, "cost_output": 0.2}},
        ),
        patch("tasks.sentiment_qa.export_output_result_task.ensure_df_schema", return_value=_BrokenDataDateFrame()),
    ):
        with pytest.raises(Exception, match="Cannot process transaction logs"):
            task._transaction_log("AI Classification", "daisyrpa", "SharePoint", df)


def test_transaction_log_raises_when_existing_log_merge_fails(task):
    task.sharepoint["control"]["transaction_log_file"] = "Control/transaction_%{DATA_DATE_YYYYMMDD}.csv"
    task.verint_access["site_name"] = "verint-site"
    task.verint_site = "verint.sharepoint.com"
    df = _export_prediction_df().copy()
    df.loc[df.index[1], "record_date"] = "20250102"
    exists_calls = {"count": 0}

    def is_item_exists(*, item_path):
        exists_calls["count"] += 1
        return exists_calls["count"] == 1

    task.sharepoint_control.is_item_exists.side_effect = is_item_exists
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(content=b"broken csv")

    with (
        patch("tasks.sentiment_qa.export_output_result_task.gemini_cost", return_value={}),
        patch(
            "tasks.sentiment_qa.export_output_result_task.GeminiBatchModule.cal_gemini_cost",
            side_effect=lambda usage_detail, cost_config: {
                name: {"cost_input": 0.1, "cost_output": 0.2} for name in usage_detail
            },
        ),
        patch("tasks.sentiment_qa.export_output_result_task.pd.read_csv", side_effect=RuntimeError("merge boom")),
    ):
        result = task._transaction_log("AI Classification", "daisyrpa", "SharePoint", df)

    assert len(result) == 2


def test_transaction_log_tracks_upload_and_date_processing_failures(task):
    task.sharepoint["control"]["transaction_log_file"] = "Control/transaction_%{DATA_DATE_YYYYMMDD}.csv"
    task.verint_access["site_name"] = "verint-site"
    task.verint_site = "verint.sharepoint.com"
    df = _export_prediction_df().copy()
    df.loc[df.index[1], "record_date"] = "20250102"
    exists_calls = {"count": 0}

    def is_item_exists(*, item_path):
        exists_calls["count"] += 1
        if exists_calls["count"] == 2:
            raise RuntimeError("date processing boom")
        return False

    task.sharepoint_control.is_item_exists.side_effect = is_item_exists
    task.sharepoint_control.upload_file.side_effect = RuntimeError("upload boom")

    with (
        patch("tasks.sentiment_qa.export_output_result_task.gemini_cost", return_value={}),
        patch(
            "tasks.sentiment_qa.export_output_result_task.GeminiBatchModule.cal_gemini_cost",
            side_effect=lambda usage_detail, cost_config: {
                name: {"cost_input": 0.1, "cost_output": 0.2} for name in usage_detail
            },
        ),
    ):
        result = task._transaction_log("AI Classification", "daisyrpa", "SharePoint", df)

    assert len(result) == 2
    assert task.sharepoint_control.upload_file.called


def test_performance_log_raises_when_schema_enforcement_fails(task):
    transaction_log_df = pd.DataFrame(
        [
            {
                "data_date": "20250101",
                "start_time": "2025-01-01 00:00:00+00:00",
                "load_dt": "2025-01-02 03:04:05",
                "gcp_project_id": "qa-proj",
                "gcp_project_name": "QA Project",
                "status_pass_failed_retry": "Pass",
                "latency_ms": 5000,
            }
        ]
    )

    with (
        patch(
            "tasks.sentiment_qa.export_output_result_task.ensure_df_schema",
            side_effect=RuntimeError("perf schema boom"),
        ),
        pytest.raises(Exception, match="Cannot create performance log"),
    ):
        task._performance_log(transaction_log_df)


def test_performance_log_raises_when_data_date_extraction_fails(task):
    transaction_log_df = pd.DataFrame(
        [
            {
                "data_date": "20250101",
                "start_time": "2025-01-01 00:00:00+00:00",
                "load_dt": "2025-01-02 03:04:05",
                "gcp_project_id": "qa-proj",
                "gcp_project_name": "QA Project",
                "status_pass_failed_retry": "Pass",
                "latency_ms": 5000,
            }
        ]
    )

    with patch("tasks.sentiment_qa.export_output_result_task.ensure_df_schema", return_value=_BrokenDataDateFrame()):
        with pytest.raises(Exception, match="Cannot process performance logs"):
            task._performance_log(transaction_log_df)


def test_performance_log_handles_new_logs_and_failures(task):
    task.sharepoint["control"]["performance_log_file"] = "Control/performance_%{DATA_DATE_YYYYMMDD}.csv"
    transaction_log_df = pd.DataFrame(
        [
            {
                "data_date": "20250101",
                "start_time": "2025-01-01 00:00:00+00:00",
                "load_dt": "2025-01-02 03:04:05",
                "gcp_project_id": "qa-proj",
                "gcp_project_name": "QA Project",
                "status_pass_failed_retry": "Pass",
                "latency_ms": 5000,
            },
            {
                "data_date": "20250102",
                "start_time": "2025-01-02 00:00:00+00:00",
                "load_dt": "2025-01-02 03:04:05",
                "gcp_project_id": "qa-proj",
                "gcp_project_name": "QA Project",
                "status_pass_failed_retry": "Failed",
                "latency_ms": 2000,
            },
        ]
    )
    exists_calls = {"count": 0}

    def is_item_exists(*, item_path):
        exists_calls["count"] += 1
        if exists_calls["count"] == 1:
            return False
        raise RuntimeError("perf date boom")

    task.sharepoint_control.is_item_exists.side_effect = is_item_exists
    task.sharepoint_control.upload_file.side_effect = RuntimeError("perf upload boom")

    task._performance_log(transaction_log_df)


def test_performance_log_raises_when_existing_log_merge_fails(task):
    task.sharepoint["control"]["performance_log_file"] = "Control/performance_%{DATA_DATE_YYYYMMDD}.csv"
    transaction_log_df = pd.DataFrame(
        [
            {
                "data_date": "20250101",
                "start_time": "2025-01-01 00:00:00+00:00",
                "load_dt": "2025-01-02 03:04:05",
                "gcp_project_id": "qa-proj",
                "gcp_project_name": "QA Project",
                "status_pass_failed_retry": "Pass",
                "latency_ms": 5000,
            }
        ]
    )
    task.sharepoint_control.is_item_exists.return_value = True
    task.sharepoint_control.get_item_by_path.return_value = SimpleNamespace(content=b"broken csv")

    with patch("tasks.sentiment_qa.export_output_result_task.pd.read_csv", side_effect=RuntimeError("perf merge boom")):
        task._performance_log(transaction_log_df)


def test_sync_excel_schema_returns_early_when_headers_match(task):
    wb = Workbook()
    ws = wb.active
    headers = ["alpha", "beta"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=idx, value=header)
    ws.cell(row=3, column=1, value="keep-a")
    ws.cell(row=3, column=2, value="keep-b")

    task._sync_excel_schema(ws, [("Only", headers)], headers)

    assert ws.cell(3, 1).value == "keep-a"
    assert ws.cell(3, 2).value == "keep-b"


def test_post_execute_skips_empty_partitions_and_uses_prod_environment(task, monkeypatch):
    task._cache_oper_log = {
        "process_date": [date(2025, 1, 3), date(2025, 1, 2)],
        "transaction_df": pd.DataFrame(
            [
                {
                    "start_time": "2025-01-02T00:00:00Z",
                    "end_time": "2025-01-02T00:00:03Z",
                    "gcp_project_id": "qa-proj",
                    "status_pass_failed_retry": "Pass",
                    "latency_ms": 3000,
                }
            ]
        ),
    }
    monkeypatch.setenv("ENVIRONMENT", "prod")

    with patch("tasks.sentiment_qa.export_output_result_task.logging_ai_operation") as log_ai_operation:
        assert task.post_execute("ok") == "ok"

    log_ai_operation.assert_called_once()
    assert log_ai_operation.call_args.kwargs["log_obj"]["environment"] == "production"


def test_transaction_log_uses_zero_seconds_for_explicit_none_duration(task):
    class _UniqueValues:
        def __init__(self, values):
            self._values = values

        def unique(self):
            return self

        def tolist(self):
            return list(self._values)

    class _UsageFrame:
        def __init__(self, records):
            self._records = records

        def copy(self):
            return self

        def to_dict(self, orient="records"):
            assert orient == "records"
            return list(self._records)

    class _PredictionFrame:
        def __init__(self, records):
            self._records = records
            self.columns = [
                "file_name",
                "full_path",
                "folder",
                "record_date",
                "duration",
                "token_input",
                "token_cached",
                "token_output",
                "status",
                "message",
                "processed_time",
                "create_time",
                "model_version",
                "load_dt",
            ]

        def __len__(self):
            return len(self._records)

        def __getitem__(self, key):
            if key == "model_version":
                return _UniqueValues([record["model_version"] for record in self._records])
            if key == self.columns:
                return _UsageFrame(self._records)
            raise KeyError(key)

    task.sharepoint["control"]["transaction_log_file"] = "Control/transaction_%{DATA_DATE_YYYYMMDD}.csv"
    task.verint_access["site_name"] = "verint-site"
    task.verint_site = "verint.sharepoint.com"
    prediction_df = _PredictionFrame(
        [
            {
                "file_name": "none-duration.wav",
                "full_path": "Input/Support/202501/20250101/none-duration.wav",
                "folder": "Support/20250101",
                "record_date": "20250101",
                "duration": None,
                "token_input": {"text": 10},
                "token_cached": 0,
                "token_output": {"text": 5},
                "status": "SUCCESS",
                "message": "",
                "processed_time": "2025-01-01T00:01:00Z",
                "create_time": "2025-01-01T00:00:00Z",
                "model_version": "gemini-2.5-flash",
                "load_dt": "2025-01-02 03:04:05",
            }
        ]
    )
    task.sharepoint_control.is_item_exists.return_value = False
    captured = {}
    original_from_dict = qa_export_module.TransactionLogSchema.from_dict

    def from_dict_side_effect(payload):
        captured["file_metadata_sec"] = payload.get("file_metadata_sec")
        return original_from_dict(payload)

    with (
        patch("tasks.sentiment_qa.export_output_result_task.gemini_cost", return_value={}),
        patch(
            "tasks.sentiment_qa.export_output_result_task.GeminiBatchModule.cal_gemini_cost",
            return_value={"none-duration.wav": {"cost_input": 0.1, "cost_output": 0.2}},
        ),
        patch(
            "tasks.sentiment_qa.export_output_result_task.TransactionLogSchema.from_dict",
            side_effect=from_dict_side_effect,
        ),
    ):
        result = task._transaction_log("AI Classification", "daisyrpa", "SharePoint", prediction_df)

    assert captured["file_metadata_sec"] == 0
    assert len(result) == 1
    assert task.sharepoint_control.upload_file.called


def test_ground_truth_schema_coerces_blank_strings_to_null():
    from tasks.sentiment_qa.schemas.ground_truth_schema import GroundTruthSchema

    data = {column: [pd.NA] for column in GroundTruthSchema.__annotations__}
    data.update(
        {
            "filename": ["call-1.wav"],
            "greeting_standard": ["   "],
            "call_type": ["\t"],
        }
    )

    validated = GroundTruthSchema.validate(pd.DataFrame(data))

    assert pd.isna(validated.loc[0, "greeting_standard"])
    assert pd.isna(validated.loc[0, "call_type"])
    assert validated.loc[0, "filename"] == "call-1.wav"
