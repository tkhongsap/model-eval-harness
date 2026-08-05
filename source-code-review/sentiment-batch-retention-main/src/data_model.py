# Process into User requirement file after append_csv
import io
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
import requests
from src.log import logger
from src.sharepoint import (
    get_access_token,
    get_item_download_url_by_path,
    list_file_name_in_sharepoint_folder,
    upload_file_to_sharepoint
)
from openpyxl import Workbook
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
import src.prompt as prompt
from src.gemini import (
    summarize
)

reason_colors = {
    'network': '#A6CEE3',
    'promotion related': "#1F78B4",
    'device promotion related': '#B2DF8A',
    'save cost': '#33A02C',
    'contract end': '#FB9A99',
    'sale upsell problem': '#E31A1C',
    'dissatisfied service': '#FDBF6F',
    'other': "#FF7F00",
    'post to pre': "#CAB2D6",
    'customer reason': '#6A3D9A',
    'true point, dtac reward': "#67673C",
    'down sell not success': '#B15928'
}

call_event_detection = [
    'Market-Driven Events (เหตุการณ์ทางการตลาด)',
    'Crisis & Emergency Events (เหตุการณ์วิกฤตหรือภัยพิบัติ)',
    'Campaign-Drvien Events (เหตุการณ์ด้านเคมเปญต่างๆของบริษัท)',
    'Technology & Service Events (เหตุการณ์ด้านเทคโนโลยี/บริการ)',
    'True-DTAC Merger(การรวมกิจการของ True และ ดีแทค)',
    'Emerging or Undefined Events (เหตุผลที่ยังไม่สามารถจัดกลุ่มได้)'
]

# Header grouping definitions
group_headers = [
    ("Customer_Information", ["call_id", "phone_number", "call_month", "call_date", "call_time", "cost(thb)", "call_duration(min)", "file_name", "user"]),
    ("Call_Summary", ["call_result", "product"]),
    ("Reason", ["main", "secondary", "third"]),
    ("", ["keyword(main)", "keyword(secondary)", "keyword(third)", "call_event_detection", "AI_recommendation"]),
    ("", ["input_folder", "output_folder", "run_date", "status", "message"]),
    ("Additional", ["issue_type", "sub_reason", "problem_statement", "churn_probability", "area_tag_province", "area_tag_district", "area_tag_sub_district", "area_tag_landmark", "product_type"])
]

# Header grouping definitions
group_headers_retention = [
    ("Customer_Information", ["call_id", "phone_number", "call_month", "call_date", "call_time", "cost(thb)", "call_duration(min)", "file_name", "user"]),
    ("Call_Summary", ["call_result", "product"]),
    ("Reason", ["main", "secondary", "third"]),
    ("", ["keyword(main)", "keyword(secondary)", "keyword(third)", "call_event_detection", "AI_recommendation"]),
    ("", ["input_folder", "output_folder", "run_date", "status", "message"])
]

def sync_excel_schema(ws, group_headers, all_columns):
    """
    Ensure Excel columns match latest schema.
    Add missing columns and rebuild header rows if needed.
    """
    # Read existing header row (row 2)
    existing_headers = [cell.value for cell in ws[2]]

    if existing_headers == all_columns:
        return  # schema already correct

    # Read existing data into memory (skip first 2 rows)
    data = list(ws.iter_rows(min_row=3, values_only=True))
    existing_df = pd.DataFrame(data, columns=existing_headers)

    # Add missing columns
    for col in all_columns:
        if col not in existing_df.columns:
            existing_df[col] = ""

    # Reorder columns to match new schema
    existing_df = existing_df[all_columns]

    # 1. Unmerge all merged cells first (IMPORTANT)
    if ws.merged_cells.ranges:
        for merged_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged_range))
        
    # Clear worksheet
    ws.delete_rows(1, ws.max_row)

    # --- Recreate headers ---
    col_idx = 1
    for group_name, cols in group_headers:
        start_col = col_idx
        col_idx += len(cols)
        end_col = col_idx - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=group_name)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.font = Font(bold=True)

    col_idx = 1
    for _, cols in group_headers:
        for col_name in cols:
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(bold=True)
            col_idx += 1

    # Rewrite old data
    for r_idx, row in enumerate(existing_df.itertuples(index=False, name=None), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def upload_master_files(data_list: list[dict], **args):
    """Appends a list of data dicts to XLSX master files based on input folder structure."""
    logger.info("Update Master file")
    PROCESSING_PREFIX = args.get("PROCESSING_PREFIX")
    OUTPUT_EXCEL_PREFIX = args.get("OUTPUT_EXCEL_PREFIX")

    VERINT_SITE_BASE_ROOT = args.get("VERINT_SITE_BASE_ROOT")
    PRODUCT_OUTPUT = args.get("PRODUCT_OUTPUT")

    if not data_list:
        logger.info("No data to append to Master Excel.")
        return
    
    # Group by output folder path
    grouped_data = {}
    for item in data_list:
        file_uri = item.get('input_folder') or item.get('file_uri') or ''

        # Extract the date folder from URI (e.g., processing/202511)
        match = re.search(fr'{re.escape(PROCESSING_PREFIX)}/(\d{{6}})', file_uri)

        if match:
            date_folder = match.group(1)  # e.g. 202511
            folder_path = f"{date_folder}"
        else:
            folder_path = f"unknown_date"

        grouped_data.setdefault(folder_path, []).append(item)

    # Flatten all expected column names
    all_columns = [col for _, cols in group_headers for col in cols]

    master_df = pd.DataFrame()
    for folder_path, records in grouped_data.items(): # yyyymm, [records]
        df = pd.DataFrame(records)

        # ✅ Ensure consistent column order
        for col in all_columns:
            if col not in df.columns:
                df[col] = ""  # fill missing columns

        df = df[all_columns]  # reorder columns
        master_path = f"{VERINT_SITE_BASE_ROOT}/{OUTPUT_EXCEL_PREFIX}/{PRODUCT_OUTPUT}/{folder_path}/master.xlsx"
        try:
            # Try downloading existing XLSX
            access_token = get_access_token("verint")
            master_download_url = get_item_download_url_by_path("verint", master_path)
            master_file = requests.get(master_download_url, headers={"Authorization": f"Bearer {access_token}"})
            master_file.raise_for_status()
            master_file_bytes = master_file.content

            with io.BytesIO(master_file_bytes) as existing_file:
                existing_wb = load_workbook(existing_file)
                ws = existing_wb.active           
                sync_excel_schema(ws, group_headers, all_columns)    
                start_row = ws.max_row + 1
                for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=start_row):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
        except Exception:
            # Create a new workbook
            existing_wb = Workbook()
            ws = existing_wb.active

            # First row (group headers)
            col_idx = 1
            for group_name, cols in group_headers:
                start_col = col_idx
                col_idx += len(cols)
                end_col = col_idx - 1
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws.cell(row=1, column=start_col, value=group_name)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = Font(bold=True)

            # Second row (actual headers)
            col_idx = 1
            for _, cols in group_headers:
                for col_name in cols:
                    cell = ws.cell(row=2, column=col_idx, value=col_name)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.font = Font(bold=True)
                    col_idx += 1

            # Data rows start from row 3
            for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=3):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        # Auto-fit column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)

        # ✅ MOVED INSIDE LOOP - Save and upload for EACH folder
        output_stream = io.BytesIO()
        existing_wb.save(output_stream)
        output_stream.seek(0)

        upload_file_to_sharepoint("verint", output_stream, master_path)

        output_ws = existing_wb.active

        # Convert existing sheet into DataFrame (skip header rows)
        output_data = output_ws.values
        next(output_data)  # skip group headers row 1
        cols = next(output_data)  # get column headers row 2
        output_df = pd.DataFrame(output_data, columns=cols).dropna(how="all")
        master_df = pd.concat([master_df, output_df], ignore_index=True)
    return master_df

def upload_daily_files(master_df, grouped_series, team, **args):
    """Appends a list of data dicts to XLSX daily files based on input folder structure."""
    global group_headers_retention
    global group_headers
    logger.info("Update Daily file")
    PROCESSING_PREFIX = args.get("PROCESSING_PREFIX")
    OUTPUT_EXCEL_PREFIX = args.get("OUTPUT_EXCEL_PREFIX")

    VERINT_SITE_BASE_ROOT = args.get("VERINT_SITE_BASE_ROOT")
    PRODUCT_OUTPUT = args.get("PRODUCT_OUTPUT")
    
    SANDBOX_SITE_BASE_ROOT = args.get("SANDBOX_SITE_BASE_ROOT")

    if len(grouped_series.keys()) == 0:
        logger.info("No data to append to Daily Excel.")
        return

    if team == 'retention':
        headers = group_headers_retention
        site_name = "verint"
    else:
        headers = group_headers
        site_name = "sandbox"
        
    # Flatten all expected column names
    all_columns = [col for _, cols in headers for col in cols]

    for yyyymm in sorted(grouped_series.keys()):
        for yyyymmdd in sorted(grouped_series[yyyymm]):            
            daily_df = master_df[(master_df['call_date'] == yyyymmdd)].copy()

            df_upload = (
                daily_df.drop_duplicates(subset=["call_id", "phone_number", "call_date","product"], keep="last")
            )
            
            if team == "network":
                daily_path = f"{SANDBOX_SITE_BASE_ROOT}/{OUTPUT_EXCEL_PREFIX}/{PRODUCT_OUTPUT}/{yyyymm}/Result_Sentiment_Analysis_Retention_{yyyymmdd}.xlsx"
            else:
                daily_path = f"{VERINT_SITE_BASE_ROOT}/{OUTPUT_EXCEL_PREFIX}/{PRODUCT_OUTPUT}/{yyyymm}/{yyyymmdd}/Result_Sentiment_Analysis_Retention_{yyyymmdd}.xlsx"
            
            df_upload = df_upload.reindex(columns=all_columns)
            
            existing_wb = Workbook()
            ws = existing_wb.active

            # First row (group headers)
            col_idx = 1
            for group_name, cols in headers:
                start_col = col_idx
                col_idx += len(cols)
                end_col = col_idx - 1
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws.cell(row=1, column=start_col, value=group_name)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = Font(bold=True)

            # Second row (actual headers)
            col_idx = 1
            for _, cols in headers:
                for col_name in cols:
                    cell = ws.cell(row=2, column=col_idx, value=col_name)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.font = Font(bold=True)
                    col_idx += 1

            # Data rows start from row 3
            for r_idx, row in enumerate(df_upload.itertuples(index=False, name=None), start=3):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            logger.info(f"✅ Write file with {len(df_upload)} records: {daily_path}")

            # Auto-fit column widths
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)

            # Save and upload
            output_stream = io.BytesIO()
            existing_wb.save(output_stream)
            output_stream.seek(0)

            upload_file_to_sharepoint(site_name, output_stream, daily_path)

def create_daily_bar_chart_daily(df, grouped_series, **args):
    logger.info("Update Daily bar chart")
    PROCESSING_PREFIX = args.get("PROCESSING_PREFIX")
    OUTPUT_EXCEL_PREFIX = args.get("OUTPUT_EXCEL_PREFIX")

    VERINT_SITE_BASE_ROOT = args.get("VERINT_SITE_BASE_ROOT")
    PRODUCT_OUTPUT = args.get("PRODUCT_OUTPUT")

    try:
        # ========= Setup Thai Font Support (Set size once globally) =========
        font_path = 'fonts/Sarabun-Regular.ttf'
        matplotlib.font_manager.fontManager.addfont(font_path)
        matplotlib.rc('font', family='Sarabun', size=8) 
        
        # ========= Rest of your code (Data Prep) =========
        # Normalize date
        df["run_date"] = pd.to_datetime(df["run_date"], errors="coerce").dt.tz_localize(None)

        # Deduplication: Keep latest record per call_id
        df = (
            df.drop_duplicates(subset=["call_id", "phone_number", "call_date", "product"], keep="last")
        ).copy()
        
        df['call_month'] = df['call_month'].astype(str)
        df['call_date'] = df['call_date'].astype(str)
        
        for yyyymm in sorted(grouped_series.keys()):
            for yyyymmdd in sorted(grouped_series[yyyymm]):
                # Filter cumulative data up to current date
                accum_graph_df = df[(df['call_month'] == yyyymm) & (df['call_date'] <= yyyymmdd) & (df['status'] == "success")].copy()
                if len(accum_graph_df) == 0:
                    logger.info(f"No data for {yyyymm}/{yyyymmdd}, skipping...")
                    continue

                # Group by the original date and reason, then sum the count, sort  
                accum_graph_melted_df = accum_graph_df.melt(id_vars=['call_date', 'call_result', 'product'], value_vars=['main', 'secondary', 'third'], value_name='all_reason')
                accum_graph_melted_df = accum_graph_melted_df[accum_graph_melted_df['all_reason'].astype(str).str.strip() != '']
                accum_graph_group_df = accum_graph_melted_df.groupby(['call_date', 'product', 'all_reason']).size().unstack(fill_value=0)
                accum_graph_group_df = accum_graph_group_df.sort_index(axis=1)
                existing_reasons = accum_graph_group_df.columns
                plot_colors = [reason_colors.get(reason, 'gray') for reason in existing_reasons] 
                # ======== gen accum stacked bar chart image save rate by date, reason ===============================
                # Melt the 'save' data to count 'save' reasons per date
                call_result_melted_df = accum_graph_df.melt(id_vars=['call_date', 'call_result', 'call_id', 'product'], 
                                                value_vars=['main', 'secondary', 'third'], 
                                                value_name='reason')
                # Filter out rows where the 'reason' is empty
                call_result_melted_df = call_result_melted_df[call_result_melted_df['reason'].astype(str).str.strip() != '']
                # Calculate total count for each output_date and reason
                total_counts = call_result_melted_df.groupby(['call_date', 'product', 'reason']).size().reset_index(name='total_count')              
                # Calculate 'save' count for each output_date and reason
                save_counts = call_result_melted_df[call_result_melted_df['call_result'] == 'save'].groupby(['call_date', 'product', 'reason']).size().reset_index(name='save_count')
                # Merge the two dataframes
                result_df = pd.merge(total_counts, save_counts, on=['call_date', 'product', 'reason'], how='left').fillna(0)
                # Calculate the save percentage
                result_df['save_percentage'] = np.where(result_df['total_count'] > 0, (result_df['save_count'] / result_df['total_count']) * 100, 0)
                # Prepare data for stacked bar chart
                plot_df = result_df.pivot_table(index=['call_date', 'product'], columns='reason', values='save_percentage').fillna(0)
                # Create a list of colors for the existing reasons in the plot_df
                existing_reasons = plot_df.columns
                plot_colors = [reason_colors.get(reason, 'gray') for reason in existing_reasons]             
                # ============ plot 2 chart =========================================================                
                # --- New variables for manual positioning ---
                bar_width_manual = 1.0 # The full width of a bar group (date, product)
                date_gap_factor = 1.5 # How many extra bar widths to add for a date change

                # --- Calculate manual x-positions and tick labels, inserting a gap when the date changes ---
                x_positions = []
                cleaned_labels = []
                current_position = 0
                prev_date = None

                for date, product in accum_graph_group_df.index:
                    # Insert a gap if the date has changed
                    if prev_date is not None and date != prev_date:
                        current_position += bar_width_manual * date_gap_factor

                    # Add the current bar's position
                    x_positions.append(current_position)
                    # Store the label for the tick (Day and Product Class)
                    cleaned_labels.append(f"{date.split('/')[-1]}\n({product})")
                    # Move to the next position for the next bar (no gap between products on same date)
                    current_position += bar_width_manual
                    # Update the previous date
                    prev_date = date

                # --- Dynamic Figure Width based on manual positions ---
                # Add a small buffer to the final position for figure width
                fig_width = (current_position + 1) * 1.5 
                fig1, ax1 = plt.subplots(figsize=(fig_width, 15))
                fig2, ax2 = plt.subplots(figsize=(fig_width, 15))

                existing_reasons_1 = accum_graph_group_df.columns
                plot_colors_1 = [reason_colors.get(reason, 'gray') for reason in existing_reasons_1]
                existing_reasons_2 = plot_df.columns
                plot_colors_2 = [reason_colors.get(reason, 'gray') for reason in existing_reasons_2]

                # --- Plot 1: Counts (Manual Bar Plotting) ---
                bottoms_1 = {x: 0 for x in x_positions}
                bar_containers_1 = []

                # Iterate over each reason (column) to plot stacked bars
                for reason, color in zip(existing_reasons_1, plot_colors_1):
                    counts = accum_graph_group_df[reason].values
                    bars = ax1.bar(x_positions, counts, 
                                        bottom=[bottoms_1[x] for x in x_positions], 
                                        width=bar_width_manual, 
                                        label=reason, 
                                        color=color, 
                                        align='center')
                    bar_containers_1.append(bars)
    
                    # Update the bottoms for the next reason
                    for x, count in zip(x_positions, counts):
                        bottoms_1[x] += count

                # Add labels to the bars
                bottoms_1 = {x: 0 for x in x_positions} # Reset bottoms for label positioning
                TEXT_HEIGHT_THRESHOLD = 5
                MIN_SPACE_FOR_TEXT = 3
                for reason_index, (reason, color) in enumerate(zip(existing_reasons_1, plot_colors_1)):
                    counts = accum_graph_group_df[reason].values
                    bars = bar_containers_1[reason_index]
                    for bar, count, x_pos in zip(bars, counts, x_positions):
                        height = bar.get_height()
                        if height > 0:
                            y_text = bottoms_1[x_pos] + height / 2
                            ax1.text(bar.get_x() + bar.get_width() / 2, y_text, 
                                    f'{int(height)}', 
                                    ha='center', 
                                    va='center', 
                                    color='white', 
                                    fontsize=12)
                            
                        bottoms_1[x_pos] += count

                ax1.set_title(f'Count of Reasons by Date and Product Class (Cumulative to {yyyymmdd})', fontsize=14)
                ax1.set_xlabel('Date - Product')
                ax1.set_ylabel('Count of Reasons')
                ax1.legend(title='Reason', bbox_to_anchor=(1.05, 1), loc='upper left')

                # Custom x-axis labels to show 'Day' and 'Product Class' clearly
                ax1.set_xticks(x_positions)
                ax1.set_xticklabels(cleaned_labels, rotation=0, ha='center', fontsize=10)
                plt.tight_layout()

                output_buffer_1 = io.BytesIO()
                fig1.savefig(output_buffer_1, format='png')
                plt.close(fig1)
                output_buffer_1.seek(0)

                # --- Plot 2: Percentages (Manual Bar Plotting) ---
                bottoms_2 = {x: 0 for x in x_positions}
                bar_containers_2 = []

                # Iterate over each reason (column) to plot stacked bars
                for reason, color in zip(existing_reasons_2, plot_colors_2):
                    percentages = plot_df[reason].values
                    bars = ax2.bar(x_positions, percentages, 
                   bottom=[bottoms_2[x] for x in x_positions], 
                   width=bar_width_manual, 
                   label=reason, 
                   color=color, 
                   align='center')
                    bar_containers_2.append(bars)
    
                    # Update the bottoms for the next reason
                    for x, percentage in zip(x_positions, percentages):
                        bottoms_2[x] += percentage

                # Add labels to the bars
                bottoms_2 = {x: 0 for x in x_positions} # Reset bottoms for label positioning
                TEXT_PERCENT_THRESHOLD = 5
                for reason_index, (reason, color) in enumerate(zip(existing_reasons_2, plot_colors_2)):
                    percentages = plot_df[reason].values
                    bars = bar_containers_2[reason_index]
                    for bar, percentage, x_pos in zip(bars, percentages, x_positions):
                        height = bar.get_height()
                        if height > 0:
                            y_text = bottoms_2[x_pos] + height / 2
                            ax2.text(bar.get_x() + bar.get_width() / 2, y_text, 
                                f'{int(height)}%', 
                                ha='center', 
                                va='center', 
                                color='white', 
                                fontsize=12)
                        bottoms_2[x_pos] += percentage
        
                ax2.set_title(f'Save Rate Percentage by Reason, Date, and Product Class (Cumulative to {yyyymmdd})', fontsize=14)
                ax2.set_xlabel('Date - Product Class')
                ax2.set_ylabel('Save Rate Percentage')
                ax2.legend(title='Reason', bbox_to_anchor=(1.05, 1), loc='upper left')

                # Custom x-axis labels (reuse labels from Plot 1)
                ax2.set_xticks(x_positions)
                ax2.set_xticklabels(cleaned_labels, rotation=0, ha='center', fontsize=10)
                plt.tight_layout()

                output_buffer_2 = io.BytesIO()
                fig2.savefig(output_buffer_2, format='png')
                plt.close(fig2)
                output_buffer_2.seek(0)

                # --- Upload logic starts here ---
                daily_path = f"{VERINT_SITE_BASE_ROOT}/{OUTPUT_EXCEL_PREFIX}/{PRODUCT_OUTPUT}/{yyyymm}/{yyyymmdd}"
                
                # upload requests_by_reason
                requests_by_reason_path = f"{daily_path}/requests_by_reason.png"
                upload_file_to_sharepoint("verint", output_buffer_1, requests_by_reason_path)

                # upload downsell_by_reason
                save_rate_by_reason_path = f"{daily_path}/save_rate_by_reason.png"
                upload_file_to_sharepoint("verint", output_buffer_2, save_rate_by_reason_path)

    except Exception as e:
        logger.error(f"Error in create_daily_bar_chart_daily: {str(e)}")
        
def create_daily_summary_files(df, grouped_series, **args):
    
    OUTPUT_EXCEL_PREFIX = args.get("OUTPUT_EXCEL_PREFIX")
    VERINT_SITE_BASE_ROOT = args.get("VERINT_SITE_BASE_ROOT")
    PRODUCT_OUTPUT = args.get("PRODUCT_OUTPUT")
    CONTROL_SITE_BASE_ROOT = args.get("CONTROL_SITE_BASE_ROOT")
    CONTROL_SITE_PROMPTS_ROOT = args.get("CONTROL_SITE_PROMPTS_ROOT")
    
    input_path = f"{CONTROL_SITE_BASE_ROOT}/{CONTROL_SITE_PROMPTS_ROOT}/summary"
    
    try:
        file_name_list = list_file_name_in_sharepoint_folder("control", input_path)
        file_name_list.sort() # latest file version will be last index
        prompt_url = f"{input_path}/{file_name_list[-1]}"
        prompt_download_url = get_item_download_url_by_path("control", prompt_url)
    
        access_token = get_access_token("control")
        resp = requests.get(prompt_download_url, headers={"Authorization": f"Bearer {access_token}"})
        resp.raise_for_status()
        prompt_file_bytes = resp.content 
        prompt_text = prompt_file_bytes.decode("utf-8").strip()
        
        df = (
                df.drop_duplicates(subset=["call_id", "phone_number", "call_date", "product"], keep="last")        
            ).copy()
        for yyyymm in sorted(grouped_series.keys()):
            for yyyymmdd in sorted(grouped_series[yyyymm]):
                prompt_summary = prompt_text
                prompt_summary = prompt_summary.replace("{date}", yyyymmdd)
                daily_df = df[(df['call_date'] == yyyymmdd)].copy()

                melt_daily_df = daily_df.melt(id_vars=['call_date', 'call_id', 'call_result', 'call_event_detection'], value_vars=['main', 'secondary', 'third'], value_name='all_reason')
                for reason in reason_colors.keys():
                    reason_count = ((melt_daily_df['all_reason'] == reason)).sum()
                    reason_save = ((melt_daily_df['all_reason'] == reason) & (melt_daily_df['call_result'] == 'save')).sum()
                    reason_churn = ((melt_daily_df['all_reason'] == reason) & (melt_daily_df['call_result'] == 'churn')).sum()
                    reason_unknown = ((melt_daily_df['all_reason'] == reason) & (melt_daily_df['call_result'] == 'unknown')).sum()
                    
                    prompt_summary = prompt_summary.replace("{"+f"{reason}"+"}", str(reason_count))
                    prompt_summary = prompt_summary.replace("{"+f"{reason}_save"+"}", str(reason_save))
                    prompt_summary = prompt_summary.replace("{"+f"{reason}_churn"+"}", str(reason_churn))
                    prompt_summary = prompt_summary.replace("{"+f"{reason}_unknown"+"}", str(reason_unknown))

                total_save = ((daily_df['call_result'] == 'save')).sum()
                total_churn = ((daily_df['call_result'] == 'churn')).sum()
                total_unknown = ((daily_df['call_result'] == 'unknown')).sum()
                prompt_summary = prompt_summary.replace("{total_save}", str(total_save))
                prompt_summary = prompt_summary.replace("{total_churn}", str(total_churn))
                prompt_summary = prompt_summary.replace("{total_unknown}", str(total_unknown))

                for call_event in call_event_detection:
                    call_event_count = ((melt_daily_df['call_event_detection'] == call_event)).sum()
                    prompt_summary = prompt_summary.replace("{"+f"{call_event}"+"}", str(call_event_count))

                number_of_call = daily_df.shape[0]
                prompt_summary = prompt_summary.replace("{number_of_call}", str(number_of_call))
            
                result = summarize(prompt_summary, **args)
                result_bytes = result.encode('utf-8')

                byte_buffer = io.BytesIO()
                byte_buffer.write(result_bytes)
                byte_buffer.seek(0)
            
                daily_path = f"{VERINT_SITE_BASE_ROOT}/{OUTPUT_EXCEL_PREFIX}/{PRODUCT_OUTPUT}/{yyyymm}/{yyyymmdd}/summary.txt"
                upload_file_to_sharepoint("verint", byte_buffer, daily_path)
                
    except Exception as e:
        logger.error(f"Error in create_daily_bar_chart_daily: {str(e)}")     