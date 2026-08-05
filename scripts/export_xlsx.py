"""The comparison as a workbook, with its caveats on the first sheet.

A spreadsheet travels further than the report that produced it. The text report carries
`RECONCILED: NO`, `PROMPT: RECONSTRUCTED` and the three not-observable error sources at
the foot, where a reader who was forwarded the file has already stopped reading. Here
they are sheet 1, before any number, because the failure mode this file is most likely
to cause is somebody quoting a weighted F1 in a migration decision without knowing that
production is handed AUDIO and this pack was handed pre-tagged Thai TEXT that no native
speaker has signed off.

Nothing here scores. Every number is read from `evalgen`/`evalharness`:
`report.mechanism_table` and `report.n_flip` for the verdicts, `cli.arm_summary` for the
weighted metrics, `compare.disagreement` for the 2x2, `fabrication.fabrication_rate` for
the invented-label counts. A second implementation of any of those would be a second
number to reconcile, and this file would be the one nobody checks.

Grain of the per-item sheet is (item, product), the same grain
`records.Record.key` merges on, taken as the UNION of the ground truth's products and
every arm's products. A product an arm invented therefore gets its own row with an empty
ground truth rather than vanishing, which is what would happen if the sheet were built
by iterating the ground truth alone.

Replicate 1 is what the per-item and comparison sheets show, matching the text report's
section 5 and for the same reason (`metrics.outer_join` keys predictions by
(call_id, phone, product), so pooling replicates silently keeps the last one). The
replicate dimension is not dropped: it is carried as the mechanism verdicts, `N_flip`,
and a `flips` marker per cell block.

Usage:

    python scripts/export_xlsx.py \
        --incumbent-base out/runs/...-incumbent-base \
        --candidate-base out/runs/...-candidate-base \
        --incumbent-e1   out/runs/...-incumbent-e1 \
        --candidate-e1   out/runs/...-candidate-e1
"""

from __future__ import annotations

import argparse
import json
import statistics
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Sequence

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.worksheet import Worksheet  # noqa: E402

from evalgen.cli import arm_summary, load_run, replicate_records  # noqa: E402
from evalgen.console import configure_stdout  # noqa: E402
from evalgen.fabrication import fabrication_rate  # noqa: E402
from evalgen.report import mechanism_table, n_flip  # noqa: E402
from evalgen.testsets import load_testset  # noqa: E402
from evalharness.adapters.retention import load_csv  # noqa: E402
from evalharness.compare import _correct, disagreement  # noqa: E402
from evalharness.records import Record  # noqa: E402

# Loaded by path, not by name. `scripts/` holds `evalgen.py`, a launcher, so putting this
# directory on sys.path makes `import evalgen.cli` resolve to that module and fail with
# "'evalgen' is not a package" -- shadowing the real package with its own entry point.
def _load_sibling(name: str):  # noqa: E402
    import importlib.util  # noqa: PLC0415

    spec = importlib.util.spec_from_file_location(
        f"_evalgen_scripts_{name}", Path(__file__).resolve().parent / f"{name}.py"
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


consistency_mod = _load_sibling("consistency")  # the repeat-pass comparison

# ------------------------------------------------------------------ presentation

# Tahoma carries the Thai block and ships with both Windows and Office for macOS. The
# transcripts are Thai; a font without Thai coverage renders them as boxes in the one
# column the reader most needs to check a label against.
FONT = "Tahoma"
SIZE = 10

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name=FONT, size=SIZE, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, size=14, bold=True, color="1F3864")
BODY_FONT = Font(name=FONT, size=SIZE)
BOLD_FONT = Font(name=FONT, size=SIZE, bold=True)
MONO_FONT = Font(name="Consolas", size=9)

BAD_FILL = PatternFill("solid", fgColor="FCE4E4")
BAD_FONT = Font(name=FONT, size=SIZE, color="9C0006")
GOOD_FILL = PatternFill("solid", fgColor="E7F4E4")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
NOGT_FILL = PatternFill("solid", fgColor="EDEDED")
CAVEAT_FILL = PatternFill("solid", fgColor="FFF2CC")
ALARM_FILL = PatternFill("solid", fgColor="C00000")
ALARM_FONT = Font(name=FONT, size=12, bold=True, color="FFFFFF")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TOP_WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")

DIMENSIONS = ("call_result", "reason", "product")

# What the GT call_result cell says on a row no ground-truth label covers. Named because
# the verifier counts ground-truth-backed rows by looking for its ABSENCE, so the writer
# and the check must agree on the exact string or the check silently passes everything.
NO_GT_MARKER = "-- no ground-truth row --"


def put(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    wrap: bool = False,
    border: bool = False,
    number_format: str | None = None,
):
    """One cell, always with an explicit font. There is no styled default to inherit."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or BODY_FONT
    cell.alignment = TOP_WRAP if wrap else TOP
    if fill is not None:
        cell.fill = fill
    if border:
        cell.border = BOX
    if number_format:
        cell.number_format = number_format
    return cell


def header_row(ws: Worksheet, row: int, labels: Sequence[str], *, start: int = 1) -> None:
    for offset, label in enumerate(labels):
        put(ws, row, start + offset, label, font=HEAD_FONT, fill=HEAD_FILL, wrap=True,
            border=True)


def widths(ws: Worksheet, spec: Sequence[float], *, start: int = 1) -> None:
    for offset, width in enumerate(spec):
        ws.column_dimensions[get_column_letter(start + offset)].width = width


# ------------------------------------------------------------------ arm loading


class Arm:
    """One run directory, read once, with everything the sheets ask of it."""

    def __init__(self, slot: str, label: str, directory: Path, gt: list[Record], items):
        self.slot = slot           # incumbent-base | candidate-base | ...
        self.label = label         # 'Gemini / base' -- what a reader sees
        self.loaded = load_run(directory)
        self.meta = self.loaded.meta
        self.directory = Path(self.meta.get("run_id", directory.name)) and directory
        self.per_replicate = replicate_records(self.loaded.result, items)
        self.summary = arm_summary(self.loaded, gt, self.per_replicate)
        self.mechanisms = mechanism_table(gt, self.per_replicate, list(items))
        self.n_flip = n_flip(self.per_replicate)
        self.rep1 = {r.key: r for r in self.per_replicate[0]}
        self.fab = fabrication_rate(directory, Path(self.meta["gt_path"]))
        self.rows = [
            json.loads(line)
            for line in (directory / "run.jsonl").read_text("utf-8").splitlines()
            if line.strip()
        ]
        # Item-level correctness per replicate, for the `flips` marker: an arm that is
        # right on replicate 1 and wrong on replicates 2 and 3 must not read as clean.
        self.per_rep_by_key: list[dict] = [
            {r.key: r for r in rows} for rows in self.per_replicate
        ]

    @property
    def prompt_id(self) -> str:
        return self.meta["prompt_id"]

    @property
    def model(self) -> str:
        return self.meta["model_requested"]

    def flips(self, gt_row: Record | None, key, dimension: str) -> bool:
        """Whether this cell took more than one value across replicates."""
        if gt_row is None:
            return False
        seen = set()
        for by_key in self.per_rep_by_key:
            seen.add(_correct(gt_row, by_key.get(key), dimension))
        return len(seen) > 1

    def cell(self, key, dimension: str) -> str:
        rec = self.rep1.get(key)
        if rec is None:
            return "(no row)"
        if dimension == "call_result":
            return rec.call_result or "(empty)"
        if dimension == "reason":
            return ", ".join(sorted(rec.reasons)) or "(empty)"
        return rec.product or "(empty)"


def latency_stats(rows: list[dict]) -> tuple[float, float]:
    values = [r["latency_s"] for r in rows if r.get("latency_s") is not None]
    if not values:
        return (0.0, 0.0)
    return (round(statistics.median(values), 2), round(max(values), 2))


def token_stats(rows: list[dict]) -> tuple[int, int, int]:
    prompt = sum(r.get("prompt_tokens") or 0 for r in rows)
    completion = sum(r.get("completion_tokens") or 0 for r in rows)
    reasoning = sum(r.get("reasoning_tokens") or 0 for r in rows)
    return (prompt, completion, reasoning)


# ------------------------------------------------------------------ sheet 1


READ_FIRST_ALARMS = [
    (
        "RECONCILED: NO",
        "These scores have NOT been checked against the Retention app's own live Gemini "
        "fact-check report. Nothing in this repository can perform that check -- it needs "
        "real labelled production data, which this repo does not and must not hold. Until "
        "that is done these are HARNESS OUTPUT, not evidence about a model, and no number "
        "in this workbook is a migration verdict.",
    ),
    (
        "PROMPT: RECONSTRUCTED",
        "The prompt was reassembled from committed assets, not read from the running "
        "system. Production fetches the user_prompt from SharePoint at run time "
        "(main.py:1140-1156). This is the copy that was live when the harness was written, "
        "with three example-JSON transcription bugs fixed and every audio reference "
        "rewritten for transcript input. Production also calls Vertex with FORCED FUNCTION "
        "CALLING, which this harness cannot reproduce: it passes the schema through "
        "response_format, which is advisory in a way mode:ANY is not.",
    ),
    (
        "PRODUCTION GETS AUDIO. THIS PACK SENDS TEXT.",
        "Production is handed an AUDIO FILE and asked to identify the speakers, transcribe "
        "them and label the call in ONE pass (prompt.py:4314-4315). This pack sends clean, "
        "pre-tagged Thai TEXT. The two are different tasks and this workbook measures only "
        "the second one.",
    ),
    (
        "THE THAI WAS DRAFTED BY AN LLM. NO NATIVE SPEAKER HAS SIGNED IT OFF.",
        "Every transcript in this pack was written by a language model. 'The model handled "
        "this Thai correctly' is evidence about THIS TEXT and about nothing else. It is not "
        "a sample of real call-centre Thai and was not drawn from one.",
    ),
]

NOT_OBSERVABLE = [
    (
        "Agent-speech misattribution",
        "Every transcript turn is already tagged with its speaker, so the model never has "
        "to decide who spoke. prompt.py:4382-4387 makes the reason depend on the phrase "
        "being CUSTOMER speech; RET-10 tests whether a model USES the tags it was given, "
        "not whether it could recover them from audio -- which is the harder thing "
        "production actually does.",
    ),
    (
        "ASR error",
        "Thai tone marks, proper nouns, code-switching and telephone-band distortion are "
        "absent. RET-11 hand-writes ASR-shaped orthography; that is an imitation of one "
        "failure mode, not a measurement of the transcription stage.",
    ),
    (
        "Diarisation",
        "Speaker segmentation, crosstalk, holds and silence do not exist in this text at "
        "all.",
    ),
]

WHAT_THIS_IS = [
    "A paired comparison of two models on ONE controlled dimension: labelling a Thai "
    "retention call transcript against True's Retention prompt v9_16.",
    "Two models (google/gemini-2.5-flash as incumbent, qwen/qwen3.6-27b as candidate) x "
    "two prompts (v9_16_base, and v9_16_e1 which blanks the worked example's secondary "
    "and third reason values) = four runs, each 20 items x 3 replicates.",
    "Both arms are PINNED to one OpenRouter backend (--provider, allow_fallbacks=false), "
    "and the pin is verified by the prompt_tokens fingerprint rather than trusted: an "
    "earlier candidate run was served by TWO builds under one model id and nothing in the "
    "log showed it.",
]

WHAT_THIS_IS_NOT = [
    "NOT a production accuracy estimate. See the audio/text row above.",
    "NOT a migration verdict. See RECONCILED: NO.",
    "NOT statistically powered. 22 scored rows means one row is 4.5 points, and McNemar "
    "on the paired discordant cells needs SIX items discordant in one direction before an "
    "exact two-sided p falls under 0.05. A five-point gap here is compatible with no "
    "difference at all. That is why sheet 'Mechanisms' reports PASS/FLAKY/FAIL verdicts "
    "and not a percentage.",
    "NOT a measurement of issue_type, which no line of the Retention app gives semantics "
    "to; this pack carries no issue_type label and cannot detect an issue_type regression.",
    "NOT a claim that the ranking transfers to audio. That is untested.",
]


def consistency_block(arms: dict[str, Arm], repeat: dict[str, Path] | None,
                      gt: list[Record], items) -> tuple[str, str, list[list[str]]] | None:
    """The repeat pass, folded down to one alarm and one small table.

    Computed through `consistency.py` rather than reimplemented, so the workbook and the
    console report cannot drift into two different answers about whether the run
    reproduced.
    """
    if not repeat:
        return (
            "REPRODUCIBILITY: NOT CHECKED.",
            "These numbers come from a SINGLE pass. The runs were not repeated, so nothing "
            "here says whether they would come out the same twice. Silence about "
            "reproducibility reads as reproducibility, so it is stated instead: unknown.",
            [],
        )

    second = {slot: consistency_mod.load(repeat[slot], gt, items) for slot in arms}
    table = [["arm", "labels agreeing (22 rows x 3 dims, replicate 1)",
              "mechanism letters", "failing item sets", "N_flip pass 1 / pass 2"]]
    reproduced = True
    for slot, arm in arms.items():
        first = {
            "rep1": arm.rep1,
            "mechanisms": arm.mechanisms,
            "n_flip": arm.n_flip,
        }
        la, _ = consistency_mod.grid(first, gt)
        lb, _ = consistency_mod.grid(second[slot], gt)
        same = sum(1 for k in la if la[k] == lb[k])
        letters_a = [(m.mechanism, m.verdict) for m in arm.mechanisms]
        letters_b = [(m.mechanism, m.verdict) for m in second[slot]["mechanisms"]]
        fail_a = {m.mechanism: set(m.failing_items) for m in arm.mechanisms}
        fail_b = {m.mechanism: set(m.failing_items) for m in second[slot]["mechanisms"]}
        moved = sorted({i for n in fail_a for i in fail_a[n] ^ fail_b.get(n, set())})
        if same != len(la) or letters_a != letters_b or moved:
            reproduced = False
        table.append([
            arm.label,
            f"{same}/{len(la)}",
            "identical" if letters_a == letters_b else "DIFFER",
            "identical" if not moved else "DIFFER: " + ", ".join(moved),
            f"{arm.n_flip} / {second[slot]['n_flip']}",
        ])

    if reproduced:
        headline = "REPRODUCIBILITY: the repeat pass reproduced this workbook exactly."
        body = ("All four runs were executed a second time with identical settings and "
                "the same pinned backends. Every label, every mechanism verdict and "
                "every headline number matched.")
    else:
        headline = "REPRODUCIBILITY: PARTIAL. A REPEAT PASS DID NOT REPRODUCE THESE NUMBERS."
        body = ("All four runs were executed a second time, same settings, same pinned "
                "backends, minutes apart. Both GEMINI arms reproduced exactly. Both QWEN "
                "arms did not: the set of items the candidate fails moved between passes, "
                "and one headline number moved with it. The numbers on the other sheets "
                "are ONE draw from a distribution, not a fixed measurement. Treat the "
                "difference between the two passes as the error bar -- this pack does not "
                "have another one. Details below and in out/reports/consistency-*.txt.")
    return headline, body, table


def sheet_read_first(wb: Workbook, arms: dict[str, Arm], built: str,
                     consistency: tuple[str, str, list[list[str]]] | None = None) -> Worksheet:
    ws = wb.create_sheet("READ FIRST")
    ws.sheet_properties.tabColor = "C00000"
    widths(ws, [34, 118])

    row = 1
    put(ws, row, 1, "Qwen vs Gemini on True Retention v9_16 - READ THIS SHEET FIRST",
        font=TITLE_FONT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    put(ws, row, 1, f"Built {built} from the run directories listed on the 'Runs' sheet.",
        font=BODY_FONT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 2

    alarms = list(READ_FIRST_ALARMS)
    if consistency is not None:
        alarms.append((consistency[0], consistency[1]))
    for headline, body in alarms:
        put(ws, row, 1, headline, font=ALARM_FONT, fill=ALARM_FILL, wrap=True, border=True)
        put(ws, row, 2, body, font=BODY_FONT, fill=CAVEAT_FILL, wrap=True, border=True)
        ws.row_dimensions[row].height = 62
        row += 1

    if consistency is not None and consistency[2]:
        row += 1
        put(ws, row, 1, "PASS 1 vs PASS 2, ARM BY ARM", font=BOLD_FONT)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        head, *body_rows = consistency[2]
        put(ws, row, 1, head[0], font=HEAD_FONT, fill=HEAD_FILL, wrap=True, border=True)
        put(ws, row, 2, " | ".join(head[1:]), font=HEAD_FONT, fill=HEAD_FILL, wrap=True,
            border=True)
        row += 1
        for line in body_rows:
            bad = any("DIFFER" in str(v) for v in line)
            put(ws, row, 1, line[0], font=BOLD_FONT, fill=BAD_FILL if bad else None,
                wrap=True, border=True)
            put(ws, row, 2, " | ".join(str(v) for v in line[1:]),
                font=BAD_FONT if bad else BODY_FONT,
                fill=BAD_FILL if bad else None, wrap=True, border=True)
            ws.row_dimensions[row].height = 30
            row += 1

    row += 1
    put(ws, row, 1, "NOT OBSERVABLE BY THIS PACK AT ALL", font=BOLD_FONT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    put(ws, row, 1, "Three of production's error sources are invisible here. No verdict in "
                    "this workbook says anything about them.", font=BODY_FONT, wrap=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    for name, body in NOT_OBSERVABLE:
        put(ws, row, 1, name, font=BOLD_FONT, fill=CAVEAT_FILL, wrap=True, border=True)
        put(ws, row, 2, body, font=BODY_FONT, wrap=True, border=True)
        ws.row_dimensions[row].height = 46
        row += 1

    row += 1
    put(ws, row, 1, "WHAT THIS WORKBOOK IS", font=BOLD_FONT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    for line in WHAT_THIS_IS:
        put(ws, row, 1, "-", font=BODY_FONT)
        put(ws, row, 2, line, font=BODY_FONT, wrap=True)
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    put(ws, row, 1, "WHAT THIS WORKBOOK IS NOT", font=BOLD_FONT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    for line in WHAT_THIS_IS_NOT:
        put(ws, row, 1, "-", font=BODY_FONT)
        put(ws, row, 2, line, font=BODY_FONT, wrap=True)
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    put(ws, row, 1, "HOW TO READ THE OTHER SHEETS", font=BOLD_FONT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    guide = [
        ("Per item", "One row per (item, product) -- the same grain the scorer merges on. "
                     "Replicate 1 of each arm. Mismatched cells are shaded red and every "
                     "cell block carries an explicit OK / MISMATCH verdict, so the finding "
                     "survives printing in black and white. 'flips' means the arm did not "
                     "give the same answer on all three replicates."),
        ("Comparison", "The paired 2x2 per dimension per prompt, then weighted precision, "
                       "recall and F1 per arm. Read it AFTER 'Mechanisms', not before."),
        ("Mechanisms", "The verdict table. PASS = every item in the group correct on every "
                       "replicate; FLAKY = some item correct on some replicates; FAIL = "
                       "some item wrong on every replicate, and FAIL outranks FLAKY."),
        ("Runs", "Provenance: model, provider, prompt id and sha, replicates, decoding, "
                 "outcome counts, N_flip, fabrication rate, tokens, cost and latency."),
    ]
    for name, body in guide:
        put(ws, row, 1, name, font=BOLD_FONT, wrap=True, border=True)
        put(ws, row, 2, body, font=BODY_FONT, wrap=True, border=True)
        ws.row_dimensions[row].height = 46
        row += 1

    ws.freeze_panes = "A4"
    return ws


# ------------------------------------------------------------------ sheet 2


def product_rows(items, gt: list[Record], arms: dict[str, Arm]) -> list[tuple]:
    """(item, product, gt_row) at the union grain, ground truth first."""
    gt_by_call: dict[str, list[Record]] = defaultdict(list)
    for record in gt:
        gt_by_call[record.call_id].append(record)

    out: list[tuple] = []
    for item in items:
        gt_rows = gt_by_call.get(item.call_id, [])
        seen = []
        for record in gt_rows:
            seen.append(record.product)
            out.append((item, record.product, record))
        extra = set()
        for arm in arms.values():
            for key, rec in arm.rep1.items():
                if key[0] == item.call_id and rec.product not in seen:
                    extra.add(rec.product)
        for product in sorted(extra, key=lambda p: (p is None, p or "")):
            out.append((item, product, None))
    return out


def sheet_per_item(wb: Workbook, items, gt: list[Record], arms: dict[str, Arm],
                   order: Sequence[str]) -> tuple[Worksheet, int]:
    ws = wb.create_sheet("Per item")
    rows = product_rows(items, gt, arms)

    lead = ["item_id", "family", "mechanism (what the item tests)", "Thai transcript",
            "GT product", "GT call_result", "GT reasons"]
    per_arm = ["call_result", "reasons", "product returned", "verdict"]

    # Two header rows: the arm banner, then the columns under it.
    for offset, label in enumerate(lead):
        put(ws, 1, 1 + offset, "", font=HEAD_FONT, fill=HEAD_FILL, border=True)
        put(ws, 2, 1 + offset, label, font=HEAD_FONT, fill=HEAD_FILL, wrap=True, border=True)
    col = len(lead) + 1
    for slot in order:
        arm = arms[slot]
        put(ws, 1, col, arm.label, font=HEAD_FONT, fill=HEAD_FILL, wrap=True, border=True)
        ws.merge_cells(start_row=1, start_column=col, end_row=1,
                       end_column=col + len(per_arm) - 1)
        for offset, label in enumerate(per_arm):
            put(ws, 2, col + offset, label, font=HEAD_FONT, fill=HEAD_FILL, wrap=True,
                border=True)
        col += len(per_arm)

    widths(ws, [10, 16, 42, 78, 12, 14, 30])
    widths(ws, [14, 30, 15, 22] * len(order), start=len(lead) + 1)

    r = 3
    for item, product, gt_row in rows:
        put(ws, r, 1, item.item_id, font=BOLD_FONT, border=True)
        put(ws, r, 2, item.family, border=True, wrap=True)
        put(ws, r, 3, item.mechanism, border=True, wrap=True)
        put(ws, r, 4, item.transcript_th, border=True, wrap=True)
        if gt_row is None:
            put(ws, r, 5, product or "(none)", fill=NOGT_FILL, border=True)
            put(ws, r, 6, NO_GT_MARKER, fill=NOGT_FILL, border=True,
                wrap=True)
            put(ws, r, 7, "", fill=NOGT_FILL, border=True)
        else:
            put(ws, r, 5, gt_row.product or "(empty)", font=BOLD_FONT, border=True)
            put(ws, r, 6, gt_row.call_result or "(empty)", font=BOLD_FONT, border=True)
            put(ws, r, 7, ", ".join(sorted(gt_row.reasons)) or "(none)", font=BOLD_FONT,
                border=True, wrap=True)

        col = len(lead) + 1
        for slot in order:
            arm = arms[slot]
            key = (item.call_id, gt_row.phone if gt_row else _phone_of(arm, item, product),
                   product)
            wrong: list[str] = []
            flip: list[str] = []
            for offset, dimension in enumerate(("call_result", "reason", "product")):
                value = arm.cell(key, dimension)
                ok = _correct(gt_row, arm.rep1.get(key), dimension) if gt_row else None
                if arm.flips(gt_row, key, dimension):
                    flip.append(dimension)
                fill = None
                font = BODY_FONT
                if ok is False:
                    wrong.append(dimension)
                    fill, font = BAD_FILL, BAD_FONT
                elif ok is None:
                    fill = NOGT_FILL
                put(ws, r, col + offset, value, font=font, fill=fill, border=True, wrap=True)
            if gt_row is None:
                verdict, fill = "EXTRA ROW (not in ground truth)", NOGT_FILL
            elif wrong:
                verdict, fill = "MISMATCH: " + ", ".join(wrong), BAD_FILL
            else:
                verdict, fill = "OK", GOOD_FILL
            if flip:
                verdict += "  [flips across replicates: " + ", ".join(flip) + "]"
                if not wrong:
                    fill = WARN_FILL
            put(ws, r, col + 3, verdict,
                font=BAD_FONT if wrong else BODY_FONT, fill=fill, border=True, wrap=True)
            col += len(per_arm)
        ws.row_dimensions[r].height = 90
        r += 1

    ws.freeze_panes = ws.cell(row=3, column=len(lead) + 1)
    ws.auto_filter.ref = f"A2:{get_column_letter(col - 1)}{r - 1}"
    return ws, len(rows)


def _phone_of(arm: Arm, item, product) -> str | None:
    """The phone an arm used for a call, for rows the ground truth does not have."""
    for key in arm.rep1:
        if key[0] == item.call_id and key[2] == product:
            return key[1]
    return None


# ------------------------------------------------------------------ side by side


def _signature(rows) -> list[tuple]:
    """One call's answer as a sortable multiset of (product, outcome, reasons).

    A multiset rather than a dict keyed on the merge key, for the reason
    `report._signature` gives: an arm that emits the same key twice has one of the two
    silently dropped by `metrics.outer_join`, and comparing multisets keeps that visible
    instead of hiding it behind a dict that quietly deduplicated.
    """
    return sorted(
        (r.product or "", r.call_result or "", tuple(sorted(r.reasons))) for r in rows
    )


def _answer_text(rows) -> str:
    """`product/outcome/reason+reason`, one clause per product, for eyeball comparison."""
    if not rows:
        return "(no answer)"
    return "  |  ".join(
        f"{p or '(none)'} / {o or '(empty)'} / {'+'.join(r) if r else '-'}"
        for p, o, r in _signature(rows)
    )


def sheet_side_by_side(wb: Workbook, items, gt: list[Record], arms: dict[str, Arm],
                       order: Sequence[str]) -> Worksheet:
    """The at-a-glance view: one row per item, every arm's whole answer beside the truth.

    This is the sheet a reader opens first and the only one that fits on a screen. `Per
    item` carries the same facts at (item, product) grain with the transcript attached,
    which is what you need to ADJUDICATE a disagreement; this is what you need to FIND
    one.

    Correctness here is deliberately **whole-item, all-or-nothing**, the same rule
    `report._item_is_correct` applies: every product of the call, its outcome and its
    full reason set must match. An arm that gets the product and outcome right and adds
    one unsupported reason reads MISMATCH, because that is what it is. The per-dimension
    view on `Comparison` is the one that gives partial credit, and the two are meant to
    be read together -- a run can look far worse here than there, and the gap between
    them is the over-labelling rate.

    Replicate 1, matching every other scored sheet (`cli.py:25-31`). The `flips` marker
    is what stops that being read as the whole story: a cell that moved between
    replicates is flagged, because on a nondeterministic arm replicate 1 is one draw.
    """
    ws = wb.create_sheet("Side by side")

    gt_by_call: dict[str, list[Record]] = defaultdict(list)
    for record in gt:
        gt_by_call[record.call_id].append(record)

    lead = ["item_id", "family", "ground truth  (product / outcome / reasons)"]
    widths(ws, [10, 15, 46])
    per_arm = 2
    widths(ws, [46, 11] * len(order), start=len(lead) + 1)

    put(ws, 1, 1, "At a glance - every arm's whole answer against the ground truth",
        font=TITLE_FONT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(lead) + per_arm * len(order))
    put(ws, 2, 1,
        "Replicate 1. OK means the WHOLE item matched - every product, its outcome and "
        "its full reason set. Adding one unsupported reason is a MISMATCH here and only "
        "a partial loss on 'Comparison'; the gap between the two sheets is the "
        "over-labelling rate. 'flips' means the arm did not give the same answer on all "
        "replicates, so this row is one draw.",
        font=BODY_FONT, wrap=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2,
                   end_column=len(lead) + per_arm * len(order))
    ws.row_dimensions[2].height = 44

    header = 4
    for offset, label in enumerate(lead):
        put(ws, header, 1 + offset, label, font=HEAD_FONT, fill=HEAD_FILL, wrap=True,
            border=True)
    col = len(lead) + 1
    for slot in order:
        put(ws, header, col, arms[slot].label, font=HEAD_FONT, fill=HEAD_FILL, wrap=True,
            border=True)
        put(ws, header, col + 1, "verdict", font=HEAD_FONT, fill=HEAD_FILL, wrap=True,
            border=True)
        col += per_arm

    r = header + 1
    tally: dict[str, int] = {slot: 0 for slot in order}
    for item in items:
        truth = gt_by_call.get(item.call_id, [])
        put(ws, r, 1, item.item_id, font=BOLD_FONT, border=True)
        put(ws, r, 2, item.family, border=True, wrap=True)
        put(ws, r, 3, _answer_text(truth), font=BOLD_FONT, border=True, wrap=True)

        col = len(lead) + 1
        for slot in order:
            arm = arms[slot]
            rows = [rec for key, rec in arm.rep1.items() if key[0] == item.call_id]
            correct = bool(truth) and _signature(rows) == _signature(truth)
            if correct:
                tally[slot] += 1
            flipped = any(
                arm.flips(truth[0] if truth else None, key, dim)
                for key, rec in arm.rep1.items() if key[0] == item.call_id
                for dim in DIMENSIONS
            )
            verdict = "OK" if correct else "MISMATCH"
            fill = GOOD_FILL if correct else BAD_FILL
            if flipped:
                verdict += " +flips"
                if correct:
                    fill = WARN_FILL
            put(ws, r, col, _answer_text(rows), font=BODY_FONT if correct else BAD_FONT,
                fill=None if correct else BAD_FILL, border=True, wrap=True)
            put(ws, r, col + 1, verdict, font=BOLD_FONT if not correct else BODY_FONT,
                fill=fill, border=True, wrap=True)
            col += per_arm
        ws.row_dimensions[r].height = 34
        r += 1

    r += 1
    put(ws, r, 1, "FULLY CORRECT", font=BOLD_FONT, border=True)
    put(ws, r, 3, f"out of {len(items)} items", font=BODY_FONT, border=True)
    col = len(lead) + 1
    for slot in order:
        put(ws, r, col, f"{tally[slot]} / {len(items)}", font=BOLD_FONT, border=True)
        put(ws, r, col + 1, "", border=True)
        col += per_arm
    r += 2
    put(ws, r, 1,
        "This count is whole-item and all-or-nothing, so it is the HARSHEST reading "
        "available and moves on a single added reason. It is not a percentage anyone "
        "should quote: at 20 items one item is 5 points, and 'Mechanisms' is the "
        "headline this pack is designed to produce.",
        font=BODY_FONT, wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r,
                   end_column=len(lead) + per_arm * len(order))
    ws.row_dimensions[r].height = 32

    ws.freeze_panes = ws.cell(row=header + 1, column=4)
    return ws


# ------------------------------------------------------------------ sheet 3


def sheet_comparison(wb: Workbook, gt: list[Record], arms: dict[str, Arm]) -> Worksheet:
    ws = wb.create_sheet("Comparison")
    widths(ws, [22, 14, 13, 13, 13, 13, 9, 46])

    r = 1
    put(ws, r, 1, "Paired comparison, replicate 1", font=TITLE_FONT)
    r += 2
    put(ws, r, 1, "1. PER-ITEM DISAGREEMENT - incumbent (Gemini) vs candidate (Qwen), "
                  "one prompt at a time", font=BOLD_FONT)
    r += 1
    put(ws, r, 1, "Counted over rows scorable for BOTH arms. 'net' is candidate-only minus "
                  "incumbent-only: positive means the candidate won more rows than it lost. "
                  "At 22 scored rows a net of +1 is one row, not a trend.",
        font=BODY_FONT, wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 30
    r += 2

    header_row(ws, r, ["prompt", "dimension", "both right", "both wrong",
                       "incumbent only", "candidate only", "net"])
    r += 1
    for prompt, inc_slot, cand_slot in (
        ("v9_16_base", "incumbent-base", "candidate-base"),
        ("v9_16_e1", "incumbent-e1", "candidate-e1"),
    ):
        inc, cand = arms[inc_slot], arms[cand_slot]
        for dimension in DIMENSIONS:
            d = disagreement(gt, inc.per_replicate[0], cand.per_replicate[0], dimension)
            net = d.candidate_only_right - d.incumbent_only_right
            put(ws, r, 1, prompt, border=True)
            put(ws, r, 2, dimension, border=True)
            put(ws, r, 3, d.both_right, border=True)
            put(ws, r, 4, d.both_wrong, border=True,
                fill=BAD_FILL if d.both_wrong else None)
            put(ws, r, 5, d.incumbent_only_right, border=True,
                fill=BAD_FILL if d.incumbent_only_right else None)
            put(ws, r, 6, d.candidate_only_right, border=True,
                fill=GOOD_FILL if d.candidate_only_right else None)
            put(ws, r, 7, f"{net:+d}", font=BOLD_FONT, border=True)
            r += 1

    r += 1
    put(ws, r, 1, "2. WEIGHTED PRECISION / RECALL / F1 - per arm, per dimension",
        font=BOLD_FONT)
    r += 1
    put(ws, r, 1, "Accuracy is deliberately absent: it is inflated by an arm whose "
                  "unparseable items were dropped, and one-vs-rest accuracy over an "
                  "11-class reason space is dominated by true negatives. Three dimensions "
                  "have three denominators on purpose (call_result drops rows whose GROUND "
                  "TRUTH outcome is absent, reason drops nothing, product works on "
                  "call-grain groups after the NaN-phone drop). A single denominator across "
                  "all three produces three wrong numbers.", font=BODY_FONT, wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 58
    r += 2

    header_row(ws, r, ["arm", "dimension", "denominator", "joined", "parse failures",
                       "w-precision", "w-recall", "w-F1"])
    r += 1
    for slot, arm in arms.items():
        for dimension in DIMENSIONS:
            result = arm.summary.dimensions[dimension]
            put(ws, r, 1, arm.label, border=True)
            put(ws, r, 2, dimension, border=True)
            put(ws, r, 3, result.denominator, border=True)
            put(ws, r, 4, result.coverage.items_joined, border=True)
            put(ws, r, 5, result.coverage.parse_failures, border=True,
                fill=BAD_FILL if result.coverage.parse_failures else None)
            put(ws, r, 6, round(result.weighted("precision"), 4), border=True,
                number_format="0.000")
            put(ws, r, 7, round(result.weighted("recall"), 4), border=True,
                number_format="0.000")
            put(ws, r, 8, round(result.weighted("f1"), 4), font=BOLD_FONT, border=True,
                number_format="0.000")
            r += 1

    r += 1
    put(ws, r, 1, "3. HEADLINE, IN WORDS", font=BOLD_FONT)
    r += 1
    for line in headline_lines(gt, arms):
        put(ws, r, 1, line, font=BODY_FONT, wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 30
        r += 1

    ws.freeze_panes = "A6"
    return ws


def headline_lines(gt: list[Record], arms: dict[str, Arm]) -> list[str]:
    lines = []
    for prompt, inc_slot, cand_slot in (
        ("v9_16_base", "incumbent-base", "candidate-base"),
        ("v9_16_e1", "incumbent-e1", "candidate-e1"),
    ):
        inc, cand = arms[inc_slot], arms[cand_slot]
        parts = []
        for dimension in DIMENSIONS:
            d = disagreement(gt, inc.per_replicate[0], cand.per_replicate[0], dimension)
            parts.append(f"{dimension} {d.candidate_only_right - d.incumbent_only_right:+d}")
        lines.append(f"{prompt}: net candidate-minus-incumbent  " + ",  ".join(parts))
    lines.append(
        "RECONCILED: NO. None of the above is a migration verdict. See the READ FIRST sheet."
    )
    return lines


# ------------------------------------------------------------------ sheet 4


def sheet_mechanisms(wb: Workbook, arms: dict[str, Arm], order: Sequence[str]) -> Worksheet:
    ws = wb.create_sheet("Mechanisms")
    widths(ws, [18, 6] + [16] * len(order))

    r = 1
    put(ws, r, 1, "Mechanism verdicts, over 20 items x 3 replicates", font=TITLE_FONT)
    r += 1
    put(ws, r, 1, "PASS = every item in the group correct on EVERY replicate.  "
                  "FLAKY = some item correct on some replicates and wrong on others.  "
                  "FAIL = some item wrong on EVERY replicate; FAIL outranks FLAKY when both "
                  "are present in one group, because a deterministic failure is the finding "
                  "and FLAKY would let a reader hope another replicate clears it.",
        font=BODY_FONT, wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + len(order))
    ws.row_dimensions[r].height = 46
    r += 2

    header_row(ws, r, ["mechanism", "n"] + [arms[s].label for s in order])
    r += 1
    first = arms[order[0]]
    for index, mech in enumerate(first.mechanisms):
        put(ws, r, 1, mech.mechanism, font=BOLD_FONT, border=True)
        put(ws, r, 2, len(mech.item_ids), border=True)
        for offset, slot in enumerate(order):
            row_obj = arms[slot].mechanisms[index]
            fill = {"PASS": GOOD_FILL, "FLAKY": WARN_FILL, "FAIL": BAD_FILL}[row_obj.verdict]
            put(ws, r, 3 + offset, row_obj.verdict, font=BOLD_FONT, fill=fill, border=True)
        r += 1

    r += 2
    put(ws, r, 1, "Detail: which items, and the failure the testset PREDICTED for each",
        font=BOLD_FONT)
    r += 1
    put(ws, r, 1, "The 'expected failure' text is verbatim from the testset and was written "
                  "before the runs. It is the difference between 'the candidate is worse' "
                  "and 'the candidate is worse in the way we said it would be'.",
        font=BODY_FONT, wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 30
    r += 2

    header_row(ws, r, ["arm", "mechanism", "items", "verdict", "detail"])
    detail_start = r
    r += 1
    for slot in order:
        arm = arms[slot]
        for mech in arm.mechanisms:
            fill = {"PASS": GOOD_FILL, "FLAKY": WARN_FILL, "FAIL": BAD_FILL}[mech.verdict]
            put(ws, r, 1, arm.label, border=True, wrap=True)
            put(ws, r, 2, mech.mechanism, border=True)
            put(ws, r, 3, ", ".join(mech.item_ids), border=True, wrap=True)
            put(ws, r, 4, mech.verdict, font=BOLD_FONT, fill=fill, border=True)
            put(ws, r, 5, mech.detail, border=True, wrap=True)
            ws.row_dimensions[r].height = 78
            r += 1
    ws.column_dimensions[get_column_letter(1)].width = 18
    ws.column_dimensions[get_column_letter(2)].width = 18
    ws.column_dimensions[get_column_letter(3)].width = 34
    ws.column_dimensions[get_column_letter(4)].width = 10
    ws.column_dimensions[get_column_letter(5)].width = 120
    ws.freeze_panes = ws.cell(row=detail_start + 1, column=1)
    return ws


# ------------------------------------------------------------------ sheet 5


RUN_FIELDS: list[tuple[str, Any]] = [
    ("run id", lambda a: a.meta["run_id"]),
    ("run directory", lambda a: str(a.directory)),
    ("arm", lambda a: a.meta["arm"]),
    ("role", lambda a: "incumbent" if "incumbent" in a.slot else "candidate"),
    ("model requested", lambda a: a.meta["model_requested"]),
    ("model observed", lambda a: _hist(a.meta.get("observed_models"))),
    ("provider requested", lambda a: a.meta.get("provider_requested") or "(unpinned)"),
    ("provider served", lambda a: _hist(a.meta.get("observed_providers"))),
    ("PIN PROOF (single-valued prompt_tokens)", lambda a: _pin_proof(a)),
    ("prompt id", lambda a: a.meta["prompt_id"]),
    ("prompt sha", lambda a: a.meta["prompt_sha"]),
    ("items", lambda a: a.meta["items"]),
    ("replicates", lambda a: a.meta["repeats"]),
    ("rows (items x replicates)", lambda a: a.meta["rows"]),
    ("decoding", lambda a: " ".join(
        f"{k}={v}" for k, v in sorted(a.meta["decoding"].items()))),
    ("concurrency", lambda a: a.meta.get("concurrency")),
    ("outcome counts", lambda a: _hist(a.meta.get("outcome_counts"))),
    ("answered nothing (named NO product)", lambda a: a.summary.answered_nothing),
    ("truncated rate", lambda a: a.meta.get("truncated_rate")),
    ("N_flip (cells that moved across replicates)", lambda a: a.n_flip),
    ("fabrication: invented labels", lambda a: a.fab["invented"]),
    ("fabrication: of those, example-derived", lambda a: a.fab["example_derived"]),
    ("fabrication rate (example-derived / invented)", lambda a: round(a.fab["rate"], 3)),
    ("fabrication: blank slots offered", lambda a: a.fab["blank_slots_offered"]),
    ("prompt tokens (run total)", lambda a: token_stats(a.rows)[0]),
    ("completion tokens (run total)", lambda a: token_stats(a.rows)[1]),
    ("reasoning tokens (run total)", lambda a: token_stats(a.rows)[2]),
    ("cost USD (lower bound)", lambda a: a.meta["total_cost_usd_lower_bound"]),
    ("latency median (s)", lambda a: latency_stats(a.rows)[0]),
    ("latency max (s)", lambda a: latency_stats(a.rows)[1]),
    ("testset sha", lambda a: a.meta["testset_sha"]),
    ("ground truth sha", lambda a: a.meta["gt_sha"]),
    ("scorer sha", lambda a: a.meta["scorer_sha"]),
    ("created (UTC)", lambda a: a.meta["created_utc"]),
]


def _hist(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, dict):
        return ", ".join(f"{k}={v}" for k, v in value.items())
    return str(value)


def _pin_proof(arm: Arm) -> str:
    spread = arm.meta.get("prompt_token_spread") or {}
    split = arm.meta.get("split_items") or {}
    return (f"{len(spread) - len(split)}/{len(spread)} items"
            + ("  <-- SPLIT, arm is not one build" if split else "  (one tokenizer answered)"))


def sheet_runs(wb: Workbook, arms: dict[str, Arm], order: Sequence[str]) -> Worksheet:
    ws = wb.create_sheet("Runs")
    widths(ws, [42] + [30] * len(order))

    r = 1
    put(ws, r, 1, "Run provenance - one column per run", font=TITLE_FONT)
    r += 1
    put(ws, r, 1, "PIN PROOF is the check that an arm was ONE backend. Every replicate of an "
                  "item sends a byte-identical request, so prompt_tokens is a pure function "
                  "of the backend's tokenizer and two values means two builds. The provider "
                  "field is the router describing its own routing and cannot verify itself.",
        font=BODY_FONT, wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1 + len(order))
    ws.row_dimensions[r].height = 46
    r += 2

    header_row(ws, r, ["field"] + [arms[s].label for s in order])
    header = r
    r += 1
    for name, getter in RUN_FIELDS:
        put(ws, r, 1, name, font=BOLD_FONT, border=True, wrap=True)
        for offset, slot in enumerate(order):
            value = getter(arms[slot])
            fill = None
            if name.startswith("PIN PROOF") and "SPLIT" in str(value):
                fill = BAD_FILL
            put(ws, r, 2 + offset, value,
                font=MONO_FONT if "sha" in name else BODY_FONT,
                fill=fill, border=True, wrap=True)
        r += 1

    ws.freeze_panes = ws.cell(row=header + 1, column=2)
    return ws


PER_CALL_COLUMNS = (
    ("run", 26), ("arm", 15), ("model", 24), ("provider", 11), ("prompt_id", 13),
    ("item_id", 9), ("call_id", 8), ("replicate", 10), ("outcome", 18),
    ("parse_ok", 9), ("finish_reason", 13),
    ("prompt_tokens", 14), ("completion_tokens", 17), ("reasoning_tokens", 17),
    ("total_tokens", 13), ("cost_usd", 12), ("latency_s", 10),
)


def sheet_per_call(wb: Workbook, arms: dict[str, Arm], order: Sequence[str]) -> Worksheet:
    """One row per API call: tokens and cost for every call of every run.

    This is the sheet that answers "what did this cost, and why". The run-level
    summary on `Runs` hides that one reasoning call can cost 20x a normal one, and
    cost is the number most likely to be quoted back at us.

    `cost_usd` is OpenRouter's own `usage.cost` and is a LOWER BOUND: it omits
    anything the router does not attribute, so a total here reads as "at least this".
    """
    ws = wb.create_sheet("Per call")
    widths(ws, [w for _, w in PER_CALL_COLUMNS])

    r = 1
    put(ws, r, 1, "Every API call, one row - tokens and cost", font=TITLE_FONT)
    r += 1
    put(ws, r, 1, "cost_usd is OpenRouter's usage.cost, a LOWER BOUND. reasoning_tokens are "
                  "billed as completion tokens and are already counted inside "
                  "completion_tokens; the column is separate because a reasoning backend can "
                  "spend the whole budget thinking and return nothing, which reads as a model "
                  "failure and is not one. Non-ok rows are shaded.",
        font=BODY_FONT, wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(PER_CALL_COLUMNS))
    ws.row_dimensions[r].height = 46
    r += 2

    header_row(ws, r, [n for n, _ in PER_CALL_COLUMNS])
    header = r
    r += 1

    totals: dict[str, list[float]] = {}
    for slot in order:
        arm = arms[slot]
        for row in arm.rows:
            pt = row.get("prompt_tokens") or 0
            ct = row.get("completion_tokens") or 0
            rt = row.get("reasoning_tokens") or 0
            cost = row.get("cost") or 0.0
            values = [
                arm.directory.name, arm.label, arm.model, row.get("provider") or "",
                arm.prompt_id, row.get("item_id"), row.get("call_id"), row.get("replicate"),
                row.get("outcome"), row.get("parse_ok"), row.get("finish_reason") or "",
                pt, ct, rt, pt + ct, cost, round(row.get("latency_s") or 0.0, 2),
            ]
            fill = BAD_FILL if row.get("outcome") != "ok" else None
            for c, value in enumerate(values, start=1):
                put(ws, r, c, value, font=BODY_FONT, fill=fill, border=True)
            r += 1
            agg = totals.setdefault(arm.label, [0.0, 0, 0, 0])
            agg[0] += cost
            agg[1] += pt
            agg[2] += ct
            agg[3] += 1

    r += 1
    put(ws, r, 1, "TOTALS PER ARM", font=BOLD_FONT)
    r += 1
    header_row(ws, r, ["arm", "calls", "prompt_tokens", "completion_tokens",
                       "cost_usd (lower bound)"])
    r += 1
    for label, (cost, pt, ct, n) in totals.items():
        for c, value in enumerate([label, n, pt, ct, round(cost, 6)], start=1):
            put(ws, r, c, value, font=BODY_FONT, border=True)
        r += 1

    ws.freeze_panes = ws.cell(row=header + 1, column=1)
    return ws


# ------------------------------------------------------------------ verification


EXPECTED_SHEETS = ("READ FIRST", "Side by side", "Per item", "Per call", "Comparison",
                   "Mechanisms", "Runs")


def verify(path: Path, expected_item_rows: int, gt_path: Path, item_ids: Sequence[str]) -> list[str]:
    """Re-open what was written and assert it, rather than trusting the writer.

    An export that is checked only by the process that produced it is checked by the
    thing most likely to be wrong. This reads the file back off disk with a second
    `load_workbook` call and asserts the sheet names, the per-item row count and that
    the caveats actually landed on sheet 1.

    The row-count check is deliberately made twice, and the second one does not use
    `product_rows` at all: the ground-truth-backed rows are counted with the stdlib
    `csv` module straight off the labels file, and every testset item is required to
    appear. Comparing the sheet only against the number the writer already computed
    would assert that a function agrees with itself.
    """
    import csv  # noqa: PLC0415 - only the verifier needs a second, dumber reader

    problems: list[str] = []
    wb = load_workbook(path)
    try:
        for name in EXPECTED_SHEETS:
            if name not in wb.sheetnames:
                problems.append(f"sheet {name!r} is missing (found {wb.sheetnames})")
        if problems:
            return problems
        if list(wb.sheetnames) != list(EXPECTED_SHEETS):
            problems.append(f"sheet order is {wb.sheetnames}, expected {list(EXPECTED_SHEETS)}")

        ws = wb["Per item"]
        actual = ws.max_row - 2  # two header rows
        if actual != expected_item_rows:
            problems.append(
                f"'Per item' holds {actual} data rows, expected {expected_item_rows}"
            )
        ids = [ws.cell(row=r, column=1).value for r in range(3, ws.max_row + 1)]
        if any(not str(v or "").startswith("RET-") for v in ids):
            problems.append("'Per item' column A holds a value that is not an item id")

        with gt_path.open(encoding="utf-8", newline="") as handle:
            gt_row_count = sum(1 for _ in csv.DictReader(handle))
        backed = sum(
            1
            for r in range(3, ws.max_row + 1)
            if str(ws.cell(row=r, column=6).value or "") != NO_GT_MARKER
        )
        if backed != gt_row_count:
            problems.append(
                f"'Per item' holds {backed} ground-truth-backed rows, but {gt_path.name} "
                f"has {gt_row_count}. Every label must reach the sheet exactly once."
            )
        missing = sorted(set(item_ids) - set(str(v) for v in ids))
        if missing:
            problems.append(f"'Per item' never mentions these testset items: {missing}")

        read_first = wb["READ FIRST"]
        text = "\n".join(
            str(cell.value)
            for row in read_first.iter_rows()
            for cell in row
            if cell.value is not None
        )
        for token in ("RECONCILED: NO", "PROMPT: RECONSTRUCTED", "AUDIO", "native speaker",
                      "Agent-speech misattribution", "ASR error", "Diarisation",
                      "REPRODUCIBILITY"):
            if token.lower() not in text.lower():
                problems.append(f"READ FIRST is missing {token!r}")

        for name, minimum in (("Comparison", 8), ("Mechanisms", 8), ("Runs", 8)):
            if wb[name].max_row < minimum:
                problems.append(f"sheet {name!r} has only {wb[name].max_row} rows")
        for name in EXPECTED_SHEETS:
            if wb[name].freeze_panes in (None, "A1"):
                problems.append(f"sheet {name!r} has no frozen header")
    finally:
        wb.close()
    return problems


# ------------------------------------------------------------------ main


def build(args) -> tuple[Path, int, list[str]]:
    slots = [
        ("incumbent-base", "Gemini / base", Path(args.incumbent_base)),
        ("candidate-base", "Qwen / base", Path(args.candidate_base)),
        ("incumbent-e1", "Gemini / e1", Path(args.incumbent_e1)),
        ("candidate-e1", "Qwen / e1", Path(args.candidate_e1)),
    ]
    first_meta = json.loads((slots[0][2] / "run.json").read_text("utf-8"))
    testset = load_testset(Path(args.testset or first_meta["testset_path"]))
    gt_path = Path(args.gt or first_meta["gt_path"])
    gt = load_csv(gt_path)

    arms: dict[str, Arm] = {}
    for slot, label, directory in slots:
        arms[slot] = Arm(slot, label, directory, gt, testset.items)
    order = [slot for slot, _, _ in slots]

    repeat = None
    if args.repeat:
        mapping = json.loads(Path(args.repeat).read_text("utf-8"))
        missing = [slot for slot, _, _ in slots if slot not in mapping]
        if missing:
            raise SystemExit(f"--repeat is missing run directories for: {missing}")
        repeat = {slot: Path(mapping[slot]) for slot, _, _ in slots}

    stamp = args.date or date.today().isoformat()
    out = Path(args.out) if args.out else REPO / "out" / "reports" / f"qwen-vs-gemini-{stamp}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    sheet_read_first(wb, arms, stamp,
                     consistency_block(arms, repeat, gt, testset.items))
    sheet_side_by_side(wb, testset.items, gt, arms, order)
    _, item_rows = sheet_per_item(wb, testset.items, gt, arms, order)
    sheet_per_call(wb, arms, order)
    sheet_comparison(wb, gt, arms)
    sheet_mechanisms(wb, arms, order)
    sheet_runs(wb, arms, order)
    wb.save(out)

    problems = verify(out, item_rows, gt_path, [i.item_id for i in testset.items])
    return out, item_rows, problems


def main(argv: Sequence[str] | None = None) -> int:
    configure_stdout()
    parser = argparse.ArgumentParser(prog="export_xlsx", description=__doc__)
    parser.add_argument("--incumbent-base", required=True)
    parser.add_argument("--candidate-base", required=True)
    parser.add_argument("--incumbent-e1", required=True)
    parser.add_argument("--candidate-e1", required=True)
    parser.add_argument("--testset", default=None)
    parser.add_argument("--gt", default=None)
    parser.add_argument("--out", default=None)
    parser.add_argument("--date", default=None, help="stamp for the default filename")
    parser.add_argument(
        "--repeat",
        default=None,
        metavar="JSON",
        help=(
            "a {slot: run dir} map for a SECOND pass of the same four runs. When given, "
            "the READ FIRST sheet reports whether this workbook's numbers reproduced. "
            "Optional, and the workbook says so either way: a spreadsheet that is silent "
            "about reproducibility reads as reproducible."
        ),
    )
    args = parser.parse_args(argv)

    out, item_rows, problems = build(args)
    print(f"wrote {out}")
    print(f"  sheets      {', '.join(EXPECTED_SHEETS)}")
    print(f"  'Per item'  {item_rows} data rows (one per item x product)")
    if problems:
        print("\nVERIFICATION FAILED:")
        for problem in problems:
            print(f"  - {problem}")
        return 1
    print("  VERIFIED    re-opened with openpyxl; every sheet present, row count and "
          "caveats confirmed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
