from __future__ import annotations

from dataclasses import replace
import hashlib
import json
import sys
from pathlib import Path
from types import SimpleNamespace

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from evalgen import cli  # noqa: E402
from evalgen.cli import EXIT_OK, EXIT_REFUSED, _write_experiment_xlsx, main
from evalgen.experiments import (
    canonical_sha,
    decision,
    item_stability,
    load_plan,
    logical_call_budget,
    operational_summary,
    projected_execution_budget,
    qualification,
    qualification_contract_sha,
    reliability_gate,
    runtime_gate,
    stability_disagreement,
    validate_plan,
)
from evalgen.prompts import validate_manifest
from evalgen.request import build_request
from evalgen.runner import ItemResult, RunConfig, RunResult
from evalgen.testsets import load_testset
from evalharness.compare import Disagreement, PairedVerdict, exact_band, paired_verdict

PLAN = ROOT / "experiments" / "retention-e5.plan.json"
EVIDENCE = ROOT / "experiments" / "evidence" / "retention-e5"


def _row(
    item: str,
    replicate: int,
    *,
    outcome: str = "ok",
    payload: dict | None = None,
    model: str | None = "qwen/qwen3.6-27b",
    provider: str | None = "CoreWeave",
    prompt_tokens: int | None = 100,
    reasoning_tokens: int | None = 0,
    http_status: int | None = None,
) -> ItemResult:
    if payload is None and outcome == "ok":
        payload = {"product": {"Postpaid": {"retention_outcome": "save"}}}
    return ItemResult(
        item_id=item,
        call_id=f"5{int(item.split('-')[1]):03d}",
        replicate=replicate,
        outcome=outcome,
        parse_ok=outcome == "ok",
        truncated=False,
        payload=payload,
        repairs=(),
        observed_model=model,
        provider=provider,
        generation_id=f"g-{item}-{replicate}",
        finish_reason="stop",
        raw_content="{}" if model else None,
        latency_s=0.25,
        prompt_tokens=prompt_tokens,
        completion_tokens=20 if model else None,
        reasoning_tokens=reasoning_tokens,
        cost=0.001 if model else None,
        error=None if model else "transport",
        http_status=http_status,
    )


def _run(*rows: ItemResult, repeats: int = 2) -> RunResult:
    return RunResult(
        config=RunConfig(
            model="qwen/qwen3.6-27b",
            arm="qwen3.6-27b",
            prompt_id="v9_16_base",
            repeats=repeats,
            provider="CoreWeave",
            reasoning_effort="none",
            max_attempts=1,
        ),
        prompt_sha="prompt",
        testset_sha="testset",
        results=tuple(rows),
    )


def _qualified_run() -> RunResult:
    rows = [
        _row(item, replicate, prompt_tokens=100 + index)
        for index, item in enumerate(("RET-01", "RET-109", "RET-138"))
        for replicate in (1, 2)
    ]
    return _run(*rows)


def _lock_with_fake_qualifications(plan, directory: Path):
    """Replace live evidence paths with a complete deterministic test evidence set."""
    plan["status"] = "locked"
    draft_plan_sha = "a" * 64
    provider_results = 0
    for arm_index, arm in enumerate(plan["arms"]):
        evidence = {}
        selected_artifact = None
        for provider_index, provider in enumerate(arm["provider_candidates"]):
            artifact = {
                "schema_version": 1,
                "experiment_id": plan["experiment_id"],
                "plan_sha": draft_plan_sha,
                "arm": arm["id"],
                "model": arm["model"],
                "provider": provider,
                "qualification_contract_sha": qualification_contract_sha(
                    plan, arm_id=arm["id"], provider=provider
                ),
                "run_id": f"qualification-{arm_index}-{provider_index}",
                "qualification": {
                    "status": "QUALIFIED",
                    "passed": True,
                    "calls": 6,
                },
            }
            artifact["qualification_sha"] = canonical_sha(artifact)
            path = directory / f"qualification-{arm_index}-{provider_index}.json"
            path.write_text(json.dumps(artifact), encoding="utf-8")
            evidence[provider] = str(path)
            provider_results += 1
            if provider_index == 0:
                selected_artifact = (provider, path, artifact)
        provider, path, artifact = selected_artifact
        arm.update({
            "availability": "QUALIFIED",
            "selected_provider": provider,
            "qualification_artifact": str(path),
            "qualification_sha": artifact["qualification_sha"],
            "qualification_evidence": evidence,
            "unavailability_evidence": {},
        })
    plan["qualification_evidence_summary"] = {
        "draft_plan_sha": draft_plan_sha,
        "provider_results": provider_results,
        "logical_calls": provider_results * 6,
        "status_counts": {"QUALIFIED": provider_results},
        "raw_model_output_committed": False,
    }
    return plan


def test_committed_experiment_plan_and_prompt_library_are_in_sync():
    plan = load_plan(PLAN)

    assert validate_plan(plan, root=ROOT) == []
    assert validate_manifest() == []
    assert logical_call_budget(plan) == {
        "qualification_max": 108,
        "full": 1242,
        "load": 216,
        "total": 1458,
        "grand_total_max": 1566,
    }
    assert len(canonical_sha(plan)) == 64

    cost = projected_execution_budget(plan, root=ROOT)
    assert cost["full_calls_per_arm"] == 414
    assert cost["load_calls_per_arm"] == 72
    assert cost["grand_maximum_usd"] > cost["full_load_maximum_usd"] > 0


def test_experiment_7_can_repeat_the_pinned_retention_contract():
    plan = json.loads(json.dumps(load_plan(PLAN)))
    plan["experiment_id"] = "retention-e7"
    plan["status"] = "draft"

    assert validate_plan(plan, root=ROOT) == []


def test_enterprise_plan_refuses_an_unregistered_experiment_id():
    plan = json.loads(json.dumps(load_plan(PLAN)))
    plan["experiment_id"] = "retention-e8"
    plan["status"] = "draft"

    problems = validate_plan(plan, root=ROOT)

    assert problems == [
        "experiment_id: expected 'retention-e5' or 'retention-e7', "
        "found 'retention-e8'"
    ]


def test_committed_gate2_execution_and_reports_are_complete_and_untampered():
    approval = json.loads(
        (EVIDENCE / "gate2-approval.json").read_text(encoding="utf-8")
    )
    approval_sha = approval.pop("approval_sha")
    assert approval_sha == canonical_sha(approval)

    evidence = json.loads((EVIDENCE / "execution.json").read_text(encoding="utf-8"))
    evidence_sha = evidence.pop("execution_evidence_sha")
    assert evidence_sha == canonical_sha(evidence)
    assert evidence["plan_sha"] == canonical_sha(load_plan(PLAN))
    assert evidence["gate2_approval_sha"] == approval_sha

    execution = evidence["execution"]
    assert execution == {
        "logical_calls": 1458,
        "full_calls": 1242,
        "load_calls": 216,
        "run_count": 12,
        "reported_cost_usd_lower_bound": pytest.approx(1.507460937),
        "wall_time_s_sum": pytest.approx(1480.425995700003),
        "maximum_api_attempts_per_logical_call": 1,
        "raw_model_output_committed": False,
        "raw_logs_retained_locally": True,
        "reported_cost_is_lower_bound": True,
    }

    runs = evidence["runs"]
    assert sum(run["logical_calls"] for run in runs) == 1458
    assert sum(run["logical_calls"] for run in runs if run["mode"] == "full") == 1242
    assert sum(run["logical_calls"] for run in runs if run["mode"] == "load") == 216
    assert {run["concurrency"] for run in runs if run["mode"] == "load"} == {1, 4, 8}
    assert all(run["max_attempts"] == 1 for run in runs)

    report_dir = EVIDENCE / "report"
    assert {path.name for path in report_dir.iterdir()} == set(
        evidence["reporting"]["files"]
    )
    for name, expected_sha in evidence["reporting"]["files"].items():
        actual_sha = hashlib.sha256((report_dir / name).read_bytes()).hexdigest()
        assert actual_sha == expected_sha
    # These fingerprints identify the historical code that generated the immutable
    # report.  Pin those historical values rather than requiring them to equal the
    # evolving checkout, which would force every later policy fix to rewrite old
    # evidence.  legacy_v1 tests below preserve the decision semantics explicitly.
    assert evidence["reporting"]["report_code_sha256"] == {
        "src/evalgen/cli.py": (
            "9cc3da19f2cbb6ce5bcd24fee0d179771b94feb15a477e773110d92095c3f1ec"
        ),
        "src/evalgen/experiments.py": (
            "828002ad6798895bc9bbf5b50641bacae97a6d519a62189238a63898caa16544"
        ),
    }
    assert all(
        (ROOT / name).is_file()
        for name in evidence["reporting"]["report_code_sha256"]
    )
    assert not list(EVIDENCE.rglob("run.jsonl"))

    summary = json.loads((report_dir / "summary.json").read_text(encoding="utf-8"))
    assert summary["reconciled"] is False
    assert summary["comparisons"]["qwen3.6-27b"]["decision"]["status"] == "FAIL"
    assert summary["comparisons"]["qwen3.6-35b-a3b"]["decision"]["status"] == "FAIL"


def test_locked_plan_requires_untampered_qualified_artifacts(tmp_path):
    plan = _lock_with_fake_qualifications(
        json.loads(json.dumps(load_plan(PLAN))), tmp_path
    )

    assert validate_plan(plan, root=ROOT) == []

    first = Path(plan["arms"][0]["qualification_artifact"])
    tampered = json.loads(first.read_text(encoding="utf-8"))
    tampered["qualification"]["status"] = "REQUEST_INCOMPATIBLE"
    first.write_text(json.dumps(tampered), encoding="utf-8")

    problems = validate_plan(plan, root=ROOT)
    assert any("invalid self-hash" in problem for problem in problems)
    assert any("qualification status" in problem for problem in problems)


def test_locked_plan_requires_evidence_for_unselected_candidates(tmp_path):
    plan = _lock_with_fake_qualifications(
        json.loads(json.dumps(load_plan(PLAN))), tmp_path
    )
    arm = plan["arms"][1]
    unselected = arm["provider_candidates"][1]
    del arm["qualification_evidence"][unselected]

    problems = validate_plan(plan, root=ROOT)

    assert any(
        "qualification_evidence must cover every candidate provider" in problem
        for problem in problems
    )


def test_validate_plan_refuses_a_transposed_incumbent_role():
    """Audit finding, 2026-08-07: nothing checked `role` anywhere, not in the always-run
    schema block and not in the locked-status deep audit that otherwise re-verifies
    `selected_provider`/`availability`/`qualification_sha` against evidence. `role`
    decides which arm `cmd_experiment_report` treats as the reference for every paired
    comparison (`cli.py`'s `incumbent_id = next(... if arm.get("role") == "incumbent")`),
    so a transposed role silently flips AHEAD/BEHIND -- and therefore PASS/FAIL -- for
    every dimension with no error anywhere in the pipeline. This is the regression test.
    """
    plan = json.loads(json.dumps(load_plan(PLAN)))
    plan["arms"][0]["role"] = "candidate"  # was "incumbent" -- now there are zero

    problems = validate_plan(plan, root=ROOT)

    assert any("role" in problem and "incumbent" in problem for problem in problems)


def test_validate_plan_refuses_two_arms_both_claiming_incumbent():
    plan = json.loads(json.dumps(load_plan(PLAN)))
    plan["arms"][1]["role"] = "incumbent"  # now there are two

    problems = validate_plan(plan, root=ROOT)

    assert any("role" in problem and "incumbent" in problem for problem in problems)


def test_validate_plan_refuses_a_role_outside_the_closed_set():
    plan = json.loads(json.dumps(load_plan(PLAN)))
    plan["arms"][1]["role"] = "reference"  # not "incumbent" or "candidate"

    problems = validate_plan(plan, root=ROOT)

    assert any("role" in problem for problem in problems)


def test_explicit_reasoning_off_is_in_the_exact_pinned_request():
    request = build_request(
        model="qwen/qwen3.6-27b",
        messages=[{"role": "user", "content": "test"}],
        max_tokens=8000,
        temperature=0.0,
        response_format={"type": "json_schema"},
        provider="Morph",
        reasoning_effort="none",
    )

    assert request["extra_body"]["reasoning"] == {"effort": "none"}
    assert request["extra_body"]["provider"] == {
        "order": ["Morph"],
        "allow_fallbacks": False,
        "require_parameters": True,
    }


def test_provider_default_preserves_historical_request_and_bad_effort_refuses():
    request = build_request(
        model="vendor/model",
        messages=[],
        max_tokens=10,
        temperature=0,
    )
    assert "reasoning" not in request["extra_body"]

    with pytest.raises(ValueError, match="unknown reasoning effort"):
        build_request(
            model="vendor/model",
            messages=[],
            max_tokens=10,
            temperature=0,
            reasoning_effort="off-ish",
        )


@pytest.mark.parametrize(
    ("discordant", "band"), [(0, None), (5, None), (6, 6), (24, 12), (34, 14)]
)
def test_exact_verdict_band_is_calibrated_from_observed_discordance(discordant, band):
    assert exact_band(discordant) == band


def test_paired_verdict_never_calls_no_disagreement_a_tie():
    none = paired_verdict(Disagreement("product", 100, 0, 0, 0))
    ahead = paired_verdict(Disagreement("reason", 80, 0, 0, 6))
    behind = paired_verdict(Disagreement("reason", 80, 0, 6, 0))

    assert none.verdict == "UNDERPOWERED"
    assert ahead.verdict == "AHEAD"
    assert behind.verdict == "BEHIND"


def test_qualification_requires_all_six_calls_identity_usage_and_zero_reasoning():
    result = qualification(
        _qualified_run(),
        expected_model="qwen/qwen3.6-27b",
        expected_provider="CoreWeave",
    )

    assert result.passed
    assert result.status == "QUALIFIED"
    assert result.parse_ok == 6
    assert result.reasons == ()


@pytest.mark.parametrize("http_status", [400, 404, 422])
def test_exact_request_rejection_is_request_incompatible_not_identity(
    http_status,
):
    rows = [
        _row(
            item,
            replicate,
            outcome="transport_error",
            payload=None,
            model=None,
            provider=None,
            prompt_tokens=None,
            reasoning_tokens=None,
            http_status=http_status,
        )
        for item in ("RET-01", "RET-109", "RET-138")
        for replicate in (1, 2)
    ]
    result = qualification(
        _run(*rows),
        expected_model="qwen/qwen3.6-27b",
        expected_provider="Morph",
    )

    assert not result.passed
    assert result.status == "REQUEST_INCOMPATIBLE"
    assert result.http_statuses == {str(http_status): 6}
    assert result.reasons[-1] == (
        f"exact request rejected with HTTP statuses {{'{http_status}': 6}}"
    )


def test_scalar_or_invalid_json_is_schema_incompatible():
    rows = [
        _row(item, replicate, outcome="schema_violation", payload=None)
        for item in ("RET-01", "RET-109", "RET-138")
        for replicate in (1, 2)
    ]
    result = qualification(
        _run(*rows),
        expected_model="qwen/qwen3.6-27b",
        expected_provider="CoreWeave",
    )

    assert result.status == "SCHEMA_INCOMPATIBLE"


def test_qualification_report_reclassifies_paid_rows_without_a_client(
    tmp_path, monkeypatch, capsys
):
    rows = [
        _row(
            item,
            replicate,
            outcome="transport_error",
            payload=None,
            model=None,
            provider=None,
            prompt_tokens=None,
            reasoning_tokens=None,
            http_status=404,
        )
        for item in ("RET-01", "RET-109", "RET-138")
        for replicate in (1, 2)
    ]
    plan = load_plan(PLAN)
    loaded = SimpleNamespace(
        directory=tmp_path,
        meta={
            "experiment_mode": "qualification",
            "experiment_plan_sha": canonical_sha(plan),
            "arm": "qwen3.6-27b",
            "provider_requested": "Morph",
            "model_requested": "qwen/qwen3.6-27b",
        },
        result=_run(*rows),
    )
    monkeypatch.setattr(cli, "load_run", lambda path: loaded)

    code = main(
        [
            "qualification-report",
            "--plan",
            str(PLAN),
            "--run",
            str(tmp_path),
        ]
    )

    assert code == 1
    assert "No model call was made and no key was read." in capsys.readouterr().out
    artifact = json.loads((tmp_path / "qualification.json").read_text())
    assert artifact["qualification"]["status"] == "REQUEST_INCOMPATIBLE"
    recorded_sha = artifact.pop("qualification_sha")
    assert recorded_sha == canonical_sha(artifact)


def test_reasoning_tokens_make_an_otherwise_valid_provider_regime_incompatible():
    run = _qualified_run()
    rows = tuple(replace(row, reasoning_tokens=1) for row in run.results)
    result = qualification(
        replace(run, results=rows),
        expected_model="qwen/qwen3.6-27b",
        expected_provider="CoreWeave",
    )

    assert result.status == "REGIME_INCOMPATIBLE"


def test_reliability_gate_uses_unrounded_counts():
    assert reliability_gate(410, 414)
    assert not reliability_gate(409, 414)


def test_item_level_stability_is_paired_and_not_an_aggregate_flip_count():
    stable_payload = {"x": 1}
    changed_payload = {"x": 2}
    incumbent = _run(
        _row("RET-01", 1, payload=stable_payload),
        _row("RET-01", 2, payload=stable_payload),
        _row("RET-02", 1, payload=stable_payload),
        _row("RET-02", 2, payload=stable_payload),
    )
    candidate = _run(
        _row("RET-01", 1, payload=stable_payload),
        _row("RET-01", 2, payload=changed_payload),
        _row("RET-02", 1, payload=stable_payload),
        _row("RET-02", 2, payload=stable_payload),
    )

    assert item_stability(incumbent) == {"RET-01": True, "RET-02": True}
    assert item_stability(candidate) == {"RET-01": False, "RET-02": True}
    assert stability_disagreement(incumbent, candidate) == Disagreement(
        "stability", 1, 0, 1, 0
    )


def test_operational_summary_retains_missing_cost_and_percentiles():
    run = _qualified_run()
    rows = list(run.results)
    rows[0] = replace(rows[0], cost=None, latency_s=9.0)

    facts = operational_summary(replace(run, results=tuple(rows)))

    assert facts["calls"] == 6
    assert facts["parse_valid_rate"] == 1.0
    assert facts["latency_s"]["max"] == 9.0
    assert facts["cost_usd"]["calls_missing"] == 1


def test_full_runtime_gate_rechecks_identity_reasoning_and_usage_after_qualification():
    run = _qualified_run()
    assert runtime_gate(
        run,
        expected_model="qwen/qwen3.6-27b",
        expected_provider="CoreWeave",
    ) == ()

    changed = replace(
        run,
        results=(replace(run.results[0], reasoning_tokens=None), *run.results[1:]),
    )
    assert "reasoning_tokens" in runtime_gate(
        changed,
        expected_model="qwen/qwen3.6-27b",
        expected_provider="CoreWeave",
    )[0]

    failed = replace(
        run.results[0],
        outcome="transport_error",
        parse_ok=False,
        payload=None,
        observed_model=None,
        provider=None,
        prompt_tokens=None,
        completion_tokens=None,
        reasoning_tokens=None,
        cost=None,
        raw_content=None,
        error="transport",
    )
    assert runtime_gate(
        replace(run, results=(failed, *run.results[1:])),
        expected_model="qwen/qwen3.6-27b",
        expected_provider="CoreWeave",
    ) == ()


def test_decision_applies_quality_before_operations_and_keeps_underpowered_visible():
    underpowered = paired_verdict(Disagreement("product", 100, 0, 0, 0))
    call_result = paired_verdict(Disagreement("call_result", 100, 0, 6, 6))
    reason = paired_verdict(Disagreement("reason", 100, 0, 6, 6))

    result = decision(
        qualification_status="QUALIFIED",
        parse_ok=414,
        total=414,
        quality_verdicts=[call_result, reason, underpowered],
        stability_verdict=paired_verdict(Disagreement("stability", 100, 0, 6, 6)),
    )

    assert result.status == "INCONCLUSIVE"
    assert "product" in result.reasons[0]


def test_decision_does_not_let_an_underpowered_stability_verdict_fall_through_to_pass():
    """Audit finding, 2026-08-07: `decision()` only collected UNDERPOWERED from the
    three quality dimensions. A stability comparison with too few discordant items
    (d < 6 at alpha=1/64, exactly `compare.exact_band`'s documented floor) returned
    verdict='UNDERPOWERED' from the identical `paired_verdict` machinery, but nothing
    routed that to INCONCLUSIVE -- with three clean quality verdicts and no BEHIND
    anywhere, execution fell through to `Decision('PASS', ...)` despite there being no
    statistical evidence that stability held. Reproduced directly before the fix: this
    exact input returned PASS.
    """
    quality = [
        paired_verdict(Disagreement(dimension, 100, 0, 6, 6))
        for dimension in ("call_result", "reason", "product")
    ]
    underpowered_stability = paired_verdict(Disagreement("stability", 90, 0, 1, 1))  # d=2
    assert underpowered_stability.verdict == "UNDERPOWERED"

    result = decision(
        qualification_status="QUALIFIED",
        parse_ok=414,
        total=414,
        quality_verdicts=quality,
        stability_verdict=underpowered_stability,
    )

    assert result.status == "INCONCLUSIVE"
    assert any("stability" in reason for reason in result.reasons)


def test_decision_still_fails_on_a_behind_stability_verdict_ahead_of_underpowered():
    """The existing BEHIND-stability path must still win over the new UNDERPOWERED
    check -- a candidate that is both unstable AND underpowered elsewhere is a FAIL,
    not an INCONCLUSIVE, and this must not regress when the branch above it changes."""
    quality = [
        paired_verdict(Disagreement(dimension, 100, 0, 6, 6))
        for dimension in ("call_result", "reason", "product")
    ]
    behind_stability = paired_verdict(Disagreement("stability", 90, 30, 33, 3))

    result = decision(
        qualification_status="QUALIFIED",
        parse_ok=414,
        total=414,
        quality_verdicts=quality,
        stability_verdict=behind_stability,
    )

    assert result.status == "FAIL"
    assert any("stability" in reason.lower() for reason in result.reasons)


def test_decision_refuses_missing_duplicate_and_unexpected_dimensions():
    call_result = paired_verdict(Disagreement("call_result", 100, 0, 6, 6))
    duplicate = paired_verdict(Disagreement("call_result", 100, 0, 6, 6))
    unexpected = paired_verdict(Disagreement("sentiment", 100, 0, 6, 6))

    result = decision(
        qualification_status="QUALIFIED",
        parse_ok=414,
        total=414,
        quality_verdicts=[call_result, duplicate, unexpected],
        stability_verdict=paired_verdict(Disagreement("stability", 100, 0, 6, 6)),
    )

    assert result.status == "INCONCLUSIVE"
    joined = " ".join(result.reasons)
    assert "duplicate" in joined
    assert "missing" in joined
    assert "unexpected" in joined


def test_decision_treats_runtime_identity_or_provenance_defects_as_inconclusive():
    quality = [
        paired_verdict(Disagreement(dimension, 100, 0, 6, 6))
        for dimension in ("call_result", "reason", "product")
    ]

    result = decision(
        qualification_status="QUALIFIED",
        parse_ok=414,
        total=414,
        quality_verdicts=quality,
        stability_verdict=paired_verdict(Disagreement("stability", 100, 0, 6, 6)),
        runtime_problems=("observed model does not match the locked plan",),
    )

    assert result.status == "INCONCLUSIVE"
    assert "invalid candidate runtime/provenance" in result.reasons[0]


def test_decision_does_not_pass_an_indistinguishable_material_net_loss():
    worse = paired_verdict(Disagreement("call_result", 80, 0, 15, 5))
    assert worse.verdict == "INDISTINGUISHABLE"
    level = {
        dimension: paired_verdict(Disagreement(dimension, 80, 0, 10, 10))
        for dimension in ("reason", "product", "stability")
    }

    strict = decision(
        qualification_status="QUALIFIED",
        parse_ok=414,
        total=414,
        quality_verdicts=[worse, level["reason"], level["product"]],
        stability_verdict=level["stability"],
    )
    legacy = decision(
        qualification_status="QUALIFIED",
        parse_ok=414,
        total=414,
        quality_verdicts=[worse, level["reason"], level["product"]],
        stability_verdict=level["stability"],
        policy="legacy_v1",
    )

    assert strict.status == "INCONCLUSIVE"
    assert "non-inferiority" in strict.reasons[0]
    assert legacy.status == "PASS"


def test_legacy_policy_reproduces_the_committed_experiment_5_decision():
    comparison = json.loads(
        (
            EVIDENCE
            / "report"
            / "comparison-gemini-2.5-flash-vs-qwen3.6-27b.json"
        ).read_text(encoding="utf-8")
    )

    result = decision(
        qualification_status="QUALIFIED",
        parse_ok=359,
        total=414,
        quality_verdicts=[
            PairedVerdict(**row) for row in comparison["slices"]["full"]
        ],
        stability_verdict=PairedVerdict(**comparison["stability"]),
        reference_parse_ok=413,
        reference_total=414,
        policy="legacy_v1",
    )

    assert result.status == comparison["decision"]["status"]
    assert list(result.reasons) == comparison["decision"]["reasons"]


def test_decision_net_loss_margin_is_explicit_and_dimension_specific():
    quality = [
        paired_verdict(Disagreement("call_result", 80, 0, 15, 5)),
        paired_verdict(Disagreement("reason", 80, 0, 10, 10)),
        paired_verdict(Disagreement("product", 80, 0, 10, 10)),
    ]

    result = decision(
        qualification_status="QUALIFIED",
        parse_ok=414,
        total=414,
        quality_verdicts=quality,
        stability_verdict=paired_verdict(
            Disagreement("stability", 80, 0, 10, 10)
        ),
        maximum_net_losses={"call_result": 10},
    )

    assert result.status == "PASS"


@pytest.mark.parametrize(
    "forged",
    [
        lambda row: replace(row, verdict="AHEAD"),
        lambda row: replace(row, band=(row.band or 0) + 2),
        lambda row: replace(row, alpha_per_side=0.10),
        lambda row: replace(row, discordant=True),
        lambda row: replace(row, net=0.0),
    ],
)
def test_decision_recomputes_and_refuses_forged_paired_evidence(forged):
    canonical = {
        dimension: paired_verdict(Disagreement(dimension, 80, 0, 10, 10))
        for dimension in ("call_result", "reason", "product", "stability")
    }
    quality = [
        forged(canonical["call_result"]),
        canonical["reason"],
        canonical["product"],
    ]

    result = decision(
        qualification_status="QUALIFIED",
        parse_ok=414,
        total=414,
        quality_verdicts=quality,
        stability_verdict=canonical["stability"],
    )

    assert result.status == "INCONCLUSIVE"
    assert any(
        token in " ".join(result.reasons)
        for token in ("canonical", "alpha_per_side", "integers")
    )


def test_decision_expected_dimensions_are_configurable_for_a_future_application():
    intent = paired_verdict(Disagreement("intent", 80, 0, 10, 10))

    result = decision(
        qualification_status="QUALIFIED",
        parse_ok=100,
        total=100,
        quality_verdicts=[intent],
        stability_verdict=paired_verdict(
            Disagreement("stability", 80, 0, 10, 10)
        ),
        expected_dimensions=("intent",),
    )

    assert result.status == "PASS"


def test_experiment_workbook_carries_quality_regressions_and_load(tmp_path):
    latency = {"p50": 1.0, "p95": 2.0, "p99": 3.0, "max": 4.0}
    cost = {"lower_bound": 0.1, "calls_reported": 6, "calls_missing": 0}
    facts = {
        "parse_ok": 6,
        "calls": 6,
        "parse_valid_rate": 1.0,
        "latency_s": latency,
        "cost_usd": cost,
        "throughput_calls_per_s": 2.0,
    }
    summary = {
        "experiment_id": "retention-e5",
        "plan_sha": "a" * 64,
        "arms": {"gemini": facts, "qwen": facts},
        "comparisons": {
            "qwen": {
                "decision": {"status": "PASS", "reasons": ["passed"]},
                "slices": {
                    "full": [{
                        "dimension": "reason", "discordant": 6, "net": 6,
                        "band": 6, "verdict": "AHEAD",
                    }]
                },
                "regressions": [{
                    "item_key": "hashed", "dimension": "product",
                    "gt_label": "TOL", "incumbent_label": "TOL",
                    "candidate_label": "Postpaid", "error_type": "wrong_label",
                }],
            }
        },
        "load": {"gemini": {"1": facts}, "qwen": {"1": facts}},
    }
    path = tmp_path / "summary.xlsx"

    _write_experiment_xlsx(summary, path)

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True)
    assert workbook.sheetnames == ["Overview", "Quality", "Regressions", "Load"]
    assert workbook["Regressions"]["B2"].value == "hashed"


def test_cli_checks_plan_offline_and_refuses_full_calls_while_draft(
    tmp_path, capsys
):
    plan = load_plan(PLAN)

    assert main(["experiment-check", "--plan", str(PLAN)]) == EXIT_OK
    assert main(["experiment-budget", "--plan", str(PLAN)]) == EXIT_OK
    draft = json.loads(json.dumps(plan))
    draft["status"] = "draft"
    draft_path = tmp_path / "draft.plan.json"
    draft_path.write_text(json.dumps(draft), encoding="utf-8")
    assert main([
        "experiment-run",
        "--plan", str(draft_path),
        "--arm", "qwen3.6-27b",
        "--confirm-plan-sha", canonical_sha(draft),
    ]) == EXIT_REFUSED
    output = capsys.readouterr()
    assert "not 'locked'" in output.err
    assert "grand maximum" in output.out


def test_locked_experiment_runs_and_reports_are_reproducible_without_network(
    monkeypatch, tmp_path, capsys
):
    """The complete paid-call shape, driven by deterministic local completions.

    This is deliberately larger than a parser smoke test: three 414-call full arms and
    nine 24-call load runs pass through the real command gates, runner, logs, loader,
    scorers, paired verdicts and JSON/Markdown/XLSX writers. If those pieces drift apart,
    this is where a plausible-looking final report must fail before an invoice exists.
    """
    plan = _lock_with_fake_qualifications(
        json.loads(json.dumps(load_plan(PLAN))), tmp_path
    )

    plan_path = tmp_path / "locked.plan.json"
    plan_path.write_text(json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8")
    assert validate_plan(plan, root=ROOT) == []
    plan_sha = canonical_sha(plan)

    testset = load_testset(ROOT / plan["assets"]["testset"]["path"])
    by_transcript = {item.transcript_th: item for item in testset.items}

    class DeterministicClient:
        def complete(
            self,
            *,
            model,
            messages,
            max_tokens,
            temperature,
            top_p=None,
            seed=None,
            response_format=None,
            provider=None,
            reasoning_effort="provider-default",
        ):
            assert reasoning_effort == "none"
            item = by_transcript[messages[1]["content"]]
            payload = {
                "product": {
                    row["product"]: {
                        "main": {"reason": row.get("main") or "", "keyword": ""},
                        "secondary": {
                            "reason": row.get("secondary") or "",
                            "keyword": "",
                        },
                        "third": {"reason": row.get("third") or "", "keyword": ""},
                        "retention_outcome": row.get("call_result") or "",
                    }
                    for row in item.gt
                },
                "call_event_detection": "Emerging or Undefined Events",
                "recommendation": "retain the exact current offer",
            }
            return SimpleNamespace(
                content=json.dumps(payload, ensure_ascii=False),
                finish_reason="stop",
                observed_model=model,
                generation_id=f"fake-{model}-{item.item_id}",
                provider=provider,
                prompt_tokens=1000 + len(item.transcript_th),
                completion_tokens=100,
                    reasoning_tokens=0,
                    cost=0.0001,
                    latency_s=0.001,
                    runtime_id="openrouter",
                    runtime_backend="openrouter",
                    runtime_fingerprint=cli.OPENROUTER_RUNTIME.fingerprint(),
                    system_fingerprint="fake-openrouter-build",
                )

    monkeypatch.setenv("OPENROUTER_API_KEY", "sk-fake-not-a-real-key")
    monkeypatch.setenv("EVAL_HARNESS_KEY_HMAC", "deterministic-report-key")
    monkeypatch.setattr(cli, "build_client", lambda *args, **kwargs: DeterministicClient())

    run_root = tmp_path / "runs"
    full_runs: dict[str, Path] = {}
    load_runs: dict[str, Path] = {}

    def execute_and_find(arguments: list[str]) -> Path:
        before = set(run_root.iterdir()) if run_root.exists() else set()
        assert main(arguments) == EXIT_OK
        created = sorted(set(run_root.iterdir()) - before)
        assert len(created) == 1
        capsys.readouterr()
        return created[0]

    for arm in plan["arms"]:
        common = [
            "experiment-run",
            "--plan", str(plan_path),
            "--arm", arm["id"],
            "--confirm-plan-sha", plan_sha,
            "--out", str(run_root),
        ]
        full_runs[arm["id"]] = execute_and_find(common)
        for concurrency in (1, 4, 8):
            load_runs[f"{arm['id']}@{concurrency}"] = execute_and_find(
                common + [
                    "--mode", "load",
                    "--concurrency-level", str(concurrency),
                ]
            )

    def report_arguments(output: Path, *, include_load: bool = True) -> list[str]:
        arguments = [
            "experiment-report", "--plan", str(plan_path), "--out", str(output)
        ]
        for arm_id, path in full_runs.items():
            arguments += ["--run", f"{arm_id}={path}"]
        if include_load:
            for name, path in load_runs.items():
                arguments += ["--load-run", f"{name}={path}"]
        return arguments

    first = tmp_path / "report-first"
    second = tmp_path / "report-second"
    assert main(report_arguments(first)) == EXIT_OK
    capsys.readouterr()
    assert main(report_arguments(second)) == EXIT_OK
    capsys.readouterr()

    assert (first / "summary.json").read_bytes() == (second / "summary.json").read_bytes()
    assert (first / "summary.md").read_bytes() == (second / "summary.md").read_bytes()
    summary = json.loads((first / "summary.json").read_text(encoding="utf-8"))
    assert set(summary["arms"]) == {arm["id"] for arm in plan["arms"]}
    assert set(summary["comparisons"]) == {"qwen3.6-27b", "qwen3.6-35b-a3b"}
    assert all(
        comparison["decision"]["status"] == "INCONCLUSIVE"
        for comparison in summary["comparisons"].values()
    ), "zero discordant pairs is UNDERPOWERED, never silently called a tie/pass"
    assert all(
        facts["runtime_gate_problems"] == [] for facts in summary["arms"].values()
    )
    assert all(
        set(levels) == {"1", "4", "8"} for levels in summary["load"].values()
    )

    full_only = tmp_path / "report-full-only"
    assert main(report_arguments(full_only, include_load=False)) == EXIT_OK
    capsys.readouterr()
    full_only_summary = json.loads(
        (full_only / "summary.json").read_text(encoding="utf-8")
    )
    assert all(levels == {} for levels in full_only_summary["load"].values())

    from openpyxl import load_workbook

    workbook = load_workbook(first / "summary.xlsx", read_only=True, data_only=True)
    assert workbook.sheetnames == ["Overview", "Quality", "Regressions", "Load"]
    assert workbook["Load"].max_row == 10  # header + 3 arms x 3 concurrency levels
    for arm in plan["arms"]:
        assert (first / f"arm-{arm['id']}.json").is_file()
        assert (first / f"arm-{arm['id']}.md").is_file()


# --- the three experiment-run safety gates, exercised with values that FAIL ----------
#
# Recorded as a High-priority gap on 2026-08-08 and still open on 2026-08-12: all three
# gates below existed and were reachable, but every test drove them with a value that
# PASSES. `--confirm-plan-sha` was only ever given the correct sha, `--concurrency-level`
# only ever a member of the allowed list, and the string "UNAVAILABLE" appeared nowhere
# in tests/ at all. A gate whose refusal branch is never taken is indistinguishable from
# an `if` that was deleted, and these three stand between a stale or tampered approval and
# real, paid calls.
#
# Each test installs a client factory that RAISES, so the assertion is not only "it
# refused" but "it refused before it could spend anything" -- the same discipline
# `check` and `--dry-run` are held to in tests/test_cli.py.


def _locked_plan_on_disk(tmp_path):
    """A locked, valid plan plus its sha. Shared by the three gate tests."""
    plan = _lock_with_fake_qualifications(
        json.loads(json.dumps(load_plan(PLAN))), tmp_path
    )
    path = tmp_path / "locked.plan.json"
    path.write_text(json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8")
    assert validate_plan(plan, root=ROOT) == []
    return plan, path, canonical_sha(plan)


@pytest.fixture
def no_calls_allowed(monkeypatch):
    """Any attempt to build a client is a test failure, not a mock."""
    def _explode(*args, **kwargs):
        raise AssertionError("a refused experiment-run must not construct a client")

    monkeypatch.setattr(cli, "build_client", _explode)


def test_experiment_run_refuses_a_confirm_sha_that_does_not_match_the_plan(
    tmp_path, no_calls_allowed, capsys
):
    """The human-approval gate. A stale approval must not execute a changed plan."""
    plan, plan_path, plan_sha = _locked_plan_on_disk(tmp_path)
    stale = "0" * 64
    assert stale != plan_sha

    code = main([
        "experiment-run",
        "--plan", str(plan_path),
        "--arm", plan["arms"][0]["id"],
        "--confirm-plan-sha", stale,
        "--out", str(tmp_path / "runs"),
    ])

    assert code == EXIT_REFUSED
    message = capsys.readouterr().err
    # Both shas must appear: a refusal that hides which plan is current cannot be acted on.
    assert stale in message and plan_sha in message


def _make_arm_validly_unavailable(plan, arm, directory: Path):
    """Turn a QUALIFIED arm into a *legitimately* UNAVAILABLE one.

    Not just `availability = "UNAVAILABLE"`: `validate_plan` (experiments.py:417-436)
    requires unavailability evidence covering every candidate provider, and refuses an
    arm that claims UNAVAILABLE while one of its artifacts says QUALIFIED. Flipping only
    the label is caught by the plan validator long before the run gate -- which is itself
    worth knowing, and is why this helper exists rather than a one-line edit.
    """
    for provider, artifact_path in arm["qualification_evidence"].items():
        path = Path(artifact_path)
        artifact = json.loads(path.read_text(encoding="utf-8"))
        artifact["qualification"] = {
            "status": "REQUEST_INCOMPATIBLE",
            "passed": False,
            "calls": 6,
        }
        artifact.pop("qualification_sha", None)
        artifact["qualification_sha"] = canonical_sha(artifact)
        path.write_text(json.dumps(artifact), encoding="utf-8")

    arm["availability"] = "UNAVAILABLE"
    arm["unavailability_evidence"] = dict(arm["qualification_evidence"])
    arm.pop("selected_provider", None)
    arm.pop("qualification_sha", None)
    arm.pop("qualification_artifact", None)

    counts: dict[str, int] = {}
    for other in plan["arms"]:
        status = (
            "REQUEST_INCOMPATIBLE"
            if other["id"] == arm["id"]
            else "QUALIFIED"
        )
        for _ in other["qualification_evidence"]:
            counts[status] = counts.get(status, 0) + 1
    plan["qualification_evidence_summary"]["status_counts"] = counts
    return plan


def test_experiment_run_refuses_an_arm_locked_as_unavailable(
    tmp_path, no_calls_allowed, capsys
):
    """UNAVAILABLE is a recorded finding, not a transient error to retry past.

    Experiment 5B locked arms this way when no provider could serve the
    production-shaped regime. Executing one anyway would either invent a provider or
    silently fall back to a different one, and the run would look like evidence.
    """
    plan, plan_path, plan_sha = _locked_plan_on_disk(tmp_path)
    _make_arm_validly_unavailable(plan, plan["arms"][0], tmp_path)
    assert validate_plan(plan, root=ROOT) == [], (
        "the plan must be VALID and unavailable, or this test proves the plan validator "
        "rather than the run gate"
    )
    plan_path.write_text(json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8")
    plan_sha = canonical_sha(plan)

    code = main([
        "experiment-run",
        "--plan", str(plan_path),
        "--arm", plan["arms"][0]["id"],
        "--confirm-plan-sha", plan_sha,
        "--out", str(tmp_path / "runs"),
    ])

    assert code == EXIT_REFUSED
    assert "UNAVAILABLE" in capsys.readouterr().err


def test_experiment_run_refuses_a_load_concurrency_outside_the_locked_list(
    tmp_path, no_calls_allowed, capsys
):
    """The load levels are part of what was reviewed and costed, not a free parameter."""
    plan, plan_path, plan_sha = _locked_plan_on_disk(tmp_path)
    allowed = plan["operations"]["concurrency_levels"]
    outside = max(allowed) + 1
    assert outside not in allowed

    code = main([
        "experiment-run",
        "--plan", str(plan_path),
        "--arm", plan["arms"][0]["id"],
        "--confirm-plan-sha", plan_sha,
        "--mode", "load",
        "--concurrency-level", str(outside),
        "--out", str(tmp_path / "runs"),
    ])

    assert code == EXIT_REFUSED
    message = capsys.readouterr().err
    assert str(allowed) in message and str(outside) in message


# --- cmd_qualify, which had no test at all ------------------------------------------
#
# Recorded High on 2026-08-08, still uncovered on 2026-08-12: `cmd_qualify` spends six
# real calls and decides QUALIFIED / REQUEST_INCOMPATIBLE / SCHEMA_INCOMPATIBLE, which is
# the gate between an unvetted endpoint and a paid qualification run -- and no test
# invoked it, through the CLI or otherwise. `qualify` and `severity` were the only two
# subcommands never driven through main().
#
# Its three refusals are tested with a client factory that RAISES, so each asserts the
# stronger property: not merely that it refused, but that it refused before it could
# spend. The fourth test drives the paid path with deterministic completions and checks
# the artifact it writes, because a probe that silently wrote the wrong verdict would be
# worse than one that crashed.


def _draft_plan_on_disk(tmp_path):
    """The committed plan is locked; qualification is only legal while it is not."""
    draft = json.loads(json.dumps(load_plan(PLAN)))
    draft["status"] = "draft"
    path = tmp_path / "draft.plan.json"
    path.write_text(json.dumps(draft, ensure_ascii=False, indent=2), encoding="utf-8")
    return draft, path


def test_qualify_refuses_a_locked_plan_without_spending(
    tmp_path, no_calls_allowed, capsys
):
    """Adding provider evidence to an approved plan is a review, not a probe."""
    code = main([
        "qualify",
        "--plan", str(PLAN),
        "--arm", "gemini-2.5-flash",
        "--provider", "Google",
        "--out", str(tmp_path / "runs"),
    ])
    assert code == EXIT_REFUSED
    assert "locked" in capsys.readouterr().err


def test_qualify_refuses_an_unknown_arm_without_spending(
    tmp_path, no_calls_allowed, capsys
):
    _draft, draft_path = _draft_plan_on_disk(tmp_path)
    code = main([
        "qualify",
        "--plan", str(draft_path),
        "--arm", "no-such-arm",
        "--provider", "Google",
        "--out", str(tmp_path / "runs"),
    ])
    assert code == EXIT_REFUSED
    assert "no-such-arm" in capsys.readouterr().err


def test_qualify_refuses_a_provider_that_was_never_preregistered(
    tmp_path, no_calls_allowed, capsys
):
    """The candidate list is part of what was reviewed.

    Probing an endpoint nobody listed produces evidence for a provider the plan does not
    know about, and the natural next step is to paste it in beside the reviewed ones.
    """
    draft, draft_path = _draft_plan_on_disk(tmp_path)
    candidates = draft["arms"][0]["provider_candidates"]
    assert "Nowhere Inc" not in candidates

    code = main([
        "qualify",
        "--plan", str(draft_path),
        "--arm", draft["arms"][0]["id"],
        "--provider", "Nowhere Inc",
        "--out", str(tmp_path / "runs"),
    ])
    assert code == EXIT_REFUSED
    message = capsys.readouterr().err
    assert "Nowhere Inc" in message and str(candidates) in message


def test_qualify_writes_a_qualified_artifact_from_six_conforming_calls(
    monkeypatch, tmp_path, capsys
):
    """The paid path, driven offline: six calls, one artifact, QUALIFIED.

    Pins the shape the locked-plan gate later depends on --
    `_lock_with_fake_qualifications` fabricates artifacts of exactly this form, so if the
    real writer drifted from it, every locked-plan test in this file would be validating
    a shape production never produces.
    """
    draft, draft_path = _draft_plan_on_disk(tmp_path)
    arm = draft["arms"][0]
    provider = arm["provider_candidates"][0]

    testset = load_testset(ROOT / draft["assets"]["testset"]["path"])
    by_transcript = {item.transcript_th: item for item in testset.items}
    calls: list[str] = []

    class QualifyingClient:
        def complete(
            self, *, model, messages, max_tokens, temperature, top_p=None, seed=None,
            response_format=None, provider=None, reasoning_effort="provider-default",
        ):
            assert reasoning_effort == "none", "qualification pins the regime explicitly"
            item = by_transcript[messages[1]["content"]]
            calls.append(item.item_id)
            payload = {
                "product": {
                    row["product"]: {
                        "main": {"reason": row.get("main") or "", "keyword": ""},
                        "secondary": {
                            "reason": row.get("secondary") or "",
                            "keyword": "",
                        },
                        "third": {"reason": row.get("third") or "", "keyword": ""},
                        "retention_outcome": row.get("call_result") or "",
                    }
                    for row in item.gt
                },
                "call_event_detection": "Emerging or Undefined Events",
                "recommendation": "retain the exact current offer",
            }
            return SimpleNamespace(
                content=json.dumps(payload, ensure_ascii=False),
                finish_reason="stop",
                observed_model=model,
                generation_id=f"qual-{item.item_id}",
                provider=provider,
                # Stable per item and positive: the qualification contract requires both,
                # because a varying count means two backends behind one model id.
                prompt_tokens=1000 + len(item.transcript_th),
                completion_tokens=100,
                reasoning_tokens=0,
                cost=0.0001,
                latency_s=0.001,
                runtime_id="openrouter",
                runtime_backend="openrouter",
                runtime_fingerprint=cli.OPENROUTER_RUNTIME.fingerprint(),
                system_fingerprint="fake-openrouter-build",
            )

    monkeypatch.setenv("OPENROUTER_API_KEY", "sk-fake-not-a-real-key")
    monkeypatch.setenv("EVAL_HARNESS_KEY_HMAC", "qualify-test-key")
    monkeypatch.setattr(cli, "build_client", lambda *a, **k: QualifyingClient())

    run_root = tmp_path / "runs"
    assert main([
        "qualify",
        "--plan", str(draft_path),
        "--arm", arm["id"],
        "--provider", provider,
        "--out", str(run_root),
    ]) == EXIT_OK
    capsys.readouterr()

    # Six logical calls: 3 preregistered items x 2 replicates, and no more.
    expected_calls = (
        len(draft["qualification"]["item_ids"]) * draft["qualification"]["replicates"]
    )
    assert len(calls) == expected_calls == 6
    assert sorted(set(calls)) == sorted(draft["qualification"]["item_ids"])

    created = sorted(run_root.iterdir())
    assert len(created) == 1
    artifact = json.loads(
        (created[0] / "qualification.json").read_text(encoding="utf-8")
    )

    assert artifact["arm"] == arm["id"]
    assert artifact["provider"] == provider
    assert artifact["model"] == arm["model"]
    assert artifact["qualification"]["status"] == "QUALIFIED"
    assert artifact["qualification"]["passed"] is True
    assert artifact["qualification"]["calls"] == 6
    assert artifact["qualification_sha"] == canonical_sha(
        {k: v for k, v in artifact.items() if k != "qualification_sha"}
    )
