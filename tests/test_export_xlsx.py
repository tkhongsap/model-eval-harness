"""`export_xlsx.verify`: the gate that re-opens the workbook, demonstrated FAILING.

`build()` calls `verify()` on every export, so the builder is checked against the gate
each time it runs. That leaves the gate itself unchecked, and a gate that returns `[]`
unconditionally would make every export look verified while asserting nothing -- which
is the exact failure this repository's version-pin gate already had once
(DEVLOG, 2026-08-04: "the one gate this build made a point of demonstrating silently
stopped running once the repo moved").

So each test here breaks the workbook in one specific way and asserts the gate notices.
The control test is the one that makes the others mean anything: if the untouched
workbook did not pass, a failure elsewhere would prove nothing about the fault injected.

The workbook is built synthetically rather than from a run directory on purpose. These
tests are about the gate, not about the sheets; the sheets are proved by the fact that
`build()` cannot write a file without `verify()` accepting it.
"""

from __future__ import annotations

import csv
import importlib.util
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))


def _load_export_module():
    """Import `scripts/export_xlsx.py` by path.

    Not by name: `scripts/` also holds `evalgen.py`, a launcher, so putting that
    directory on `sys.path` shadows the real `evalgen` package with its own entry point
    and `import evalgen.cli` fails with "'evalgen' is not a package".
    """
    spec = importlib.util.spec_from_file_location(
        "_export_xlsx_under_test", ROOT / "scripts" / "export_xlsx.py"
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


export_xlsx = _load_export_module()

ITEM_IDS = [f"RET-{n:02d}" for n in range(1, 4)]
GT_ROWS = [
    {"call_id": "5001", "phone_number": "0810000001", "product": "postpaid",
     "call_result": "save", "main": "network", "secondary": "", "third": ""},
    {"call_id": "5002", "phone_number": "0810000002", "product": "tol",
     "call_result": "churn", "main": "save cost", "secondary": "", "third": ""},
    {"call_id": "5003", "phone_number": "0810000003", "product": "tvs",
     "call_result": "churn", "main": "other", "secondary": "", "third": ""},
]

# One row per ground-truth row, plus one row an arm invented that the GT does not have.
SHEET_ROWS = [
    ("RET-01", "save"),
    ("RET-02", "churn"),
    ("RET-03", "churn"),
    ("RET-03", export_xlsx.NO_GT_MARKER),
]

CAVEAT_TEXT = (
    "RECONCILED: NO / PROMPT: RECONSTRUCTED / production is handed AUDIO and this pack "
    "sends text / the Thai has no native speaker sign-off / not observable: "
    "Agent-speech misattribution, ASR error, Diarisation / REPRODUCIBILITY: NOT CHECKED"
)


@pytest.fixture
def gt_path(tmp_path: Path) -> Path:
    path = tmp_path / "gt.csv"
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(GT_ROWS[0]))
        writer.writeheader()
        writer.writerows(GT_ROWS)
    return path


def _workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for name in export_xlsx.EXPECTED_SHEETS:
        ws = wb.create_sheet(name)
        ws.freeze_panes = "A2"
        if name == "READ FIRST":
            ws["A1"] = CAVEAT_TEXT
        elif name == "Per item":
            ws["A1"], ws["A2"] = "arm banner", "item_id"
            ws["F2"] = "GT call_result"
            for offset, (item_id, call_result) in enumerate(SHEET_ROWS):
                ws.cell(row=3 + offset, column=1, value=item_id)
                ws.cell(row=3 + offset, column=6, value=call_result)
            ws.freeze_panes = "B3"
        else:
            for row in range(1, 11):
                ws.cell(row=row, column=1, value=f"{name} row {row}")
    return wb


@pytest.fixture
def book(tmp_path: Path) -> Path:
    path = tmp_path / "book.xlsx"
    _workbook().save(path)
    return path


def _verify(path: Path, gt_path: Path, *, rows: int = len(SHEET_ROWS), ids=None):
    return export_xlsx.verify(path, rows, gt_path, ITEM_IDS if ids is None else ids)


def test_control_untouched_workbook_passes(book: Path, gt_path: Path):
    """Without this, every other test in the file proves nothing."""
    assert _verify(book, gt_path) == []


def test_missing_sheet_is_caught(book: Path, gt_path: Path, tmp_path: Path):
    wb = load_workbook(book)
    del wb["Comparison"]
    wb.save(book)
    problems = _verify(book, gt_path)
    assert any("Comparison" in p and "missing" in p for p in problems)


def test_reordered_sheets_are_caught(book: Path, gt_path: Path):
    wb = load_workbook(book)
    wb.move_sheet("READ FIRST", offset=2)
    wb.save(book)
    assert any("sheet order" in p for p in _verify(book, gt_path))


def test_wrong_row_count_is_caught(book: Path, gt_path: Path):
    assert any("data rows" in p for p in _verify(book, gt_path, rows=len(SHEET_ROWS) + 1))


def test_a_ground_truth_row_falling_off_the_sheet_is_caught(book: Path, gt_path: Path):
    """The check that does NOT trust the writer's own row count.

    Losing a label and losing an entire row are different bugs: the row count still
    matches when a populated row is quietly relabelled as having no ground truth, which
    is what a mis-keyed join produces.
    """
    wb = load_workbook(book)
    wb["Per item"].cell(row=3, column=6, value=export_xlsx.NO_GT_MARKER)
    wb.save(book)
    problems = _verify(book, gt_path)
    assert any("ground-truth-backed rows" in p for p in problems)


def test_a_missing_testset_item_is_caught(book: Path, gt_path: Path):
    assert any("never mentions" in p for p in _verify(book, gt_path, ids=[*ITEM_IDS, "RET-99"]))


def test_a_non_item_in_the_id_column_is_caught(book: Path, gt_path: Path):
    wb = load_workbook(book)
    wb["Per item"].cell(row=3, column=1, value="total")
    wb.save(book)
    assert any("not an item id" in p for p in _verify(book, gt_path))


@pytest.mark.parametrize(
    "caveat",
    ["RECONCILED: NO", "PROMPT: RECONSTRUCTED", "AUDIO", "native speaker",
     "Agent-speech misattribution", "ASR error", "Diarisation", "REPRODUCIBILITY"],
)
def test_every_caveat_is_individually_required(book: Path, gt_path: Path, caveat: str):
    """A spreadsheet travels further than its caveats. Each one is a separate gate.

    Parametrised rather than asserted as a block so that dropping exactly one of them
    fails with the name of the one that went missing.
    """
    wb = load_workbook(book)
    stripped = CAVEAT_TEXT.replace(caveat, "").replace(caveat.lower(), "")
    wb["READ FIRST"]["A1"] = stripped
    wb.save(book)
    problems = _verify(book, gt_path)
    assert any(caveat in p and "missing" in p for p in problems)


def test_an_unfrozen_header_is_caught(book: Path, gt_path: Path):
    wb = load_workbook(book)
    wb["Runs"].freeze_panes = None
    wb.save(book)
    assert any("no frozen header" in p for p in _verify(book, gt_path))


def test_a_truncated_sheet_is_caught(book: Path, gt_path: Path):
    wb = load_workbook(book)
    ws = wb["Mechanisms"]
    wb.remove(ws)
    fresh = wb.create_sheet("Mechanisms")
    fresh.freeze_panes = "A2"
    fresh["A1"] = "only one row"
    wb.move_sheet("Mechanisms", offset=-1)
    wb.save(book)
    assert any("only 1 rows" in p for p in _verify(book, gt_path))
